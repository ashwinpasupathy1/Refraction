#!/usr/bin/env python3
"""
prism_barplot_app.py
====================
Claude Plotter -- main application window (macOS, tabbed ttk layout).

Module structure
----------------
prism_barplot_app.py  -- this file; App class + PLOT_REGISTRY + icon helpers
prism_widgets.py      -- design-system tokens, PButton/PEntry/PCheckbox etc.
prism_validators.py   -- standalone spreadsheet validation functions
prism_results.py      -- results-panel population, export, and copy helpers
plotter_functions.py    -- matplotlib plot functions (29 chart types)
prism_tabs.py         -- TabState, TabManager, TabBar (plot tab system)

The App class imports from all four companion modules so each can be
developed, tested, and documented independently.
"""
import collections, json, logging, math, os, threading, traceback, uuid

_log = logging.getLogger(__name__)

_pd_module = None
def _pd():
    """Return pandas, importing and caching it on first call."""
    global _pd_module
    if _pd_module is None:
        import pandas as _p
        _pd_module = _p
    return _pd_module

# Must be set before any matplotlib import to prevent Cocoa GUI init on macOS.
os.environ.setdefault("MPLBACKEND", "Agg")

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# -- Companion-module imports --------------------------------------------------
import sys as _sys
_HERE = os.path.dirname(os.path.abspath(__file__))
if _HERE not in _sys.path:
    _sys.path.insert(0, _HERE)

try:
    from plotter_widgets import (
        _DS, PButton, PCheckbox, PRadioGroup, PEntry, PCombobox,
        section_sep, _create_tooltip, add_placeholder, _bind_scroll_recursive,
        LABELS, HINTS, label, hint, tip,
        _is_num, _non_numeric_values, _scipy_summary, _sys_bg,
        PAD,
    )
except ImportError as _e:
    print(f"[prism] warning: prism_widgets not found ({_e})")

try:
    from plotter_validators import (
        validate_flat_header, validate_bar, validate_line,
        validate_grouped_bar, validate_kaplan_meier, validate_heatmap,
        validate_two_way_anova, validate_contingency, validate_chi_square_gof,
        validate_bland_altman, validate_forest_plot, validate_pyramid,
    )
    _VALIDATORS_AVAILABLE = True
except ImportError as _e:
    print(f"[prism] warning: prism_validators not found ({_e})")
    _VALIDATORS_AVAILABLE = False

try:
    from plotter_results import populate_results, export_results_csv, copy_results_tsv
    _RESULTS_AVAILABLE = True
except ImportError as _e:
    print(f"[prism] warning: prism_results not found ({_e})")
    _RESULTS_AVAILABLE = False

try:
    from plotter_app_icons import ICON_FN
    _ICON_FN = ICON_FN  # backward-compat alias; local icon functions removed
except ImportError as _e:
    print(f"[prism] warning: plotter_app_icons not found ({_e})")
    _ICON_FN = {}

# Heavy scientific imports are deferred to first use so the window
# appears immediately. They are loaded on a background thread via
# _import_functions() and also imported lazily inside each method
# that needs them (_do_run, _load_sheets, etc.).
# matplotlib and pandas are NOT imported here at module level.

# Try to use TkinterDnD for drag-and-drop support
try:
    from tkinterdnd2 import TkinterDnD, DND_FILES
    _DND_AVAILABLE = True
except ImportError:
    _DND_AVAILABLE = False
    DND_FILES      = None

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ICON_PNG   = os.path.join(SCRIPT_DIR, "assets", "AppIcon.png")
ICON_ICNS  = os.path.join(SCRIPT_DIR, "assets", "AppIcon.icns")

# Add the script directory to sys.path so plotter_functions is importable
import sys
if SCRIPT_DIR not in sys.path:
    sys.path.insert(0, SCRIPT_DIR)

def _load_icon_nsimage():
    """Return an NSImage loaded from bytes. The white border is baked into
    AppIcon.icns and AppIcon.png so no runtime compositing is needed.
    Tries .icns first (multi-resolution), then .png.
    """
    try:
        from AppKit import NSImage
        import Foundation
        for icon_path in (ICON_ICNS, ICON_PNG):
            if not os.path.exists(icon_path):
                continue
            with open(icon_path, "rb") as fh:
                raw = fh.read()
            data = Foundation.NSData.dataWithBytes_length_(raw, len(raw))
            img  = NSImage.alloc().initWithData_(data)
            if img and img.isValid():
                return img
    except Exception:
        _log.debug("_load_icon_nsimage: AppKit icon load failed", exc_info=True)
    return None


def set_dock_icon():
    """Set the macOS dock icon from bytes  -  immune to Cocoa path-resolution
    resets that happen when matplotlib/seaborn import NSApplication internally.

    Three-pronged approach:
      1. Load icon from raw bytes (not file path)  -  works inside bundles.
      2. Explicitly set NSApplicationActivationPolicyRegular so macOS treats
         the process as a proper GUI app.
      3. Called at startup, every 250 ms during heavy imports, and once more
         after all imports finish  -  see _watch_dock_icon.
    """
    try:
        from AppKit import NSApplication, NSApplicationActivationPolicyRegular
        app = NSApplication.sharedApplication()
        # Ensure the app is registered as a regular GUI application.
        # This prevents macOS from downgrading the dock entry after import.
        try:
            app.setActivationPolicy_(NSApplicationActivationPolicyRegular)
        except Exception:
            _log.debug("set_dock_icon: setActivationPolicy_ failed", exc_info=True)
        img = _load_icon_nsimage()
        if img:
            app.setApplicationIconImage_(img)
    except Exception:
        _log.debug("set_dock_icon: AppKit dock icon setup failed", exc_info=True)


def load_tk_icon():
    """Load the app icon as a Tkinter PhotoImage for wm_iconphoto().
    The white border is baked into AppIcon.png so no runtime compositing needed.
    """
    try:
        from PIL import Image, ImageTk
        img = Image.open(ICON_PNG).convert("RGBA")
        if max(img.size) > 256:
            img = img.resize((256, 256), Image.LANCZOS)
        return ImageTk.PhotoImage(img)
    except Exception:
        _log.debug("load_tk_icon: could not load Tk icon from %r", ICON_PNG, exc_info=True)
        return None


def _lbl(parent, row, key, font_size=13):
    """Create a bold section label with tooltip, grid it, return (label, next_row)."""
    l = ttk.Label(parent, text=label(key), font=("Helvetica Neue", font_size, "bold"))
    l.grid(row=row, column=0, sticky="w", padx=PAD, pady=(4, 2))
    tip(l, key)
    return l, row + 1


# ---------------------------------------------------------------------------
# Plot type registry  -  edit prism_registry.py to add new chart types
# ---------------------------------------------------------------------------

from plotter_registry import (
    PlotTypeConfig, _REGISTRY_SPECS,
    ERROR_TYPE_MAP, STATS_TEST_MAP, MARKER_STYLE_MAP, PAD,
)
from plotter_tabs import TabState, TabManager, TabBar

PLOT_REGISTRY: list = []  # populated after App class definition so fn_name can be resolved


# ---------------------------------------------------------------------------
# Factory defaults for every UI variable.
# _reset_vars_to_defaults() reads this to restore state on chart-type switch.
# Sheet changes and in-pane edits must NOT call that method.
# ---------------------------------------------------------------------------
_VAR_DEFAULTS: dict = {
    # file / sheet
    "excel_path": "", "sheet": "",
    # data tab
    "error": "SEM (Standard Error)", "show_points": True,
    "show_n_labels": False, "show_value_labels": False, "color": "Default",
    "title": "", "xlabel": "", "ytitle": "",
    "jitter_amount": "0", "error_below_bar": False,
    # axes tab
    "yscale": "Linear", "ylim_lo": "", "ylim_hi": "",
    "figw": "", "figh": "", "font_size": "12",
    "ref_line_y": "0", "ref_line_enabled": False,
    "ylim_data_min": False, "ylim_none": True,
    "ylim_mode": 0,   # 0=Auto, 1=Data min, 2=Manual
    "xlim_mode": 0,   # 0=Auto, 1=Manual
    "xlim_lo": "", "xlim_hi": "",
    "gridlines": False, "open_points": False, "grid_style": "None",
    "horizontal": False, "show_median": False,
    "bar_alpha": "0.85", "xscale": "Linear", "ref_line_label": "",
    "bar_width": "0.6", "line_width": "1.5",
    "marker_style": "Different Markers", "marker_size": "7",
    "notch_box": False,
    # stats tab (shared)
    "show_stats": False, "show_ns": False, "show_p_values": False,
    "show_effect_size": False, "show_test_name": False,
    "show_normality_warning": True,
    "p_sig_threshold": "0.05",
    "bracket_style": "Lines",
    "stacked_horizontal": False,
    "xtick_labels_str": "",
    "twin_y_series_str": "",
    "ref_vline_enabled": False,
    "ref_vline_x": "0",
    "ref_vline_label": "",
    "stats_test": "Parametric", "n_permutations": "",
    "mc_correction": "Holm-Bonferroni", "posthoc": "Tukey HSD",
    "control": "",
    # scatter-specific
    "show_regression": False, "show_ci_band": False,
    "show_prediction_band": False, "show_correlation": False,
    "correlation_type": "Pearson",
    # kaplan-meier
    "show_ci": True, "show_censors": True, "show_at_risk": False,
    # heatmap
    "annotate": False, "cluster_rows": False, "cluster_cols": False,
    "robust": False, "heatmap_fmt": "",
    "heatmap_vmin": "", "heatmap_vmax": "", "heatmap_center": "",
    # two-way anova
    "show_posthoc": False,
    # grouped bar
    "show_anova_per_group": False,
    # histogram
    "hist_bins": "0", "hist_density": False, "hist_overlay_normal": False,
    # curve fit
    "curve_model": "4PL Sigmoidal (EC50/IC50)",
    "cf_show_ci": True, "cf_show_residuals": False,
    "cf_show_equation": True, "cf_show_r2": True, "cf_show_params": True,
    # column stats
    "cs_show_normality": True, "cs_show_ci": True, "cs_show_cv": True,
    # contingency
    "ct_show_pct": True, "ct_show_expected": False,
    # repeated measures
    "rm_show_lines": True, "rm_test_type": "Parametric",
    # ── Priority-1 styling (axes tab) ────────────────────────────────────────
    "axis_style":  "Open (default)",
    "tick_dir":    "Outward (default)",
    "minor_ticks": False,
    "point_size":  "6",
    "point_alpha": "0.80",
    "cap_size":    "4",
    "legend_pos":  "Upper right",
    "spine_width": "0.8",
    # ── Priority-2 stats ──────────────────────────────────────────────────────
    "show_regression_table": False,
    "one_sample_mu0": "0",
    # chi-square GoF
    "gof_expected_equal": True,
    # stacked bar
    "stacked_mode": "absolute", "stacked_value_labels": False,
    # bubble chart
    "bubble_scale": "1.0", "bubble_show_labels": False,
    # dot plot
    "dp_show_mean": True, "dp_show_median": False,
    # bland-altman
    "ba_show_ci": True,
    # forest plot
    "fp_ref_value": "0", "fp_show_weights": True, "fp_show_summary": True,
    # tick intervals
    "ytick_interval": "", "xtick_interval": "",
    # figure background
    "fig_bg": "White",
}


PREFS_PATH = os.path.expanduser("~/Library/Preferences/claude_plotter.json")

def _load_prefs():
    try:
        with open(PREFS_PATH) as f:
            return json.load(f)
    except Exception:
        _log.debug("_load_prefs: could not load %r", PREFS_PATH, exc_info=True)
        return {}

def _save_prefs(data):
    try:
        os.makedirs(os.path.dirname(PREFS_PATH), exist_ok=True)
        with open(PREFS_PATH, "w") as f:
            json.dump(data, f, indent=2)
    except Exception:
        _log.debug("_save_prefs: could not write %r", PREFS_PATH, exc_info=True)

def _add_recent(path):
    prefs = _load_prefs()
    recent = prefs.get("recent_files", [])
    if path in recent:
        recent.remove(path)
    recent.insert(0, path)
    prefs["recent_files"] = recent[:10]
    prefs["last_file"] = path
    _save_prefs(prefs)


# Lightweight sheet/dataframe cache  -  avoids re-reading the same file twice
_SHEET_CACHE = {}   # path -> (mtime, sheetnames)
_DF_CACHE    = {}   # (path, sheet) -> (mtime, DataFrame)

def _cached_sheets(path):
    """Return sheet names, using cache if file unchanged."""
    try:
        mtime = os.path.getmtime(path)
        if path in _SHEET_CACHE and _SHEET_CACHE[path][0] == mtime:
            return _SHEET_CACHE[path][1]
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = wb.sheetnames; wb.close()
        _SHEET_CACHE[path] = (mtime, sheets)
        return sheets
    except Exception:
        _log.debug("_cached_sheets: could not read sheets from %r", path, exc_info=True)
        return []

def _cached_df(path, sheet):
    """Return DataFrame, using cache if file unchanged."""
    pd = _pd()
    try:
        mtime = os.path.getmtime(path)
        key   = (path, str(sheet))
        if key in _DF_CACHE and _DF_CACHE[key][0] == mtime:
            return _DF_CACHE[key][1]
        df = pd.read_excel(path, sheet_name=sheet, header=None)
        _DF_CACHE[key] = (mtime, df)
        return df
    except Exception as e:
        raise



_SB_ITEM_H    = 56   # height of each sidebar item in pixels
_SB_ICON_SIZE = 28   # icon canvas size
_SB_WIDTH     = 160  # total sidebar width


def _scipy_summary(fn, max_chars: int = 280) -> str:
    """Return the opening summary of a scipy function's docstring."""
    import inspect
    doc   = inspect.getdoc(fn) or ""
    paras = [p.strip() for p in doc.split("\n\n")]
    result = []
    for p in paras:
        if any(p.startswith(h) for h in
               ("Parameters", "Returns", "See Also", "Notes",
                "Warns", "References", "Examples", "Attributes")):
            break
        clean = " ".join(p.split())
        if clean:
            result.append(clean)
        if sum(len(r) for r in result) > max_chars:
            break
    return " ".join(result)


# ── widget classes imported from plotter_widgets (see top-of-file imports) ───


class App(TkinterDnD.Tk if _DND_AVAILABLE else tk.Tk):
    def __init__(self):
        super().__init__()
        # Hide window during build to avoid visible resize animation
        self.withdraw()
        self.title("Claude Plotter")
        self.resizable(True, True)
        self._pf              = None
        self._pf_ready        = False
        self._running         = False
        self._vars            = {}
        self._file_selected   = False
        self._validated       = False
        self._validation_lock = False  # True while form is locked due to validation failure
        self._status_fade_id  = None   # handle for auto-clear timer
        self._kw_history      = collections.deque(maxlen=10)  # undo stack
        self._results_tsv_data = ""
        self._popup_count = 0   # number of open popup dialogs
        self._tab_manager:         TabManager | None = None   # set in _build_right_pane
        self._switching_tabs       = False   # True while TabManager.switch_to restores vars
        self._preview_after_id    = None   # pending after() id for live preview debounce
        self._live_preview_enabled = True  # can be toggled by user preference
        # --- Wave 2 infrastructure ---
        from plotter_events import EventBus
        self._bus = EventBus()
        from plotter_undo import UndoStack
        self._undo_stack = UndoStack(max_depth=50)
        from plotter_errors import reporter
        reporter.set_root(self)
        # --- end Wave 2 infrastructure ---
        ttk.Style().theme_use("aqua")
        set_dock_icon()
        self._set_tk_icon()
        self._build()
        self.after(200, self._build_lockable_cache)  # build after window renders
        self._setup_dnd()  # register drag-and-drop targets
        # Start locked; unlock happens in _load_sheets after first file pick
        self._run_btn.config(state="disabled", text="Generate Plot")
        self._lock_form()
        self._import_functions()
        # Keep dock icon alive during background import (Cocoa resets it)
        self.after(250, self._watch_dock_icon)
        # Wire live-preview traces after the UI is fully built
        self.after(300, self._setup_live_preview)
        # Session persistence: auto-save every 30 s, offer restore on launch
        try:
            from plotter_session import Session
            self._session = Session()
            self.after(500, self._session_restore_prompt)
        except Exception:
            _log.debug("App.__init__: plotter_session import or init failed", exc_info=True)
            self._session = None
        self.protocol("WM_DELETE_WINDOW", self._on_quit)
        self.after(30000, self._auto_save)

        # Phase 3: start FastAPI server for Plotly rendering
        try:
            from plotter_server import start_server
            start_server(app_instance=self)
            self._web_server_running = True
        except Exception:
            _log.debug("App.__init__: plotter_server start failed", exc_info=True)
            self._web_server_running = False

        # Phase 3: web view instances per tab
        self._plotly_views: dict = {}
        self._use_webview = tk.BooleanVar(value=True)

    def _auto_save(self):
        """Save session state to disk every 30 seconds."""
        try:
            if self._session is not None:
                pt = self._plot_type.get() if hasattr(self, "_plot_type") else "bar"
                state = self._session.capture(self._vars, pt, self.geometry())
                self._session.save_to_disk(state)
        except Exception:
            _log.debug("App._auto_save: session auto-save failed", exc_info=True)
        self.after(30000, self._auto_save)

    def _session_restore_prompt(self):
        """Offer to restore the previous session on startup."""
        try:
            if self._session is None:
                return
            saved = self._session.load_from_disk()
            if not saved or not saved.get("vars"):
                return
            from tkinter import messagebox
            if messagebox.askyesno("Restore Session",
                                   "Restore your previous session?"):
                self._session.restore(saved, self._vars)
        except Exception:
            _log.debug("App._session_restore_prompt: session restore failed", exc_info=True)

    def _on_quit(self):
        """Save session then destroy the window."""
        try:
            if self._session is not None:
                pt = self._plot_type.get() if hasattr(self, "_plot_type") else "bar"
                state = self._session.capture(self._vars, pt, self.geometry())
                self._session.save_to_disk(state)
        except Exception:
            _log.debug("App._on_quit: session save on quit failed", exc_info=True)
        self.destroy()

    def _set_tk_icon(self):
        photo = load_tk_icon()
        if photo:
            self._icon_photo = photo
            self.iconphoto(True, photo)

    def _watch_dock_icon(self):
        """Re-apply the dock icon on a timer while heavy imports run, then
        once more after they finish, and again 1 s later as a safety net.

        Why multiple restores?
          • matplotlib.pyplot import touches NSApplication  ->  resets icon
          • seaborn import does the same
          • openpyxl's first read can trigger a secondary Cocoa event loop tick

        The watcher fires every 200 ms (was 250) so it catches any reset
        within one animation frame.
        """
        set_dock_icon()
        if not self._pf_ready:
            self.after(200, self._watch_dock_icon)
        else:
            # Imports just finished  -  do one more restore after a short delay
            # to catch any residual Cocoa activity from seaborn/openpyxl.
            self.after(800, set_dock_icon)

    def _setup_dnd(self):
        """Register file-path entry fields as drag-and-drop targets."""
        if not _DND_AVAILABLE:
            return
        def _on_drop(event):
            # macOS wraps paths in braces if they contain spaces; strip them
            raw = event.data.strip()
            # Handle multiple files  -  take only the first
            if raw.startswith("{"):
                raw = raw[1:raw.index("}")] if "}" in raw else raw[1:]
            path = raw.strip()
            if path.lower().endswith((".xlsx", ".xls")):
                self._vars["excel_path"].set(path)
                self._load_sheets(path)
            else:
                self._set_status("Drop an .xlsx or .xls file", err=True)

        # Register every file-path entry field
        for widget in getattr(self, "_drop_targets", []):
            try:
                widget.drop_target_register(DND_FILES)
                widget.dnd_bind("<<Drop>>", _on_drop)
            except Exception:
                _log.debug("App._setup_dnd: DND binding failed for widget %r", widget, exc_info=True)

    def _import_functions(self):
        threading.Thread(target=self._do_import, daemon=True).start()

    def _do_import(self):
        try:
            import matplotlib
            matplotlib.use("Agg")   # belt-and-suspenders alongside MPLBACKEND env var
            import plotter_functions as pf
            self._pf = pf
            # ── _ensure_imports() loads matplotlib.pyplot + seaborn ──────────
            # These must complete BEFORE _pf_ready flips so that _watch_dock_icon
            # is still running if Cocoa resets the icon during seaborn init.
            pf._ensure_imports()
            self._plt = pf.plt
            self._pd  = pf.pd
            # All heavy imports done  -  safe to mark ready and stop the watcher
            self._pf_ready = True
            # One final dock-icon restore after everything settles
            self.after(0, set_dock_icon)
            # Enable Generate Plot if a file was already validated while loading
            def _check_ready():
                if self._validated:
                    self._run_btn.config(state="normal", text="Generate Plot")
                else:
                    self._run_btn.config(state="disabled", text="Generate Plot")
                # Swatches skipped their custom-palette render during import
                # (to avoid the deadlock).  Now matplotlib is loaded, refresh them.
                self._vars["color"].set(self._vars["color"].get())
                self._sync_analyze_btn()
            self.after(0, _check_ready)
        except Exception as e:
            m = str(e)
            self.after(0, lambda: self._set_status(f"Load error: {m}", err=True))

    # ── Form locking ──────────────────────────────────────────────────────────

    def _reset_vars_to_defaults(self):
        """Reset all UI variables to factory defaults.
        Called ONLY on chart-type switch  -  never on sheet change or in-pane edits."""
        for key, default in _VAR_DEFAULTS.items():
            if key in self._vars:
                try:
                    self._vars[key].set(default)
                except Exception:
                    pass
        # Clear the control group dropdown values so stale group names from a
        # previous chart type can't crash the next plot run.
        for cb in getattr(self, "_control_cbs", []):
            try:
                cb.config(values=[], state="disabled")
            except Exception:
                pass
        if "control" in self._vars:
            try: self._vars["control"].set("")
            except Exception: pass
        if "comparison_mode" in self._vars:
            try: self._vars["comparison_mode"].set(0)
            except Exception: pass
        for lbl in getattr(self, "_control_hint_lbls", []):
            try:
                lbl.config(text="Load a file to populate group names",
                           foreground="gray")
            except Exception:
                pass
        # Re-fire toggles that depend on var state
        if hasattr(self, "_yl_lo"):
            self._tog_ylim()
        if hasattr(self, "_xl_lo"):
            self._tog_xlim()

    def _reset_chart_type_state(self):
        """Called after a chart-type change (or new tab).
        Resets all UI variables to factory defaults, clears the cached spreadsheet,
        and locks the form until the user selects a new file for this chart type."""
        # Reset every UI variable to its factory default first
        self._reset_vars_to_defaults()
        # Clear file selection and validation state
        self._vars["excel_path"].set("")
        self._vars["sheet"].set("")
        for cb in getattr(self, "_sheet_cbs", []):
            cb.config(values=[], state="disabled")
        for lbl in getattr(self, "_sheet_hints", []):
            lbl.config(text="Browse for an Excel file to populate sheets")
        self._file_selected = False
        self._validated = False
        self._run_btn.config(state="disabled", text="Generate Plot")
        self._set_validate_text("", "gray")
        # Lock everything except the file browser so user must upload first
        if hasattr(self, "_file_row_frames"):
            self._lock_form()
        # Disable plot-output buttons
        for attr in ("_copy_btn", "_zoom_reset_btn", "_zoom_in_btn",
                     "_zoom_out_btn", "_export_btn", "_copy_svg_btn"):
            if hasattr(self, attr):
                getattr(self, attr).config(state="disabled")

    def _sb_select_silent(self, idx: int) -> None:
        """Update sidebar highlight and form pane without triggering a chart-type reset.

        Used by TabManager.switch_to so that switching tabs does not wipe the
        restored form state.
        """
        if hasattr(self, "_sb_select"):
            self._sb_select(idx)
        if hasattr(self, "_sb_show_pane"):
            self._sb_show_pane(idx)
        # Explicitly does NOT call _on_chart_type_change / _reset_chart_type_state

    def _sync_active_tab_label(self) -> None:
        """Keep the active tab's label in sync with the chart title field."""
        if getattr(self, "_switching_tabs", False):
            return
        if self._tab_manager is None:
            return
        tab = self._tab_manager.active
        if tab is None:
            return
        raw   = self._vars["title"].get().strip()
        label = raw if raw else "Untitled"
        tab.label = label
        self._tab_manager.update_label(tab.tab_id, label)

    def _lock_form(self):
        self._set_form_state("disabled")

    def _unlock_form(self):
        self._set_form_state("!disabled")
        self._validation_lock = False
        # _set_form_state re-enables every lockable widget unconditionally  - 
        # re-apply limit toggles to restore the correct enabled/disabled state.
        if hasattr(self, "_yl_lo"):
            self._tog_ylim()
        if hasattr(self, "_xl_lo"):
            self._tog_xlim()
        self._sync_analyze_btn()

    def _sync_analyze_btn(self):
        """Enable/disable the Help Analyze button based on validation state."""
        try:
            state = "normal" if (self._validated and self._pf_ready) else "disabled"
            self._help_analyze_btn.config(state=state)
        except Exception:
            pass

    def _build_lockable_cache(self):
        """Build a flat list of widgets that should be locked/unlocked.
        Called once after _build() and again after each lazy tab build."""
        skip_cls  = {"TLabel", "TSeparator", "TFrame", "Frame",
                     "Canvas", "Scrollbar", "TScrollbar",
                     "Listbox"}   # sidebar listbox must never be locked

        # Everything in the file row frames stays on (Browse button + path entry)
        always_on = set(self._file_row_frames)
        # All children of file row frames stay on too
        for frf in self._file_row_frames:
            try:
                for child in frf.winfo_children():
                    always_on.add(child)
            except Exception:
                pass
        # Sheet comboboxes must always stay enabled  -  user must be able to
        # switch sheets even after a validation failure on the current sheet
        for cb in getattr(self, "_sheet_cbs", []):
            always_on.add(cb)
        # The sidebar items and their parent are always on
        if hasattr(self, "_sb_sidebar_outer"):
            always_on.add(self._sb_sidebar_outer)
            try:
                def _walk_always(w):
                    always_on.add(w)
                    for ch in w.winfo_children():
                        _walk_always(ch)
                _walk_always(self._sb_sidebar_outer)
            except Exception:
                pass

        lockable = []

        def _walk(widget):
            cls = widget.winfo_class()
            if widget in always_on: return
            if cls == "Listbox": return   # never recurse into Listbox

            # Design-system custom widgets  -  lock via their .config(state=...)
            if getattr(widget, '_is_pwidget', False):
                if not getattr(widget, '_lock_exempt', False):
                    lockable.append(widget)
                return   # don't recurse into internals

            if cls in skip_cls:
                # recurse into container frames but don't add them
                for child in widget.winfo_children():
                    _walk(child)
                return

            if cls == "TRadiobutton": return
            if cls == "TButton":
                try:
                    txt = str(widget.cget("text"))
                except Exception:
                    txt = ""
                if "Browse" in txt or "Template" in txt:
                    return
            if cls not in skip_cls:
                lockable.append(widget)
            for child in widget.winfo_children():
                _walk(child)

        children = self.winfo_children()
        for child in children:
            if isinstance(child, ttk.Frame) and child is children[-1]:
                continue
            _walk(child)

        self._lockable_widgets = lockable

    def _set_form_state(self, state):
        if not hasattr(self, "_lockable_widgets"):
            self._build_lockable_cache()
        normal = (state == "!disabled")
        for widget in self._lockable_widgets:
            try:
                if getattr(widget, '_is_pwidget', False):
                    widget.config(state="normal" if normal else "disabled")
                else:
                    widget.state([state])
            except Exception:
                try:
                    widget.config(state="normal" if normal else "disabled")
                except Exception:
                    pass

    # ── Build ─────────────────────────────────────────────────────────────────

    def _build(self):
        # Build the native menu bar first (File / Edit / View / Window)
        self._build_menubar()

        # ── PanedWindow: sash disabled so user cannot accidentally drag it ────
        # We set sashwidth=0 and bind <Button-1> on the sash to absorb clicks.
        # The plot is always sized to fill the right pane via _embed_plot.
        paned = tk.PanedWindow(self, orient="horizontal",
                               sashwidth=0, sashrelief="flat", bg="#d1d1d6")
        self._paned = paned
        self._build_toolbar()
        paned.pack(fill="both", expand=True)
        self._left_pane = left = ttk.Frame(paned)
        paned.add(left, minsize=560)
        self._plot_type = tk.StringVar(value="bar")
        self._build_sidebar(left)
        self._build_right_pane(paned)

        # ── Suppress the spurious resize event at startup ─────────────────────
        # Tkinter fires a <Configure> event the moment the window is shown.
        # We gate _invalidate_pane_cache with a flag so the first event is
        # a no-op, preventing a flash of mismatched geometry.
        self._startup_done = False

        self.update_idletasks()
        self.geometry("1280x820")
        try:
            self.attributes("-zoomed", True)   # macOS / Linux
        except Exception:
            try:
                self.state("zoomed")           # Windows
            except Exception:
                pass
        self.deiconify()
        self._startup_done = True

        self.bind("<Command-r>", lambda e: self._run())
        self.bind("<Command-z>", lambda e: self._undo())
        self.bind("<Command-Z>", lambda e: self._do_redo())
        self.bind("<Command-s>", lambda e: self._save_project())
        self.bind("<Command-e>", lambda e: self._download_png())
        self.bind("<Command-o>", lambda e: self._browse_file())
        self.bind("<Command-c>", lambda e: self._copy_to_clipboard())
        self.bind("<Command-equal>", lambda e: self._zoom_plot_factor(1.1))
        self.bind("<Command-plus>",  lambda e: self._zoom_plot_factor(1.1))
        self.bind("<Command-minus>", lambda e: self._zoom_plot_factor(0.9))
        self.bind("<Command-0>",     lambda e: self._reset_zoom())

        # Set sash to 50/50 after the window settles  -  not during startup
        def _set_sash():
            w = self.winfo_width()
            try:
                paned.sash_place(0, w // 2, 0)
            except Exception:
                pass
        self.after(200, _set_sash)

    def _build_menubar(self):
        """Build the native macOS / cross-platform menu bar.

        Implements File, Edit, View, Window menus with reasonable actions.
        Falls back gracefully on platforms where some items are not applicable.
        """
        menubar = tk.Menu(self)

        # ── File menu ────────────────────────────────────────────────────────
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="Open Excel File…",
                              accelerator="⌘O",
                              command=self._browse_file)
        file_menu.add_command(label="Open .cplot Project…",
                              command=self._open_cplot)
        file_menu.add_command(label="Import .pzfx File…",
                              command=self._open_pzfx)
        file_menu.add_separator()
        file_menu.add_command(label="Save Project…",
                              accelerator="⌘S",
                              command=self._save_project)
        file_menu.add_separator()
        file_menu.add_command(label="Generate Plot",
                              accelerator="⌘R",
                              command=self._run)
        file_menu.add_command(label="Undo Last Plot",
                              accelerator="⌘Z",
                              command=self._undo)
        file_menu.add_separator()
        file_menu.add_command(label="Export Plot…",
                              accelerator="⌘E",
                              command=self._download_png)
        file_menu.add_command(label="Export All Charts (PDF Report)…",
                              command=self._export_all_pdf)
        file_menu.add_separator()
        file_menu.add_command(label="Copy PNG",
                              accelerator="⌘C",
                              command=self._copy_to_clipboard)
        file_menu.add_command(label="Copy Transparent PNG",
                              command=self._copy_transparent)
        file_menu.add_command(label="Copy SVG",
                              command=self._copy_as_svg)
        menubar.add_cascade(label="File", menu=file_menu)

        # ── Edit menu ─────────────────────────────────────────────────────────
        edit_menu = tk.Menu(menubar, tearoff=0)
        edit_menu.add_command(label="Select All",
                              accelerator="⌘A",
                              command=lambda: self.focus_get() and
                              self.focus_get().event_generate("<<SelectAll>>"))
        edit_menu.add_separator()
        edit_menu.add_command(label="Reset to Defaults",
                              command=self._reset_vars_to_defaults)
        menubar.add_cascade(label="Edit", menu=edit_menu)

        # ── View menu ─────────────────────────────────────────────────────────
        view_menu = tk.Menu(menubar, tearoff=0)
        view_menu.add_command(label="Zoom In",
                              accelerator="⌘+",
                              command=lambda: self._zoom_plot_factor(1.1))
        view_menu.add_command(label="Zoom Out",
                              accelerator="⌘−",
                              command=lambda: self._zoom_plot_factor(0.9))
        view_menu.add_command(label="Reset Zoom",
                              accelerator="⌘0",
                              command=self._reset_zoom)
        view_menu.add_separator()
        view_menu.add_checkbutton(label="Use Web Renderer (Plotly)",
                                  variable=self._use_webview)
        view_menu.add_separator()
        view_menu.add_command(label="Reference Wiki",
                              command=self._open_wiki_popup)
        menubar.add_cascade(label="View", menu=view_menu)

        # ── Window menu ───────────────────────────────────────────────────────
        window_menu = tk.Menu(menubar, tearoff=0)
        window_menu.add_command(label="Minimize",
                                accelerator="⌘M",
                                command=self.iconify)
        window_menu.add_command(label="Bring All to Front",
                                command=self.lift)
        menubar.add_cascade(label="Window", menu=window_menu)

        # ── Help menu ────────────────────────────────────────────────────────
        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="Help Analyze…",
                              command=self._help_analyze)
        help_menu.add_separator()
        help_menu.add_command(label="About Claude Plotter",
                              command=self._show_about)
        menubar.add_cascade(label="Help", menu=help_menu)

        self.config(menu=menubar)

        # ── Chart-type keyboard shortcuts: ⌘1–⌘9 ─────────────────────────────
        try:
            from plotter_registry import KEYBOARD_SHORTCUTS, _REGISTRY_SPECS
            _key_to_idx = {spec.key: i for i, spec in enumerate(_REGISTRY_SPECS)}
            for num, chart_key in KEYBOARD_SHORTCUTS.items():
                idx = _key_to_idx.get(chart_key)
                if idx is not None:
                    self.bind_all(f"<Command-Key-{num}>",
                                  lambda e, i=idx: self._jump_to_chart(i))
                    self.bind_all(f"<Control-Key-{num}>",
                                  lambda e, i=idx: self._jump_to_chart(i))
        except Exception:
            _log.debug("App._build_menubar: keyboard shortcut binding failed", exc_info=True)

    def _jump_to_chart(self, idx: int):
        """Switch to chart type by registry index (used by ⌘1–⌘9)."""
        try:
            if hasattr(self, "_sb_select"):
                self._sb_select(idx)
            if hasattr(self, "_sb_show_pane"):
                self._sb_show_pane(idx)
        except Exception:
            _log.debug("App._jump_to_chart: failed to jump to chart index %d", idx, exc_info=True)

    def _build_toolbar(self):
        """Build the bottom status/button bar.

        Changes from v6:
          • Copy PNG and Copy (α) consolidated into one "Copy" button with a
            smaller "Copy α" ghost button for the transparent variant.
          • Keyboard hint strip shortened and less cluttered.
          • MIT license credit added bottom-left in gray.
        """
        ttk.Separator(self).pack(side="bottom", fill="x")
        bot = ttk.Frame(self)
        bot.pack(side="bottom", fill="x", padx=10, pady=5)

        # ── Left side: status + wiki + MIT note ───────────────────────────────
        left_grp = ttk.Frame(bot)
        left_grp.pack(side="left", fill="y")

        self._status_var = tk.StringVar(value="")
        self._status_lbl = ttk.Label(left_grp, textvariable=self._status_var,
                                     font=("Helvetica Neue", 11))
        self._status_lbl.pack(side="top", anchor="w")

        self._wiki_btn = PButton(left_grp, text="📖 Reference Wiki", style="ghost",
                                 command=self._open_wiki_popup)
        self._wiki_btn._lock_exempt = True
        self._wiki_btn.pack(side="top", anchor="w", pady=(1, 0))

        # MIT license note
        ttk.Label(left_grp,
                  text="MIT License · Designed & implemented by Claude (Anthropic) · Ashwin Pasupathy",
                  foreground="#cccccc", font=("Helvetica Neue", 9)
                  ).pack(side="top", anchor="w")

        # Compact keyboard shortcut strip
        ttk.Label(bot,
                  text="⌘R Generate  ·  ⌘Z Undo  ·  ⌘C Copy  ·  ⌘E Export  ·  ⌘O Open  ·  ⌘+/− Zoom",
                  foreground="#cccccc", font=("Helvetica Neue", 9)
                  ).pack(side="left", padx=(12, 0))

        # ── Right side: action buttons ────────────────────────────────────────
        self._export_btn = PButton(bot, text="Export…", style="secondary",
                                   command=self._download_png, state="disabled")
        self._export_btn.pack(side="right", padx=(4, 0))
        _create_tooltip(self._export_btn, "Export PNG / TIFF / EPS / SVG / PDF  (⌘E)")

        # Single consolidated "Copy" button
        self._copy_btn = PButton(bot, text="Copy", style="secondary",
                                 command=self._copy_to_clipboard, state="disabled")
        self._copy_btn.pack(side="right", padx=(4, 0))
        _create_tooltip(self._copy_btn, "Copy PNG to clipboard  (⌘C)")

        self._copy_transparent_btn = PButton(bot, text="Copy α", style="ghost",
                                             command=self._copy_transparent, state="disabled")
        self._copy_transparent_btn.pack(side="right", padx=(1, 0))
        _create_tooltip(self._copy_transparent_btn, "Copy with transparent background")

        self._copy_svg_btn = PButton(bot, text="SVG", style="ghost",
                                     command=self._copy_as_svg, state="disabled")
        self._copy_svg_btn.pack(side="right", padx=(1, 0))
        _create_tooltip(self._copy_svg_btn, "Copy as SVG")

        # Zoom controls
        self._zoom_reset_btn = PButton(bot, text="1:1", style="ghost",
                                       command=self._reset_zoom, state="disabled", width=3)
        self._zoom_reset_btn.pack(side="right", padx=(2, 0))
        self._zoom_out_btn = PButton(bot, text="−", style="ghost",
                                     command=lambda: self._zoom_plot_factor(0.9),
                                     state="disabled", width=2)
        self._zoom_out_btn.pack(side="right", padx=(1, 0))
        self._zoom_in_btn = PButton(bot, text="+", style="ghost",
                                    command=lambda: self._zoom_plot_factor(1.1),
                                    state="disabled", width=2)
        self._zoom_in_btn.pack(side="right", padx=(4, 0))

        self._run_btn = PButton(bot, text="Generate Plot", style="primary", command=self._run)
        self._run_btn.pack(side="right", padx=(6, 0))
        _create_tooltip(self._run_btn, "Generate Plot  (⌘R)")

        self._undo_btn = PButton(bot, text="Undo", style="ghost",
                                 command=self._undo, state="disabled")
        self._undo_btn.pack(side="right", padx=(2, 0))
        _create_tooltip(self._undo_btn, "Undo last plot  (⌘Z)")

        self._help_analyze_btn = PButton(bot, text="✦ Help Analyze",
                                         style="secondary", command=self._help_analyze,
                                         state="disabled")
        self._help_analyze_btn._lock_exempt = True   # managed separately via _sync_analyze_btn
        self._help_analyze_btn.pack(side="right", padx=(4, 8))

    def _build_sidebar(self, left):
        """Build the chart-type icon sidebar and the scrollable content panes."""
        SB_BG       = "#f0f0f0"
        SB_SEL_BG   = "#2274A5"
        SB_SEL_FG   = "white"
        SB_NORM_FG  = "#333333"
        SB_HOVER_BG = "#dde8f5"

        sidebar_outer = tk.Frame(left, width=_SB_WIDTH, bg=SB_BG)
        sidebar_outer.pack(side="left", fill="y")
        sidebar_outer.pack_propagate(False)
        self._sb_sidebar_outer = sidebar_outer

        hdr_lbl = tk.Label(sidebar_outer, text="Chart Type",
                           font=("Helvetica Neue", 11, "bold"),
                           fg="#555555", bg=SB_BG, anchor="w")
        hdr_lbl.pack(fill="x", padx=10, pady=(10, 4))
        tk.Frame(sidebar_outer, bg="#cccccc", height=1).pack(fill="x", padx=6)

        _sb_vsb    = ttk.Scrollbar(sidebar_outer, orient="vertical")
        _sb_vsb.pack(side="right", fill="y")
        _sb_canvas = tk.Canvas(sidebar_outer, bg=SB_BG, highlightthickness=0,
                               yscrollcommand=_sb_vsb.set, width=_SB_WIDTH)
        _sb_canvas.pack(side="left", fill="both", expand=True)
        _sb_vsb.config(command=_sb_canvas.yview)
        _sb_inner = tk.Frame(_sb_canvas, bg=SB_BG)
        _sb_win   = _sb_canvas.create_window((0, 0), window=_sb_inner, anchor="nw")
        _sb_inner.bind("<Configure>",
            lambda e: _sb_canvas.configure(scrollregion=_sb_canvas.bbox("all")))
        _sb_canvas.bind("<Configure>",
            lambda e: _sb_canvas.itemconfig(_sb_win, width=e.width))

        _sb_items    = []
        _sb_selected = [0]

        def _sb_paint(idx, selected):
            if idx >= len(_sb_items): return
            row, ic, lbl = _sb_items[idx]
            key = _SB_KEYS[idx]
            bg  = SB_SEL_BG if selected else SB_BG
            fg  = SB_SEL_FG if selected else SB_NORM_FG
            row.config(bg=bg); ic.config(bg=bg); lbl.config(bg=bg, fg=fg)
            ic.delete("all")
            _ICON_FN.get(key, _icon_bar)(ic, SB_SEL_FG if selected else "#2274A5", bg)

        def _sb_select(idx):
            old = _sb_selected[0]
            _sb_selected[0] = idx
            _sb_paint(old, False)
            _sb_paint(idx, True)

        def _sb_wheel(e):
            raw = e.delta if e.delta else 0
            if not raw: return
            bbox = _sb_canvas.bbox("all")
            if not bbox: return
            content_h = bbox[3] - bbox[1]
            view_h    = _sb_canvas.winfo_height()
            if content_h <= view_h: return
            import sys as _sys_sb
            _scale_sb = 8.0 if _sys_sb.platform == "darwin" and abs(raw) < 40 else 4.0
            frac = -raw * _scale_sb / content_h
            cur  = _sb_canvas.yview()[0]
            _sb_canvas.yview_moveto(max(0.0, min(1.0 - view_h/content_h, cur + frac)))
        _sb_canvas.bind("<MouseWheel>", _sb_wheel)
        _sb_inner.bind("<MouseWheel>",  _sb_wheel)

        _SB_KEYS = []

        def _sb_add_item(key, label_text, idx):
            _SB_KEYS.append(key)
            row = tk.Frame(_sb_inner, bg=SB_BG, cursor="hand2", height=_SB_ITEM_H)
            row.pack(fill="x", pady=0)
            row.pack_propagate(False)
            ic = tk.Canvas(row, width=_SB_ICON_SIZE, height=_SB_ICON_SIZE,
                           bg=SB_BG, highlightthickness=0)
            ic.place(x=8, y=(_SB_ITEM_H - _SB_ICON_SIZE) // 2)
            _ICON_FN.get(key, _icon_bar)(ic, "#2274A5", SB_BG)
            lbl = tk.Label(row, text=label_text, font=("Helvetica Neue", 11),
                           fg=SB_NORM_FG, bg=SB_BG, anchor="w",
                           justify="left", wraplength=_SB_WIDTH - _SB_ICON_SIZE - 22)
            lbl.place(x=_SB_ICON_SIZE + 14, y=0,
                      width=_SB_WIDTH - _SB_ICON_SIZE - 20, height=_SB_ITEM_H)

            def _enter(e, i=idx):
                if _sb_selected[0] != i:
                    for w in (_sb_items[i][0], _sb_items[i][1], _sb_items[i][2]):
                        w.config(bg=SB_HOVER_BG)
            def _leave(e, i=idx):
                if _sb_selected[0] != i:
                    for w in (_sb_items[i][0], _sb_items[i][1], _sb_items[i][2]):
                        w.config(bg=SB_BG)
            def _click(e, i=idx):
                _sb_select(i)
                _show_pane(i)
                _on_tab_change()

            for w in (row, ic, lbl):
                w.bind("<Enter>",    _enter)
                w.bind("<Leave>",    _leave)
                w.bind("<Button-1>", _click)
                w.bind("<MouseWheel>", _sb_wheel)

            tk.Frame(_sb_inner, bg="#e0e0e0", height=1).pack(fill="x")
            _sb_items.append((row, ic, lbl))

        class _SbShim:
            def curselection(self):     return (_sb_selected[0],)
            def selection_set(self, i): _sb_select(i)
            def activate(self, i):      pass
            def insert(self, pos, txt): pass
            @property
            def master(self):           return sidebar_outer

        self._chart_lb = _SbShim()

        ttk.Separator(left, orient="vertical").pack(side="left", fill="y")

        # Outer wrapper holds the navigator strip + scrollable content
        _content_outer = ttk.Frame(left)
        _content_outer.pack(side="left", fill="both", expand=True)

        # ── Section jump navigator ──────────────────────────────────────────
        _nav_bar = tk.Frame(_content_outer, bg="#f0f4fa", height=30)
        _nav_bar.pack(side="top", fill="x")
        _nav_bar.pack_propagate(False)
        tk.Label(_nav_bar, text="Jump to:", bg="#f0f4fa", fg="#666666",
                 font=("Helvetica Neue", 10)).pack(side="left", padx=(10, 6))
        self._nav_btns = {}
        for _sec_name, _sec_icon in (("Data", "📊"), ("Axes", "📐"), ("Statistics", "📈")):
            _b = tk.Label(_nav_bar, text=f"{_sec_icon} {_sec_name}",
                          bg="#f0f4fa", fg="#2274A5",
                          font=("Helvetica Neue", 10, "bold"),
                          cursor="hand2", padx=8, pady=4)
            _b.pack(side="left", padx=2)
            _b.bind("<Enter>", lambda e, b=_b: b.config(bg="#dde8f5"))
            _b.bind("<Leave>", lambda e, b=_b: b.config(bg="#f0f4fa"))
            self._nav_btns[_sec_name] = _b

        self._section_anchors = {}  # key -> {section_name: y_fraction}

        def _jump_to_section(section_name):
            idx = _current_idx[0]
            chart_key = _tab_map.get(idx, "")
            anchors = self._section_anchors.get(chart_key, {})
            canvas = self._tab_canvases.get((idx, 0))
            if canvas and section_name in anchors:
                canvas.yview_moveto(anchors[section_name])

        for _sec_name in ("Data", "Axes", "Statistics"):
            self._nav_btns[_sec_name].bind(
                "<Button-1>",
                lambda e, s=_sec_name: _jump_to_section(s)
            )

        content_frame = ttk.Frame(_content_outer)
        content_frame.pack(side="top", fill="both", expand=True)

        _panes       = {}
        _current_idx = [0]

        def _show_pane(idx):
            for p in _panes.values(): p.place_forget()
            if idx in _panes:
                _panes[idx].place(relx=0, rely=0, relwidth=1, relheight=1)
            _current_idx[0] = idx

        class _FakePlotNB:
            def select(self, idx=None):
                if idx is not None: _show_pane(idx)
            def index(self, what=None): return _current_idx[0]
            def bind(self, seq, fn): pass

        self._plot_nb = _FakePlotNB()

        def _make_scrollable(parent):
            vsb    = ttk.Scrollbar(parent, orient="vertical")
            vsb.pack(side="right", fill="y")
            canvas = tk.Canvas(parent, highlightthickness=0, yscrollcommand=vsb.set)
            canvas.pack(side="left", fill="both", expand=True)
            vsb.config(command=canvas.yview)
            inner = ttk.Frame(canvas)
            win   = canvas.create_window((0, 0), window=inner, anchor="nw")
            inner.bind("<Configure>",
                lambda e, c=canvas: c.configure(scrollregion=c.bbox("all")))
            canvas.bind("<Configure>",
                lambda e, c=canvas, w=win: c.itemconfig(w, width=e.width))
            return inner, canvas

        self._tab_canvases = {}

        def _add_plot_type(nb_attr, label_text, key, mode, main_idx):
            pane = ttk.Frame(content_frame)
            _panes[main_idx] = pane
            inner, canvas = _make_scrollable(pane)
            for sub in range(3):
                self._tab_canvases[(main_idx, sub)] = canvas
            class _StubNB:
                def select(self, i=None): pass
                def index(self, w=None):  return 0
            setattr(self, nb_attr, _StubNB())
            tabs = [(inner, canvas)] * 3
            _sb_add_item(key, label_text, main_idx)
            return tabs

        _all_tabs  = []
        _tab_inner = {}
        for spec in _REGISTRY_SPECS:
            tabs = _add_plot_type(f"_{spec.key.replace('-','_')}_nb",
                                  spec.label, spec.key, spec.tab_mode, len(_all_tabs))
            _all_tabs.append(tabs)
            _tab_inner[spec.key] = tabs

        _tab_map  = {i: spec.key for i, spec in enumerate(_REGISTRY_SPECS)}

        # Wiki is no longer a sidebar tab  -  it lives in a bottom-bar popup button.
        # Keep wiki_pane as a stub so old code paths don't crash.
        wiki_pane = ttk.Frame(content_frame)

        _show_pane(0)
        _sb_select(0)

        # Expose closures so _sb_select_silent can call them directly
        self._sb_select    = _sb_select
        self._sb_show_pane = _show_pane

        def _on_chart_type_change(e=None, from_tab_switch=False):
            idx = _current_idx[0]
            key = _tab_map.get(idx, "bar")
            self._plot_type.set(key)
            if key in self._all_tabs_map:
                self._build_tab_content(key)
            if not from_tab_switch:
                self.after(0, self._reset_chart_type_state)

        # Keep backward-compat alias used in a few other places
        _on_tab_change = _on_chart_type_change

        _pane_cache = {}
        def _invalidate_pane_cache(e=None):
            # Skip the spurious resize that fires during window creation
            if not getattr(self, "_startup_done", False):
                return
            _pane_cache.clear()
        self.bind("<Configure>", _invalidate_pane_cache)

        def _smart_wheel(e):
            """Route scroll events strictly to the pane the cursor is over.

            Guard: if any popup dialog is open, return immediately so the
            background pane does not scroll while a popup has focus.

            Fix: exact hit-test with fresh coords each time to prevent cross-pane
            scroll bleed that reappeared when the pane cache was stale.
            """
            if getattr(self, "_popup_count", 0) > 0:
                return
            wx, wy = e.x_root, e.y_root

            def _fresh_over(w):
                """Always query live geometry - never use stale cache for hit test."""
                try:
                    x1 = w.winfo_rootx()
                    y1 = w.winfo_rooty()
                    x2 = x1 + w.winfo_width()
                    y2 = y1 + w.winfo_height()
                    return x1 <= wx <= x2 and y1 <= wy <= y2
                except Exception:
                    return False

            def _over(w):
                """Cache geometry  -  only used for non-critical hit tests."""
                wid = id(w)
                if wid not in _pane_cache:
                    try:
                        _pane_cache[wid] = (w.winfo_rootx(), w.winfo_rooty(),
                                            w.winfo_rootx() + w.winfo_width(),
                                            w.winfo_rooty() + w.winfo_height())
                    except Exception:
                        return False
                x1, y1, x2, y2 = _pane_cache[wid]
                return x1 <= wx <= x2 and y1 <= wy <= y2

            def _scroll_canvas(canvas, raw):
                import sys
                bbox = canvas.bbox("all")
                if not bbox: return
                content_h = bbox[3] - bbox[1]
                view_h    = canvas.winfo_height()
                if content_h <= view_h: return
                # macOS trackpad sends tiny fractional deltas; scale up
                # macOS trackpad: small deltas need bigger scale for responsive feel
                scale = 8.0 if sys.platform == "darwin" and abs(raw) < 40 else 4.0
                frac = -raw * scale / content_h
                cur  = canvas.yview()[0]
                canvas.yview_moveto(max(0.0, min(1.0 - view_h/content_h, cur + frac)))

            raw = e.delta if e.delta else 0
            if not raw: return

            # Use _fresh_over for the two main panes to get exact real-time geometry
            in_right = _fresh_over(self._right_pane)
            in_left  = _fresh_over(self._left_pane)

            if in_right:
                if e.state & 0x8 or e.state & 0x4:
                    self._zoom_plot(raw); return
                if e.state & 0x1:
                    import sys as _sys
                    bbox = self._plot_canvas.bbox("all")
                    if bbox:
                        content_w = bbox[2] - bbox[0]
                        view_w    = self._plot_canvas.winfo_width()
                        if content_w > view_w:
                            _hscale = 8.0 if _sys.platform == "darwin" and abs(raw) < 40 else 4.0
                            cur = self._plot_canvas.xview()[0]
                            self._plot_canvas.xview_moveto(
                                max(0.0, min(1.0, cur + (-raw * _hscale / content_w))))
                else:
                    bbox = self._plot_canvas.bbox("all")
                    if bbox:
                        content_h = bbox[3] - bbox[1]
                        view_h    = self._plot_canvas.winfo_height()
                        if content_h > view_h:
                            cur = self._plot_canvas.yview()[0]
                            self._plot_canvas.yview_moveto(
                                max(0.0, min(1.0, cur + (-raw * 4 / content_h))))
                return

            if in_left and not _fresh_over(sidebar_outer):
                main_idx = _current_idx[0]
                canvas   = self._tab_canvases.get((main_idx, 0))
                if canvas:
                    _scroll_canvas(canvas, raw)
            # If cursor is over sidebar, do nothing (sidebar has fixed content)

        self.bind_all("<MouseWheel>", _smart_wheel)

        try:
            def _on_magnify(e):
                wx, wy = e.x_root, e.y_root
                try:
                    x1 = self._right_pane.winfo_rootx()
                    y1 = self._right_pane.winfo_rooty()
                    x2 = x1 + self._right_pane.winfo_width()
                    y2 = y1 + self._right_pane.winfo_height()
                except Exception:
                    return
                if x1 <= wx <= x2 and y1 <= wy <= y2:
                    self._zoom_plot_factor(1.0 + e.delta)
            self.bind_all("<Magnify>", _on_magnify)
        except Exception:
            _log.debug("App._build: <Magnify> event binding failed (non-macOS?)", exc_info=True)

        self._tabs_built   = set()
        self._all_tabs_map = {spec.key: (spec, tabs)
                              for spec, tabs in zip(_REGISTRY_SPECS, _all_tabs)}

        def _build_tab_content(key):
            if key in self._tabs_built: return
            self._tabs_built.add(key)
            spec, tabs = self._all_tabs_map[key]
            outer = tabs[0][0]

            def _section(title):
                hdr = tk.Frame(outer, bg="#dde7f3")
                hdr.pack(fill="x", padx=0, pady=(18, 2))
                tk.Frame(hdr, bg="#2274A5", width=4).pack(side="left", fill="y", padx=(0, 10))
                tk.Label(hdr, text=title.upper(), bg="#dde7f3", fg="#1a5a8a",
                         font=("Helvetica Neue", 10, "bold"), anchor="w"
                         ).pack(side="left", pady=6)
                sub = ttk.Frame(outer)
                sub.pack(fill="x", expand=False)
                sub.columnconfigure(0, weight=1)
                return sub

            data_frame = ttk.Frame(outer)
            data_frame.pack(fill="x", expand=False)
            data_frame.columnconfigure(0, weight=1)
            self._tab_data(data_frame, mode=spec.tab_mode)

            # Record section anchor for the navigator
            if key not in self._section_anchors:
                self._section_anchors[key] = {}
            self._section_anchors[key]["Data"] = 0.0  # always at top

            axes_frame  = _section("Axes")
            self._tab_axes(axes_frame, mode=spec.tab_mode)

            stats_frame = _section("Statistics")
            _stats_dispatch = {
                "grouped_bar":      self._tab_stats_grouped_bar,
                "scatter":          self._tab_stats_scatter,
                "kaplan_meier":     self._tab_stats_kaplan_meier,
                "heatmap":          self._tab_stats_heatmap,
                "two_way_anova":    self._tab_stats_two_way_anova,
                "before_after":     self._tab_stats_before_after,
                "histogram":        self._tab_stats_histogram,
                "curve_fit":        self._tab_stats_curve_fit,
                "column_stats":     self._tab_stats_column_stats,
                "contingency":      self._tab_stats_contingency,
                "repeated_measures":self._tab_stats_repeated_measures,
                "chi_square_gof":   self._tab_stats_chi_square_gof,
                "stacked_bar":      self._tab_stats_stacked_bar,
                "bubble":           self._tab_stats_bubble,
                "dot_plot":         self._tab_stats_dot_plot,
                "bland_altman":     self._tab_stats_bland_altman,
                "forest_plot":      self._tab_stats_forest_plot,
            }
            _stats_dispatch.get(spec.stats_tab, self._tab_stats)(stats_frame)

            ttk.Frame(outer, height=20).pack()
            self._build_lockable_cache()

            # Measure Axes and Statistics section y-positions after layout settles
            def _measure_anchors(key=key, axes_frame=axes_frame, stats_frame=stats_frame, outer=outer):
                canvas = self._tab_canvases.get((_tab_map.get(list(_tab_map.keys())[
                    list(_tab_map.values()).index(key)], -1) 
                    if key in _tab_map.values() else -1, 0))
                # Find pane index for this key
                for idx, k in _tab_map.items():
                    if k == key:
                        canvas = self._tab_canvases.get((idx, 0))
                        break
                if canvas is None:
                    return
                try:
                    canvas.update_idletasks()
                    total_h = canvas.bbox("all")
                    if not total_h: return
                    content_h = total_h[3] - total_h[1]
                    if content_h <= 0: return
                    # axes_frame y relative to the canvas window
                    axes_y = axes_frame.winfo_y()
                    stats_y = stats_frame.winfo_y()
                    self._section_anchors[key]["Axes"] = max(0.0, (axes_y - 10) / content_h)
                    self._section_anchors[key]["Statistics"] = max(0.0, (stats_y - 10) / content_h)
                except Exception:
                    _log.debug("App._build: _measure_anchors failed for key %r", key, exc_info=True)
            self.after(300, _measure_anchors)

        self._build_tab_content = _build_tab_content
        _build_tab_content("bar")

    def _build_right_pane(self, paned):
        """Build the tab bar and scrollable plot canvas on the right side of the split."""
        self._right_pane = right_outer = ttk.Frame(paned)
        paned.add(right_outer, minsize=300)

        # ── Tab bar ───────────────────────────────────────────────────────────
        self._tab_bar_widget = TabBar(
            right_outer,
            on_select=lambda tab_id: self._tab_manager.switch_to(tab_id),
            on_close=lambda tab_id: self._tab_manager.close_tab(tab_id),
            on_new=lambda: self._tab_manager.new_tab("bar"),
            on_reorder=lambda f, t: self._tab_manager.reorder(f, t),
        )
        self._tab_bar_widget.pack(fill="x", side="top")

        # ── Results panel  -  pinned at the bottom, hidden until first run ──────
        self._results_visible = False
        self._results_strip   = tk.Frame(right_outer, bg="#f4f7fb", height=0)
        self._results_strip.pack(side="bottom", fill="x")
        self._results_strip.pack_propagate(False)

        # Toggle header bar
        _rh = tk.Frame(self._results_strip, bg="#dde8f5", height=28, cursor="hand2")
        _rh.pack(fill="x")
        _rh.pack_propagate(False)
        self._results_toggle_arrow = tk.Label(_rh, text="▲ Results", bg="#dde8f5",
                                              fg="#2274A5", font=("Helvetica Neue", 11, "bold"),
                                              cursor="hand2")
        self._results_toggle_arrow.pack(side="left", padx=10)
        self._results_copy_btn = tk.Label(_rh, text="Export CSV", bg="#dde8f5",
                                          fg="#2274A5", font=("Helvetica Neue", 10),
                                          cursor="hand2")
        self._results_copy_btn.pack(side="right", padx=10)
        self._results_copy_btn.bind("<Button-1>", lambda e: self._export_results_csv())

        for w in (_rh, self._results_toggle_arrow):
            w.bind("<Button-1>", lambda e: self._toggle_results_panel())

        # Scrollable body
        _rb = tk.Frame(self._results_strip, bg="#f4f7fb")
        _rb.pack(fill="both", expand=True)
        _vsb = ttk.Scrollbar(_rb, orient="vertical")
        _vsb.pack(side="right", fill="y")
        _hsb = ttk.Scrollbar(_rb, orient="horizontal")
        _hsb.pack(side="bottom", fill="x")
        self._results_canvas = tk.Canvas(_rb, bg="#f4f7fb", highlightthickness=0,
                                         yscrollcommand=_vsb.set, xscrollcommand=_hsb.set)
        self._results_canvas.pack(side="left", fill="both", expand=True)
        _vsb.config(command=self._results_canvas.yview)
        _hsb.config(command=self._results_canvas.xview)
        self._results_inner = tk.Frame(self._results_canvas, bg="#f4f7fb")
        self._results_inner_id = self._results_canvas.create_window(
            (0, 0), window=self._results_inner, anchor="nw")
        self._results_inner.bind("<Configure>", lambda e: self._results_canvas.configure(
            scrollregion=self._results_canvas.bbox("all")))
        self._results_canvas.bind("<Configure>", lambda e: self._results_canvas.itemconfig(
            self._results_inner_id, width=e.width))
        # mousewheel on results
        def _rscroll(e):
            self._results_canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        self._results_canvas.bind("<MouseWheel>", _rscroll)
        self._results_inner.bind("<MouseWheel>", _rscroll)

        # ── Main plot scrollable canvas ───────────────────────────────────────
        right_vsb = ttk.Scrollbar(right_outer, orient="vertical")
        right_vsb.pack(side="right", fill="y")
        right_hsb = ttk.Scrollbar(right_outer, orient="horizontal")
        right_hsb.pack(side="bottom", fill="x")
        self._plot_canvas = tk.Canvas(right_outer, highlightthickness=0,
                                      yscrollcommand=right_vsb.set,
                                      xscrollcommand=right_hsb.set)
        self._plot_canvas.pack(side="left", fill="both", expand=True)
        right_vsb.config(command=self._plot_canvas.yview)
        right_hsb.config(command=self._plot_canvas.xview)

        self._plot_frame    = None   # points to active tab's plot_frame; set by TabManager
        self._canvas_widget = None   # FigureCanvasTkAgg; updated on tab switch
        self._fig           = None   # matplotlib Figure; updated on tab switch
        self._zoom_level    = 1.0

        self._empty_state_frame = ttk.Frame(self._plot_canvas)
        self._empty_state_id    = self._plot_canvas.create_window(
            0, 0, window=self._empty_state_frame, anchor="center")

        def _centre_empty(e=None):
            self._plot_canvas.coords(
                self._empty_state_id,
                self._plot_canvas.winfo_width() // 2,
                self._plot_canvas.winfo_height() // 2)
        self._plot_canvas.bind("<Configure>", _centre_empty)

        ttk.Label(self._empty_state_frame, text="📊",
                  font=("Helvetica Neue", 48)).pack(pady=(0, 8))
        ttk.Label(self._empty_state_frame, text="No plot yet",
                  font=("Helvetica Neue", 18, "bold"),
                  foreground="#aaaaaa").pack()
        ttk.Label(self._empty_state_frame,
                  text="Browse for an Excel file on the left,\nthen click Generate Plot",
                  font=("Helvetica Neue", 13), foreground="#bbbbbb",
                  justify="center").pack(pady=(6, 0))

        # ── Tab manager ───────────────────────────────────────────────────────
        # Created after _plot_canvas exists; opens with one bar-chart tab.
        self._tab_manager = TabManager(self, self._tab_bar_widget, self._plot_canvas)
        self._tab_manager.new_tab("bar")

    # ── Tab builders ──────────────────────────────────────────────────────────

    # ── Wiki ──────────────────────────────────────────────────────────────────

    def _track_popup(self, dlg):
        """Increment popup counter when dlg opens, decrement when it closes.
        While _popup_count > 0 the main-window scroll handler is suspended,
        preventing the background pane from scrolling while a popup is active.
        """
        self._popup_count = getattr(self, "_popup_count", 0) + 1
        def _on_close():
            self._popup_count = max(0, getattr(self, "_popup_count", 1) - 1)
        dlg.protocol("WM_DELETE_WINDOW", lambda: (_on_close(), dlg.destroy()))
        dlg.bind("<Destroy>", lambda e: _on_close() if e.widget is dlg else None)

    def _open_wiki_popup(self):
        """Open the Reference Wiki in a floating popup window."""
        try:
            from plotter_app_wiki import open_wiki_popup
            open_wiki_popup(self,
                            track_popup_fn=self._track_popup,
                            bind_scroll_fn=None)
        except Exception as e:
            from tkinter import messagebox
            messagebox.showinfo("Wiki", f"Wiki unavailable: {e}")

    def _open_wiki_popup_inline(self):
        """Inline fallback wiki — kept for reference, not called directly."""
        dlg = tk.Toplevel(self)
        dlg.title("Reference Wiki")
        dlg.resizable(True, True)
        dlg.geometry("680x700")
        dlg.configure(bg="#ffffff")
        # non-modal so the user can keep the app visible alongside
        dlg.transient(self)
        self._track_popup(dlg)  # suspend background scroll while open

        # Header
        hdr = tk.Frame(dlg, bg="#2274A5", height=52)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="📖  Reference Wiki",
                 bg="#2274A5", fg="white",
                 font=("Helvetica Neue", 15, "bold")).pack(side="left", padx=18, pady=12)
        tk.Label(hdr, text="scipy · GraphPad Prism 11 Statistics Guide",
                 bg="#2274A5", fg="#a8cce0",
                 font=("Helvetica Neue", 11)).pack(side="left", padx=(0, 18), pady=12)

        # Two-pane layout: sidebar TOC (left) + scrollable content (right)
        body_outer = tk.Frame(dlg, bg="#ffffff")
        body_outer.pack(fill="both", expand=True)

        # TOC sidebar
        toc_frame = tk.Frame(body_outer, bg="#f0f4fa", width=180)
        toc_frame.pack(side="left", fill="y")
        toc_frame.pack_propagate(False)
        tk.Label(toc_frame, text="CONTENTS", bg="#f0f4fa", fg="#2274A5",
                 font=("Helvetica Neue", 9, "bold"), anchor="w"
                 ).pack(fill="x", padx=10, pady=(12, 4))
        tk.Frame(toc_frame, bg="#c8d8ea", height=1).pack(fill="x", padx=6)
        toc_listbox = tk.Listbox(toc_frame, bg="#f0f4fa", fg="#333333",
                                 font=("Helvetica Neue", 11), selectmode="single",
                                 selectbackground="#2274A5", selectforeground="white",
                                 activestyle="none", relief="flat",
                                 highlightthickness=0, borderwidth=0, cursor="hand2")
        toc_listbox.pack(fill="both", expand=True, padx=4, pady=4)
        # Disable scroll on TOC listbox from bleeding to content
        toc_listbox.bind("<MouseWheel>", lambda e: "break")
        tk.Frame(body_outer, bg="#dde4ee", width=1).pack(side="left", fill="y")

        # Content area
        content_area = tk.Frame(body_outer, bg="#ffffff")
        content_area.pack(side="left", fill="both", expand=True)
        vsb    = ttk.Scrollbar(content_area, orient="vertical")
        vsb.pack(side="right", fill="y")
        canvas = tk.Canvas(content_area, highlightthickness=0,
                           yscrollcommand=vsb.set, background="white")
        canvas.pack(side="left", fill="both", expand=True)
        vsb.config(command=canvas.yview)
        frame = ttk.Frame(canvas)
        win   = canvas.create_window((0, 0), window=frame, anchor="nw")
        frame.bind("<Configure>", lambda e: canvas.configure(
            scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(win, width=e.width))

        # TOC anchor registry  -  {section_title: y_pixel_offset}
        _toc_anchors = {}
        _toc_entries = []  # ordered list of section titles

        def _scroll(e):
            import sys
            raw = getattr(e, "delta", 0) or (120 if getattr(e, "num", 0) == 4 else -120)
            if not raw: return
            bbox = canvas.bbox("all")
            if not bbox: return
            content_h = bbox[3] - bbox[1]
            view_h    = canvas.winfo_height()
            if content_h <= view_h: return
            # macOS trackpad sends small deltas (±3–10); mouse sends ±120
            scale = 8.0 if sys.platform == "darwin" and abs(raw) < 40 else 4.0
            frac = -raw * scale / content_h
            cur  = canvas.yview()[0]
            canvas.yview_moveto(max(0.0, min(1.0 - view_h/content_h, cur + frac)))

        def _scroll_linux(e):
            raw = 120 if e.num == 4 else -120
            _scroll(type("E", (), {"delta": raw, "num": e.num})())

        # Bind only to this popup's widget tree (not bind_all which leaks to parent)
        dlg.after(50, lambda: _bind_scroll_recursive(dlg, _scroll, _scroll_linux, _scroll_linux))

        def _toc_select(e):
            sel = toc_listbox.curselection()
            if not sel: return
            title = _toc_entries[sel[0]]
            lbl_widget = _toc_anchors.get(title)
            if lbl_widget is None: return
            try:
                # Get the label's y position relative to the scrolled canvas content
                lbl_widget.update_idletasks()
                lbl_y = lbl_widget.winfo_y()
                bbox = canvas.bbox("all")
                if bbox:
                    content_h = bbox[3] - bbox[1]
                    if content_h > 0:
                        canvas.yview_moveto(max(0.0, (lbl_y - 10) / content_h))
            except Exception:
                pass
        toc_listbox.bind("<<ListboxSelect>>", _toc_select)

        # ── Content helpers ───────────────────────────────────────────────────
        def h1(text):
            lbl = ttk.Label(frame, text=text,
                      font=("Helvetica Neue", 16, "bold"),
                      foreground="#2274A5", background="white")
            lbl.pack(anchor="w", padx=20, pady=(20, 4))
            ttk.Separator(frame).pack(fill="x", padx=20, pady=(0, 8))
            # Register TOC entry
            _toc_entries.append(text)
            toc_listbox.insert("end", f"  {text}")
            # Store a reference to measure y-position after layout
            _toc_anchors[text] = lbl

        def h2(text):
            ttk.Label(frame, text=text,
                      font=("Helvetica Neue", 13, "bold"),
                      foreground="#444444", background="white"
                      ).pack(anchor="w", padx=20, pady=(12, 2))

        def body(text):
            import re
            text = re.sub(r'\*\*(.+?)\*\*', r'\1', text)
            text = re.sub(r'\*(.+?)\*',     r'\1', text)
            text = re.sub(r'``(.+?)``',     r'\1', text)
            text = re.sub(r':math:`(.+?)`', r'\1', text)
            text = re.sub(r'.. versionadded.*', '', text)
            text = re.sub(r'.. versionchanged.*', '', text)
            text = text.strip()
            if not text: return
            ttk.Label(frame, text=text,
                      font=("Helvetica Neue", 12),
                      foreground="#333333", background="white",
                      wraplength=580, justify="left"
                      ).pack(anchor="w", padx=28, pady=(0, 6))

        def tag(text, color="#2274A5"):
            lbl = tk.Label(frame, text=f" {text} ",
                           font=("Helvetica Neue", 10, "bold"),
                           fg="white", bg=color, padx=4, pady=1)
            lbl.pack(anchor="w", padx=28, pady=(0, 4))

        def plotter_link(page_slug, link_text):
            """Clickable hyperlink to a GraphPad Prism 11 Statistics Guide page."""
            import webbrowser
            url = f"https://www.graphpad.com/guides/prism/latest/statistics/{page_slug}"
            lbl = tk.Label(frame, text=f"📎 {link_text}",
                           font=("Helvetica Neue", 11, "underline"),
                           fg="#2274A5", bg="white", cursor="hand2")
            lbl.pack(anchor="w", padx=28, pady=(0, 6))
            lbl.bind("<Button-1>", lambda e: webbrowser.open(url))

        def scipy_entry(fn, label_text, use_sections=("Notes",)):
            import inspect, re
            doc = inspect.getdoc(fn) or ""
            paras = doc.split("\n\n")
            if paras:
                body(paras[0].replace("\n", " "))
            current_section = None
            section_buf     = []
            for para in paras[1:]:
                lines = para.strip().splitlines()
                if (len(lines) >= 2 and
                        set(lines[1].strip()) <= set("-=~^") and
                        len(lines[1].strip()) >= 3):
                    if current_section in use_sections and section_buf:
                        sentences = re.split(r'(?<=[.!?])\s+', " ".join(section_buf).strip())
                        body(" ".join(sentences[:6]))
                    current_section = lines[0].strip()
                    section_buf     = []
                elif current_section in use_sections:
                    clean = " ".join(para.replace("\n", " ").split())
                    if clean and not clean.startswith(">>>"):
                        section_buf.append(clean)
            if current_section in use_sections and section_buf:
                sentences = re.split(r'(?<=[.!?])\s+', " ".join(section_buf).strip())
                body(" ".join(sentences[:6]))

        from scipy import stats as _st

        # ── Statistical Tests ─────────────────────────────────────────────────
        h1("Statistical Tests")

        h2("Independent Samples t-test  (Parametric, 2 groups)")
        tag("scipy.stats.ttest_ind", "#2274A5")
        scipy_entry(_st.ttest_ind, "ttest_ind")
        body("Welch's t-test (unequal variances) is the default in GraphPad Prism "
             "since version 8. It does not assume equal standard deviations and is "
             "preferred over Student's t-test in most practical situations.")
        plotter_link("stat_qa_choosing_a_test_to_compare_.htm",
                   "Prism 11: Choosing a test to compare two groups")

        h2("Paired t-test  (Parametric, 2 matched groups)")
        tag("scipy.stats.ttest_rel", "#2274A5")
        scipy_entry(_st.ttest_rel, "ttest_rel")
        plotter_link("stat_checklist_pairedt.htm",
                   "Prism 11: Checklist for paired t-test")

        h2("One-Way ANOVA  (Parametric, 3+ groups)")
        tag("scipy.stats.f_oneway", "#2274A5")
        scipy_entry(_st.f_oneway, "f_oneway")
        body("Prism 8+ also offers Welch's ANOVA and Brown-Forsythe ANOVA when equal "
             "variances cannot be assumed. These adjust the F-ratio and degrees of freedom "
             "for heteroscedasticity  -  recommended when Levene's or Bartlett's test is significant.")
        plotter_link("stat_checklist_1wayanova.htm",
                   "Prism 11: Checklist for one-way ANOVA")
        plotter_link("stat_how_to_multiple_comparisons_af.htm",
                   "Prism 11: Multiple comparisons after one-way ANOVA")

        h2("Repeated-Measures One-Way ANOVA  (Parametric, matched)")
        tag("pingouin.rm_anova", "#2274A5")
        body("Tests differences across 3+ conditions when the same subjects are measured "
             "in each condition. More powerful than one-way ANOVA when matching is effective. "
             "Prism 11 applies the Geisser-Greenhouse correction by default when sphericity "
             "cannot be assumed.")
        plotter_link("stat_checklist_1wayanova_rm.htm",
                   "Prism 11: Repeated-measures one-way ANOVA")

        h2("Two-Way ANOVA  (Parametric, 2 factors)")
        tag("pingouin.anova (Type II SS)", "#2274A5")
        body("Tests main effects of Factor A and Factor B, plus their interaction  -  "
             "whether the effect of one factor depends on the level of the other. "
             "Type II Sum of Squares tests each main effect after controlling for the other. "
             "Post-hoc pairwise comparisons are corrected with Holm-Bonferroni.")
        plotter_link("stat_howto_twowayanova.htm",
                   "Prism 11: Two-way ANOVA")

        h2("Mann-Whitney U  (Non-parametric, 2 groups)")
        tag("scipy.stats.mannwhitneyu", "#32936F")
        scipy_entry(_st.mannwhitneyu, "mannwhitneyu")
        plotter_link("stat_checklist_mwt.htm",
                   "Prism 11: Checklist for Mann-Whitney test")

        h2("Wilcoxon Signed-Rank  (Non-parametric, paired)")
        tag("scipy.stats.wilcoxon", "#32936F")
        scipy_entry(_st.wilcoxon, "wilcoxon")
        plotter_link("stat_checklist_wilcoxon_signed_rank.htm",
                   "Prism 11: Wilcoxon signed-rank test")

        h2("Kruskal-Wallis H  (Non-parametric, 3+ groups)")
        tag("scipy.stats.kruskal", "#32936F")
        scipy_entry(_st.kruskal, "kruskal")
        plotter_link("stat_checklist_kw.htm",
                   "Prism 11: Checklist for Kruskal-Wallis test")

        h2("Friedman Test  (Non-parametric, repeated measures)")
        tag("scipy.stats.friedmanchisquare", "#32936F")
        body("Non-parametric analogue of repeated-measures one-way ANOVA. Ranks values "
             "within each subject (row) and tests whether ranks differ across conditions. "
             "Follow with Dunn's post-hoc + Holm correction for pairwise comparisons.")
        plotter_link("stat_checklist_friedman.htm",
                   "Prism 11: Friedman test")

        h2("Tukey HSD  (Post-hoc, all-pairwise, parametric)")
        tag("scipy.stats.tukey_hsd", "#F18F01")
        try:
            scipy_entry(_st.tukey_hsd, "tukey_hsd")
        except Exception:
            body("Tukey's Honestly Significant Difference test compares all pairs of means "
                 "while maintaining the family-wise error rate. The Prism 11 default for "
                 "all-pairwise comparisons after one-way ANOVA when equal SDs are assumed.")
        body("Holm-Sidak is more powerful than Tukey HSD (can find significant differences "
             "where Tukey cannot) but does not produce confidence intervals. Prism 11 "
             "recommends Tukey for its combination of power and interpretability.")
        plotter_link("stat_options_tab_1wayanova.htm",
                   "Prism 11: Post-hoc test options for one-way ANOVA")

        h2("Dunn's Test  (Post-hoc, non-parametric)")
        tag("scikit_posthocs.posthoc_dunn", "#F18F01")
        body("Performs pairwise comparisons on ranked data after a significant "
             "Kruskal-Wallis or Friedman test. Holm-Bonferroni correction controls the "
             "family-wise error rate across all comparisons. Prism 11 default for "
             "non-parametric pairwise post-hoc analysis.")
        plotter_link("stat_checklist_kw.htm",
                   "Prism 11: Multiple comparisons after Kruskal-Wallis")

        # ── Normality & Variance Tests ────────────────────────────────────────
        h1("Normality & Variance Tests")

        h2("Shapiro-Wilk Normality Test")
        tag("scipy.stats.shapiro", "#A846A0")
        scipy_entry(_st.shapiro, "shapiro")
        body("Prism 11 runs Shapiro-Wilk automatically. However, Prism explicitly warns "
             "that 'it is not a good idea to base your decision solely on the normality test'  -  "
             "small samples have low power to detect non-normality, and large samples "
             "detect trivially small deviations. Use it as one input among several.")
        plotter_link("stat_checklist_1wayanova.htm",
                   "Prism 11: Normality assumption in ANOVA")

        h2("Levene's Test for Equal Variances")
        tag("scipy.stats.levene", "#A846A0")
        scipy_entry(_st.levene, "levene")
        body("Prism 11 uses Brown-Forsythe and Bartlett's tests for homoscedasticity, "
             "not Levene's. Brown-Forsythe (median-centred) is more robust to non-normality. "
             "Claude Plotter uses Levene's as an equivalent diagnostic.")
        plotter_link("stat_checklist_1wayanova.htm",
                   "Prism 11: Equal variance assumption (Brown-Forsythe / Bartlett)")

        # ── Effect Sizes ─────────────────────────────────────────────────────
        h1("Effect Sizes")

        h2("Cohen's d  (2-group standardised difference)")
        tag("d = (μ₁ − μ₂) / s_pooled", "#6B4226")
        body("Standardised mean difference. d = 0.2: small, 0.5: medium, 0.8: large "
             "(Cohen 1988). Prism 11 reports Cohen's d for t-tests and partial η² for ANOVA.")
        plotter_link("stat_options_tab_one-way_anova.htm",
                   "Prism 11: Effect size for ANOVA")

        h2("Eta Squared / Partial Eta Squared  (ANOVA)")
        tag("η² = SS_effect / SS_total", "#6B4226")
        body("Proportion of total variance explained by a factor. η² = 0.01: small, "
             "0.06: medium, 0.14: large (Cohen). For one-way ANOVA η² = partial η². "
             "For two-way ANOVA Claude Plotter reports partial η² (effect SS / (effect SS + error SS)).")
        plotter_link("stat_options_tab_one-way_anova.htm",
                   "Prism 11: Eta squared and omega squared")

        # ── Choosing the Right Test ───────────────────────────────────────────
        h1("Choosing the Right Test")

        body("Use parametric tests (t-test, ANOVA) when data are approximately normally "
             "distributed. Shapiro-Wilk p > 0.05 is one indicator, but also inspect "
             "histograms and Q-Q plots for small samples.")
        body("Use non-parametric tests (Mann-Whitney, Kruskal-Wallis) when normality "
             "fails, samples are small (n < 5), or data are ordinal. They compare "
             "ranks rather than means, losing less power than commonly assumed.")
        body("Use paired tests (paired t-test, Wilcoxon) when observations are matched  -  "
             "before/after on the same subject, or left-vs-right measurements. Pairing "
             "removes between-subject variability and greatly increases power.")
        body("After a significant ANOVA, always apply a post-hoc test with "
             "multiple-comparison correction. Prism 11 recommends Tukey HSD for "
             "all-pairwise comparisons and Dunnett for comparing groups to a control. "
             "Fisher's LSD provides no correction and inflates Type I error  -  "
             "only use it with a very specific justification.")
        plotter_link("stat_how_to_multiple_comparisons_af.htm",
                   "Prism 11: Guide to multiple comparisons")
        plotter_link("stat_---_principles_of_statistics_-.htm",
                   "Prism 11: Principles of Statistics (overview)")

        # ── Chart Types ───────────────────────────────────────────────────────
        h1("Chart Types")
        chart_info = [
            ("Bar Chart", "#2274A5",
             "Displays group means as vertical bars with error bars (SEM, SD, or 95% CI). "
             "Significance brackets show pairwise statistical comparisons. "
             "Best for comparing means across independent groups.",
             "stat_choosing_test_one-way_anova.htm",
             "Prism 11: One-way ANOVA (the analysis behind bar charts)"),
            ("Grouped Bar", "#F18F01",
             "Bar chart with two grouping factors  -  bars grouped by Category, colored "
             "by Subgroup. Best for 2×2 or 2×3 factorial designs. Stats compare subgroups "
             "within each category cluster.",
             "stat_howto_twowayanova.htm",
             "Prism 11: Two-way ANOVA"),
            ("Scatter Plot", "#048A81",
             "Plots individual (X, Y) data points. Optional linear regression with 95% CI "
             "band and Pearson or Spearman correlation annotation.",
             "stat_checklist_linearregression.htm",
             "Prism 11: Checklist for linear regression"),
            ("Survival Curve", "#D4AC0D",
             "Kaplan-Meier survival curves. Log-rank (Mantel-Cox) test compares groups. "
             "Tick marks show censored observations.",
             "stat_howto_km.htm",
             "Prism 11: Kaplan-Meier survival analysis"),
            ("Before / After", "#32936F",
             "Paired dot plot. Each subject appears in every condition connected by a line. "
             "Paired t-test or Wilcoxon signed-rank when statistics are enabled.",
             "stat_checklist_pairedt.htm",
             "Prism 11: Paired t-test"),
            ("Repeated Measures", "#D4AC0D",
             "Within-subjects design. Parametric: RM one-way ANOVA (Geisser-Greenhouse "
             "correction available). Non-parametric: Friedman + Dunn's post-hoc.",
             "stat_checklist_1wayanova_rm.htm",
             "Prism 11: Repeated-measures one-way ANOVA"),
            ("Bland-Altman", "#6B4226",
             "Assesses agreement between two measurement methods. Mean difference = bias; "
             "±1.96 SD lines = limits of agreement. Not a hypothesis test  -  "
             "clinical acceptability is the criterion.",
             "stat_howto_bland_altman.htm",
             "Prism 11: Bland-Altman plot"),
        ]
        for chart_name, color, desc, slug, link_text in chart_info:
            h2(chart_name)
            tag(chart_name, color)
            body(desc)
            plotter_link(slug, link_text)

        ttk.Frame(frame, height=40).pack()

        # Footer with single Close button
        ttk.Separator(dlg).pack(fill="x")
        foot = tk.Frame(dlg, bg="#ffffff")
        foot.pack(fill="x", padx=20, pady=10)
        PButton(foot, text="Close", style="primary",
                command=dlg.destroy).pack(side="right")

    # Keep _build_wiki as a no-op stub so nothing crashes if called
    def _build_wiki(self, parent):
        pass

    def _tab_data(self, f, mode="bar"):
        for k, v in [("excel_path", ""), ("sheet", ""),
                     ("error", "SEM (Standard Error)"), ("show_points", True),
                     ("show_n_labels", False), ("show_value_labels", False),
                     ("stacked_value_labels", False), ("error_below_bar", False),
                     ("jitter_amount", "0"),
                     ("color", "Default"), ("title", ""),
                     ("xlabel", ""), ("ytitle", "")]:
            if k not in self._vars:
                self._vars[k] = (tk.BooleanVar(value=v)
                                 if isinstance(v, bool) else tk.StringVar(value=v))

        # Initialise per-tab widget lists on first call
        if not hasattr(self, "_drop_targets"): self._drop_targets = []
        for attr in ("_file_row_frames", "_validate_lbls", "_sheet_hints", "_sheet_cbs"):
            if not hasattr(self, attr):
                setattr(self, attr, [])

        g = f; r = 0

        _l = ttk.Label(g, text=label("excel_path"), font=("Helvetica Neue", 13, "bold"))
        _l.grid(row=r, column=0, sticky="w", padx=PAD, pady=(12, 2)); tip(_l, "excel_path"); r += 1

        # Single-file row
        file_row_frame = ttk.Frame(g)
        file_row_frame.grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=2)
        _e_path = PEntry(file_row_frame, textvariable=self._vars["excel_path"],
                  font=("Menlo", 12)); _e_path.pack(side="left", fill="x", expand=True)
        add_placeholder(_e_path, self._vars["excel_path"], "Browse or drop an Excel file…")
        PButton(file_row_frame, text="Browse", style="ghost", lock_exempt=True,
                   command=self._browse_file).pack(side="left", padx=(8, 0))
        self._recent_btn = PButton(file_row_frame, text="Recent", style="ghost", lock_exempt=True,
                   command=self._show_recent_files)
        self._recent_btn.pack(side="left", padx=(4, 0))
        self._file_row_frames.append(file_row_frame)
        self._file_row_frame = file_row_frame
        self._drop_targets.append(_e_path)
        r += 1

        # Format hint + template button
        _FORMAT_HINTS = {
            "bar":         "Row 1: group names  |  Rows 2+: numeric values",
            "line":        "Row 1: col 1 = X label, cols 2+ = series names  |  Rows 2+: col 1 = X value, cols 2+ = Y values",
            "grouped_bar": "Row 1: category names  |  Row 2: subgroup names  |  Rows 3+: numeric values",
            "box":         "Row 1: group names  |  Rows 2+: numeric values",
            "scatter":     "Row 1: col 1 = X label, cols 2+ = series names  |  Rows 2+: col 1 = X value, cols 2+ = Y values",
            "violin":      "Row 1: group names  |  Rows 2+: numeric values",
            "kaplan_meier":"Row 1: group names (each spans 2 cols)  |  Row 2: 'Time' / 'Event'  |  Rows 3+: time value, event (1=occurred, 0=censored)",
            "heatmap":     "Row 1: blank/label in A1, then column labels  |  Rows 2+: row label in col A, then numeric values",
            "two_way_anova": "Row 1: column headers (e.g. Factor_A, Factor_B, Value)  |  Rows 2+: one observation per row (long format)",
            "before_after":  "Row 1: condition names (e.g. Before, After)  |  Rows 2+: one subject per row (matched by row order)",
            "histogram":     "Row 1: series names  |  Rows 2+: raw numeric values (columns may differ in length)",
            "subcolumn_scatter": "Row 1: group names  |  Rows 2+: numeric values",
            "curve_fit":   "Row 1: col 1 = X label, cols 2+ = series names  |  Rows 2+: X value, Y replicates",
            "column_stats":"Row 1: group names  |  Rows 2+: numeric values",
            "contingency": "Row 1: col A blank, cols B+ = outcome labels  |  Rows 2+: col A = group name, then counts",
            "repeated_measures": "Row 1: condition/timepoint names  |  Rows 2+: one row per subject",
        }
        ttk.Label(g, text=_FORMAT_HINTS.get(mode, ""),
                  foreground="#888", font=("Helvetica Neue", 10),
                  wraplength=400, justify="left"
                  ).grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=(2, 2)); r += 1
        PButton(g, text="Template", style="ghost", lock_exempt=True,
                   command=lambda m=mode: self._download_template(m)
                   ).grid(row=r, column=0, sticky="w", padx=PAD, pady=(0, 4)); r += 1

        vr = ttk.Frame(g); vr.grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=(4, 2)); r += 1
        validate_lbl = ttk.Label(vr, text="", foreground="#aaaaaa",
                                 font=("Helvetica Neue", 11))
        validate_lbl.pack(side="left")
        self._validate_lbls.append(validate_lbl)
        self._validate_lbl = validate_lbl

        section_sep(g, r, "Sheet"); r += 1

        _l, r = _lbl(g, r, "sheet")
        sheet_hint = ttk.Label(g, text="Browse for an Excel file to populate sheets",
                               foreground="gray")
        sheet_hint.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD); r += 1
        self._sheet_hints.append(sheet_hint)
        self._sheet_hint = sheet_hint

        sheet_cb = PCombobox(g, textvariable=self._vars["sheet"],
                                state="disabled", width=36, font=("Menlo", 12))
        sheet_cb.grid(row=r, column=0, columnspan=2, sticky="w", padx=PAD, pady=4); r += 1
        self._sheet_cbs.append(sheet_cb)
        self._sheet_cb = sheet_cb
        # When user picks a different sheet, revalidate without resetting sheet selection
        sheet_cb.bind("<<ComboboxSelected>>", lambda e: self._load_sheets(
            self._vars["excel_path"].get().strip(), reset_sheet=False)
            if self._vars["excel_path"].get().strip() else None)

        section_sep(g, r, "Error Bars"); r += 1

        _l, r = _lbl(g, r, "error")
        PCombobox(g, textvariable=self._vars["error"],
                     values=["SEM (Standard Error)", "SD (Standard Deviation)", "95% CI"],
                     state="readonly", width=26, font=("Helvetica Neue", 12)
                     ).grid(row=r, column=0, sticky="w", padx=PAD, pady=4); r += 1

        fr = ttk.Frame(g); fr.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=(6, 2)); r += 1
        PCheckbox(fr, variable=self._vars["show_points"],
                        text=f" {label('show_points')}").pack(side="left")

        r = self._checkbox(g, r, "show_n_labels", " Show n= sample size on x-axis")
        if mode in ("bar", "grouped_bar"):
            r = self._checkbox(g, r, "show_value_labels", " Show value label on each bar")
        if mode == "stacked_bar":
            r = self._checkbox(g, r, "stacked_value_labels", " Show value label in each segment")

        r = self._checkbox(g, r, "error_below_bar", " Extend error bar below bar (to zero)")
        section_sep(g, r, "Data Points"); r += 1

        _l, r = _lbl(g, r, "jitter_amount")
        jrow = ttk.Frame(g); jrow.grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=4); r += 1
        jlbl = ttk.Label(jrow, text="0.00", width=5, anchor="e"); jlbl.pack(side="right")
        def _ju(val): v=round(float(val),2); self._vars["jitter_amount"].set(str(v)); jlbl.config(text=f"{v:.2f}")
        ttk.Scale(jrow, from_=0.0, to=0.5, orient="horizontal",
                  variable=self._vars["jitter_amount"], command=_ju
                  ).pack(side="left", fill="x", expand=True, padx=(0, 8))
        section_sep(g, r, "Color"); r += 1

        _l, r = _lbl(g, r, "color")
        _color_cb = PCombobox(g, textvariable=self._vars["color"],
                     values=[
                         # ── Prism default ──────────────────────────────────
                         "Default",
                         # ── Curated presets ────────────────────────────────
                         "Pastel",          # soft pastels, great for posters
                         "Vivid",           # saturated, high-contrast
                         "CB-safe",         # colorblind-safe (Wong 2011)
                         "Grayscale",       # publication B&W
                         "Blues",           # sequential blue ramp
                         "Reds",            # sequential red ramp
                         "Greens",          # sequential green ramp
                         "RdBu",            # diverging red-blue
                         "mako",            # perceptually uniform blue-green (restored)
                         "rocket",          # dark-to-light red sequential
                         "flare",           # warm sequential
                         "crest",           # cool sequential
                         # ── Seaborn qualitative ────────────────────────────
                         "deep", "muted", "pastel", "bright", "dark",
                         "colorblind", "tab10",
                         # ── Matplotlib named sets ──────────────────────────
                         "Set1", "Set2", "Set3", "husl", "hls",
                         # ── Single colors ─────────────────────────────────
                         "steelblue", "tomato", "mediumseagreen", "mediumpurple",
                         "coral", "goldenrod", "slategray", "darkorange",
                     ],
                     state="normal", width=26, font=("Menlo", 12))
        _color_cb.grid(row=r, column=0, columnspan=2, sticky="w", padx=PAD, pady=4)
        add_placeholder(_color_cb, self._vars["color"], "Default")

        # Color swatch  -  5 squares that update live as palette name changes
        _swatch_frame = ttk.Frame(g)
        _swatch_frame.grid(row=r, column=2, sticky="w", padx=(4, PAD), pady=4)
        _swatch_canvases = []
        for _si in range(5):
            _sc = tk.Canvas(_swatch_frame, width=16, height=16,
                            highlightthickness=1, highlightbackground="#cccccc")
            _sc.pack(side="left", padx=2)
            _swatch_canvases.append(_sc)

        def _update_swatches(*_):
            import sys
            pal_name = self._vars["color"].get().strip()
            # Curated presets defined inline (mirrors _assign_colors in plotter_functions)
            _PRESET_SWATCHES = {
                "Pastel":    ["#AEC6CF", "#FFD1DC", "#B5EAD7", "#FFDAC1", "#C9C9FF"],
                "Vivid":     ["#E8453C", "#2274A5", "#32936F", "#F18F01", "#A846A0"],
                "CB-safe":   ["#E69F00", "#56B4E9", "#009E73", "#F0E442", "#0072B2"],
                "Grayscale": ["#222222", "#555555", "#888888", "#aaaaaa", "#cccccc"],
                "Blues":     ["#084594", "#2171b5", "#4292c6", "#6baed6", "#9ecae1"],
                "Reds":      ["#99000d", "#cb181d", "#ef3b2c", "#fb6a4a", "#fc9272"],
                "Greens":    ["#005a32", "#238b45", "#41ae76", "#74c476", "#a1d99b"],
                "RdBu":      ["#d73027", "#f46d43", "#e0f3f8", "#74add1", "#4575b4"],
            }
            if not pal_name or pal_name.lower() in ("none", "default (prism)"):
                pf_mod = sys.modules.get("plotter_functions")
                hex_cols = (pf_mod.PRISM_PALETTE[:5] if pf_mod
                            else ["#E8453C", "#2274A5", "#32936F", "#F18F01", "#A846A0"])
            elif pal_name in _PRESET_SWATCHES:
                hex_cols = _PRESET_SWATCHES[pal_name]
            else:
                if "matplotlib" not in sys.modules:
                    return
                import matplotlib.colors as _mcolors
                _SEABORN_ALIASES = {"Blues": "Blues_d", "Reds": "Reds_d",
                                    "Greens": "Greens_d", "RdBu": "RdBu_r"}
                resolved = _SEABORN_ALIASES.get(pal_name, pal_name)
                try:
                    import seaborn as _sns
                    pal = _sns.color_palette(resolved, 5)
                    hex_cols = ["#%02x%02x%02x" % (int(r*255), int(g_*255), int(b*255))
                                for r, g_, b in pal]
                except Exception:
                    _log.debug("_update_swatches: seaborn palette %r failed, trying mcolors", pal_name, exc_info=True)
                    try:
                        c = _mcolors.to_hex(pal_name)
                        hex_cols = [c] * 5
                    except Exception:
                        _log.debug("_update_swatches: mcolors fallback for %r failed", pal_name, exc_info=True)
                        hex_cols = ["#cccccc"] * 5
            for _sci, _sc in enumerate(_swatch_canvases):
                _sc.delete("all")
                _sc.create_rectangle(1, 1, 15, 15,
                                     fill=hex_cols[_sci % len(hex_cols)],
                                     outline="")

        self._vars["color"].trace_add("write", _update_swatches)
        _color_cb.bind("<<ComboboxSelected>>", _update_swatches)
        self.after(100, _update_swatches)
        r += 1

        # ── Style Presets ──────────────────────────────────────────────────────
        section_sep(g, r, "Style Presets"); r += 1
        try:
            from plotter_presets import BUILT_IN_PRESETS
            _preset_names = ["— none —"] + list(BUILT_IN_PRESETS.keys())
        except Exception:
            _log.debug("App._tab_data: plotter_presets import failed", exc_info=True)
            _preset_names = ["— none —"]
            BUILT_IN_PRESETS = {}
        if "preset" not in self._vars:
            self._vars["preset"] = tk.StringVar(value="— none —")
        _preset_row = ttk.Frame(g)
        _preset_row.grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=4)
        _preset_cb = PCombobox(_preset_row, textvariable=self._vars["preset"],
                               values=_preset_names, state="readonly", width=26,
                               font=("Menlo", 12))
        _preset_cb.pack(side="left")

        def _apply_preset(*_):
            name = self._vars["preset"].get()
            preset = BUILT_IN_PRESETS.get(name, {})
            for k, v in preset.items():
                if k in self._vars:
                    try:
                        self._vars[k].set(v)
                    except Exception:
                        pass

        PButton(_preset_row, text="Apply", style="ghost",
                command=_apply_preset).pack(side="left", padx=(6, 0))
        _preset_cb.bind("<<ComboboxSelected>>", _apply_preset)
        r += 1

        section_sep(g, r, "Labels"); r += 1

        # Label fields  -  deliberately NO placeholder text.
        # Empty = no label on the chart, which is the correct default behavior.
        for key in ("title", "xlabel", "ytitle"):
            _l = ttk.Label(g, text=label(key), font=("Helvetica Neue", 13, "bold"))
            _l.grid(row=r, column=0, sticky="w", padx=PAD, pady=(4, 2)); tip(_l, key); r += 1
            # Row frame holds entry + Clear × button side-by-side
            _row_f = ttk.Frame(g)
            _row_f.grid(row=r, column=0, columnspan=2, sticky="ew", padx=PAD, pady=4)
            _row_f.columnconfigure(0, weight=1)
            _e = PEntry(_row_f, textvariable=self._vars[key], width=34, font=("Menlo", 12))
            _e.grid(row=0, column=0, sticky="ew")
            # Clear button — clears the StringVar with one click
            _var_ref = self._vars[key]
            def _make_clear(v):
                return lambda e=None: v.set("")
            _clr = tk.Label(_row_f, text="×", bg=_DS.BG, fg="#aaaaaa",
                             font=("Helvetica Neue", 14), cursor="hand2",
                             padx=4, pady=0)
            _clr.grid(row=0, column=1, sticky="e", padx=(4, 0))
            _clr.bind("<Button-1>", _make_clear(_var_ref))
            _clr.bind("<Enter>",  lambda e, w=_clr: w.config(fg="#cc0000"))
            _clr.bind("<Leave>",  lambda e, w=_clr: w.config(fg="#aaaaaa"))
            r += 1

        g.columnconfigure(0, weight=1)
        ttk.Frame(g).grid(row=r, pady=8)

    def _tab_axes(self, f, mode="bar"):
        self._init_vars({
            "yscale": "Linear", "ylim_lo": "", "ylim_hi": "",
            "figw": "", "figh": "", "font_size": "12", "ref_line_y": "0",
            "ref_line_enabled": False, "ylim_data_min": False, "ylim_none": True,
            "ylim_mode": 0, "xlim_mode": 0, "xlim_lo": "", "xlim_hi": "",
            "gridlines": False, "open_points": False,
            "horizontal": False, "show_median": False,
            "bar_alpha": "0.85", "xscale": "Linear", "ref_line_label": "",
        })
        # Resolve capability flags from the registry (single source of truth).
        # This replaces the hardcoded _POINTS_MODES / _CAP_MODES / etc. sets.
        _spec = next((s for s in _REGISTRY_SPECS if s.tab_mode == mode), None)
        _has_points    = _spec.has_points     if _spec else (mode not in {"heatmap", "kaplan_meier", "histogram", "column_stats", "contingency", "two_way_anova", "chi_square_gof"})
        _has_error_bars= _spec.has_error_bars if _spec else (mode in {"bar", "line", "grouped_bar", "box", "before_after", "subcolumn_scatter", "repeated_measures"})
        _has_legend    = _spec.has_legend     if _spec else (mode in {"line", "grouped_bar", "scatter", "kaplan_meier", "histogram"})
        _x_continuous  = _spec.x_continuous   if _spec else (mode in {"line", "scatter", "kaplan_meier", "histogram"})
        if mode in ("bar", "grouped_bar", "box"):
            self._init_vars({"bar_width": "0.6"})
        if mode in ("line", "scatter"):
            self._init_vars({"line_width": "1.5", "marker_style": "Different Markers",
                             "marker_size": "7"})
        if mode == "box":
            self._init_vars({"notch_box": False})

        g = f; r = 0

        _l = ttk.Label(g, text=label("yscale"), font=("Helvetica Neue", 13, "bold"))
        _l.grid(row=r, column=0, sticky="w", padx=PAD, pady=(12, 2)); tip(_l, "yscale"); r += 1
        PCombobox(g, textvariable=self._vars["yscale"], values=["Linear", "Log"],
                     state="readonly", width=14, font=("Helvetica Neue", 12)
                     ).grid(row=r, column=0, sticky="w", padx=PAD, pady=4); r += 1
        ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=8); r += 1
        section_sep(g, r, "Y Limits"); r += 1

        _l, r = _lbl(g, r, "ylim")
        # ── Y limit radio group ───────────────────────────────────────────────
        _ylim_rg = PRadioGroup(g, variable=self._vars["ylim_mode"],
                               options=["Auto", "Start at data min", "Manual range"],
                               command=self._tog_ylim)
        _ylim_rg.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=4); r += 1

        fr2 = ttk.Frame(g); fr2.grid(row=r, column=0, columnspan=3, sticky="w",
                                     padx=(PAD + 4, PAD), pady=(0, 6)); r += 1
        ttk.Label(fr2, text="from").pack(side="left")
        self._yl_lo = PEntry(fr2, textvariable=self._vars["ylim_lo"],
                             width=10, font=("Menlo", 12))
        self._yl_lo.pack(side="left", padx=6)
        add_placeholder(self._yl_lo, self._vars["ylim_lo"], "e.g. 0")
        ttk.Label(fr2, text="to").pack(side="left")
        self._yl_hi = PEntry(fr2, textvariable=self._vars["ylim_hi"],
                             width=10, font=("Menlo", 12))
        self._yl_hi.pack(side="left", padx=6)
        add_placeholder(self._yl_hi, self._vars["ylim_hi"], "e.g. 100")
        self._tog_ylim()
        ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=8); r += 1

        # ── X Limits (only meaningful for continuous-x charts) ────────────────
        section_sep(g, r, "X Limits"); r += 1
        if _x_continuous:
            _xlim_rg = PRadioGroup(g, variable=self._vars["xlim_mode"],
                                   options=["Auto", "Manual range"],
                                   command=self._tog_xlim)
            _xlim_rg.grid(row=r, column=0, columnspan=3, sticky="w",
                          padx=PAD, pady=4); r += 1
            frx = ttk.Frame(g)
            frx.grid(row=r, column=0, columnspan=3, sticky="w",
                     padx=(PAD + 4, PAD), pady=(0, 6)); r += 1
            ttk.Label(frx, text="from").pack(side="left")
            self._xl_lo = PEntry(frx, textvariable=self._vars["xlim_lo"],
                                 width=10, font=("Menlo", 12))
            self._xl_lo.pack(side="left", padx=6)
            add_placeholder(self._xl_lo, self._vars["xlim_lo"], "e.g. 0")
            ttk.Label(frx, text="to").pack(side="left")
            self._xl_hi = PEntry(frx, textvariable=self._vars["xlim_hi"],
                                 width=10, font=("Menlo", 12))
            self._xl_hi.pack(side="left", padx=6)
            add_placeholder(self._xl_hi, self._vars["xlim_hi"], "e.g. 100")
            self._tog_xlim()
        else:
            ttk.Label(g, text="X axis is categorical  -  limits not applicable",
                      foreground="#aaaaaa", font=("Helvetica Neue", 11)
                      ).grid(row=r, column=0, columnspan=3, sticky="w",
                             padx=PAD, pady=4); r += 1
        ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew",
                              padx=PAD, pady=8); r += 1
        section_sep(g, r, "Figure Size"); r += 1

        _l, r = _lbl(g, r, "figsize")
        fr3 = ttk.Frame(g); fr3.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=4); r += 1
        _e_figw = PEntry(fr3, textvariable=self._vars["figw"], width=6, font=("Menlo", 12))
        _e_figw.pack(side="left")
        ttk.Label(fr3, text="  x  ").pack(side="left")
        _e_figh = PEntry(fr3, textvariable=self._vars["figh"], width=6, font=("Menlo", 12))
        _e_figh.pack(side="left")
        ttk.Label(fr3, text="  inches").pack(side="left")
        ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=8); r += 1

        if mode in ("bar", "grouped_bar", "box"):
            r = self._slider(g, r, "bar_width",  label("bar_width"),  0.1, 1.0, ".2f", 0.6)
            ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=8); r += 1

        if mode == "line":
            r = self._slider(g, r, "line_width", label("line_width"), 0.5, 5.0, ".1f", 1.5)
            ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=8); r += 1

        if mode in ("line", "scatter"):
            _l = ttk.Label(g, text=label("marker_style"), font=("Helvetica Neue", 13, "bold"))
            _l.grid(row=r, column=0, sticky="w", padx=PAD, pady=(4, 2)); tip(_l, "marker_style"); r += 1
            PCombobox(g, textvariable=self._vars["marker_style"],
                         values=["Different Markers", "Circle (o)", "Square (s)", "Triangle (^)",
                                 "Diamond (D)", "Down-Triangle (v)", "Star (*)", "Plus (P)",
                                 "Cross (X)", "Hexagon (h)"],
                         state="readonly", width=22, font=("Menlo", 12)
                         ).grid(row=r, column=0, sticky="w", padx=PAD, pady=4); r += 1
            ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=8); r += 1
            r = self._slider(g, r, "marker_size", label("marker_size"), 2, 20, "d", 7)
            ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=8); r += 1

        section_sep(g, r, "Font"); r += 1
        r = self._slider(g, r, "font_size", label("font_size"), 6, 24, "d", 12)
        ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=8); r += 1
        section_sep(g, r, "Reference Line"); r += 1

        ref_row = ttk.Frame(g)
        ref_row.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=(4, 2)); r += 1
        PCheckbox(ref_row, variable=self._vars["ref_line_enabled"],
                        text=" Show at Y =").pack(side="left")
        _e_ref = PEntry(ref_row, textvariable=self._vars["ref_line_y"],
                           width=8, font=("Menlo", 12))
        _e_ref.pack(side="left", padx=(6, 0))
        add_placeholder(_e_ref, self._vars["ref_line_y"], "e.g. 0")
        ref_lbl_row = ttk.Frame(g)
        ref_lbl_row.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=(0, 4)); r += 1
        ttk.Label(ref_lbl_row, text="Label:", font=("Helvetica Neue", 11)).pack(side="left")
        _e_ref_lbl = PEntry(ref_lbl_row, textvariable=self._vars["ref_line_label"],
                               width=20, font=("Menlo", 12))
        _e_ref_lbl.pack(side="left", padx=(6, 0))
        add_placeholder(_e_ref_lbl, self._vars["ref_line_label"], "blank = show y=…")

        # ── Vertical reference line (where meaningful: line, scatter, curve_fit) ──
        if mode in ("line", "scatter", "curve_fit"):
            self._init_vars({"ref_vline_enabled": False, "ref_vline_x": "0",
                             "ref_vline_label": ""})
            section_sep(g, r, "Vertical Reference Line"); r += 1
            ref_v_row = ttk.Frame(g)
            ref_v_row.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=(4, 2)); r += 1
            PCheckbox(ref_v_row, variable=self._vars["ref_vline_enabled"],
                      text=" Show at X =").pack(side="left")
            _e_vref = PEntry(ref_v_row, textvariable=self._vars["ref_vline_x"],
                             width=8, font=("Menlo", 12))
            _e_vref.pack(side="left", padx=(6, 0))
            add_placeholder(_e_vref, self._vars["ref_vline_x"], "e.g. 50")
            ref_vl_row = ttk.Frame(g)
            ref_vl_row.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=(0, 4)); r += 1
            ttk.Label(ref_vl_row, text="Label:", font=("Helvetica Neue", 11)).pack(side="left")
            _e_vref_lbl = PEntry(ref_vl_row, textvariable=self._vars["ref_vline_label"],
                                 width=20, font=("Menlo", 12))
            _e_vref_lbl.pack(side="left", padx=(6, 0))
            add_placeholder(_e_vref_lbl, self._vars["ref_vline_label"], "blank = show x=…")

        # ── P19: Twin Y-axis (line only) ──────────────────────────────────────
        if mode == "line":
            self._init_vars({"twin_y_series_str": ""})
            section_sep(g, r, "Secondary Y-Axis (P19)"); r += 1
            ttk.Label(g, text="Series on right Y-axis (comma-separated series names):",
                      font=("Helvetica Neue", 12)).grid(row=r, column=0, columnspan=3,
                      sticky="w", padx=PAD); r += 1
            _tw_e = PEntry(g, textvariable=self._vars["twin_y_series_str"], width=38)
            _tw_e.grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=2); r += 1
            add_placeholder(_tw_e, self._vars["twin_y_series_str"], "e.g. Temperature, Pressure")
            r += 1  # spacing

        # ── P18: Custom x-tick labels (categorical charts) ────────────────────
        if mode in ("bar", "box", "violin", "dot_plot", "subcolumn_scatter",
                    "before_after", "repeated_measures"):
            self._init_vars({"xtick_labels_str": ""})
            section_sep(g, r, "Custom Category Labels (P18)"); r += 1
            ttk.Label(g, text="Override labels (comma-separated):",
                      font=("Helvetica Neue", 12)).grid(row=r, column=0, columnspan=3,
                      sticky="w", padx=PAD); r += 1
            _xt_e = PEntry(g, textvariable=self._vars["xtick_labels_str"], width=38)
            _xt_e.grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=2); r += 1
            add_placeholder(_xt_e, self._vars["xtick_labels_str"],
                            "e.g. Control, Drug A, Drug B")
            r += 1

        display_opts = [
            ("open_points", " Open/hollow data points"),
        ]
        if mode == "bar":
            display_opts += [("horizontal", " Horizontal bars"), ("show_median", " Show median line on bar")]
        if mode == "grouped_bar":
            display_opts += [("horizontal", " Horizontal bars")]
        if mode == "box":
            display_opts.append(("notch_box", " Notched box (95% CI on median)"))

        for key, text in display_opts:
            r = self._checkbox(g, r, key, text)

        # Grid style  -  radio group replaces old gridlines checkbox
        self._init_vars({"grid_style": "None"})
        ttk.Label(g, text="Grid Lines", font=("Helvetica Neue", 13, "bold")
                  ).grid(row=r, column=0, sticky="w", padx=PAD, pady=(8, 2)); r += 1
        PRadioGroup(g, variable=self._vars["grid_style"],
                    options=["None", "Horizontal", "Full (H + V)"],
                    ).grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=4); r += 1

        if mode == "bar":
            ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=6); r += 1
            r = self._slider(g, r, "bar_alpha", "Bar Transparency", 0.1, 1.0, ".2f", 0.85)
            ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=6); r += 1
            ttk.Label(g, text="X Scale", font=("Helvetica Neue", 13, "bold")
                      ).grid(row=r, column=0, sticky="w", padx=PAD, pady=(4, 2)); r += 1
            _xscale_row = ttk.Frame(g)
            _xscale_row.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=4); r += 1
            PCombobox(_xscale_row, textvariable=self._vars["xscale"], values=["Linear", "Log"],
                         state="disabled", width=14, font=("Helvetica Neue", 12)
                         ).pack(side="left")
            ttk.Label(_xscale_row, text="  (categorical axis  -  not applicable)",
                      foreground="#aaaaaa", font=("Helvetica Neue", 11)
                      ).pack(side="left")

        # ── Priority-1: Axis Frame / Tick / Point / Cap styling ──────────────
        ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=8); r += 1
        r = self._section_label(g, r, "Axis & Tick Style")

        # Axis Frame
        self._init_vars({"axis_style": "Open (default)"})
        ttk.Label(g, text="Axis Frame", font=("Helvetica Neue", 13, "bold")
                  ).grid(row=r, column=0, sticky="w", padx=PAD, pady=(4, 2)); r += 1
        PCombobox(g, textvariable=self._vars["axis_style"],
                     values=["Open (default)", "Closed box", "Floating", "None"],
                     state="readonly", width=22, font=("Helvetica Neue", 12)
                     ).grid(row=r, column=0, sticky="w", padx=PAD, pady=4); r += 1

        # Background color
        self._init_vars({"fig_bg": "White"})
        ttk.Label(g, text="Background", font=("Helvetica Neue", 13, "bold")
                  ).grid(row=r, column=0, sticky="w", padx=PAD, pady=(8, 2)); r += 1
        PCombobox(g, textvariable=self._vars["fig_bg"],
                     values=["White", "Light gray", "Transparent", "Black"],
                     state="readonly", width=22, font=("Helvetica Neue", 12)
                     ).grid(row=r, column=0, sticky="w", padx=PAD, pady=4); r += 1

        # Tick Direction (radio group)
        self._init_vars({"tick_dir": "Outward (default)"})
        ttk.Label(g, text="Tick Direction", font=("Helvetica Neue", 13, "bold")
                  ).grid(row=r, column=0, sticky="w", padx=PAD, pady=(8, 2)); r += 1
        PRadioGroup(g, variable=self._vars["tick_dir"],
                    options=["Outward (default)", "Inward", "Both", "None"],
                    ).grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=4); r += 1

        # Minor ticks
        self._init_vars({"minor_ticks": False})
        r = self._checkbox(g, r, "minor_ticks", " Show minor tick marks")

        # Spine / axis line width
        self._init_vars({"spine_width": "0.8"})
        r = self._slider(g, r, "spine_width", "Axis Line Width", 0.3, 3.0, ".1f", 0.8)

        ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=8); r += 1
        r = self._section_label(g, r, "Data Points & Error Bars")

        # Point size slider (for chart types that show points)
        if _has_points:
            self._init_vars({"point_size": "6", "point_alpha": "0.80"})
            r = self._slider(g, r, "point_size",  "Point Size (pts)",        2.0, 20.0, ".0f", 6.0)
            ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=4); r += 1
            r = self._slider(g, r, "point_alpha", "Point Transparency",      0.1,  1.0, ".2f", 0.80)
            ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=4); r += 1

        # Cap width slider (for chart types with error bars)
        if _has_error_bars:
            self._init_vars({"cap_size": "4"})
            r = self._slider(g, r, "cap_size", "Error Bar Cap Width", 0.0, 12.0, ".1f", 4.0)
            ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=4); r += 1

        # Legend position (for chart types that have a legend)
        if _has_legend:
            self._init_vars({"legend_pos": "Upper right"})
            ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=8); r += 1
            r = self._section_label(g, r, "Legend")
            ttk.Label(g, text="Legend Position", font=("Helvetica Neue", 13, "bold")
                      ).grid(row=r, column=0, sticky="w", padx=PAD, pady=(4, 2)); r += 1
            PCombobox(g, textvariable=self._vars["legend_pos"],
                         values=["Auto (best fit)", "Upper right", "Upper left",
                                 "Lower right", "Lower left", "Outside right", "None (hidden)"],
                         state="readonly", width=22, font=("Helvetica Neue", 12)
                         ).grid(row=r, column=0, sticky="w", padx=PAD, pady=4); r += 1

        # ── Tick Interval controls ────────────────────────────────────────────
        ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=8); r += 1
        r = self._section_label(g, r, "Tick Intervals")
        self._init_vars({"ytick_interval": "", "xtick_interval": ""})

        _tick_hint = ttk.Label(g, text="Leave blank for automatic ticks.",
                               foreground="#888888", font=("Helvetica Neue", 11))
        _tick_hint.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=(0, 6)); r += 1

        fr_ytick = ttk.Frame(g)
        fr_ytick.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=(0, 4)); r += 1
        ttk.Label(fr_ytick, text="Y-axis interval:", font=("Helvetica Neue", 12)
                  ).pack(side="left")
        _e_ytick = PEntry(fr_ytick, textvariable=self._vars["ytick_interval"],
                          width=10, font=("Menlo", 12))
        _e_ytick.pack(side="left", padx=(8, 0))
        add_placeholder(_e_ytick, self._vars["ytick_interval"], "e.g. 10")

        if _x_continuous:
            fr_xtick = ttk.Frame(g)
            fr_xtick.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=(0, 4)); r += 1
            ttk.Label(fr_xtick, text="X-axis interval:", font=("Helvetica Neue", 12)
                      ).pack(side="left")
            _e_xtick = PEntry(fr_xtick, textvariable=self._vars["xtick_interval"],
                              width=10, font=("Menlo", 12))
            _e_xtick.pack(side="left", padx=(8, 0))
            add_placeholder(_e_xtick, self._vars["xtick_interval"], "e.g. 5")

        g.columnconfigure(0, weight=1)
        ttk.Frame(g).grid(row=r + 1, pady=8)

    def _tab_stats(self, f):
        self._init_vars({
            "show_stats": False, "show_ns": False, "show_p_values": False,
            "show_effect_size": False, "show_test_name": False,
            "show_normality_warning": True,
            "stats_test": "Parametric", "n_permutations": "",
            "mc_correction": "Holm-Bonferroni", "posthoc": "Tukey HSD",
            "control": "", "p_sig_threshold": "0.05",
        })

        try:
            from scipy import stats as _st
            sep = "\n\n"
            TEST_INFO = {
                "Paired":        "Paired t-test (2 groups) or repeated pairwise paired t-tests (3+ groups). Requires equal n per group  -  values are matched by row order.",
                "Parametric":    "2 groups: " + _scipy_summary(_st.ttest_ind, 160) + sep +
                                 "3+ groups: " + _scipy_summary(_st.f_oneway, 160),
                "Non-parametric":"2 groups: " + _scipy_summary(_st.mannwhitneyu, 160) + sep +
                                 "3+ groups: " + _scipy_summary(_st.kruskal, 180),
                "Permutation":   _scipy_summary(_st.permutation_test, 320),
            }
        except Exception:
            _log.debug("App._tab_stats: scipy TEST_INFO build failed", exc_info=True)
            TEST_INFO = {
                "Paired":         "Paired t-test (matched observations, same n per group).",
                "Parametric":     "Welch's t-test (2 groups, unequal variance OK) or one-way ANOVA + Tukey HSD (3+ groups).",
                "Non-parametric": "Mann-Whitney U (2 groups) or Kruskal-Wallis + Dunn's (3+ groups).",
                "Permutation":    "Permutation test on difference of means + Holm-Bonferroni.",
            }

        g = f; r = 0

        r = self._checkbox(g, r, "show_stats",       f" {label('show_stats')}")
        r = self._sep(g, r)
        r = self._checkbox(g, r, "show_ns",          " Show 'ns' brackets (non-significant)")
        r = self._sep(g, r)

        # ── Significance threshold ─────────────────────────────────────────────
        r = self._section_label(g, r, "Significance Threshold")
        thresh_row = ttk.Frame(g)
        thresh_row.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=(2, 0)); r += 1
        ttk.Label(thresh_row, text="Show brackets only when p ≤",
                  font=("Helvetica Neue", 12)).pack(side="left")
        _thresh_cb = PCombobox(thresh_row, textvariable=self._vars["p_sig_threshold"],
                               values=["0.05", "0.01", "0.001", "0.0001"],
                               state="normal", width=8, font=("Menlo", 12))
        _thresh_cb.pack(side="left", padx=(8, 0))
        add_placeholder(_thresh_cb, self._vars["p_sig_threshold"], "0.05")
        r = self._hint(g, r, "Default 0.05 (one star). Set to 0.01 to show only ** *** **** "
                             "brackets, hiding borderline * results. Matches Prism's α setting. "
                             "Type any value or choose from the list.")
        r = self._sep(g, r)

        # ── Bracket style (P16) ────────────────────────────────────────────────
        r = self._section_label(g, r, "Bracket Style")
        _bs_row = ttk.Frame(g)
        _bs_row.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=(2, 4)); r += 1
        if "bracket_style" not in self._vars:
            self._vars["bracket_style"] = tk.StringVar(value="Lines")
        for _bs_val, _bs_lbl in (("Lines", "Lines (default)"),
                                  ("Bracket", "Square brackets"),
                                  ("Asterisks only", "Asterisks only")):
            tk.Radiobutton(_bs_row, text=_bs_lbl,
                           variable=self._vars["bracket_style"],
                           value=_bs_val,
                           font=("Helvetica Neue", 12),
                           bg="#ffffff" if hasattr(self, "_bg") else None
                           ).pack(side="left", padx=(0, 10))
        r = self._sep(g, r)

        r = self._checkbox(g, r, "show_p_values",    " Show raw p-values (instead of stars)")
        r = self._sep(g, r)
        r = self._checkbox(g, r, "show_effect_size", " Show effect size (Cohen's d)")
        r = self._sep(g, r)
        r = self._checkbox(g, r, "show_test_name",   " Show test name below plot")
        r = self._sep(g, r)

        r = self._section_label(g, r, "Normality Check")
        r = self._checkbox(g, r, "show_normality_warning", " Show normality warning on plot")
        r = self._hint(g, r, "When Show Significance Brackets is on, a Shapiro-Wilk test "
                             "is run automatically. An orange warning appears on the plot "
                             "if any group fails normality (p\u22640.05) and Parametric is selected.")
        r = self._sep(g, r)

        _l, r = _lbl(g, r, "stats_test")
        cb2 = PCombobox(g, textvariable=self._vars["stats_test"],
                           values=["Paired", "Parametric", "Non-parametric", "Permutation",
                                   "One-sample"],
                           state="readonly", width=22, font=("Helvetica Neue", 12))
        cb2.grid(row=r, column=0, sticky="w", padx=PAD, pady=4); r += 1

        self._test_info_lbl = ttk.Label(g, text=TEST_INFO.get("Parametric", ""),
                                        foreground="gray", wraplength=520,
                                        justify="left", font=("Helvetica Neue", 11))
        self._test_info_lbl.grid(row=r, column=0, columnspan=3, sticky="ew",
                                 padx=PAD, pady=(0, 4)); r += 1
        self._test_info_lbl.bind("<Configure>",
            lambda e: self._test_info_lbl.config(wraplength=max(200, e.width - 8)))
        self._TEST_INFO = TEST_INFO

        # One-sample μ₀ field (shown only when One-sample is selected)
        self._init_vars({"one_sample_mu0": "0"})
        _mu0_frame = ttk.Frame(g)
        _mu0_frame.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=(0, 4))
        ttk.Label(_mu0_frame, text="Null hypothesis μ₀ =",
                  font=("Helvetica Neue", 12)).pack(side="left")
        _mu0_entry = PEntry(_mu0_frame, textvariable=self._vars["one_sample_mu0"],
                            width=10, font=("Menlo", 12))
        _mu0_entry.pack(side="left", padx=(6, 0))
        add_placeholder(_mu0_entry, self._vars["one_sample_mu0"], "e.g. 0")
        self._mu0_frame = _mu0_frame
        r += 1

        def _on_test_change(e=None):
            test = self._vars["stats_test"].get()
            self._test_info_lbl.config(text=TEST_INFO.get(test, ""))
            self._tog_perm()
            self._tog_posthoc()
            self._tog_control()
            # Show/hide μ₀ field
            if test == "One-sample":
                self._mu0_frame.grid()
            else:
                self._mu0_frame.grid_remove()
        cb2.bind("<<ComboboxSelected>>", _on_test_change)
        # Initialise visibility
        if self._get_var("stats_test", "Parametric") != "One-sample":
            _mu0_frame.grid_remove()
        r = self._sep(g, r)

        # Post-hoc (shown/hidden by _tog_posthoc)
        self._posthoc_label = ttk.Label(g, text=label("posthoc"),
                                        font=("Helvetica Neue", 13, "bold"))
        self._posthoc_cb    = PCombobox(g, textvariable=self._vars["posthoc"],
                                           values=["Tukey HSD", "Bonferroni", "Sidak",
                                                   "Fisher LSD", "Dunnett (vs control)"],
                                           state="readonly", width=26,
                                           font=("Helvetica Neue", 12))
        self._posthoc_sep   = ttk.Separator(g)
        # Non-parametric hint: Dunn's test is always used, no selector needed
        self._posthoc_dunn_hint = ttk.Label(
            g,
            text="Post-hoc: Dunn's test  (Kruskal-Wallis pairwise rank-sum comparisons)\n"
                 "Correction method is controlled by the Multiple Comparisons setting below.",
            foreground="#555555", wraplength=480, justify="left",
            font=("Helvetica Neue", 11))
        self._posthoc_r     = r; r += 3

        # Wire posthoc change to _tog_control so Dunnett activates the control selector
        self._posthoc_cb.bind("<<ComboboxSelected>>",
                              lambda e: (self._tog_control(), self._tog_posthoc()))

        _l, r = _lbl(g, r, "mc_correction")
        PCombobox(g, textvariable=self._vars["mc_correction"],
                     values=["Holm-Bonferroni", "Bonferroni",
                             "Benjamini-Hochberg (FDR)", "None (uncorrected)"],
                     state="readonly", width=26, font=("Helvetica Neue", 12)
                     ).grid(row=r, column=0, sticky="w", padx=PAD, pady=4); r += 1
        r = self._sep(g, r)

        # Permutation entry (shown/hidden by _tog_perm)
        self._perm_label   = ttk.Label(g, text=label("n_permutations"),
                                       font=("Helvetica Neue", 13, "bold"))
        self._perm_hint    = ttk.Label(g, text=hint("n_permutations"))
        self._perm_entry   = PEntry(g, textvariable=self._vars["n_permutations"],
                                       width=14, font=("Menlo", 12))
        add_placeholder(self._perm_entry, self._vars["n_permutations"], "e.g. 9999")
        self._perm_sep     = ttk.Separator(g)
        self._perm_r_start = r; r += 4

        # ── Comparison Mode  -  Prism-style: all-pairs vs. vs-control ──────
        r = self._section_label(g, r, "Comparison Mode")
        self._init_vars({"comparison_mode": 0})  # 0=all-pairwise, 1=vs-control
        _cmode_rg = PRadioGroup(g, variable=self._vars["comparison_mode"],
                                options=["All pairwise", "Each group vs. control"],
                                command=self._tog_control)
        _cmode_rg.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=4)
        # Store ref so _tog_control can disable it when the mode is forced
        self._cmode_rg_widget = _cmode_rg
        r += 1
        r = self._hint(g, r,
                       "All pairwise: every group compared to every other group.\n"
                       "Each group vs. control: only brackets between the selected "
                       "control and each other group. Use Dunnett for the most "
                       "statistically appropriate control comparison.")
        r = self._sep(g, r)

        # Control group dropdown
        r = self._section_label(g, r, "Control Group")
        control_cb = PCombobox(g, textvariable=self._vars["control"],
                                  values=["(none \u2014 all pairwise)"],
                                  state="disabled", width=28, font=("Menlo", 12))
        control_cb.grid(row=r, column=0, columnspan=2, sticky="w", padx=PAD, pady=4)
        tip(control_cb, "control")
        control_hint_lbl = ttk.Label(g,
                                     text="Load a file to populate group names",
                                     foreground="gray", font=("Helvetica Neue", 10),
                                     wraplength=300)
        control_hint_lbl.grid(row=r+1, column=0, columnspan=3, sticky="w",
                               padx=PAD, pady=(0, 4))
        r += 2

        if not hasattr(self, "_control_cbs"):
            self._control_cbs       = []
            self._control_hint_lbls = []
        self._control_cbs.append(control_cb)
        self._control_hint_lbls.append(control_hint_lbl)
        self._control_cb       = control_cb
        self._control_hint_lbl = control_hint_lbl

        # Wire comparison_mode changes to toggle the control dropdown state
        self._vars["comparison_mode"].trace_add("write",
            lambda *_: self._tog_control())

        self._perm_visible    = False
        self._posthoc_visible = False
        self._tog_perm()
        self._tog_posthoc()
        self._tog_control()
        self._tab_footer(g, r)

    def _tab_stats_grouped_bar(self, f):
        """Stats tab for Grouped Bar  -  standard pairwise options plus per-category ANOVA."""
        self._init_vars({"show_anova_per_group": False})
        # Render the full shared stats UI (uses grid on f)
        self._tab_stats(f)
        # Append the ANOVA section using grid, after whatever rows _tab_stats used.
        # grid_size() returns (cols, rows)  -  rows is the count of occupied rows.
        r = f.grid_size()[1]
        section_sep(f, r, "Per-Category One-Way ANOVA"); r += 1
        PCheckbox(f,
                  variable=self._vars["show_anova_per_group"],
                  text=" Run one-way ANOVA on each category cluster"
                  ).grid(row=r, column=0, columnspan=3, sticky="w",
                         padx=PAD, pady=(0, 4)); r += 1
        ttk.Label(f,
                  text="Shows F-statistic and p-value above each category cluster, "
                       "testing whether the subgroups within that cluster differ. "
                       "Independent of the pairwise brackets above.",
                  foreground="gray", wraplength=380, justify="left",
                  font=("Helvetica Neue", 11)
                  ).grid(row=r, column=0, columnspan=3, sticky="w",
                         padx=PAD, pady=(0, 8))

    def _tab_stats_scatter(self, f):
        """Stats tab specific to scatter plot  -  correlation + regression options."""
        self._init_vars({
            "show_regression": False, "show_ci_band": False,
            "show_prediction_band": False, "show_correlation": False,
            "correlation_type": "Pearson", "show_regression_table": False,
        })

        g = f; r = 0

        r = self._section_label(g, r, "Regression")
        r = self._checkbox(g, r, "show_regression",       " Show linear regression line")
        r = self._checkbox(g, r, "show_ci_band",          " Show 95% confidence band")
        r = self._checkbox(g, r, "show_prediction_band",  " Show 95% prediction band")
        r = self._checkbox(g, r, "show_regression_table", " Show full regression table (slope, F, runs test)")
        r = self._hint(g, r, "Full table: slope ± SE, 95% CI, intercept ± SE, "
                             "F-statistic, R, R², adjusted R², runs test for linearity. "
                             "Single-series only.")
        r = self._sep(g, r)

        r = self._section_label(g, r, "Correlation")
        r = self._checkbox(g, r, "show_correlation", " Show correlation annotation")

        ttk.Label(g, text="Correlation Type",
                  font=("Helvetica Neue", 13, "bold")
                  ).grid(row=r, column=0, sticky="w", padx=PAD, pady=(4, 2)); r += 1
        PCombobox(g, textvariable=self._vars["correlation_type"],
                     values=["Pearson", "Spearman"],
                     state="readonly", width=18, font=("Helvetica Neue", 12)
                     ).grid(row=r, column=0, sticky="w", padx=PAD, pady=4); r += 1
        r = self._hint(g, r, "Pearson: linear correlation (assumes normality).\n"
                             "Spearman: rank-based, non-parametric.")
        r = self._sep(g, r)

        r = self._section_label(g, r, "Reference Line")
        ref_row = ttk.Frame(g)
        ref_row.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=(4, 2)); r += 1
        PCheckbox(ref_row, variable=self._vars["ref_line_enabled"],
                        text=" Show at Y =").pack(side="left")
        _e_ref = PEntry(ref_row, textvariable=self._vars["ref_line_y"],
                           width=8, font=("Menlo", 12))
        _e_ref.pack(side="left", padx=(6, 0))
        add_placeholder(_e_ref, self._vars["ref_line_y"], "e.g. 0")
        ref_lbl_row = ttk.Frame(g)
        ref_lbl_row.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=(0, 4)); r += 1
        ttk.Label(ref_lbl_row, text="Label:", font=("Helvetica Neue", 11)).pack(side="left")
        _e_ref_lbl = PEntry(ref_lbl_row, textvariable=self._vars["ref_line_label"],
                               width=20, font=("Menlo", 12))
        _e_ref_lbl.pack(side="left", padx=(6, 0))
        add_placeholder(_e_ref_lbl, self._vars["ref_line_label"], "blank = show y=…")

        self._tab_footer(g, r)

    # ── Helpers ───────────────────────────────────────────────────────────────

    # ── UI widget helpers  -  eliminate repeated boilerplate in tab builders ────

    def _get_float(self, key: str, default: float) -> float:
        """Get a float from a tkvar, returning default on any conversion failure."""
        try:
            return float(self._get_var(key, str(default)))
        except (ValueError, TypeError):
            return default

    def _slider(self, g, r: int, var_key: str, label_text: str,
                lo: float, hi: float, fmt: str, default) -> int:
        """Grid a label + Scale + live value label row; return next row index.

        fmt   -  Python format spec for the value label, e.g. '.2f', '.1f', 'd'
        """
        ttk.Label(g, text=label_text, font=("Helvetica Neue", 13, "bold")
                  ).grid(row=r, column=0, sticky="w", padx=PAD, pady=(4, 2))
        r += 1
        row_fr = ttk.Frame(g)
        row_fr.grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=4)
        r += 1
        val_lbl = ttk.Label(row_fr, text=format(default, fmt), width=5, anchor="e")
        val_lbl.pack(side="right")

        def _on_change(raw, _vk=var_key, _fmt=fmt, _lbl=val_lbl):
            v = round(float(raw), 6)
            self._vars[_vk].set(str(v))
            _lbl.config(text=format(v, _fmt))

        ttk.Scale(row_fr, from_=lo, to=hi, orient="horizontal",
                  variable=self._vars[var_key], command=_on_change
                  ).pack(side="left", fill="x", expand=True, padx=(0, 8))
        return r

    def _init_vars(self, defaults: dict):
        """Ensure each key exists in self._vars; create it from its default if not.
        Booleans  ->  BooleanVar, ints  ->  IntVar, everything else  ->  StringVar."""
        for k, v in defaults.items():
            if k not in self._vars:
                if isinstance(v, bool):
                    self._vars[k] = tk.BooleanVar(value=v)
                elif isinstance(v, int):
                    self._vars[k] = tk.IntVar(value=v)
                else:
                    self._vars[k] = tk.StringVar(value=str(v))

    def _sep(self, g, r: int) -> int:
        """Grid a horizontal separator and return the next row index."""
        ttk.Separator(g).grid(row=r, column=0, columnspan=3,
                              sticky="ew", padx=PAD, pady=8)
        return r + 1

    def _section_label(self, g, r: int, text: str) -> int:
        """Grid a styled section-band header and return the next row index."""
        outer = tk.Frame(g, bg="#edf2f8")
        outer.grid(row=r, column=0, columnspan=3, sticky="ew", padx=0, pady=(14, 2))
        tk.Frame(outer, bg="#2274A5", width=3).pack(side="left", fill="y", padx=(0, 9))
        tk.Label(outer, text=text.upper(), bg="#edf2f8", fg="#2274A5",
                 font=("Helvetica Neue", 9, "bold"), anchor="w").pack(side="left", pady=5)
        return r + 1

    def _checkbox(self, g, r: int, key: str, text: str) -> int:
        """Grid a single PCheckbox row and return the next row index.
        Auto-creates a BooleanVar(False) if the key is not yet in self._vars,
        so tab-build order cannot cause a KeyError.
        """
        if key not in self._vars:
            self._vars[key] = tk.BooleanVar(value=False)
        cb = PCheckbox(g, variable=self._vars[key], text=text)
        cb.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=(0, 4))
        return r + 1

    def _hint(self, g, r: int, text: str) -> int:
        """Grid a small gray explanatory label and return the next row index."""
        ttk.Label(g, text=text,
                  foreground="gray", wraplength=380, justify="left",
                  font=("Helvetica Neue", 11)
                  ).grid(row=r, column=0, columnspan=3, sticky="w",
                         padx=PAD, pady=(0, 8))
        return r + 1

    def _tab_footer(self, g, r: int):
        """Finish a tab frame: configure column weight and add bottom spacer."""
        g.columnconfigure(0, weight=1)
        ttk.Frame(g).grid(row=r + 1, pady=10)

    # ── Toggle helpers ─────────────────────────────────────────────────────────

    def _tog_ylim(self):
        """Enable the manual from/to fields only when mode==2 (Manual range)."""
        manual = (self._get_var("ylim_mode", 0) == 2)
        s = ["!disabled"] if manual else ["disabled"]
        if hasattr(self, "_yl_lo"):
            self._yl_lo.state(s); self._yl_hi.state(s)

    def _tog_xlim(self):
        """Enable the x manual from/to fields only when xlim_mode==1."""
        manual = (self._get_var("xlim_mode", 0) == 1)
        s = ["!disabled"] if manual else ["disabled"]
        if hasattr(self, "_xl_lo"):
            self._xl_lo.state(s); self._xl_hi.state(s)

    def _tog_posthoc(self):
        """Show post-hoc selector for Parametric only.
        Non-parametric always uses Dunn's test  -  no choice needed, just show a hint.
        Paired and Permutation don't use post-hoc selectors."""
        test = self._vars["stats_test"].get()
        show_parametric = (test == "Parametric")
        show_nonparam   = (test == "Non-parametric")
        r = self._posthoc_r
        if show_parametric and not self._posthoc_visible:
            self._posthoc_label.grid(row=r,   column=0, sticky="w",  padx=PAD, pady=(4, 2))
            self._posthoc_cb   .grid(row=r+1, column=0, sticky="w",  padx=PAD, pady=4)
            self._posthoc_sep  .grid(row=r+2, column=0, columnspan=3, sticky="ew", padx=PAD, pady=6)
            if hasattr(self, "_posthoc_dunn_hint"):
                self._posthoc_dunn_hint.grid_remove()
            self._posthoc_visible = True
        elif not show_parametric and self._posthoc_visible:
            self._posthoc_label.grid_remove()
            self._posthoc_cb   .grid_remove()
            self._posthoc_sep  .grid_remove()
            self._posthoc_visible = False

        # Non-parametric: always Dunn's test  -  show informational hint only
        if show_nonparam:
            if hasattr(self, "_posthoc_dunn_hint"):
                self._posthoc_dunn_hint.grid(row=r, column=0, columnspan=3,
                                             sticky="w", padx=PAD, pady=(4, 6))
                self._posthoc_sep.grid(row=r+2, column=0, columnspan=3,
                                       sticky="ew", padx=PAD, pady=6)
        else:
            if hasattr(self, "_posthoc_dunn_hint"):
                self._posthoc_dunn_hint.grid_remove()
            # Only hide sep when not parametric either
            if not show_parametric:
                self._posthoc_sep.grid_remove()

    def _tog_control(self):
        """Synchronize the Comparison Mode radio buttons and Control Group dropdown
        so every visible state is consistent and logically valid.

        Consistency rules (mirrors GraphPad Prism behaviour):

        HIDDEN entirely (comparison_mode radio + control group section invisible):
          - One-sample test  ->  each group vs μ₀, no group-to-group comparison
          - Paired test       ->  direct pair comparison, no control concept
          - Only 1 or 2 groups loaded  ->  no choice needed (single pair or nothing)

        FORCED "vs control" (radio locked, required):
          - Dunnett (vs control) posthoc selected  ->  inherently a control test

        FORCED "all pairwise" (radio + dropdown both disabled):
          - Non-parametric with k>2  ->  Dunn's test is always all-pairwise
          - Permutation test          ->  all-pairwise only implementation

        NORMAL (radio choosable, dropdown enabled when mode=vs-control):
          - Parametric, k>=3, non-Dunnett posthoc

        DROPDOWN CONTENT:
          - Mode=all-pairwise  ->  dropdown disabled, value cleared to ""
          - Mode=vs-control, groups loaded  ->  shows ONLY real group names
            (no "(none)" sentinel — that would be a contradictory state)
          - No file loaded  ->  disabled with "Load a file…" hint
        """
        if not hasattr(self, "_control_cbs"):
            return

        # ── Gather current state ──────────────────────────────────────────────
        test       = self._get_var("stats_test", "Parametric")
        posthoc    = self._get_var("posthoc", "Tukey HSD")
        is_dunnett = (posthoc == "Dunnett (vs control)")
        cmode      = self._get_var("comparison_mode", 0)  # 0=all-pairs, 1=vs-ctrl

        # Extract actual group names from the first control combobox
        # The list was populated by _populate_control_dropdown during validation.
        raw_values = []
        try:
            raw_values = list(self._control_cbs[0].cget("values"))
        except Exception:
            pass
        # Filter out any sentinel "(none…)" entries to get real group names
        group_names = [v for v in raw_values if v and not v.startswith("(none")]
        n_groups    = len(group_names)
        has_groups  = n_groups > 0

        # ── Decide visibility / enabled state ─────────────────────────────────
        # Tests where comparison mode + control make no sense at all
        hide_section = test in ("One-sample", "Paired")

        # 2-group case: nothing to choose — always a single pair
        if has_groups and n_groups <= 2:
            hide_section = True

        # Tests that are all-pairwise-only (no control comparison implemented)
        force_pairwise = (test in ("Permutation",) or
                          (test == "Non-parametric" and n_groups > 2))

        # Dunnett forces vs-control mode
        if is_dunnett and not hide_section:
            force_pairwise = False
            # Silently set comparison_mode to 1 (vs control) if it isn't already
            if self._get_var("comparison_mode", 0) != 1:
                try:
                    self._vars["comparison_mode"].set(1)
                    cmode = 1
                except Exception:
                    pass

        # ── Apply visibility to the comparison mode radio group ───────────────
        if hasattr(self, "_cmode_rg_widget"):
            try:
                if hide_section or force_pairwise:
                    self._cmode_rg_widget.config(state="disabled")
                else:
                    self._cmode_rg_widget.config(state="normal")
            except Exception:
                pass

        # ── Apply state to control dropdown ───────────────────────────────────
        # Decide whether the dropdown should be enabled
        want_control_enabled = (
            not hide_section
            and not force_pairwise
            and has_groups
            and cmode == 1       # "vs control" mode selected
        )

        for cb in self._control_cbs:
            try:
                if want_control_enabled:
                    # Populate with ONLY real group names (no sentinel)
                    cb.config(values=group_names, state="readonly")
                    # If current value is a sentinel or empty, auto-select first group
                    cur = self._vars["control"].get()
                    if not cur or cur.startswith("(none") or cur not in group_names:
                        self._vars["control"].set(group_names[0])
                else:
                    # Clear to empty and disable
                    cb.config(values=group_names if group_names else [""], state="disabled")
                    if not want_control_enabled:
                        self._vars["control"].set("")
            except Exception:
                pass

        # ── Update hint label ─────────────────────────────────────────────────
        for lbl in getattr(self, "_control_hint_lbls", []):
            try:
                if hide_section:
                    lbl.config(
                        text="Not applicable for this test type.",
                        foreground="#aaaaaa")
                elif force_pairwise:
                    lbl.config(
                        text="This test always compares all groups pairwise.",
                        foreground="#aaaaaa")
                elif not has_groups:
                    lbl.config(text="Load a file to populate group names.",
                               foreground="#888888")
                elif cmode == 0:  # all pairwise
                    lbl.config(
                        text=f"{n_groups} groups — all pairwise brackets.",
                        foreground="#666666")
                elif want_control_enabled:
                    cur = self._vars["control"].get()
                    if is_dunnett:
                        lbl.config(
                            text=f"Dunnett test: comparing all groups vs. '{cur}'  "
                                 f"(control required).",
                            foreground="#2274A5")
                    else:
                        lbl.config(
                            text=f"Showing brackets vs. '{cur}' only.  "
                                 f"Other pairs hidden.",
                            foreground="#2274A5")
            except Exception:
                pass

        # Re-fire when control selection changes so hint stays live
        for cb in self._control_cbs:
            try:
                cb.bind("<<ComboboxSelected>>",
                        lambda e: self._tog_control(), add="+")
            except Exception:
                pass

    def _tog_perm(self):
        is_perm = self._vars["stats_test"].get() == "Permutation"
        r = self._perm_r_start
        if is_perm and not self._perm_visible:
            self._perm_label.grid(row=r,   column=0, sticky="w",  padx=PAD, pady=(4, 2))
            self._perm_hint .grid(row=r+1, column=0, columnspan=3, sticky="w", padx=PAD)
            self._perm_entry.grid(row=r+2, column=0, sticky="w",  padx=PAD, pady=4)
            self._perm_sep  .grid(row=r+3, column=0, columnspan=3, sticky="ew", padx=PAD, pady=6)
            self._perm_visible = True
        elif not is_perm and self._perm_visible:
            for w in (self._perm_label, self._perm_hint, self._perm_entry, self._perm_sep):
                w.grid_remove()
            self._perm_visible = False

    # ── File loading ──────────────────────────────────────────────────────────

    def _browse_file(self):
        path = filedialog.askopenfilename(
            title="Select Excel file",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")])
        if path:
            self._vars["excel_path"].set(path)
            self._load_sheets(path)

    def _download_template(self, mode):
        """Write an example Excel template for the given plot type to the Desktop."""
        import os
        from datetime import datetime
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            messagebox.showerror("Missing package", "openpyxl is required to create templates.")
            return

        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
        fname   = f"claude_plotter_template_{mode}_{ts}.xlsx"
        path    = os.path.join(desktop, fname)

        wb = openpyxl.Workbook()
        ws = wb.active

        # Styles
        hdr_font    = Font(bold=True, color="FFFFFF")
        hdr_fill    = PatternFill("solid", fgColor="4472C4")
        sub_fill    = PatternFill("solid", fgColor="8EAADB")
        hint_font   = Font(italic=True, color="888888", size=9)
        center      = Alignment(horizontal="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"))

        def hdr(ws, row, col, val):
            c = ws.cell(row=row, column=col, value=val)
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = center; c.border = thin_border

        def sub(ws, row, col, val):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(bold=True); c.fill = sub_fill
            c.alignment = center; c.border = thin_border

        def dat(ws, row, col, val):
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = center; c.border = thin_border

        def hint(ws, row, col, val):
            c = ws.cell(row=row, column=col, value=val)
            c.font = hint_font

        if mode in ("bar", "box"):
            ws.title = "Bar_Box Template"
            hint(ws, 1, 1, "Row 1: group names | Rows 2+: numeric replicates")
            for ci, grp in enumerate(["Control", "Treatment A", "Treatment B"], 1):
                hdr(ws, 2, ci, grp)
            data = [[2.1, 3.4, 4.2], [1.9, 3.8, 4.6], [2.3, 3.1, 3.9],
                    [2.0, 3.5, 4.0], [2.2, 3.6, 4.4]]
            for ri, row_vals in enumerate(data, 3):
                for ci, v in enumerate(row_vals, 1):
                    dat(ws, ri, ci, v)
            for ci in range(1, 4): ws.column_dimensions[chr(64+ci)].width = 14

        elif mode == "grouped_bar":
            ws.title = "Grouped Bar Template"
            hint(ws, 1, 1, "Row 1: category names | Row 2: subgroup names | Rows 3+: replicates")
            cats = ["Condition A", "Condition A", "Condition B", "Condition B"]
            subs = ["Male", "Female", "Male", "Female"]
            for ci, (cat, s) in enumerate(zip(cats, subs), 1):
                hdr(ws, 2, ci, cat); sub(ws, 3, ci, s)
            data = [[2.1, 2.8, 3.4, 4.1], [1.9, 2.6, 3.6, 3.9],
                    [2.3, 2.9, 3.1, 4.3], [2.0, 2.7, 3.5, 4.0]]
            for ri, row_vals in enumerate(data, 4):
                for ci, v in enumerate(row_vals, 1):
                    dat(ws, ri, ci, v)
            for ci in range(1, 5): ws.column_dimensions[chr(64+ci)].width = 14

        elif mode in ("line", "scatter"):
            ws.title = "Line_Scatter Template"
            hint(ws, 1, 1, "Row 1: col 1=X label, cols 2+ = series names (repeat per replicate) | Rows 2+: X value then Y replicates")
            headers = ["Time (h)", "Control", "Control", "Control", "Drug", "Drug", "Drug"]
            for ci, h in enumerate(headers, 1):
                hdr(ws, 2, ci, h)
            data = [
                [0,  1.0, 1.1, 0.9,  1.0, 1.2, 0.8],
                [6,  1.5, 1.4, 1.6,  2.1, 2.3, 1.9],
                [12, 2.0, 2.2, 1.9,  3.4, 3.6, 3.2],
                [24, 2.8, 2.7, 3.0,  5.1, 4.9, 5.3],
                [48, 3.1, 3.3, 2.9,  6.2, 6.0, 6.4],
            ]
            for ri, row_vals in enumerate(data, 3):
                for ci, v in enumerate(row_vals, 1):
                    dat(ws, ri, ci, v)
            ws.column_dimensions["A"].width = 12
            for ci in range(2, 8): ws.column_dimensions[chr(64+ci)].width = 10

        elif mode == "before_after":
            ws.title = "Before_After Template"
            hint(ws, 1, 1, "Row 1: condition names | Rows 2+: one subject per row (matched by row order)")
            for ci, grp in enumerate(["Before", "After"], 1):
                hdr(ws, 2, ci, grp)
            data = [[2.1, 3.4], [1.9, 3.8], [2.3, 2.9], [2.0, 3.5], [2.2, 4.1],
                    [1.8, 3.2], [2.5, 4.4], [2.1, 3.7]]
            for ri, row_vals in enumerate(data, 3):
                for ci, v in enumerate(row_vals, 1):
                    dat(ws, ri, ci, v)
            for ci in range(1, 3): ws.column_dimensions[chr(64+ci)].width = 12

        elif mode == "histogram":
            ws.title = "Histogram Template"
            hint(ws, 1, 1, "Row 1: series names | Rows 2+: raw numeric values (columns may have different lengths)")
            for ci, grp in enumerate(["Group A", "Group B"], 1):
                hdr(ws, 2, ci, grp)
            import random; random.seed(42)
            a_vals = [round(random.gauss(5, 1), 2) for _ in range(30)]
            b_vals = [round(random.gauss(6.5, 1.2), 2) for _ in range(30)]
            for ri, (a, b) in enumerate(zip(a_vals, b_vals), 3):
                dat(ws, ri, 1, a); dat(ws, ri, 2, b)
            for ci in range(1, 3): ws.column_dimensions[chr(64+ci)].width = 12

        elif mode in ("subcolumn_scatter", "violin"):
            ws.title = "Subcolumn_Violin Template"
            hint(ws, 1, 1, "Row 1: group names | Rows 2+: numeric replicates")
            for ci, grp in enumerate(["Control", "Low Dose", "High Dose"], 1):
                hdr(ws, 2, ci, grp)
            data = [[2.1, 3.4, 4.8], [1.9, 3.1, 5.2], [2.4, 3.7, 4.5],
                    [2.0, 2.9, 5.0], [2.3, 3.5, 4.7], [1.8, 3.2, 5.3]]
            for ri, row_vals in enumerate(data, 3):
                for ci, v in enumerate(row_vals, 1):
                    dat(ws, ri, ci, v)
            for ci in range(1, 4): ws.column_dimensions[chr(64+ci)].width = 14

        elif mode == "kaplan_meier":
            ws.title = "Survival Template"
            hint(ws, 1, 1, "Row 1: group names (each spans 2 cols) | Row 2: 'Time' / 'Event' | Rows 3+: time, event (1=occurred, 0=censored)")
            groups_km = ["Control", "Treatment"]
            for gi, grp in enumerate(groups_km):
                c1 = gi * 2 + 1
                hdr(ws, 2, c1, grp); hdr(ws, 2, c1 + 1, grp)
                sub(ws, 3, c1, "Time"); sub(ws, 3, c1 + 1, "Event")
            km_data = [
                [5,  1,  3,  1],
                [8,  0,  6,  1],
                [10, 1,  8,  0],
                [12, 1, 10,  1],
                [15, 0, 12,  1],
                [18, 1, 14,  0],
                [20, 1, 16,  1],
                [22, 0, 18,  1],
            ]
            for ri, row_vals in enumerate(km_data, 4):
                for ci, v in enumerate(row_vals, 1):
                    dat(ws, ri, ci, v)
            for ci in range(1, 5): ws.column_dimensions[chr(64+ci)].width = 12

        elif mode == "heatmap":
            ws.title = "Heatmap Template"
            hint(ws, 1, 1, "Row 1: blank in A1, then column labels | Rows 2+: row label in col A, then numeric values")
            col_labels = ["Sample 1", "Sample 2", "Sample 3", "Sample 4"]
            row_labels = ["Gene A", "Gene B", "Gene C", "Gene D", "Gene E"]
            for ci, lbl in enumerate(col_labels, 2):
                hdr(ws, 2, ci, lbl)
            import random; random.seed(7)
            for ri, rlbl in enumerate(row_labels, 3):
                sub(ws, ri, 1, rlbl)
                for ci in range(2, 6):
                    dat(ws, ri, ci, round(random.gauss(0, 2), 3))
            ws.column_dimensions["A"].width = 12
            for ci in range(2, 6): ws.column_dimensions[chr(64+ci)].width = 11

        elif mode == "two_way_anova":
            ws.title = "Two-Way ANOVA Template"
            hint(ws, 1, 1, "Row 1: column headers | Rows 2+: one observation per row (long format)")
            for ci, h in enumerate(["Factor_A", "Factor_B", "Value"], 1):
                hdr(ws, 2, ci, h)
            twa_data = [
                ["Drug",    "Male",   3.2], ["Drug",    "Male",   2.9], ["Drug",    "Male",   3.5],
                ["Drug",    "Female", 4.1], ["Drug",    "Female", 3.8], ["Drug",    "Female", 4.4],
                ["Control", "Male",   1.8], ["Control", "Male",   2.1], ["Control", "Male",   1.6],
                ["Control", "Female", 2.3], ["Control", "Female", 2.0], ["Control", "Female", 2.5],
            ]
            for ri, row_vals in enumerate(twa_data, 3):
                for ci, v in enumerate(row_vals, 1):
                    dat(ws, ri, ci, v)
            for ci in range(1, 4): ws.column_dimensions[chr(64+ci)].width = 14

        elif mode == "curve_fit":
            ws.title = "Curve Fit Template"
            hint(ws, 1, 1, "Row 1: col 1=X label, cols 2+ = series names (repeat per replicate) | Rows 2+: X value then Y replicates")
            headers = ["Concentration", "Control", "Control", "Control", "Drug", "Drug", "Drug"]
            for ci, h in enumerate(headers, 1):
                hdr(ws, 2, ci, h)
            cf_data = [
                [0.001, 0.02, 0.03, 0.01,  0.03, 0.02, 0.04],
                [0.01,  0.08, 0.07, 0.09,  0.12, 0.14, 0.11],
                [0.1,   0.31, 0.29, 0.33,  0.55, 0.52, 0.58],
                [1,     0.62, 0.60, 0.65,  0.88, 0.85, 0.91],
                [10,    0.81, 0.79, 0.83,  0.97, 0.95, 0.98],
                [100,   0.91, 0.89, 0.93,  0.99, 0.98, 1.00],
                [1000,  0.95, 0.94, 0.96,  1.00, 0.99, 1.00],
            ]
            for ri, row_vals in enumerate(cf_data, 3):
                for ci, v in enumerate(row_vals, 1):
                    dat(ws, ri, ci, v)
            ws.column_dimensions["A"].width = 15
            for ci in range(2, 8): ws.column_dimensions[chr(64+ci)].width = 10

        elif mode == "column_stats":
            ws.title = "Column Stats Template"
            hint(ws, 1, 1, "Row 1: group names | Rows 2+: numeric replicates")
            for ci, grp in enumerate(["Group A", "Group B", "Group C"], 1):
                hdr(ws, 2, ci, grp)
            import random; random.seed(99)
            cs_data = [[round(random.gauss(5+ci*1.5, 1), 2) for ci in range(3)]
                       for _ in range(12)]
            for ri, row_vals in enumerate(cs_data, 3):
                for ci, v in enumerate(row_vals, 1):
                    dat(ws, ri, ci, v)
            for ci in range(1, 4): ws.column_dimensions[chr(64+ci)].width = 12

        elif mode == "contingency":
            ws.title = "Contingency Template"
            hint(ws, 1, 1, "Row 1: col A blank, cols B+ = outcome labels | Rows 2+: col A = group name, then counts")
            for ci, lbl in enumerate(["Survived", "Died"], 2):
                hdr(ws, 2, ci, lbl)
            ct_data = [["Drug", 45, 5], ["Control", 20, 30]]
            for ri, row_vals in enumerate(ct_data, 3):
                sub(ws, ri, 1, row_vals[0])
                for ci, v in enumerate(row_vals[1:], 2):
                    dat(ws, ri, ci, v)
            ws.column_dimensions["A"].width = 12
            for ci in range(2, 4): ws.column_dimensions[chr(64+ci)].width = 12

        elif mode == "repeated_measures":
            ws.title = "Repeated Measures Template"
            hint(ws, 1, 1, "Row 1: condition/timepoint names | Rows 2+: one row per subject")
            for ci, cond in enumerate(["Baseline", "Week 2", "Week 4", "Week 8"], 1):
                hdr(ws, 2, ci, cond)
            rm_data = [
                [2.1, 2.8, 3.5, 4.2],
                [1.9, 2.5, 3.1, 3.8],
                [2.4, 3.0, 3.7, 4.5],
                [1.8, 2.4, 2.9, 3.6],
                [2.2, 2.9, 3.4, 4.0],
                [2.0, 2.6, 3.2, 3.9],
            ]
            for ri, row_vals in enumerate(rm_data, 3):
                for ci, v in enumerate(row_vals, 1):
                    dat(ws, ri, ci, v)
            for ci in range(1, 5): ws.column_dimensions[chr(64+ci)].width = 12

        wb.save(path)
        self._set_status(f"Template saved to Desktop: {fname}")
        messagebox.showinfo("Template saved",
                            f"Template saved to your Desktop:\n{fname}\n\n"
                            "Open it in Excel, fill in your data, then browse for it in the app.")

    def _load_sheets(self, path, reset_sheet=True):
        """Load sheet names for the given file.
        reset_sheet=True   ->  set sheet var to first sheet (new file load)
        reset_sheet=False  ->  keep current sheet selection (sheet change trigger)
        """
        # openpyxl first import can trigger a Cocoa tick that resets the dock icon.
        self.after(100, set_dock_icon)
        try:
            import openpyxl
            sheets = _cached_sheets(path)
            if not sheets: raise ValueError("No sheets found")
            for cb in self._sheet_cbs:
                cb.config(values=sheets, state="readonly")
            if reset_sheet:
                self._vars["sheet"].set(sheets[0])
            elif self._vars["sheet"].get() not in sheets:
                # Current selection no longer valid  -  fall back to first
                self._vars["sheet"].set(sheets[0])
            for lbl in self._sheet_hints:
                lbl.config(text=f"{len(sheets)} sheet(s) found")
        except Exception as e:
            for cb in self._sheet_cbs:
                cb.config(values=[], state="readonly")
            for lbl in self._sheet_hints:
                lbl.config(text=f"Could not read sheets: {e}")

        if not self._file_selected:
            self._file_selected = True
            self._unlock_form()

        # Always keep sheet comboboxes enabled regardless of validation outcome
        for cb in self._sheet_cbs:
            try: cb.config(state="readonly")
            except Exception: pass

        self._validated = False
        self._run_btn.config(state="disabled", text="Generate Plot")
        self._set_validate_text("Validating...", "gray")
        # Run validation automatically after a short delay so the UI updates first
        self.after(50, self._validate_spreadsheet)

    def _set_validate_text(self, text, color="gray"):
        """Update the validation label on all tabs with a status dot prefix."""
        _color_map = {
            "green":  "#2a8a2a",
            "orange": "#cc6600",
            "red":    "#cc0000",
            "gray":   "#aaaaaa",
        }
        hex_color = _color_map.get(color, color)
        dot       = "● " if text else ""
        for lbl in self._validate_lbls:
            lbl.config(text=f"{dot}{text}", foreground=hex_color)

    # ── Validation ────────────────────────────────────────────────────────────

    def _validate_spreadsheet(self):
        path = self._vars["excel_path"].get().strip()
        if not path:
            self._set_validate_text("", "gray"); return
        if not os.path.exists(path):
            # File moved or deleted  -  clear the stale path silently rather than
            # flashing a red error (common on startup if the file was renamed)
            self._vars["excel_path"].set("")
            self._set_validate_text("", "gray")
            return

        try:
            pd = _pd()
            sheet     = self._vars["sheet"].get().strip()
            sheet_arg = int(sheet) if sheet.lstrip("-").isdigit() else (sheet or 0)
            df        = pd.read_excel(path, sheet_name=sheet_arg, header=None)
        except Exception as e:
            self._set_validate_text(f"Could not read file: {e}", "red")
            messagebox.showerror("Validation Failed", f"Could not read the Excel file:\n\n{e}")
            self._validated = False
            self._run_btn.config(state="disabled", text="✗  Generate Plot")
            return

        mode = self._plot_type.get()
        spec = next((s for s in _REGISTRY_SPECS if s.key == mode), _REGISTRY_SPECS[0])

        # Prefer standalone validators from plotter_validators (pure functions).
        # Fall back to the App instance-method versions for any that weren't extracted.
        _STANDALONE_VALIDATORS = {
            "_validate_bar":           validate_bar           if _VALIDATORS_AVAILABLE else None,
            "_validate_line":          validate_line          if _VALIDATORS_AVAILABLE else None,
            "_validate_grouped_bar":   validate_grouped_bar   if _VALIDATORS_AVAILABLE else None,
            "_validate_kaplan_meier":  validate_kaplan_meier  if _VALIDATORS_AVAILABLE else None,
            "_validate_heatmap":       validate_heatmap       if _VALIDATORS_AVAILABLE else None,
            "_validate_two_way_anova": validate_two_way_anova if _VALIDATORS_AVAILABLE else None,
            "_validate_contingency":   validate_contingency   if _VALIDATORS_AVAILABLE else None,
            "_validate_chi_square_gof":validate_chi_square_gof if _VALIDATORS_AVAILABLE else None,
            "_validate_bland_altman":  validate_bland_altman  if _VALIDATORS_AVAILABLE else None,
            "_validate_forest_plot":   validate_forest_plot   if _VALIDATORS_AVAILABLE else None,
            "_validate_pyramid":       validate_pyramid       if _VALIDATORS_AVAILABLE else None,
        }
        standalone = _STANDALONE_VALIDATORS.get(spec.validate)
        if standalone is not None:
            errors, warnings = standalone(df)
        else:
            errors, warnings = getattr(self, spec.validate)(df)
        tw = spec.label  # full label e.g. "Bar Chart"

        if errors:
            msg = "\n\n".join(errors)
            self._set_validate_text(f"{len(errors)} error(s)", "red")
            messagebox.showerror(f"Validation Failed - {tw} Graph",
                                 "The spreadsheet has the following issues:\n\n" + msg)
            self._validated = False
            self._run_btn.config(state="disabled", text="✗  Generate Plot")
            # Lock all controls except file upload, sheet selector, and chart switcher
            if not self._validation_lock:
                self._validation_lock = True
                self._lock_form()
                # Ensure always-on items stay on after lock
                for cb in getattr(self, "_sheet_cbs", []):
                    try: cb.config(state="readonly")
                    except Exception: pass
        elif warnings:
            msg = "\n\n".join(warnings)
            self._set_validate_text(f"{len(warnings)} warning(s) - OK to run", "orange")
            messagebox.showwarning(f"Validation Warnings - {tw} Graph",
                                   "Spreadsheet looks OK but has some warnings:\n\n" + msg)
            self._validated = True
            if self._validation_lock:
                self._unlock_form()
            if self._pf_ready:
                self._run_btn.config(state="normal", text="Generate Plot")
            else:
                self._run_btn.config(state="disabled", text="Generate Plot")
        else:
            pd = _pd()
            n_cols, n_rows = df.shape[1], df.shape[0] - 1
            if mode == "line":
                hdrs  = [str(h) if pd.notna(h) else "" for h in df.iloc[0, 1:]]
                uniq  = list(dict.fromkeys(h for h in hdrs if h))
                mstr  = f"line ({len(uniq)} series, {n_rows} X points)"
            elif mode == "grouped_bar":
                cats  = list(dict.fromkeys(str(h) for h in df.iloc[0] if pd.notna(h) and str(h).strip()))
                subs  = list(dict.fromkeys(str(h) for h in df.iloc[1] if pd.notna(h) and str(h).strip()))
                mstr  = f"grouped bar ({len(cats)} categories, {len(subs)} subgroups)"
            elif mode == "box":
                mstr = f"box ({n_cols} groups, {n_rows} row(s))"
            elif mode == "scatter":
                hdrs  = [str(h) if pd.notna(h) else "" for h in df.iloc[0, 1:]]
                uniq  = list(dict.fromkeys(h for h in hdrs if h))
                mstr  = f"scatter ({len(uniq)} series, {n_rows} X points)"
            elif mode == "kaplan_meier":
                n_groups = df.shape[1] // 2
                mstr  = f"survival ({n_groups} group(s), {n_rows-1} observations)"
            elif mode == "heatmap":
                mstr = f"heatmap ({df.shape[1]-1} columns × {df.shape[0]-1} rows)"
            elif mode == "two_way_anova":
                mstr = f"two-way ({df.shape[0]-1} observations, {df.shape[1]} columns)"
            elif mode == "violin":
                mstr = f"violin ({n_cols} groups, {n_rows} row(s))"
            elif mode == "histogram":
                mstr = f"histogram ({n_cols} series, {n_rows} value(s))"
            elif mode == "before_after":
                mstr = f"before/after ({n_cols} conditions, {n_rows} subject(s))"
            elif mode == "subcolumn_scatter":
                mstr = f"subcolumn ({n_cols} groups, {n_rows} row(s))"
            elif mode == "curve_fit":
                hdrs  = [str(h) if pd.notna(h) else "" for h in df.iloc[0, 1:]]
                uniq  = list(dict.fromkeys(h for h in hdrs if h))
                mstr  = f"curve fit ({len(uniq)} series, {n_rows} X points)"
            elif mode == "column_stats":
                mstr = f"column stats ({n_cols} groups, {n_rows} row(s))"
            elif mode == "contingency":
                mstr = f"contingency ({df.shape[0]-1} rows × {df.shape[1]-1} outcomes)"
            elif mode == "repeated_measures":
                mstr = f"repeated measures ({n_cols} conditions, {n_rows} subject(s))"
            else:
                mstr = f"bar ({n_cols} groups)"
            self._set_validate_text(f"Valid {mstr}", "green")
            self._validated = True
            # Unlock any validation-triggered lock now that the sheet is valid
            if self._validation_lock:
                self._unlock_form()
            if self._pf_ready:
                self._run_btn.config(state="normal", text="Generate Plot")
            else:
                self._run_btn.config(state="disabled", text="Generate Plot")
            # Populate control group dropdown immediately from the sheet  - 
            # Prism lets you set the control before running the plot.
            # Extract group names according to each chart type's layout.
            if hasattr(self, "_control_cbs"):
                try:
                    pd = _pd()
                    grp_names = []

                    if mode in {"bar", "box", "violin", "subcolumn_scatter",
                                "before_after", "repeated_measures",
                                "histogram", "column_stats", "dot_plot"}:
                        # Row 1 = flat group headers
                        grp_names = [str(h) for h in df.iloc[0]
                                     if pd.notna(h) and str(h).strip()]

                    elif mode == "grouped_bar":
                        # Row 1 = category names (may repeat), Row 2 = subgroup names
                        # Control applies to subgroups (row 2 unique values)
                        grp_names = list(dict.fromkeys(
                            str(h) for h in df.iloc[1]
                            if pd.notna(h) and str(h).strip()
                        ))

                    elif mode in {"line", "scatter", "curve_fit"}:
                        # Row 1: col0 = x-label, cols 1+ = series names
                        grp_names = list(dict.fromkeys(
                            str(h) for h in df.iloc[0, 1:]
                            if pd.notna(h) and str(h).strip()
                        ))

                    elif mode == "two_way_anova":
                        # Long-format; Factor_A column values are the "groups"
                        try:
                            df_long = df.copy()
                            df_long.columns = df_long.iloc[0]
                            df_long = df_long.iloc[1:].reset_index(drop=True)
                            fa_col = next(
                                (c for c in df_long.columns
                                 if str(c).strip().lower() in ("factor_a", "factor a", "groupa", "group_a")),
                                df_long.columns[0] if len(df_long.columns) > 0 else None
                            )
                            if fa_col is not None:
                                grp_names = list(dict.fromkeys(
                                    str(v) for v in df_long[fa_col].dropna()
                                ))
                        except Exception:
                            pass

                    if grp_names:
                        # Store ONLY real group names — no "(none)" sentinel.
                        # _tog_control() decides enabled/disabled state and
                        # whether the dropdown shows these names or is locked.
                        for cb in self._control_cbs:
                            cb.config(values=grp_names)
                        # Preserve current selection if still valid; else clear
                        cur = self._vars["control"].get()
                        if cur not in grp_names:
                            self._vars["control"].set("")
                        # Let _tog_control apply all visibility/consistency rules
                        self._tog_control()
                except Exception:
                    pass

        # Sync the Help Analyze button state with current validation result
        self._sync_analyze_btn()

    # ── Shared validator helpers ──────────────────────────────────────────────

    def _validate_flat_header(self, df,
                              min_groups: int = 2,
                              min_rows: int = 3,
                              chart_name: str = "bar graph") -> tuple:
        """Validate a flat-header layout (Row 1 = group names, Rows 2+ = values).

        Used by: bar, box, violin, subcolumn_scatter, before_after,
                 repeated_measures, histogram, dot_plot, and any future chart
                 that uses the same row-1-header / rows-2+-data layout.

        Parameters
        ----------
        min_groups : minimum number of named columns to avoid a warning
        min_rows   : minimum number of data rows to support stats (warning only)
        chart_name : human-readable chart type for error messages

        Returns
        -------
        (errors, warnings)  -  both are lists of strings
        """
        pd = _pd()
        errors, warnings = [], []

        if df.shape[0] < 2:
            errors.append(
                f"Sheet has fewer than 2 rows. "
                f"Expected: Row 1 = group names, Rows 2+ = data values.")
            return errors, warnings
        if df.shape[1] < 1:
            errors.append("Sheet has no columns.")
            return errors, warnings

        headers    = df.iloc[0]
        empty_hdrs = [i for i, h in enumerate(headers)
                      if pd.isna(h) or str(h).strip() == ""]

        if len(empty_hdrs) == df.shape[1]:
            errors.append(
                "Row 1 is entirely empty. "
                "Row 1 should contain group names (e.g. 'Control', 'Treatment').")
        elif empty_hdrs:
            warnings.append(
                f"Row 1 has {len(empty_hdrs)} empty cell(s) in "
                f"column(s) {[i+1 for i in empty_hdrs]}. "
                "Each column should have a group name in row 1.")

        data = df.iloc[1:]
        for col_i in range(df.shape[1]):
            ch       = (str(headers.iloc[col_i])
                        if pd.notna(headers.iloc[col_i]) else f"Column {col_i+1}")
            col_vals = data.iloc[:, col_i].dropna()
            if len(col_vals) == 0:
                warnings.append(
                    f"Column '{ch}' has no data values (all empty below row 1).")
                continue
            non_num = _non_numeric_values(col_vals)
            if non_num:
                errors.append(
                    f"Column '{ch}' contains non-numeric values: "
                    f"{non_num}. All data values (rows 2+) must be numbers.")

        if df.shape[1] < min_groups:
            warnings.append(
                f"Only {df.shape[1]} group/column(s) found. "
                f"{chart_name.capitalize()} work best with {min_groups}+ groups.")
        if df.shape[0] < min_rows + 1:  # +1 for header row
            warnings.append(
                f"Only {df.shape[0] - 1} data row(s) per group. "
                "Statistical tests require at least 3 replicates per group.")

        return errors, warnings

    def _validate_bar(self, df):
        return self._validate_flat_header(df, min_groups=2, min_rows=3,
                                          chart_name="bar graph")

    def _validate_line(self, df):
        pd = _pd()
        errors, warnings = [], []

        # Minimum shape
        if df.shape[0] < 2:
            errors.append("Sheet has fewer than 2 rows. "
                          "Expected: Row 1 = series names, Rows 2+ = data values.")
            return errors, warnings
        if df.shape[1] < 2:
            errors.append("Sheet needs at least 2 columns: "
                          "Col 1 = numeric X values, Col 2+ = Y replicates.")
            return errors, warnings

        # Row 1: col 1 can be anything (X axis title or blank), cols 2+ = series names
        series_hdrs = [str(h) if pd.notna(h) else ""
                       for h in df.iloc[0, 1:]]
        empty_series = [i+2 for i, h in enumerate(series_hdrs) if not h.strip()]
        if len(empty_series) == len(series_hdrs):
            errors.append("Row 1 (columns 2+) must contain series names. "
                          "All series name cells are empty.")
        elif empty_series:
            warnings.append(f"Row 1 has empty series name(s) in column(s) "
                            f"{empty_series}. Each Y column should have a series name.")

        # Col 1 (rows 2+): must be numeric X values
        x_col = df.iloc[1:, 0]
        non_num_x = _non_numeric_values(x_col)
        if non_num_x:
            errors.append(f"Column 1 must contain numeric X values, but found: "
                          f"{non_num_x}.")
        empty_x = x_col.isna().sum()
        if empty_x > 0:
            warnings.append(f"Column 1 (X values) has {empty_x} empty cell(s). "
                            "Each row should have a numeric X value.")

        # Cols 2+ (rows 2+): must be numeric Y values
        for col_i in range(1, df.shape[1]):
            ch = series_hdrs[col_i - 1] if series_hdrs[col_i - 1] else f"Column {col_i+1}"
            col_data = df.iloc[1:, col_i].dropna()
            non_num = _non_numeric_values(col_data)
            if non_num:
                errors.append(f"Column '{ch}' (col {col_i+1}) contains non-numeric values: "
                              f"{non_num}.")

        # Warn if fewer than 2 X points
        if df.shape[0] < 3:
            warnings.append("Only 1 data row found. "
                            "A line graph needs at least 2 X points to draw a line.")

        # Warn if any series has fewer than 3 replicates
        from collections import Counter
        counts = Counter(h for h in series_hdrs if h.strip())
        for s, n in counts.items():
            if n < 3:
                warnings.append(f"Series '{s}' has only {n} replicate column(s). "
                                "At least 3 replicates are recommended for error bars.")

        return errors, warnings

    def _validate_grouped_bar(self, df):
        pd = _pd()
        from collections import Counter
        errors, warnings = [], []

        if df.shape[0] < 3:
            errors.append("Sheet needs at least 3 rows: "
                          "Row 1 = category names, Row 2 = subgroup names, Rows 3+ = data values.")
            return errors, warnings
        if df.shape[1] < 2:
            errors.append("Sheet needs at least 2 columns.")
            return errors, warnings

        row1 = [str(h) if pd.notna(h) else "" for h in df.iloc[0]]
        row2 = [str(h) if pd.notna(h) else "" for h in df.iloc[1]]

        # Row 1: category names
        empty_cats = [i+1 for i, h in enumerate(row1) if not h.strip()]
        if len(empty_cats) == len(row1):
            errors.append("Row 1 must contain category names  -  all cells are empty.")
        elif empty_cats:
            warnings.append(f"Row 1 has empty category name(s) in column(s) {empty_cats}.")

        # Row 2: subgroup names
        empty_subs = [i+1 for i, h in enumerate(row2) if not h.strip()]
        if len(empty_subs) == len(row2):
            errors.append("Row 2 must contain subgroup names  -  all cells are empty.")
        elif empty_subs:
            warnings.append(f"Row 2 has empty subgroup name(s) in column(s) {empty_subs}.")

        # Rows 3+: must be numeric
        data = df.iloc[2:]
        for col_i in range(df.shape[1]):
            ch = row2[col_i] if row2[col_i] else (row1[col_i] if row1[col_i] else f"Column {col_i+1}")
            non_num = _non_numeric_values(data.iloc[:, col_i])
            if non_num:
                errors.append(f"Column '{ch}' (col {col_i+1}) contains non-numeric values: "
                              f"{non_num}.")

        if df.shape[0] < 4:
            warnings.append("Only 1 data row  -  at least 3 replicates are recommended per subgroup.")

        # Check at least 2 subgroups per category (otherwise grouped chart is pointless)
        cats = list(dict.fromkeys(h for h in row1 if h.strip()))
        subs = list(dict.fromkeys(h for h in row2 if h.strip()))
        if len(subs) < 2:
            warnings.append("Only 1 subgroup found. "
                            "A grouped bar chart works best with 2+ subgroups per category.")

        # Warn if replicate counts are unequal across subgroups
        for cat in cats:
            sub_counts = Counter()
            for col_i in range(df.shape[1]):
                if row1[col_i] == cat and row2[col_i].strip():
                    sub_counts[row2[col_i]] += 1
            counts = list(sub_counts.values())
            if counts and max(counts) != min(counts):
                warnings.append(f"Category '{cat}' has unequal replicate counts across subgroups "
                                f"{dict(sub_counts)}  -  this is allowed but worth checking.")

        return errors, warnings

    def _validate_kaplan_meier(self, df):
        pd = _pd()
        errors, warnings = [], []
        if df.shape[0] < 3:
            errors.append("Sheet needs at least 3 rows: Row 1 = group names, "
                          "Row 2 = 'Time'/'Event' headers, Rows 3+ = data.")
            return errors, warnings
        if df.shape[1] < 2:
            errors.append("Sheet needs at least 2 columns (one Time + one Event column).")
            return errors, warnings

        row2 = [str(v).strip().lower() if pd.notna(v) else "" for v in df.iloc[1]]
        if not any("time" in h for h in row2):
            warnings.append("Row 2 should contain 'Time' / 'Event' column headers. "
                            "Could not find 'Time'  -  check the format.")

        data = df.iloc[2:]
        non_num_cols = []
        for ci in range(df.shape[1]):
            if _non_numeric_values(data.iloc[:, ci]):
                non_num_cols.append(ci + 1)
        if non_num_cols:
            errors.append(f"Non-numeric values found in column(s): {non_num_cols}. "
                          f"All time and event values must be numeric (events: 1=occurred, 0=censored).")

        if df.shape[1] % 2 != 0:
            warnings.append("Odd number of columns  -  each group needs exactly 2 columns "
                            "(Time and Event). Check column pairing.")

        return errors, warnings

    def _tab_stats_kaplan_meier(self, f):
        """Stats tab specific to Kaplan-Meier  -  log-rank options."""
        self._init_vars({
            "show_stats": False, "show_p_values": False,
            "show_ci": True, "show_censors": True, "show_at_risk": False,
        })
        g = f; r = 0

        r = self._section_label(g, r, "Survival Curve Options")
        r = self._checkbox(g, r, "show_ci",      " Show 95% confidence band")
        r = self._checkbox(g, r, "show_censors", " Show censoring tick marks")
        r = self._checkbox(g, r, "show_at_risk", " Show at-risk table below plot")
        r = self._sep(g, r)

        r = self._section_label(g, r, "Log-Rank Test")
        r = self._checkbox(g, r, "show_stats",    " Show log-rank p-values on plot")
        r = self._checkbox(g, r, "show_p_values", " Show raw p-values (instead of stars)")
        r = self._hint(g, r, "Mantel-Cox log-rank test. Each pair of groups is compared.\n"
                             "Censored observations (Event=0) are handled automatically.")
        r = self._sep(g, r)

        r = self._section_label(g, r, "Reference Line")
        ref_row = ttk.Frame(g)
        ref_row.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=(4, 2)); r += 1
        PCheckbox(ref_row, variable=self._vars["ref_line_enabled"],
                        text=" Show at Y =").pack(side="left")
        _e_ref = PEntry(ref_row, textvariable=self._vars["ref_line_y"],
                           width=8, font=("Menlo", 12))
        _e_ref.pack(side="left", padx=(6, 0))
        add_placeholder(_e_ref, self._vars["ref_line_y"], "e.g. 0.5")
        ref_lbl_row = ttk.Frame(g)
        ref_lbl_row.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=(0, 4)); r += 1
        ttk.Label(ref_lbl_row, text="Label:", font=("Helvetica Neue", 11)).pack(side="left")
        _e_lbl = PEntry(ref_lbl_row, textvariable=self._vars["ref_line_label"],
                           width=20, font=("Menlo", 12))
        _e_lbl.pack(side="left", padx=(6, 0))
        add_placeholder(_e_lbl, self._vars["ref_line_label"], "blank = show y=…")
        self._tab_footer(g, r)

    # ── Stats tab: Before/After ───────────────────────────────────────────────

    def _tab_stats_before_after(self, f):
        """Options for the Before/After paired plot."""
        self._init_vars({"show_stats": False, "show_p_values": False,
                         "show_n_labels": False, "show_normality_warning": True})
        g = f; r = 0

        r = self._section_label(g, r, "Paired Statistics")
        r = self._checkbox(g, r, "show_stats",    " Show paired t-test bracket (2 groups only)")
        r = self._checkbox(g, r, "show_p_values", " Show raw p-value (instead of stars)")
        r = self._checkbox(g, r, "show_n_labels", " Show n= on x-axis")
        r = self._sep(g, r)
        r = self._section_label(g, r, "Normality Check")
        r = self._checkbox(g, r, "show_normality_warning", " Show normality warning on plot")
        r = self._sep(g, r)
        r = self._hint(g, r, "Each subject is connected by a gray line. "
                             "The mean ± SD bar is drawn per group. "
                             "Paired t-test requires equal n in both groups "
                             "(matched by row order).")
        self._tab_footer(g, r)

    # ── Stats tab: Histogram ──────────────────────────────────────────────────

    def _tab_stats_histogram(self, f):
        """Bin and display options for Histogram."""
        self._init_vars({"hist_bins": "0", "hist_density": False, "hist_overlay_normal": False})
        g = f; r = 0

        r = self._section_label(g, r, "Bins")
        fr_bins = ttk.Frame(g)
        fr_bins.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=4); r += 1
        ttk.Label(fr_bins, text="Number of bins:", font=("Helvetica Neue", 12)).pack(side="left")
        _e_bins = PEntry(fr_bins, textvariable=self._vars["hist_bins"],
                            width=6, font=("Menlo", 12))
        _e_bins.pack(side="left", padx=(6, 0))
        add_placeholder(_e_bins, self._vars["hist_bins"], "0 = auto")
        ttk.Label(fr_bins, text="  (0 = auto via Sturges)",
                  foreground="gray", font=("Helvetica Neue", 11)).pack(side="left")
        r = self._sep(g, r)

        r = self._section_label(g, r, "Display")
        r = self._checkbox(g, r, "hist_density",        " Density mode (area = 1, not count)")
        r = self._checkbox(g, r, "hist_overlay_normal", " Overlay fitted normal curve")
        r = self._sep(g, r)
        r = self._hint(g, r, "All series share the same bin edges so bars "
                             "are directly comparable across groups.")
        self._tab_footer(g, r)

    # ── Stats tab: Curve Fit ──────────────────────────────────────────────────

    def _tab_stats_curve_fit(self, f):
        """Model selector and display options for nonlinear curve fitting."""
        _default_model = "4PL Sigmoidal (EC50/IC50)"
        self._init_vars({
            "curve_model":       _default_model,
            "cf_show_ci":        True,  "cf_show_residuals": False,
            "cf_show_equation":  True,  "cf_show_r2":        True,
            "cf_show_params":    True,
        })
        g = f; r = 0

        r = self._section_label(g, r, "Model")
        _model_names = list(getattr(
            __import__("sys").modules.get(
                "plotter_functions", type("M", (), {"CURVE_MODELS": {}})()),
            "CURVE_MODELS", {}
        ).keys()) or [
            "4PL Sigmoidal (EC50/IC50)", "3PL Sigmoidal (no bottom)",
            "One-phase exponential decay", "One-phase exponential growth",
            "Two-phase exponential decay", "Michaelis-Menten",
            "Hill equation", "Gaussian (bell curve)", "Log-normal",
            "Linear", "Polynomial (2nd order)",
        ]
        _model_cb = PCombobox(g, textvariable=self._vars["curve_model"],
                                 values=_model_names, state="readonly",
                                 width=34, font=("Helvetica Neue", 12))
        _model_cb.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=4); r += 1
        _desc_lbl = ttk.Label(g, text="", foreground="#666",
                              wraplength=380, justify="left", font=("Helvetica Neue", 10))
        _desc_lbl.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=(0, 6)); r += 1

        _MODEL_DESC = {
            "4PL Sigmoidal (EC50/IC50)":   "4-parameter logistic. Bottom, Top, EC50 (inflection), HillSlope. Standard for dose-response curves.",
            "3PL Sigmoidal (no bottom)":   "3-parameter logistic. Bottom fixed at 0. Simpler dose-response.",
            "One-phase exponential decay": "Y = Plateau + (Y0 − Plateau)·e^(−K·X). Single-phase decay to a plateau.",
            "One-phase exponential growth":"Y = Plateau − (Plateau − Y0)·e^(−K·X). Single-phase association.",
            "Two-phase exponential decay": "Biphasic decay  -  fast and slow components. Adds Fraction_fast parameter.",
            "Michaelis-Menten":            "Y = Vmax·X / (Km + X). Enzyme kinetics, receptor binding.",
            "Hill equation":               "Generalised Michaelis-Menten with cooperativity exponent n.",
            "Gaussian (bell curve)":       "Symmetric bell. Amplitude, Mean, SD.",
            "Log-normal":                  "Log-normal bell curve. Peak position on log scale.",
            "Linear":                      "Y = Slope·X + Intercept.",
            "Polynomial (2nd order)":      "Y = A·X² + B·X + C.",
        }
        def _on_model_change(e=None):
            _desc_lbl.config(text=_MODEL_DESC.get(self._vars["curve_model"].get(), ""))
        _model_cb.bind("<<ComboboxSelected>>", _on_model_change)
        _on_model_change()
        r = self._sep(g, r)

        r = self._section_label(g, r, "Display")
        r = self._checkbox(g, r, "cf_show_ci",        " Show 95% confidence band")
        r = self._checkbox(g, r, "cf_show_r2",        " Show R² on plot")
        r = self._checkbox(g, r, "cf_show_params",    " Show parameter estimates on plot")
        r = self._checkbox(g, r, "cf_show_equation",  " Show equation on plot")
        r = self._checkbox(g, r, "cf_show_residuals", " Show residuals panel below plot")
        r = self._sep(g, r)
        r = self._hint(g, r, "Uses the same Excel format as Scatter  -  Col 1 is X, "
                             "remaining columns are Y replicates with series names in Row 1. "
                             "Each series is fitted independently.")
        self._tab_footer(g, r)

    # ── Stats tab: Column Statistics ─────────────────────────────────────────

    def _tab_stats_column_stats(self, f):
        self._init_vars({"cs_show_normality": True, "cs_show_ci": True, "cs_show_cv": True})
        g = f; r = 0

        r = self._section_label(g, r, "Table Contents")
        r = self._checkbox(g, r, "cs_show_ci",        " 95% Confidence Interval (low / high)")
        r = self._checkbox(g, r, "cs_show_cv",        " Coefficient of Variation (CV%)")
        r = self._checkbox(g, r, "cs_show_normality", " Shapiro-Wilk normality test")
        r = self._sep(g, r)
        r = self._hint(g, r, "Always shown: N, Mean, SD, SEM, Median, Min, Max.\n"
                             "Table rows = groups; columns = statistics.")
        self._tab_footer(g, r)

    # ── Stats tab: Contingency ───────────────────────────────────────────────

    def _tab_stats_contingency(self, f):
        self._init_vars({"ct_show_pct": True, "ct_show_expected": False})
        g = f; r = 0

        r = self._section_label(g, r, "Display")
        r = self._checkbox(g, r, "ct_show_pct",      " Show row % on bars")
        r = self._checkbox(g, r, "ct_show_expected", " Show expected counts (dashed line)")
        r = self._sep(g, r)
        r = self._hint(g, r, "2×2 tables: Fisher's exact test + Odds Ratio.\n"
                             "Larger tables: Chi-square test of independence.\n"
                             "Statistics are shown automatically on the plot.")
        self._tab_footer(g, r)

    def _validate_heatmap(self, df):
        pd = _pd()
        errors, warnings = [], []
        if df.shape[0] < 2:
            errors.append("Sheet needs at least 2 rows: Row 1 = column labels, Rows 2+ = data.")
            return errors, warnings
        if df.shape[1] < 2:
            errors.append("Sheet needs at least 2 columns: Column A = row labels, Cols B+ = data.")
            return errors, warnings
        data = df.iloc[1:, 1:]
        non_num_cells = sum(
            len([v for v in data.iloc[:, ci].dropna() if not _is_num(v)])
            for ci in range(data.shape[1])
        )
        if non_num_cells > 0:
            errors.append(f"Data region contains {non_num_cells} non-numeric value(s). "
                          "All values except row/column labels must be numeric.")
        if data.apply(pd.to_numeric, errors="coerce").isna().all().all():
            errors.append("No numeric data found in the data region.")
        if df.shape[0] < 3:
            warnings.append("Only 1 data row  -  heatmaps work best with multiple rows.")
        return errors, warnings

    def _validate_two_way_anova(self, df):
        pd = _pd()
        errors, warnings = [], []
        if df.shape[0] < 2:
            errors.append("Sheet needs at least 2 rows: Row 1 = headers, Rows 2+ = data.")
            return errors, warnings
        if df.shape[1] < 3:
            errors.append("Need at least 3 columns: Factor_A, Factor_B, Value.")
            return errors, warnings
        num_cols = [c for c in df.columns
                    if pd.to_numeric(df[c], errors="coerce").notna().any()]
        if not num_cols:
            errors.append("No numeric column found. The last column should contain numeric values.")
            return errors, warnings
        cat_cols = [c for c in df.columns if c not in num_cols]
        if len(cat_cols) < 2:
            errors.append("Need at least 2 categorical columns for Factor A and Factor B.")
            return errors, warnings
        fa, fb, dv = cat_cols[0], cat_cols[1], num_cols[-1]
        data = df[[fa, fb, dv]].dropna()
        if len(data) < 4:
            errors.append(f"Too few observations ({len(data)}). "
                          "Need at least 4 (ideally 2+ per cell).")
            return errors, warnings
        a_levels = sorted(data[fa].unique())
        b_levels = sorted(data[fb].unique())
        if len(a_levels) < 2:
            errors.append(f"Factor A ('{fa}') has only 1 level  -  need at least 2.")
        if len(b_levels) < 2:
            errors.append(f"Factor B ('{fb}') has only 1 level  -  need at least 2.")
        thin_cells = []
        for a in a_levels:
            for b in b_levels:
                n = len(data[(data[fa] == a) & (data[fb] == b)])
                if n == 0:
                    errors.append(f"Missing cell: {fa}={a}, {fb}={b}. "
                                  "All factor combinations must have data.")
                elif n == 1:
                    thin_cells.append(f"{fa}={a}/{fb}={b}")
        if thin_cells:
            warnings.append(f"Cells with only 1 observation: {', '.join(thin_cells)}. "
                            "At least 2 replicates per cell are recommended.")
        return errors, warnings

    def _validate_contingency(self, df):
        pd = _pd()
        errors, warnings = [], []
        if df.shape[0] < 3:
            errors.append("Need at least 2 data rows + 1 header row.\n"
                          "Row 1: column labels  |  Col A rows 2+: row labels  |  "
                          "Remaining cells: counts.")
            return errors, warnings
        if df.shape[1] < 3:
            errors.append("Need at least 2 outcome columns (plus row-label column A).")
            return errors, warnings
        data = df.iloc[1:, 1:]
        non_num = sum(1 for ci in range(data.shape[1])
                      for v in data.iloc[:, ci].dropna() if not _is_num(v))
        if non_num > 0:
            errors.append(f"Data cells must be numeric counts ({non_num} non-numeric found).")
        if data.apply(pd.to_numeric, errors="coerce").lt(0).any().any():
            warnings.append("Negative counts found  -  counts should be non-negative integers.")
        return errors, warnings

    def _validate_chi_square_gof(self, df):
        """Validate Chi-Square GoF layout.

        Row 1 : category names
        Row 2 : observed counts  (required, all numeric, all ≥ 0)
        Row 3 : expected proportions or counts (optional)
        """
        pd = _pd()
        errors, warnings = [], []
        if df.shape[0] < 2:
            errors.append("Need at least Row 1 (category names) and Row 2 (observed counts).")
            return errors, warnings
        if df.shape[1] < 2:
            errors.append("Need at least 2 categories (columns).")
            return errors, warnings

        headers = df.iloc[0]
        empty_hdrs = [i for i, h in enumerate(headers)
                      if pd.isna(h) or str(h).strip() == ""]
        if empty_hdrs:
            warnings.append(f"Row 1 has {len(empty_hdrs)} empty category name(s).")

        obs_row = pd.to_numeric(df.iloc[1], errors="coerce")
        if obs_row.isna().any():
            errors.append("Row 2 (observed counts) must contain only numeric values.")
        elif (obs_row < 0).any():
            errors.append("Observed counts (Row 2) must be non-negative.")
        elif (obs_row < 5).any():
            warnings.append("Some expected counts < 5  -  chi-square approximation may be unreliable.")

        if df.shape[0] >= 3:
            exp_row = pd.to_numeric(df.iloc[2], errors="coerce")
            if exp_row.isna().any():
                errors.append("Row 3 (expected) must be numeric if provided.")
            elif (exp_row <= 0).any():
                errors.append("Expected values (Row 3) must be positive.")

        return errors, warnings


    def _validate_bland_altman(self, df):
        """Validate Bland-Altman layout: Row 1 = two method names, Rows 2+ = paired values."""
        pd = _pd()
        errors, warnings = [], []
        if df.shape[0] < 3:
            errors.append("Need Row 1 (method names) + at least 2 data rows.")
            return errors, warnings
        if df.shape[1] < 2:
            errors.append("Need exactly 2 columns: Method A and Method B.")
            return errors, warnings
        if df.shape[1] > 2:
            warnings.append("Only the first 2 columns are used (Method A, Method B).")
        for ci in range(min(2, df.shape[1])):
            col_vals = pd.to_numeric(df.iloc[1:, ci], errors="coerce").dropna()
            if len(col_vals) == 0:
                errors.append(f"Column {ci+1} has no numeric data.")
        n_a = pd.to_numeric(df.iloc[1:, 0], errors="coerce").dropna().shape[0]
        n_b = pd.to_numeric(df.iloc[1:, 1], errors="coerce").dropna().shape[0]
        if abs(n_a - n_b) > 0:
            warnings.append(f"Column lengths differ ({n_a} vs {n_b}); shorter column will be truncated.")
        if min(n_a, n_b) < 3:
            errors.append("Need at least 3 paired observations for a meaningful plot.")
        return errors, warnings

    def _validate_forest_plot(self, df):
        """Validate Forest plot layout: Row 1 = headers, Rows 2+ = studies."""
        pd = _pd()
        errors, warnings = [], []
        if df.shape[0] < 3:
            errors.append("Need a header row + at least 2 study rows.")
            return errors, warnings
        if df.shape[1] < 4:
            errors.append("Need at least 4 columns: Study, Effect, CI_lo, CI_hi.")
            return errors, warnings
        for ci, name in ((1, "Effect"), (2, "CI_lo"), (3, "CI_hi")):
            col = pd.to_numeric(df.iloc[1:, ci], errors="coerce")
            if col.isna().all():
                errors.append(f"Column '{name}' (column {ci+1}) must be numeric.")
        if not errors:
            effects = pd.to_numeric(df.iloc[1:, 1], errors="coerce")
            ci_lo   = pd.to_numeric(df.iloc[1:, 2], errors="coerce")
            ci_hi   = pd.to_numeric(df.iloc[1:, 3], errors="coerce")
            if (ci_lo > effects).any() or (ci_hi < effects).any():
                warnings.append("Some CI bounds are inverted (CI_lo > Effect or CI_hi < Effect).")
        return errors, warnings

    # ── Stats tab: Repeated Measures ─────────────────────────────────────────

    def _tab_stats_repeated_measures(self, f):
        self._init_vars({
            "rm_show_lines": True, "show_stats": False,
            "show_p_values": False, "rm_test_type": "Parametric",
            "show_normality_warning": True,
        })
        g = f; r = 0

        r = self._section_label(g, r, "Display")
        r = self._checkbox(g, r, "rm_show_lines", " Show subject connecting lines")
        r = self._checkbox(g, r, "show_stats",    " Show significance brackets")
        r = self._checkbox(g, r, "show_p_values", " Show raw p-values (instead of stars)")
        r = self._sep(g, r)

        r = self._section_label(g, r, "Statistical Test")
        PCombobox(g, textvariable=self._vars["rm_test_type"],
                     values=["Parametric", "Non-parametric"],
                     state="readonly", width=22, font=("Helvetica Neue", 12)
                     ).grid(row=r, column=0, sticky="w", padx=PAD, pady=4); r += 1
        r = self._sep(g, r)

        r = self._section_label(g, r, "Normality Check")
        r = self._checkbox(g, r, "show_normality_warning", " Show normality warning on plot")
        r = self._sep(g, r)

        r = self._hint(g, r,
                       "Parametric: RM-ANOVA (requires pingouin) with Holm post-hoc, "
                       "or pairwise paired t-tests if pingouin is unavailable.\n\n"
                       "Non-parametric: Friedman test + Dunn's post-hoc "
                       "(Holm-Bonferroni corrected).\n\n"
                       "Each row = one subject. Subjects matched by row order.")
        self._tab_footer(g, r)

    def _tab_stats_chi_square_gof(self, f):
        """Statistics tab for Chi-Square Goodness of Fit."""
        self._init_vars({"gof_expected_equal": True})
        g = f; r = 0

        r = self._section_label(g, r, "Expected Frequencies")
        r = self._checkbox(g, r, "gof_expected_equal",
                           " Equal expected proportions (ignore Row 3)")
        r = self._hint(g, r,
                       "If unchecked, Row 3 of your spreadsheet is read as expected "
                       "proportions or counts.\n\n"
                       "Dashed bars on the plot show expected values.\n\n"
                       "Layout:\n  Row 1: category names\n"
                       "  Row 2: observed counts\n"
                       "  Row 3: expected proportions (optional)")
        self._tab_footer(g, r)




    def _tab_stats_stacked_bar(self, f):
        """Stats tab for Stacked Bar  -  stack mode, orientation (P17), labels (P18)."""
        self._init_vars({"stacked_mode": "absolute", "stacked_value_labels": False,
                         "stacked_horizontal": False, "xtick_labels_str": ""})
        g = f; r = 0
        r = self._section_label(g, r, "Stack Mode")
        PRadioGroup(g, variable=self._vars["stacked_mode"],
                    options=["absolute", "percent"],
                    ).grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=4); r += 1
        r = self._hint(g, r,
                       "absolute  -  bars show raw mean values stacked.\n"
                       "percent   -  each bar normalized to 100% (proportional composition).")
        r = self._sep(g, r)
        r = self._checkbox(g, r, "stacked_value_labels", " Show value label in each segment")
        r = self._sep(g, r)
        # P17: Horizontal orientation
        r = self._section_label(g, r, "Orientation (P17)")
        r = self._checkbox(g, r, "stacked_horizontal", " Horizontal stacked bars")
        r = self._hint(g, r, "Mirrors the grouped bar horizontal option  -  bars run left-to-right.")
        r = self._sep(g, r)
        # P18: Custom x-tick labels
        r = self._section_label(g, r, "Custom Category Labels (P18)")
        ttk.Label(g, text="Override labels (comma-separated):",
                  font=("Helvetica Neue", 12)).grid(row=r, column=0, sticky="w", padx=PAD); r += 1
        _xt_e = PEntry(g, textvariable=self._vars["xtick_labels_str"], width=38)
        _xt_e.grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=2); r += 1
        add_placeholder(_xt_e, self._vars["xtick_labels_str"], "e.g. Control, Drug A, Drug B")
        r = self._hint(g, r, "Leave blank to use category names from the spreadsheet. "
                             "Must match the number of categories exactly.")
        self._tab_footer(g, r)

    def _tab_stats_bubble(self, f):
        """Stats tab for Bubble Chart  -  size scaling and labels."""
        self._init_vars({"bubble_scale": "1.0", "bubble_show_labels": False})
        g = f; r = 0
        r = self._section_label(g, r, "Bubble Sizing")
        r = self._slider(g, r, "bubble_scale", "Bubble Scale", 0.1, 5.0, ".1f", 1.0)
        r = self._sep(g, r)
        r = self._checkbox(g, r, "bubble_show_labels", " Label each bubble with its size value")
        r = self._hint(g, r,
                       "Excel layout:\n"
                       "  Row 1 : X-label | Series name(s)\n"
                       "  Rows 2+: X | Y | Size  (repeat triple per series)\n\n"
                       "Bubble area is proportional to the Size column.\n"
                       "Median size maps to a standard area; Scale adjusts all bubbles.")
        self._tab_footer(g, r)

    def _tab_stats_dot_plot(self, f):
        """Stats tab for Dot Plot  -  mean/median overlay."""
        self._init_vars({"dp_show_mean": True, "dp_show_median": False})
        g = f; r = 0
        r = self._section_label(g, r, "Summary Lines")
        r = self._checkbox(g, r, "dp_show_mean",   " Show mean line")
        r = self._checkbox(g, r, "dp_show_median", " Show median line (dashed)")
        r = self._hint(g, r,
                       "Dot plot shows individual data points jittered horizontally.\n"
                       "No bar fill  -  ideal for small samples where distribution matters.\n\n"
                       "Same Excel layout as bar chart:\n"
                       "  Row 1 : group names\n  Rows 2+: values")
        self._tab_footer(g, r)

    def _tab_stats_bland_altman(self, f):
        """Stats tab for Bland-Altman plot."""
        self._init_vars({"ba_show_ci": True})
        g = f; r = 0
        r = self._section_label(g, r, "Display")
        r = self._checkbox(g, r, "ba_show_ci",
                           " Show 95% CI bands around mean diff and limits of agreement")
        r = self._hint(g, r,
                       "Bland-Altman plot assesses agreement between two measurement methods.\n\n"
                       "X-axis: mean of the two measurements\n"
                       "Y-axis: difference (Method A − Method B)\n\n"
                       "Solid line = mean difference\n"
                       "Dashed lines = ±1.96 SD (limits of agreement)\n\n"
                       "Excel layout:\n"
                       "  Row 1 : method names (2 columns)\n"
                       "  Rows 2+: paired values per subject")
        self._tab_footer(g, r)

    def _tab_stats_forest_plot(self, f):
        """Stats tab for Forest Plot  -  reference line, weights, summary."""
        self._init_vars({
            "fp_ref_value": "0",
            "fp_show_weights": True,
            "fp_show_summary": True,
        })
        g = f; r = 0
        r = self._section_label(g, r, "Reference Line")

        ref_row = ttk.Frame(g)
        ref_row.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=4); r += 1
        ttk.Label(ref_row, text="Null value (e.g. 0 for MD, 1 for OR):",
                  font=("Helvetica Neue", 12)).pack(side="left")
        _e = PEntry(ref_row, textvariable=self._vars["fp_ref_value"],
                    width=8, font=("Menlo", 12))
        _e.pack(side="left", padx=(8, 0))
        add_placeholder(_e, self._vars["fp_ref_value"], "e.g. 0")

        r = self._sep(g, r)
        r = self._section_label(g, r, "Display")
        r = self._checkbox(g, r, "fp_show_weights", " Show weight values beside studies")
        r = self._checkbox(g, r, "fp_show_summary",
                           " Show pooled-effect summary diamond")
        r = self._hint(g, r,
                       "Excel layout (header row required):\n"
                       "  Study | Effect | CI_lo | CI_hi | Weight\n\n"
                       "Weight column is optional (uniform if absent).\n"
                       "Summary diamond = inverse-variance weighted pooled estimate.")
        self._tab_footer(g, r)

    def _tab_stats_heatmap(self, f):
        """Options tab for heatmap  -  color scale, clustering, annotations."""
        self._init_vars({
            "annotate": False, "cluster_rows": False, "cluster_cols": False,
            "robust": False, "heatmap_vmin": "", "heatmap_vmax": "",
            "heatmap_center": "", "heatmap_fmt": "",
        })
        g = f; r = 0

        r = self._section_label(g, r, "Color Map")
        _hm_color_cb = PCombobox(
            g, textvariable=self._vars["color"],
            values=["Default (Blue-Red)", "Viridis", "Mako", "Plasma",
                    "Coolwarm", "Spectral", "YlOrRd", "Blues", "Greens", "RdYlGn"],
            state="readonly", width=22, font=("Helvetica Neue", 12))
        _hm_color_cb.grid(row=r, column=0, sticky="w", padx=PAD, pady=4); r += 1
        add_placeholder(_hm_color_cb, self._vars["color"], "Default (Blue-Red)")
        r = self._sep(g, r)

        r = self._section_label(g, r, "Color Scale")
        for row_label, key, ph in [("Min", "heatmap_vmin", "auto"),
                                    ("Max", "heatmap_vmax", "auto"),
                                    ("Center", "heatmap_center", "none")]:
            fr = ttk.Frame(g)
            fr.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=2); r += 1
            ttk.Label(fr, text=f"{row_label}:", width=7,
                      font=("Helvetica Neue", 12)).pack(side="left")
            _e = PEntry(fr, textvariable=self._vars[key], width=10, font=("Menlo", 12))
            _e.pack(side="left", padx=(4, 0))
            add_placeholder(_e, self._vars[key], ph)
        r = self._checkbox(g, r, "robust", " Robust scale (2nd–98th percentile)")
        r = self._sep(g, r)

        r = self._section_label(g, r, "Annotations")
        r = self._checkbox(g, r, "annotate", " Show values in cells")
        fr_fmt = ttk.Frame(g)
        fr_fmt.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=(0, 8)); r += 1
        ttk.Label(fr_fmt, text="Format:", font=("Helvetica Neue", 12)).pack(side="left")
        _fmt_cb = PCombobox(fr_fmt, textvariable=self._vars["heatmap_fmt"],
                               values=[".2f", ".1f", ".0f", ".3f", ".2e", ".1%"],
                               state="normal", width=8, font=("Menlo", 12))
        _fmt_cb.pack(side="left", padx=(6, 0))
        add_placeholder(_fmt_cb, self._vars["heatmap_fmt"], "e.g. .2f")
        r = self._sep(g, r)

        r = self._section_label(g, r, "Clustering")
        r = self._checkbox(g, r, "cluster_rows", " Cluster rows (hierarchical)")
        r = self._checkbox(g, r, "cluster_cols", " Cluster columns (hierarchical)")
        r = self._hint(g, r, "Average-linkage Euclidean clustering. "
                             "Reorders rows/columns by similarity.")
        self._tab_footer(g, r)

    def _tab_stats_two_way_anova(self, f):
        """Stats tab for two-way ANOVA."""
        self._init_vars({
            "show_stats": False, "show_posthoc": False,
            "show_p_values": False, "show_effect_size": False,
            "show_normality_warning": True,
        })
        g = f; r = 0

        r = self._section_label(g, r, "Two-Way ANOVA")
        r = self._checkbox(g, r, "show_stats",       " Show ANOVA table on plot")
        r = self._checkbox(g, r, "show_posthoc",     " Show post-hoc brackets (pairwise t-tests, Holm corrected)")
        r = self._checkbox(g, r, "show_p_values",    " Show raw p-values (instead of stars)")
        r = self._checkbox(g, r, "show_effect_size", " Show effect size (partial η²)")
        r = self._sep(g, r)
        r = self._section_label(g, r, "Normality Check")
        r = self._checkbox(g, r, "show_normality_warning", " Show normality warning on plot")
        r = self._sep(g, r)
        r = self._hint(g, r,
                       "Type II Sum of Squares. Tests each main effect after "
                       "controlling for the other. Interaction term tests whether "
                       "the effect of Factor A differs across levels of Factor B.\n\n"
                       "Data must be in long format: one row per observation, "
                       "with separate columns for Factor A, Factor B, and Value.")
        self._tab_footer(g, r)

    # ── Help Analyze ─────────────────────────────────────────────────────────

    # Flat row-1-header charts: group names in row 0, values in rows 1+
    _HELP_ANALYZE_FLAT_MODES = {
        "bar", "box", "violin", "subcolumn_scatter", "dot_plot",
        "histogram", "column_stats", "before_after", "repeated_measures",
    }

    # Charts with full separate implementations
    _HELP_ANALYZE_FULL_MODES = {"grouped_bar", "scatter", "line", "two_way_anova"}

    # Charts where Help Analyze shows an informational panel (no live decision tree)
    _HELP_ANALYZE_INFO_MODES = {
        "kaplan_meier", "heatmap", "contingency", "chi_square_gof",
        "stacked_bar", "bubble", "bland_altman", "forest_plot", "curve_fit",
    }

    def _help_analyze(self):
        """Dispatcher: validate then route to the correct analysis path."""
        if not self._validated:
            messagebox.showwarning(
                "Help Analyze",
                "Please validate the spreadsheet first (click Validate Spreadsheet).")
            return
        mode = self._plot_type.get()
        if mode in self._HELP_ANALYZE_FLAT_MODES:
            self._help_analyze_flat()
        elif mode == "grouped_bar":
            self._help_analyze_grouped_bar()
        elif mode == "scatter":
            self._help_analyze_scatter()
        elif mode == "line":
            self._help_analyze_line()
        elif mode == "two_way_anova":
            self._help_analyze_two_way_anova()
        elif mode in self._HELP_ANALYZE_INFO_MODES:
            self._help_analyze_info_only(mode)
        else:
            label_map = {s.key: s.label for s in _REGISTRY_SPECS}
            messagebox.showinfo(
                "Help Analyze",
                f"Help Analyze is not yet implemented for "
                f"{label_map.get(mode, mode)} charts.")

    # ── Shared data / stats helpers ───────────────────────────────────────────

    def _ha_load_df(self, header=None):
        """Load the active Excel sheet.  Returns DataFrame or None on error."""
        pd = _pd()
        path      = self._vars["excel_path"].get().strip()
        sheet     = self._vars["sheet"].get().strip()
        sheet_arg = int(sheet) if sheet.lstrip("-").isdigit() else (sheet or 0)
        try:
            return pd.read_excel(path, sheet_name=sheet_arg, header=header)
        except Exception as e:
            messagebox.showerror("Help Analyze", f"Could not read file:\n{e}")
            return None

    def _ha_run_normality(self, groups: dict) -> dict:
        """Shapiro-Wilk per group.  Returns {name: {n, stat, p, normal, note}}."""
        from scipy import stats as _st
        out = {}
        for name, vals in groups.items():
            n = len(vals)
            if n < 3:
                out[name] = dict(n=n, stat=None, p=None, normal=True,
                                 note="too few values  -  assumed normal")
            elif n > 5000:
                out[name] = dict(n=n, stat=None, p=None, normal=True,
                                 note="n>5000  -  Shapiro-Wilk skipped")
            else:
                w, p = _st.shapiro(vals)
                out[name] = dict(n=n, stat=w, p=p, normal=(p > 0.05),
                                 note=f"W={w:.4f}, p={p:.4f}")
        return out

    def _ha_flat_decision_tree(self, groups, normality, mode):
        """Shared parametric/nonparametric/paired decision tree.
        Returns (rec_title, reasoning, test, posthoc, mc).

        Aligned with GraphPad Prism 11 Statistics Guide:
        - 2 groups: Welch's t-test (default since Prism 8, handles unequal variances)
        - 3+ groups, normal + equal var: One-way ANOVA + Tukey HSD (all-pairwise)
        - 3+ groups, normal + unequal var: Welch's ANOVA (Brown-Forsythe)  -  note only,
          app falls back to Parametric (one-way ANOVA) since Welch's ANOVA is not yet
          in the function library; flag is shown as a warning.
        - Non-normal: Mann-Whitney / Kruskal-Wallis + Dunn's (Holm corrected)
        - Paired: Paired t-test / Repeated pairwise / Friedman
        """
        from scipy import stats as _st
        k         = len(groups)
        ns        = {n: len(v) for n, v in groups.items()}
        all_norm  = all(d["normal"] for d in normality.values())
        equal_n   = len(set(ns.values())) == 1
        is_paired = mode in ("before_after", "repeated_measures")

        reasoning = []
        test, posthoc, mc, rec_title = "Parametric", "Tukey HSD", "Holm-Bonferroni", ""

        reasoning.append(("📊", f"{k} group(s): {', '.join(groups.keys())}"))

        if not all_norm:
            failing = [n for n, d in normality.items() if not d["normal"]]
            reasoning.append(("✗", f"Normality (Shapiro-Wilk): FAILED for "
                              f"{', '.join(failing)}."))
        else:
            reasoning.append(("✓", "Normality (Shapiro-Wilk): all groups pass (p > 0.05)."))

        # Levene equal-variance test (run for all group counts, not just k==2)
        levene_p = None
        if all_norm and len(groups) >= 2:
            try:
                _, levene_p = _st.levene(*list(groups.values()))
            except Exception:
                pass

        if is_paired:
            if equal_n:
                reasoning.append(("✓", "Equal n  -  paired comparison is appropriate."))
            else:
                reasoning.append(("⚠", "Unequal n  -  paired test requires equal n; "
                                   "defaulting to unpaired."))

        if is_paired and equal_n:
            if k == 2:
                test, rec_title = "Paired", "Paired t-test"
                reasoning.append((" -> ", "RECOMMENDATION: Paired t-test"))
                reasoning.append(("ℹ", "Each subject measured twice; differences are "
                                   "normally distributed by CLT. Paired t-test is the "
                                   "Prism default for 2-condition within-subjects designs."))
            elif all_norm:
                test, rec_title = "Paired", "Repeated-measures paired t-tests (Holm corrected)"
                reasoning.append((" -> ", "RECOMMENDATION: Repeated pairwise paired t-tests "
                                   "+ Holm-Bonferroni correction"))
                reasoning.append(("ℹ", "Multiple conditions, same subjects. Prism 11 uses "
                                   "repeated-measures one-way ANOVA with optional "
                                   "Geisser-Greenhouse correction for non-sphericity. "
                                   "Claude Plotter uses pairwise paired t-tests with "
                                   "Holm correction as a robust equivalent."))
            else:
                test, rec_title = "Non-parametric", "Friedman test + Dunn's post-hoc (Holm corrected)"
                reasoning.append((" -> ", "RECOMMENDATION: Non-parametric  -  Friedman + Dunn's"))
                reasoning.append(("ℹ", "Non-normal repeated data: Friedman test is the "
                                   "non-parametric analogue of RM-ANOVA. Dunn's post-hoc "
                                   "performs rank-based pairwise comparisons (Prism 11, "
                                   "Non-parametric comparisons section)."))
        elif k == 2:
            if all_norm:
                test, rec_title = "Parametric", "Welch's unpaired t-test"
                reasoning.append((" -> ", "RECOMMENDATION: Welch's unpaired t-test (parametric)"))
                reasoning.append(("ℹ", "Two groups, normally distributed. Welch's t-test "
                                   "is the default in GraphPad Prism since version 8  -  it "
                                   "does not assume equal variances, so it is correct "
                                   "regardless of the Levene result."))
                if levene_p is not None and levene_p < 0.05:
                    reasoning.append(("⚠", f"Levene's test: unequal variances "
                                      f"(p={levene_p:.4f}). Welch's handles this correctly "
                                      " -  no action needed."))
                else:
                    if levene_p is not None:
                        reasoning.append(("✓", f"Levene's test: equal variances "
                                          f"(p={levene_p:.4f}). Welch's t-test is still preferred."))
            else:
                test, rec_title = "Non-parametric", "Mann-Whitney U test"
                reasoning.append((" -> ", "RECOMMENDATION: Mann-Whitney U (non-parametric)"))
                reasoning.append(("ℹ", "Two groups, non-normal. Mann-Whitney U compares "
                                   "rank sums. Non-parametric alternative to Welch's t-test "
                                   "(Prism 11, Non-parametric comparisons guide)."))
        else:  # k >= 3
            if all_norm:
                test, rec_title = "Parametric", "One-way ANOVA + Tukey HSD post-hoc"
                reasoning.append((" -> ", "RECOMMENDATION: One-way ANOVA + Tukey HSD (all-pairwise)"))
                reasoning.append(("ℹ", f"{k} groups, normally distributed. One-way ANOVA "
                                   "tests for any group difference; Tukey HSD controls "
                                   "family-wise error across all pairwise comparisons  -  "
                                   "the Prism 11 default for 3+ parametric groups."))
                if levene_p is not None and levene_p < 0.05:
                    reasoning.append(("⚠",
                                      f"Levene's test: unequal variances (p={levene_p:.4f}). "
                                      "Prism 8+ offers Welch's ANOVA and Brown-Forsythe ANOVA "
                                      "as alternatives when equal variances cannot be assumed. "
                                      "Consider this if group SDs differ substantially."))
                elif levene_p is not None:
                    reasoning.append(("✓", f"Levene's test: equal variances "
                                      f"(p={levene_p:.4f}). Standard ANOVA assumption holds."))
            else:
                test, rec_title = "Non-parametric", "Kruskal-Wallis + Dunn's post-hoc (Holm corrected)"
                reasoning.append((" -> ", "RECOMMENDATION: Kruskal-Wallis + Dunn's post-hoc"))
                reasoning.append(("ℹ", f"{k} groups, non-normal. Kruskal-Wallis is the "
                                   "non-parametric one-way ANOVA equivalent (rank-based). "
                                   "Dunn's test with Holm-Bonferroni correction performs "
                                   "pairwise comparisons on ranked data "
                                   "(Prism 11, Non-parametric comparisons guide)."))

        return rec_title, reasoning, test, posthoc, mc

    def _ha_apply_standard_stats(self, test, posthoc, mc):
        """Write standard stats dropdowns and refresh the dependent widgets.

        Always resets comparison_mode to 0 (all-pairwise)  -  Help Analyze
        always recommends an omnibus comparison. User can override afterwards.
        """
        self._vars["stats_test"].set(test)
        self._vars["posthoc"].set(posthoc)
        self._vars["mc_correction"].set(mc)
        self._vars["show_stats"].set(True)
        self._vars["show_normality_warning"].set(True)
        # Reset comparison mode to all-pairwise (Prism default for omnibus tests)
        try:
            self._vars["comparison_mode"].set(0)
        except Exception:
            pass
        if hasattr(self, "_test_info_lbl") and hasattr(self, "_TEST_INFO"):
            self._test_info_lbl.config(text=self._TEST_INFO.get(test, ""))
        for fn in ("_tog_posthoc", "_tog_perm", "_tog_control"):
            try:
                getattr(self, fn)()
            except Exception:
                pass

    # ── Flat-layout analysis ──────────────────────────────────────────────────

    def _help_analyze_flat(self):
        """Full decision-tree analysis for flat row-1-header charts."""
        pd  = _pd()
        mode = self._plot_type.get()
        df   = self._ha_load_df()
        if df is None:
            return

        groups = {}
        for ci in range(df.shape[1]):
            hdr = df.iloc[0, ci]
            if pd.isna(hdr) or not str(hdr).strip():
                continue
            name = str(hdr).strip()
            vals = pd.to_numeric(df.iloc[1:, ci], errors="coerce").dropna().values
            if len(vals) >= 1:
                groups[name] = vals

        if len(groups) < 2:
            messagebox.showwarning("Help Analyze",
                                   "Need at least 2 groups with numeric data to recommend a test.")
            return

        normality = self._ha_run_normality(groups)
        rec_title, reasoning, test, posthoc, mc = self._ha_flat_decision_tree(
            groups, normality, mode)
        self._ha_apply_standard_stats(test, posthoc, mc)
        self._show_analyze_dialog(rec_title, groups, normality, reasoning, test, posthoc, mc)

    # ── Grouped Bar ───────────────────────────────────────────────────────────

    def _help_analyze_grouped_bar(self):
        """Two-factor decision tree for grouped bar charts.
        Pools values per subgroup across all categories for the normality check,
        then applies the standard parametric/non-parametric branch."""
        import numpy as np
        pd = _pd()
        df = self._ha_load_df()
        if df is None:
            return
        if df.shape[0] < 3:
            messagebox.showwarning("Help Analyze",
                                   "Need at least 3 rows: Row 1=categories, "
                                   "Row 2=subgroups, Rows 3+=data.")
            return

        row1 = [str(h) if pd.notna(h) else "" for h in df.iloc[0]]
        row2 = [str(h) if pd.notna(h) else "" for h in df.iloc[1]]
        cats = list(dict.fromkeys(h for h in row1 if h.strip()))
        subs = list(dict.fromkeys(h for h in row2 if h.strip()))

        # Pool all values per subgroup across categories for normality
        groups = {}
        for ci in range(df.shape[1]):
            sub = row2[ci] if ci < len(row2) else ""
            if not sub.strip():
                continue
            vals = pd.to_numeric(df.iloc[2:, ci], errors="coerce").dropna().values
            if len(vals) < 1:
                continue
            groups[sub] = np.concatenate([groups[sub], vals]) if sub in groups else vals

        if len(groups) < 2:
            messagebox.showwarning("Help Analyze", "Need at least 2 subgroups with data.")
            return

        normality  = self._ha_run_normality(groups)
        all_normal = all(d["normal"] for d in normality.values())
        k          = len(subs)

        reasoning = []
        reasoning.append(("📊",
            f"{len(cats)} categor{'y' if len(cats) == 1 else 'ies'} × "
            f"{k} subgroup{'s' if k != 1 else ''}  -  "
            f"cats: {', '.join(cats[:4])}{'…' if len(cats) > 4 else ''}; "
            f"subs: {', '.join(subs)}"))
        reasoning.append(("ℹ",
            "Pairwise stats compare subgroups within each category cluster. "
            "'Per-Category ANOVA' additionally tests within each cluster independently."))

        if all_normal:
            reasoning.append(("✓", "Normality (Shapiro-Wilk, pooled per subgroup): all pass."))
            test, posthoc, mc = "Parametric", "Tukey HSD", "Holm-Bonferroni"
            verb = "Welch's t-test" if k == 2 else f"One-way ANOVA + Tukey HSD"
            rec_title = f"{verb} across {k} subgroup{'s' if k != 1 else ''}"
            reasoning.append((" -> ", f"RECOMMENDATION: Parametric  -  {verb} across subgroups"))
            reasoning.append(("ℹ",
                "Also enabling 'Per-Category ANOVA' to test within each "
                "category cluster independently."))
        else:
            failing = [n for n, d in normality.items() if not d["normal"]]
            reasoning.append(("✗", f"Normality FAILED for: {', '.join(failing)}."))
            test, posthoc, mc = "Non-parametric", "Tukey HSD", "Holm-Bonferroni"
            rec_title = "Kruskal-Wallis + Dunn's post-hoc (Holm corrected) across subgroups"
            reasoning.append((" -> ",
                "RECOMMENDATION: Non-parametric  -  Kruskal-Wallis + Dunn's across subgroups"))
            reasoning.append(("ℹ",
                "'Per-Category ANOVA' is disabled for non-parametric tests."))

        self._ha_apply_standard_stats(test, posthoc, mc)
        try:
            self._vars["show_anova_per_group"].set(all_normal)
        except Exception:
            pass
        self._show_analyze_dialog(rec_title, groups, normality, reasoning, test, posthoc, mc)

    # ── Scatter ───────────────────────────────────────────────────────────────

    def _help_analyze_scatter(self):
        """Pearson vs Spearman decision based on regression-residual normality."""
        import numpy as np
        pd = _pd()
        df = self._ha_load_df()
        if df is None:
            return

        # Row 0 col 1+ = series names; col 0 = X; rows 1+ = data
        series_names = [
            str(df.iloc[0, ci]) if pd.notna(df.iloc[0, ci]) else f"Series {ci}"
            for ci in range(1, df.shape[1])
        ]
        x_raw = pd.to_numeric(df.iloc[1:, 0], errors="coerce")
        if x_raw.dropna().empty:
            messagebox.showwarning("Help Analyze",
                                   "Column 1 must contain numeric X values.")
            return

        series_data = {}
        for i, name in enumerate(series_names):
            y_raw = pd.to_numeric(df.iloc[1:, i + 1], errors="coerce")
            mask  = x_raw.notna() & y_raw.notna()
            if mask.sum() >= 3:
                series_data[name] = (x_raw[mask].values, y_raw[mask].values)

        if not series_data:
            messagebox.showwarning("Help Analyze",
                                   "No series with 3+ paired X/Y values found.")
            return

        from scipy import stats as _st
        # Normality on residuals  -  the correct Pearson assumption
        resid_groups = {}
        for name, (x, y) in series_data.items():
            slope, intercept, *_ = _st.linregress(x, y)
            resid_groups[f"{name} (residuals)"] = y - (slope * x + intercept)

        normality  = self._ha_run_normality(resid_groups)
        all_normal = all(d["normal"] for d in normality.values())

        reasoning = []
        reasoning.append(("📊", f"{len(series_data)} series: {', '.join(series_data.keys())}"))
        reasoning.append(("ℹ",
            "Pearson's assumption of bivariate normality is tested via "
            "Shapiro-Wilk on the regression residuals  -  the standard diagnostic approach."))

        if all_normal:
            reasoning.append(("✓",
                "Residual normality (Shapiro-Wilk): all series pass (p > 0.05)."))
            corr_type = "Pearson"
            rec_title = "Pearson r correlation + linear regression"
            reasoning.append((" -> ",
                "RECOMMENDATION: Pearson r  -  parametric linear correlation"))
            reasoning.append(("ℹ",
                "Pearson r measures linear association. Enable 'Regression line' "
                "and '95% CI band' for a complete GraphPad-style scatter output."))
        else:
            failing = [n.replace(" (residuals)", "")
                       for n, d in normality.items() if not d["normal"]]
            reasoning.append(("✗",
                f"Residual normality FAILED for: {', '.join(failing)}."))
            corr_type = "Spearman"
            rec_title = "Spearman ρ rank correlation"
            reasoning.append((" -> ",
                "RECOMMENDATION: Spearman ρ  -  non-parametric rank correlation"))
            reasoning.append(("ℹ",
                "Spearman ρ ranks X and Y independently and correlates the ranks. "
                "Robust to outliers and non-linearity. A regression line is still "
                "shown for visual reference."))

        # Pearson vs Spearman discrepancy hint
        for name, (x, y) in series_data.items():
            rp, _ = _st.pearsonr(x, y)
            rs, _ = _st.spearmanr(x, y)
            if abs(abs(rp) - abs(rs)) > 0.10:
                reasoning.append(("⚠",
                    f"Series '{name}': Pearson r={rp:.3f} vs Spearman ρ={rs:.3f} "
                    "(>0.10 gap). This suggests non-linearity or influential outliers  -  "
                    "inspect a residuals plot."))

        # Apply settings
        try:
            self._vars["correlation_type"].set(
                "Pearson" if corr_type == "Pearson" else "Spearman")
            self._vars["show_correlation"].set(True)
            self._vars["show_regression"].set(True)
            self._vars["show_ci_band"].set(True)
        except Exception:
            pass

        self._show_analyze_dialog(
            rec_title, resid_groups, normality, reasoning,
            corr_type, " - ", " - ",
            norm_label="Normality Check  -  Regression Residuals (Shapiro-Wilk)",
            settings_override={
                "Correlation type":         (corr_type,
                    "Pearson: linear, parametric  |  Spearman: rank-based, robust"),
                "Show correlation annotation": ("✓ Enabled", ""),
                "Show regression line":        ("✓ Enabled", ""),
                "Show 95% CI band":            ("✓ Enabled", ""),
            })

    # ── Line graph ────────────────────────────────────────────────────────────

    def _help_analyze_line(self):
        """Decision tree for line graphs  -  each Y series is treated as a group."""
        import numpy as np
        pd   = _pd()
        df   = self._ha_load_df()
        if df is None:
            return

        # Row 0 col 1+ = series names, rows 1+ = Y data per series
        series_names = [
            str(df.iloc[0, ci]) if pd.notna(df.iloc[0, ci]) else f"Series {ci}"
            for ci in range(1, df.shape[1])
        ]
        groups = {}
        for i, name in enumerate(series_names):
            vals = pd.to_numeric(df.iloc[1:, i + 1], errors="coerce").dropna().values
            if len(vals) >= 1:
                groups[name] = (np.concatenate([groups[name], vals])
                                if name in groups else vals)

        if len(groups) < 2:
            messagebox.showwarning("Help Analyze",
                                   "Need at least 2 series with numeric data.")
            return

        normality = self._ha_run_normality(groups)
        rec_title, reasoning, test, posthoc, mc = self._ha_flat_decision_tree(
            groups, normality, "line")

        # Inject line-specific context after the group count line
        reasoning.insert(1, ("ℹ",
            "Line-graph series are treated as independent groups  -  "
            "each Y-column's values are compared across series "
            "(one test per series pair or omnibus across all)."))

        self._ha_apply_standard_stats(test, posthoc, mc)
        self._show_analyze_dialog(rec_title, groups, normality, reasoning, test, posthoc, mc)

    # ── Two-Way ANOVA ─────────────────────────────────────────────────────────

    def _help_analyze_two_way_anova(self):
        """Normality + Levene checks for two-way ANOVA (long-format data)."""
        from scipy import stats as _st
        # Two-way ANOVA uses header=0 (first row is column names)
        df = self._ha_load_df(header=0)
        if df is None:
            return

        num_cols = [c for c in df.columns
                    if _pd().to_numeric(df[c], errors="coerce").notna().any()]
        cat_cols = [c for c in df.columns if c not in num_cols]

        if len(cat_cols) < 2 or not num_cols:
            messagebox.showwarning("Help Analyze",
                                   "Need ≥2 categorical columns (Factor A, Factor B) "
                                   "and 1 numeric column (Value).")
            return

        fa, fb, dv = cat_cols[0], cat_cols[1], num_cols[-1]
        data     = df[[fa, fb, dv]].dropna()
        a_levels = sorted(data[fa].unique())
        b_levels = sorted(data[fb].unique())

        # One group per cell (Factor A × Factor B)
        groups = {}
        for a in a_levels:
            for b in b_levels:
                key  = f"{a} / {b}"
                vals = data[(data[fa] == a) & (data[fb] == b)][dv].values
                if len(vals) >= 1:
                    groups[key] = vals.astype(float)

        if not groups:
            messagebox.showwarning("Help Analyze", "No data cells found after filtering.")
            return

        normality  = self._ha_run_normality(groups)
        all_normal = all(d["normal"] for d in normality.values())

        levene_p = None
        try:
            eligible = [v for v in groups.values() if len(v) >= 2]
            if len(eligible) >= 2:
                _, levene_p = _st.levene(*eligible)
        except Exception:
            pass

        reasoning = []
        n_cells = len(a_levels) * len(b_levels)
        reasoning.append(("📊",
            f"Design: {len(a_levels)} levels of '{fa}' × "
            f"{len(b_levels)} levels of '{fb}' = {n_cells} cells,  DV = '{dv}'"))
        reasoning.append(("ℹ",
            f"Two-way ANOVA partitions variance into: main effect of '{fa}', "
            f"main effect of '{fb}', and their interaction ('{fa}' × '{fb}'). "
            "Type II SS  -  each main effect is tested after controlling for the other."))

        if all_normal:
            reasoning.append(("✓",
                "Normality (Shapiro-Wilk, per cell): all cells pass (p > 0.05)."))
        else:
            failing = [n for n, d in normality.items() if not d["normal"]]
            reasoning.append(("✗",
                f"Normality FAILED in {len(failing)} cell(s): "
                f"{', '.join(failing[:4])}{'…' if len(failing) > 4 else ''}."))
            reasoning.append(("⚠",
                "Two-way ANOVA is moderately robust to non-normality when cell sizes "
                "are equal (balanced design). Consider log-transforming a right-skewed DV."))

        if levene_p is not None:
            if levene_p < 0.05:
                reasoning.append(("⚠",
                    f"Levene's test: unequal variances across cells (p={levene_p:.4f}). "
                    "Two-way ANOVA assumes homoscedasticity  -  interpret with caution "
                    "if cell sizes are very unequal."))
            else:
                reasoning.append(("✓",
                    f"Levene's test: homogeneous variances across cells (p={levene_p:.4f})."))

        rec_title = f"Two-Way ANOVA: '{fa}' × '{fb}'"
        reasoning.append((" -> ",
            f"RECOMMENDATION: Two-Way ANOVA  -  main effects of '{fa}' and '{fb}' "
            "plus interaction term"))
        reasoning.append(("ℹ",
            "Enabling post-hoc pairwise comparisons (Holm corrected) and "
            "partial η² effect sizes  -  standard Prism output for two-way ANOVA."))

        # Apply settings
        try:
            self._vars["show_stats"].set(True)
            self._vars["show_posthoc"].set(True)
            self._vars["show_effect_size"].set(True)
            self._vars["show_normality_warning"].set(True)
            self._vars["show_p_values"].set(False)
        except Exception:
            pass

        self._show_analyze_dialog(
            rec_title, groups, normality, reasoning,
            "Parametric", "Holm-Bonferroni post-hoc", "Holm-Bonferroni",
            norm_label=f"Normality Check  -  Per Cell ({fa} × {fb}, Shapiro-Wilk)",
            settings_override={
                "Test":                    ("Two-Way ANOVA (Type II SS)",
                    f"Main effects: '{fa}', '{fb}' + interaction term"),
                "Post-hoc":                ("Pairwise t-tests, Holm corrected",
                    "✓ Enabled"),
                "Effect size":             ("Partial η²",
                    "✓ Enabled"),
                "Normality warning on plot": ("✓ Enabled", ""),
            })

    # ── Informational stubs ───────────────────────────────────────────────────

    _HELP_ANALYZE_INFO = {
        "kaplan_meier": {
            "title":     "Kaplan-Meier Survival Curve",
            "purpose":   "Estimates the survival function S(t)  -  the probability of surviving "
                         "beyond time t  -  from censored time-to-event data. Each step in the "
                         "curve corresponds to one or more observed events.",
            "test":      "Log-rank test (Mantel-Cox) compares survival distributions between "
                         "groups. It is non-parametric and handles right-censored observations "
                         "correctly. Output: χ² statistic and p-value.",
            "interpret": "Curves that diverge early and stay separated indicate a large, early "
                         "survival difference. Crossing curves suggest time-varying treatment "
                         "effects  -  the log-rank test has low power in this case; consider "
                         "Breslow (Gehan) or Tarone-Ware weighted alternatives.",
            "apply":     "Enable 'Show censors' to display tick marks for censored observations. "
                         "Enable 'Show at-risk table' for a publication-quality plot. "
                         "The log-rank p-value is computed automatically when 2+ groups are present.",
        },
        "heatmap": {
            "title":     "Heatmap",
            "purpose":   "Visualises a numeric matrix as a color grid. Common uses: gene "
                         "expression data, correlation matrices, and confusion tables.",
            "test":      "No statistical test is embedded in the heatmap. For expression data, "
                         "apply differential-expression analysis (DESeq2, edgeR) upstream. "
                         "For correlation matrices, each cell is a pre-computed Pearson or "
                         "Spearman r coefficient.",
            "interpret": "Color saturation encodes magnitude. Hierarchical clustering (enable "
                         "in Options) reorders rows/columns by similarity  -  look for block "
                         "structure indicating co-regulated genes or correlated variables. "
                         "Use a diverging palette for data centred at zero (fold-changes, z-scores).",
            "apply":     "Enable 'Show values in cells' for small matrices (≤10×10). "
                         "Enable row and/or column clustering to reveal structure. "
                         "Use Coolwarm or RdBu for data centred at zero; Viridis or Mako for "
                         "magnitude-only data.",
        },
        "contingency": {
            "title":     "Contingency Table",
            "purpose":   "Displays observed counts for combinations of two categorical variables, "
                         "with optional expected counts and row/column percentages.",
            "test":      "Chi-square test of independence (Pearson χ²) when all expected cells "
                         "≥ 5, otherwise Fisher's exact test. Tests whether the row and column "
                         "variables are associated. Cramér's V gives effect size (0=no association, "
                         "1=perfect association).",
            "interpret": "p < 0.05 means the variables are not independent. Inspect the "
                         "residuals (observed − expected) to see which cells drive the association. "
                         "Fisher's exact is preferred for small N or sparse tables "
                         "(any expected cell < 5).",
            "apply":     "Enable 'Show expected counts' to verify the chi-square assumption. "
                         "Enable '% of row' or '% of column' to compare proportions across groups.",
        },
        "chi_square_gof": {
            "title":     "Chi-Square Goodness of Fit",
            "purpose":   "Tests whether observed categorical counts match a specified theoretical "
                         "distribution  -  for example, equal proportions (uniform null) or Hardy-"
                         "Weinberg equilibrium.",
            "test":      "Pearson chi-square: χ² = Σ(O − E)²/E, df = k − 1. "
                         "Expected counts can be equal (uniform) or user-specified proportions "
                         "entered in row 3 of the spreadsheet.",
            "interpret": "Reject H₀ when p < 0.05 (observed deviates from expected). "
                         "Calculate each term (O−E)²/E to see which categories contribute most. "
                         "The test requires expected count ≥ 5 per cell; merge rare categories "
                         "if this is violated.",
            "apply":     "Select 'Equal expected proportions' for a uniform null hypothesis. "
                         "Enter custom expected values in row 3 to test a different distribution.",
        },
        "stacked_bar": {
            "title":     "Stacked Bar Chart",
            "purpose":   "Shows part-to-whole composition  -  each bar is divided into segments "
                         "for sub-categories, displayed as absolute values or as percentages "
                         "of the bar total.",
            "test":      "No statistical test is performed on stacked bars directly. "
                         "For count data, use a Contingency or Chi-Square GoF analysis on the "
                         "same data. For continuous sub-category data, analyze each component "
                         "separately with a Bar or Box plot.",
            "interpret": "Use 'Percent mode' to compare proportions independent of group size. "
                         "Stacked bars are best for 3–6 sub-categories; with more segments, "
                         "a grouped bar chart is easier to read.",
            "apply":     "Switch to 'Percent' mode (Options tab) for proportional comparison. "
                         "Enable 'Value labels' to annotate each segment with its value.",
        },
        "bubble": {
            "title":     "Bubble Chart",
            "purpose":   "Extends a scatter plot with a third quantitative dimension encoded "
                         "as bubble area. Each series provides X, Y, and Size columns.",
            "test":      "No statistical test is performed automatically. "
                         "For X–Y correlation: use a Scatter plot with Pearson or Spearman. "
                         "For size vs. X or Y: run a separate scatter/regression analysis on "
                         "those two variables.",
            "interpret": "Bubble area encodes magnitude  -  ensure the scale is labelled. "
                         "Overlap obscures data; reduce 'Bubble scale' if needed. "
                         "Avoid bubble charts when precise size comparisons are critical.",
            "apply":     "Adjust 'Bubble scale' in the Options tab. Enable 'Show labels' to "
                         "identify individual observations by name.",
        },
        "bland_altman": {
            "title":     "Bland-Altman Agreement Plot",
            "purpose":   "Assesses agreement between two measurement methods by plotting the "
                         "difference (Method A − Method B) against the mean of the two methods "
                         "for each subject.",
            "test":      "No formal hypothesis test is required. Key statistics: mean difference "
                         "(systematic bias) and 95% limits of agreement (mean diff ± 1.96 SD). "
                         "A one-sample t-test on the differences tests whether bias ≠ 0.",
            "interpret": "The limits of agreement define the range within which 95% of "
                         "differences fall. They should be clinically acceptable for the methods "
                         "to be interchangeable. A trend in difference vs. mean indicates "
                         "proportional bias  -  consider log-transforming before analysis.",
            "apply":     "Enable '95% CI bands' to show uncertainty around the mean difference "
                         "and limits of agreement. Small N (< 30) makes the limits imprecise; "
                         "report them with CIs in papers.",
        },
        "forest_plot": {
            "title":     "Forest Plot (Meta-Analysis)",
            "purpose":   "Displays effect estimates and confidence intervals from multiple "
                         "studies alongside a pooled summary estimate.",
            "test":      "Pooled estimate: inverse-variance weighted mean. "
                         "Heterogeneity: Cochran's Q test + I² statistic. "
                         "I² > 50% suggests substantial between-study heterogeneity  -  a "
                         "random-effects model is then preferred over fixed-effects.",
            "interpret": "Studies whose CI excludes the null line are individually significant. "
                         "The summary diamond's width is the pooled 95% CI. "
                         "High I² means the studies may not be estimating the same effect  -  "
                         "explore subgroup analyzes or moderators.",
            "apply":     "Set the null reference line to 0 (mean difference) or 1 (odds ratio / "
                         "hazard ratio) in the Stats tab. Enable 'Show weights' to display each "
                         "study's inverse-variance contribution to the pooled estimate.",
        },
        "curve_fit": {
            "title":     "Nonlinear Curve Fit",
            "purpose":   "Fits a parametric model (sigmoidal, exponential, Michaelis-Menten, "
                         "Gaussian, etc.) to XY data using nonlinear least-squares regression.",
            "test":      "R² and RMSE measure goodness of fit. Fitted parameters are reported "
                         "with 95% confidence intervals derived from the covariance matrix. "
                         "For sigmoidal models the inflection point is EC50/IC50.",
            "interpret": "A good fit has R² > 0.95 and small, random residuals. "
                         "A systematic pattern in the residuals means the model is misspecified  -  "
                         "try a different equation. Hill coefficient n > 1 indicates positive "
                         "cooperativity; n < 1 negative cooperativity.",
            "apply":     "Select the model that matches your biological mechanism. "
                         "Enable 'Show residuals' subplot to diagnose fit quality. "
                         "Enable 'Show parameter table' for EC50/IC50 with confidence intervals "
                         " -  required for publication.",
        },
    }

    def _help_analyze_info_only(self, mode):
        """Show an informational Help Analyze panel for charts without a live decision tree."""
        info = self._HELP_ANALYZE_INFO.get(mode)
        if info is None:
            label_map = {s.key: s.label for s in _REGISTRY_SPECS}
            messagebox.showinfo("Help Analyze",
                                f"Help Analyze is not yet available for "
                                f"{label_map.get(mode, mode)} charts.")
            return
        self._show_analyze_info_dialog(info)

    def _show_analyze_info_dialog(self, info: dict):
        """Informational modal for charts without a statistical decision tree."""
        dlg = tk.Toplevel(self)
        dlg.title("Help Analyze")
        dlg.resizable(True, True)
        dlg.geometry("620x480")
        dlg.configure(bg="#ffffff")
        dlg.grab_set()
        self._track_popup(dlg)  # suspend background scroll

        # Header
        hdr = tk.Frame(dlg, bg="#1a4f7a", height=56)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="✦  Help Analyze", bg="#1a4f7a", fg="white",
                 font=("Helvetica Neue", 16, "bold")).pack(side="left", padx=18, pady=14)
        tk.Label(hdr, text=info["title"], bg="#1a4f7a", fg="#a8cce0",
                 font=("Helvetica Neue", 11)).pack(side="left", padx=(0, 18), pady=14)

        # Scrollable body
        body_outer = tk.Frame(dlg, bg="#ffffff")
        body_outer.pack(fill="both", expand=True)
        canvas    = tk.Canvas(body_outer, bg="#ffffff", highlightthickness=0)
        scrollbar = ttk.Scrollbar(body_outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        body = tk.Frame(canvas, bg="#ffffff")
        cw   = canvas.create_window((0, 0), window=body, anchor="nw")
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(cw, width=e.width))
        body.bind("<Configure>",
                  lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        def _scroll(e):
            import sys
            raw = getattr(e, "delta", 0) or (120 if getattr(e, "num", 0) == 4 else -120)
            if not raw: return
            bbox = canvas.bbox("all")
            if not bbox: return
            ch, vh = bbox[3] - bbox[1], canvas.winfo_height()
            if ch <= vh: return
            scale = 8.0 if sys.platform == "darwin" and abs(raw) < 40 else 4.0
            canvas.yview_moveto(
                max(0., min(1. - vh / ch, canvas.yview()[0] - raw * scale / ch)))

        def _scroll_linux(e):
            raw = 120 if e.num == 4 else -120
            _scroll(type("E", (), {"delta": raw, "num": e.num})())

        dlg.after(50, lambda: _bind_scroll_recursive(dlg, _scroll, _scroll_linux, _scroll_linux))

        PAD2 = 20
        for title, bg, fg, key in [
            ("What it shows",        "#f4f7fb", "#1a4f7a", "purpose"),
            ("Statistical test",     "#f4f7fb", "#1a4f7a", "test"),
            ("How to interpret",     "#f4f7fb", "#1a4f7a", "interpret"),
            ("Recommended settings", "#fffbec", "#666633", "apply"),
        ]:
            frm = tk.Frame(body, bg=bg)
            frm.pack(fill="x", padx=PAD2, pady=(10, 2))
            tk.Label(frm, text=title, bg=bg, fg=fg,
                     font=("Helvetica Neue", 12, "bold")
                     ).pack(anchor="w", padx=10, pady=(8, 4))
            tk.Label(frm, text=info[key], bg=bg, fg="#222222",
                     font=("Helvetica Neue", 11),
                     wraplength=540, justify="left", anchor="nw"
                     ).pack(anchor="w", padx=14, pady=(0, 10), fill="x")

        # Footer
        ttk.Separator(dlg).pack(fill="x")
        foot = tk.Frame(dlg, bg="#ffffff")
        foot.pack(fill="x", padx=PAD2, pady=10)
        PButton(foot, text="Close", style="primary",
                command=dlg.destroy).pack(side="right")

    def _show_analyze_dialog(self, rec_title, groups, normality,
                             reasoning, test, posthoc, mc,
                             norm_label=None, settings_override=None):
        """Show the Help Analyze explanation modal  -  mirrors Prism's Analyze dialog."""
        import tkinter.font as tkfont

        dlg = tk.Toplevel(self)
        dlg.title("Help Analyze")
        dlg.resizable(True, True)
        dlg.geometry("620x560")
        dlg.configure(bg="#ffffff")
        dlg.grab_set()
        self._track_popup(dlg)  # suspend background scroll

        # ── Header ────────────────────────────────────────────────────────────
        hdr = tk.Frame(dlg, bg="#1a4f7a", height=56)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="✦  Help Analyze", bg="#1a4f7a", fg="white",
                 font=("Helvetica Neue", 16, "bold")).pack(side="left", padx=18, pady=14)
        tk.Label(hdr, text="Recommended statistical test based on your data",
                 bg="#1a4f7a", fg="#a8cce0",
                 font=("Helvetica Neue", 11)).pack(side="left", padx=(0, 18), pady=14)

        # ── Scrollable body ───────────────────────────────────────────────────
        body_outer = tk.Frame(dlg, bg="#ffffff")
        body_outer.pack(fill="both", expand=True, padx=0, pady=0)
        canvas   = tk.Canvas(body_outer, bg="#ffffff", highlightthickness=0)
        scrollbar = ttk.Scrollbar(body_outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        body = tk.Frame(canvas, bg="#ffffff")
        canvas_win = canvas.create_window((0, 0), window=body, anchor="nw")

        def _on_resize(e):
            canvas.itemconfig(canvas_win, width=e.width)
        canvas.bind("<Configure>", _on_resize)
        body.bind("<Configure>",
                  lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        def _scroll(e):
            import sys
            raw = getattr(e, "delta", 0)
            if not raw:
                raw = 120 if getattr(e, "num", 0) == 4 else -120
            bbox = canvas.bbox("all")
            if not bbox: return
            content_h = bbox[3] - bbox[1]
            view_h    = canvas.winfo_height()
            if content_h <= view_h: return
            scale = 8.0 if sys.platform == "darwin" and abs(raw) < 40 else 4.0
            frac = -raw * scale / content_h
            cur  = canvas.yview()[0]
            canvas.yview_moveto(max(0.0, min(1.0 - view_h / content_h, cur + frac)))

        def _scroll_linux(e):
            raw = 120 if e.num == 4 else -120
            _scroll(type("E", (), {"delta": raw, "num": e.num})())

        # Bind only to this popup's widget tree  -  avoids capturing parent window scroll
        dlg.after(50, lambda: _bind_scroll_recursive(dlg, _scroll, _scroll_linux, _scroll_linux))

        body.bind("<Configure>", lambda e:
            canvas.configure(scrollregion=canvas.bbox("all")))

        PAD2 = 20

        def _section(parent, title, bg="#f4f7fb"):
            frm = tk.Frame(parent, bg=bg, bd=0)
            frm.pack(fill="x", padx=PAD2, pady=(10, 2))
            tk.Label(frm, text=title, bg=bg, fg="#1a4f7a",
                     font=("Helvetica Neue", 12, "bold")
                     ).pack(anchor="w", padx=10, pady=(8, 4))
            return frm

        def _row(parent, icon, text, fg="#222222", bg="#f4f7fb"):
            r = tk.Frame(parent, bg=bg)
            r.pack(fill="x", padx=10, pady=1)
            tk.Label(r, text=icon, bg=bg, fg=fg,
                     font=("Helvetica Neue", 12), width=2,
                     anchor="n").pack(side="left", padx=(0, 6), pady=2)
            tk.Label(r, text=text, bg=bg, fg=fg,
                     font=("Helvetica Neue", 11),
                     wraplength=500, justify="left", anchor="nw"
                     ).pack(side="left", fill="x", expand=True, pady=2)

        # ── Recommendation pill ───────────────────────────────────────────────
        pill_frm = tk.Frame(body, bg="#e8f4e8", bd=0)
        pill_frm.pack(fill="x", padx=PAD2, pady=(14, 4))
        tk.Label(pill_frm, text="Recommended Test", bg="#e8f4e8", fg="#2d6a2d",
                 font=("Helvetica Neue", 10, "bold")).pack(anchor="w", padx=14, pady=(8, 0))
        tk.Label(pill_frm, text=rec_title, bg="#e8f4e8", fg="#1a4a1a",
                 font=("Helvetica Neue", 14, "bold")).pack(anchor="w", padx=14, pady=(2, 8))

        # ── Decision reasoning ────────────────────────────────────────────────
        reas_frm = _section(body, "Decision Reasoning")
        icon_color = {
            "✓": "#2d7d2d", "✗": "#c0392b", "⚠": "#c07a00",
            " -> ": "#1a4f7a", "ℹ": "#555555", "📊": "#333333",
        }
        for icon, text in reasoning:
            fg = icon_color.get(icon, "#333333")
            _row(reas_frm, icon, text, fg=fg, bg="#f4f7fb")

        # ── Per-group normality table ─────────────────────────────────────────
        norm_frm = _section(body, norm_label or "Normality Check (Shapiro-Wilk)")
        hdr_row = tk.Frame(norm_frm, bg="#dce6f2")
        hdr_row.pack(fill="x", padx=10, pady=(0, 2))
        for txt, w in [("Group", 16), ("n", 5), ("W stat", 10), ("p-value", 10), ("Result", 8)]:
            tk.Label(hdr_row, text=txt, bg="#dce6f2", fg="#1a4f7a",
                     font=("Helvetica Neue", 10, "bold"), width=w, anchor="w"
                     ).pack(side="left", padx=2, pady=3)

        for name, d in normality.items():
            row_bg = "#fff8f8" if not d["normal"] else "#f4f7fb"
            data_row = tk.Frame(norm_frm, bg=row_bg)
            data_row.pack(fill="x", padx=10, pady=1)
            stat_str = f"{d['stat']:.4f}" if d["stat"] is not None else " - "
            p_str    = f"{d['p']:.4f}"    if d["p"]    is not None else " - "
            res_str  = ("✗ FAIL" if not d["normal"] else "✓ pass")
            res_fg   = ("#c0392b" if not d["normal"] else "#2d7d2d")
            for txt, w, fg in [
                (name[:22], 16, "#222222"),
                (str(d["n"]), 5, "#555555"),
                (stat_str, 10, "#555555"),
                (p_str, 10, "#555555"),
                (res_str, 8, res_fg),
            ]:
                tk.Label(data_row, text=txt, bg=row_bg, fg=fg,
                         font=("Helvetica Neue", 10), width=w, anchor="w"
                         ).pack(side="left", padx=2, pady=2)

        # ── Applied settings summary ──────────────────────────────────────────
        app_frm = _section(body, "Settings Applied to Stats Tab", bg="#fffbec")
        if settings_override:
            # Custom settings table for charts with non-standard controls
            for lbl, (val, note) in settings_override.items():
                r2 = tk.Frame(app_frm, bg="#fffbec")
                r2.pack(fill="x", padx=10, pady=2)
                tk.Label(r2, text=lbl + ":", bg="#fffbec", fg="#666633",
                         font=("Helvetica Neue", 10, "bold"), width=28, anchor="w"
                         ).pack(side="left")
                tk.Label(r2, text=val, bg="#fffbec", fg="#222200",
                         font=("Helvetica Neue", 10), anchor="w"
                         ).pack(side="left")
                if note:
                    tk.Label(r2, text=f"  ({note})", bg="#fffbec", fg="#999977",
                             font=("Helvetica Neue", 9), anchor="w"
                             ).pack(side="left")
        else:
            settings_map = {
                "Parametric":     ("Statistical Test", test,
                                   "2 groups  ->  Welch's t-test  |  3+ groups  ->  One-way ANOVA"),
                "Non-parametric": ("Statistical Test", test,
                                   "2 groups  ->  Mann-Whitney U  |  3+ groups  ->  Kruskal-Wallis + Dunn's"),
                "Paired":         ("Statistical Test", test,
                                   "2 groups  ->  Paired t-test  |  3+ groups  ->  Repeated pairwise paired t-tests"),
            }
            s_label, s_val, s_note = settings_map.get(test, ("Statistical Test", test, ""))
            for lbl, val, note in [
                (s_label, s_val, s_note),
                ("Post-hoc (parametric)", posthoc if test == "Parametric" else "Dunn's test",
                 "Controls which pairs are compared after omnibus test"),
                ("Multiple Comparisons", mc,
                 "Correction applied to all pairwise p-values"),
                ("Show Significance Brackets", "✓ Enabled", ""),
                ("Normality Warning on Plot", "✓ Enabled", ""),
            ]:
                r2 = tk.Frame(app_frm, bg="#fffbec")
                r2.pack(fill="x", padx=10, pady=2)
                tk.Label(r2, text=lbl + ":", bg="#fffbec", fg="#666633",
                         font=("Helvetica Neue", 10, "bold"), width=28, anchor="w"
                         ).pack(side="left")
                tk.Label(r2, text=val, bg="#fffbec", fg="#222200",
                         font=("Helvetica Neue", 10), anchor="w"
                         ).pack(side="left")
                if note:
                    tk.Label(r2, text=f"  ({note})", bg="#fffbec", fg="#999977",
                             font=("Helvetica Neue", 9), anchor="w"
                             ).pack(side="left")

        # ── Prism reference note ──────────────────────────────────────────────
        ref_frm = tk.Frame(body, bg="#f0f0f0")
        ref_frm.pack(fill="x", padx=PAD2, pady=(8, 4))
        tk.Label(ref_frm,
                 text="ℹ  Follows GraphPad Prism's 'Analyze > Statistical comparisons' "
                      "decision logic. Welch's t-test is the Prism default (v8+) as it "
                      "does not assume equal variances. You can override any setting in "
                      "the Stats tab before generating the plot.",
                 bg="#f0f0f0", fg="#666666", wraplength=560, justify="left",
                 font=("Helvetica Neue", 10)
                 ).pack(anchor="w", padx=12, pady=8)

        # ── Footer buttons ────────────────────────────────────────────────────
        ttk.Separator(dlg).pack(fill="x", pady=0)
        foot = tk.Frame(dlg, bg="#ffffff")
        foot.pack(fill="x", padx=PAD2, pady=10)
        PButton(foot, text="Apply & Close", style="primary",
                command=dlg.destroy).pack(side="right")
        PButton(foot, text="Apply & Generate Plot", style="primary",
                command=lambda: (dlg.destroy(), self._run())
                ).pack(side="right", padx=(0, 8))
        PButton(foot, text="Close", style="ghost",
                command=lambda: (
                    self._vars["show_stats"].set(False),
                    dlg.destroy()
                )).pack(side="left")

    # ── Run ───────────────────────────────────────────────────────────────────

    def _run(self):
        if self._running: return
        if not self._validated:
            messagebox.showwarning("Validate first",
                                   "Please run Validate Spreadsheet before plotting."); return
        if self._pf is None:
            messagebox.showerror("Not ready", "Functions not loaded yet."); return
        excel = self._vars["excel_path"].get().strip()
        if not excel:
            messagebox.showwarning("Missing input", "Select an Excel file first."); return
        kw = self._collect(excel)
        if kw is None: return
        # Push a snapshot to the undo history before running
        import copy
        self._kw_history.append(copy.deepcopy(kw))
        self._undo_btn.config(state="normal")

        # Tag this render job to the active tab for thread safety
        tab    = self._tab_manager.active if self._tab_manager else None
        job_id = uuid.uuid4().hex
        if tab is not None:
            tab.render_job_id = job_id

        self._running = True
        self._run_btn.config(state="disabled")
        self._set_status("Generating plot...")
        self._start_spinner()
        tab_id = tab.tab_id if tab is not None else None
        threading.Thread(
            target=self._do_run,
            args=(kw, tab_id, job_id),
            daemon=True,
        ).start()

    def _undo(self):
        """Re-run the previous plot configuration (⌘Z)."""
        if self._running or len(self._kw_history) < 2:
            return
        # Pop the current state, leaving the one before it on top
        self._kw_history.pop()
        kw = self._kw_history[-1]
        import copy
        self._running = True
        self._run_btn.config(state="disabled")
        self._set_status("Undoing…")
        self._start_spinner()
        if len(self._kw_history) < 2:
            self._undo_btn.config(state="disabled")
        tab    = self._tab_manager.active if self._tab_manager else None
        job_id = uuid.uuid4().hex
        if tab is not None:
            tab.render_job_id = job_id
        tab_id = tab.tab_id if tab is not None else None
        threading.Thread(
            target=self._do_run,
            args=(copy.deepcopy(kw), tab_id, job_id),
            daemon=True,
        ).start()

    def _do_redo(self):
        """Redo using the UndoStack (⌘⇧Z)."""
        if hasattr(self, "_undo_stack") and self._undo_stack.can_redo():
            self._undo_stack.redo(self._vars)
            self._set_status("Redone")

    def _get_var(self, key, default=None):
        """Safely get a tkvar value, returning default if not present."""
        v = self._vars.get(key)
        if v is None:
            if default is None:
                return None
            if isinstance(default, bool):
                v = tk.BooleanVar(value=default)
            elif isinstance(default, int):
                v = tk.IntVar(value=default)
            else:
                v = tk.StringVar(value=str(default))
        return v.get()

    # ── Collect helpers  -  each gathers one logical group of kwargs ────────────

    def _collect_display(self, kw: dict):
        """Populate display/rendering kwargs from UI vars."""
        kw["error"]         = ERROR_TYPE_MAP.get(self._vars["error"].get(),
                                                   self._vars["error"].get())
        kw["show_points"]      = self._vars["show_points"].get()
        kw["show_n_labels"]    = self._vars["show_n_labels"].get()
        kw["show_value_labels"] = self._get_var("show_value_labels", False)
        kw["error_below_bar"] = self._get_var("error_below_bar", False)
        kw["gridlines"]       = self._get_var("gridlines", False)  # legacy compat
        # Grid style  -  "None"/"Horizontal"/"Full (H + V)"  ->  "none"/"horizontal"/"full"
        _gs_map = {"None": "none", "Horizontal": "horizontal", "Full (H + V)": "full"}
        raw_gs = self._get_var("grid_style", "None")
        kw["grid_style"] = _gs_map.get(raw_gs, "none")
        # Keep legacy gridlines bool in sync for functions that still check it
        kw["gridlines"]  = kw["grid_style"] != "none"
        kw["open_points"]   = self._get_var("open_points", False)
        kw["horizontal"]    = self._get_var("horizontal", False)
        kw["show_median"]   = self._get_var("show_median", False)
        kw["notch"]         = self._get_var("notch_box", False)
        kw["alpha"]         = self._get_float("bar_alpha", 0.85)
        kw["xscale"]        = "log" if self._get_var("xscale", "Linear") == "Log" else "linear"

        # ── Priority-1 styling params ─────────────────────────────────────────
        from plotter_functions import AXIS_STYLES, TICK_DIRS, LEGEND_POSITIONS
        raw_axis  = self._get_var("axis_style",  "Open (default)")
        raw_tick  = self._get_var("tick_dir",    "Outward (default)")
        raw_leg   = self._get_var("legend_pos",  "Upper right")
        kw["axis_style"]  = AXIS_STYLES.get(raw_axis,  "open")
        kw["tick_dir"]    = TICK_DIRS.get(raw_tick,    "out")
        kw["minor_ticks"] = self._get_var("minor_ticks", False)
        kw["point_size"]  = self._get_float("point_size",  6.0)
        kw["point_alpha"] = self._get_float("point_alpha", 0.80)
        kw["cap_size"]    = self._get_float("cap_size",    4.0)
        kw["legend_pos"]  = LEGEND_POSITIONS.get(raw_leg, "upper right")
        kw["spine_width"] = self._get_float("spine_width", 0.8)

        # Background color
        FIG_BG_MAP = {
            "White":       "white",
            "Light gray":  "#f5f5f5",
            "Transparent": "none",
            "Black":       "black",
        }
        raw_bg = self._get_var("fig_bg", "White")
        kw["fig_bg"] = FIG_BG_MAP.get(raw_bg, raw_bg.lower() if raw_bg else "white")

        cr = self._vars["color"].get().strip()
        if not cr or cr.lower() in ("none", "default (prism)"):
            kw["color"] = None
        elif cr.startswith("["):
            try:
                kw["color"] = json.loads(cr)
            except Exception:
                messagebox.showerror("Invalid color", f"Cannot parse: {cr}")
                return False
        else:
            kw["color"] = cr.strip("\"'")
        return True

    def _collect_labels(self, kw: dict):
        """Populate title/axis label kwargs."""
        for k in ("title", "xlabel", "ytitle"):
            kw[k] = self._vars[k].get().strip()
        kw["yscale"] = "log" if self._vars["yscale"].get() == "Log" else "linear"

        ylim_mode = self._get_var("ylim_mode", 0)
        if ylim_mode == 0:   # Auto
            kw["ylim"]           = None
            kw["_ylim_data_min"] = False
        elif ylim_mode == 1:  # Start at data min
            kw["ylim"]           = None
            kw["_ylim_data_min"] = True
        else:                 # Manual range
            kw["_ylim_data_min"] = False
            try:
                kw["ylim"] = (float(self._vars["ylim_lo"].get()),
                               float(self._vars["ylim_hi"].get()))
            except ValueError:
                messagebox.showerror("Invalid Y limits", "Enter numbers for both Y limit fields.")
                return False

        # X limits  -  only collected when continuous-x tab built the fields
        xlim_mode = self._get_var("xlim_mode", 0)
        if xlim_mode == 1 and hasattr(self, "_xl_lo"):
            try:
                kw["_xlim"] = (float(self._vars["xlim_lo"].get()),
                               float(self._vars["xlim_hi"].get()))
            except ValueError:
                messagebox.showerror("Invalid X limits", "Enter numbers for both X limit fields.")
                return False
        else:
            kw["_xlim"] = None

        kw["ref_line"] = None
        if self._vars["ref_line_enabled"].get():
            try:
                kw["ref_line"] = float(self._vars["ref_line_y"].get())
            except ValueError:
                pass
        kw["ref_line_label"] = (self._vars["ref_line_label"].get().strip()
                                if "ref_line_label" in self._vars else "")

        # Tick intervals  -  convert blank/invalid entries to 0 (= auto)
        def _parse_interval(key):
            raw = self._get_var(key, "").strip()
            try:
                v = float(raw)
                return v if v > 0 else 0.0
            except (ValueError, TypeError):
                return 0.0
        kw["ytick_interval"] = _parse_interval("ytick_interval")
        kw["xtick_interval"] = _parse_interval("xtick_interval")
        return True

    def _collect_stats(self, kw: dict):
        """Populate statistics kwargs."""
        kw["show_stats"]             = self._vars["show_stats"].get()
        kw["show_ns"]                = self._vars["show_ns"].get()
        kw["show_p_values"]          = self._vars["show_p_values"].get()
        kw["show_effect_size"]       = self._vars["show_effect_size"].get()
        kw["show_test_name"]         = self._vars["show_test_name"].get()
        kw["show_normality_warning"] = self._get_var("show_normality_warning", True)
        kw["stats_test"]       = STATS_TEST_MAP.get(self._vars["stats_test"].get(),
                                                     self._vars["stats_test"].get())
        kw["mc_correction"]    = self._vars["mc_correction"].get()
        kw["posthoc"]          = self._vars["posthoc"].get()
        try:
            kw["n_permutations"] = int(self._vars["n_permutations"].get())
        except ValueError:
            kw["n_permutations"] = 9999

        # Significance threshold  -  default 0.05 (Prism standard α)
        try:
            raw = self._vars["p_sig_threshold"].get().strip()
            kw["p_sig_threshold"] = float(raw) if raw else 0.05
        except (ValueError, TypeError):
            kw["p_sig_threshold"] = 0.05

        # Comparison mode: 0 = all pairwise, 1 = vs. control
        # When mode is 0 (all pairwise) we always pass control=None regardless
        # of what the dropdown says  -  matches Prism's explicit comparison selector.
        cmode = self._get_var("comparison_mode", 0)
        if cmode == 1:
            ctrl = self._vars["control"].get().strip()
            kw["control"] = (None if (not ctrl or ctrl.startswith("(none"))
                             else ctrl)
        else:
            kw["control"] = None

        # One-sample mu0
        try:
            kw["mu0"] = float(self._get_var("one_sample_mu0", "0") or 0)
        except (ValueError, TypeError):
            kw["mu0"] = 0.0

        # P16: bracket style
        _bs_map = {"Lines": "lines", "Bracket": "bracket", "Asterisks only": "asterisks_only"}
        kw["bracket_style"] = _bs_map.get(self._get_var("bracket_style", "Lines"), "lines")

    def _collect_figsize(self, kw: dict):
        """Populate figsize and chart-geometry kwargs."""
        try:
            fw = self._vars["figw"].get().strip()
            fh = self._vars["figh"].get().strip()
            if fw and fh:
                kw["figsize"] = (float(fw), float(fh))
            else:
                pane_w = max(300, self._plot_canvas.winfo_width()  or 600)
                pane_h = max(300, self._plot_canvas.winfo_height() or 600)
                side   = round(min(pane_w * 0.88 / 120, pane_h * 0.82 / 120, 7.0), 1)
                kw["figsize"] = (side, side)
        except (ValueError, TypeError):
            kw["figsize"] = (5.0, 5.0)

        if self._plot_type.get() in ("bar", "grouped_bar", "box"):
            kw["bar_width"] = self._get_float("bar_width", 0.6)
        else:
            kw["line_width"]   = self._get_float("line_width", 1.5)
            kw["marker_style"] = MARKER_STYLE_MAP.get(
                self._get_var("marker_style", "Different Markers"), "auto")
            kw["marker_size"]  = self._get_float("marker_size", 7.0)

        kw["font_size"]     = self._get_float("font_size", 12.0)
        kw["jitter_amount"] = self._get_float("jitter_amount", 0.0)

    def _collect(self, excel):
        """Assemble the full kwargs dict for the plot function."""
        sr = self._vars["sheet"].get().strip()
        kw = {
            "excel_path": excel,
            "sheet": int(sr) if sr.lstrip("-").isdigit() else (sr if sr else 0),
        }
        if not self._collect_display(kw): return None
        if not self._collect_labels(kw):  return None
        self._collect_stats(kw)
        self._collect_figsize(kw)
        return kw

    def _compute_data_ymin(self, excel_path, sheet, plot_type):
        """Return the true minimum Y-data value for 'Start at data min'.

        Each Excel layout needs a different slicing strategy to isolate
        Y-values (avoiding X-axis values, row/column labels, or count data
        that shouldn't influence the Y scale).

        Returns float or None if the minimum cannot be determined.
        """
        pd = _pd()
        import numpy as np
        try:
            df = pd.read_excel(excel_path, sheet_name=sheet, header=None)
        except Exception:
            return None

        try:
            # Charts where col-0 is X and cols 1+ are Y replicates
            # (line, scatter, curve_fit): exclude col-0 (X values) and
            # row-0 (series names) before computing the minimum.
            if plot_type in ("line", "scatter", "curve_fit"):
                y_vals = pd.to_numeric(
                    df.iloc[1:, 1:].values.flatten(), errors="coerce")

            # Grouped bar: rows 0-1 are category/subgroup headers,
            # rows 2+ are the numeric data.
            elif plot_type == "grouped_bar":
                y_vals = pd.to_numeric(
                    df.iloc[2:].values.flatten(), errors="coerce")

            # Two-way ANOVA: long-format with a header row.
            # The "Value" column holds the Y data; other columns are factors.
            elif plot_type == "two_way_anova":
                df_long = pd.read_excel(excel_path, sheet_name=sheet, header=0)
                val_col = next(
                    (c for c in df_long.columns
                     if str(c).strip().lower() == "value"), None)
                if val_col is None:
                    # Fall back to last numeric column
                    num_cols = df_long.select_dtypes(include="number").columns
                    val_col  = num_cols[-1] if len(num_cols) else None
                if val_col is None:
                    return None
                y_vals = pd.to_numeric(df_long[val_col], errors="coerce").values

            # Heatmap: row-0 = column labels, col-0 = row labels.
            # Numeric matrix is rows 1+, cols 1+.
            elif plot_type == "heatmap":
                y_vals = pd.to_numeric(
                    df.iloc[1:, 1:].values.flatten(), errors="coerce")

            # KM survival: Y axis is survival probability (0–1) computed
            # internally; the Excel time/event values don't map to Y.
            # Contingency: count table; Y axis is counts or percentages.
            # Column stats: summary table, not a simple scatter Y axis.
            # These chart types don't benefit from "start at data min"
            # in a meaningful way  -  return None to leave the axis unchanged.
            elif plot_type in ("kaplan_meier", "contingency", "column_stats",
                               "chi_square_gof", "bland_altman", "forest_plot"):
                return None

            # Stacked bar: same 2-row header layout as grouped_bar
            elif plot_type == "stacked_bar":
                y_vals = pd.to_numeric(
                    df.iloc[2:].values.flatten(), errors="coerce")

            # Bubble: X/Y/Size triples starting col 1; skip col 0 (X label)
            elif plot_type == "bubble":
                y_vals = pd.to_numeric(
                    df.iloc[1:, 2::3].values.flatten(), errors="coerce")

            # All flat-header charts (bar, box, violin, subcolumn_scatter,
            # before_after, repeated_measures, histogram):
            # row 0 = group names, rows 1+ = numeric values.
            else:
                y_vals = pd.to_numeric(
                    df.iloc[1:].values.flatten(), errors="coerce")

            valid = y_vals[~np.isnan(y_vals)]
            return float(valid.min()) if len(valid) > 0 else None

        except Exception:
            return None

    def _do_run(self, kw, tab_id=None, job_id=None):
        try:
            if hasattr(self, "_bus"):
                self._bus.emit("plot.started", kw=kw)
            pd = _pd()  # cached  -  no overhead after first call
            self._plt.close("all")

            # Strip internal-only keys before passing to plot functions
            ylim_data_min = kw.pop("_ylim_data_min", False)
            xlim_override = kw.pop("_xlim", None)

            # Set the show_ns flag on the module directly
            show_ns = kw.pop("show_ns", False)
            self._pf.__show_ns__ = show_ns

            # Set the significance threshold (α)  -  controls which brackets appear
            p_sig_threshold = kw.pop("p_sig_threshold", 0.05)
            self._pf.__p_sig_threshold__ = float(p_sig_threshold)

            # Set the normality warning flag on the module
            show_norm_warn = kw.pop("show_normality_warning", True)
            self._pf.__show_normality_warning__ = show_norm_warn

            groups = []
            try:
                df_g   = pd.read_excel(kw["excel_path"], sheet_name=kw.get("sheet", 0),
                                       header=None, nrows=1)
                groups = [str(c) for c in df_g.iloc[0].dropna().tolist()]
            except Exception:
                _log.debug("App._do_run: could not read group names from header row", exc_info=True)

            pt   = self._plot_type.get()
            spec = next((s for s in _REGISTRY_SPECS if s.key == pt), _REGISTRY_SPECS[0])
            fn   = getattr(self._pf, spec.fn_name)

            # ── Permutation progress indicator ───────────────────────────────
            # For long permutation runs (> 999 resamples) we tick a live counter
            # in the status bar every ~200ms so the user knows we're working.
            _n_perm = kw.get("n_permutations", 0)
            _perm_ticker_id = None
            if (kw.get("show_stats") and kw.get("stats_test") == "permutation"
                    and isinstance(_n_perm, int) and _n_perm > 999):
                _perm_start = __import__("time").monotonic
                _perm_t0 = _perm_start()
                _perm_count = [0]   # mutable cell for closure

                def _perm_tick():
                    if not self._running:
                        return
                    elapsed = __import__("time").monotonic() - _perm_t0
                    _perm_count[0] += 1
                    # Estimate fraction done: each tick ≈ 200ms; total ≈ n_perm * 0.05ms
                    est_done = min(0.99, elapsed / max(_n_perm * 0.00005, 0.1))
                    pct = int(est_done * 100)
                    self.after(0, lambda p=pct: self._set_status(
                        f"Running permutation test… {p}%  ({_n_perm:,} resamples)"))
                    if self._running:
                        self.after(250, _perm_tick)

                self.after(400, _perm_tick)   # start after 400ms (short runs finish before this)

            # Call the chart-specific extra_collect hook (if any) to inject
            # params that aren't part of the shared _collect() output.
            # This replaces the old if/elif pt == "scatter" / "kaplan_meier" etc. chain.
            if spec.extra_collect is not None:
                spec.extra_collect(self, kw)

            # Filter kwargs to only what the function actually accepts,
            # derived from its signature  -  no manual strip_keys needed.
            kw = spec.filter_kwargs(kw, fn)

            fig, ax = fn(**kw)

            # Apply "Start at data min" post-plot
            if ylim_data_min and kw.get("yscale", "linear") != "log":
                true_min = self._compute_data_ymin(
                    kw["excel_path"], kw.get("sheet", 0), pt)
                if true_min is not None:
                    _, cur_hi = ax.get_ylim()
                    ax.set_ylim(bottom=true_min, top=cur_hi)
                    fig.tight_layout(pad=1.2)

            # Apply manual x limits post-plot
            if xlim_override is not None:
                ax.set_xlim(xlim_override)

            # Capture values for results panel before lambdas close over them
            _ep = kw.get("excel_path", "")
            _sh = kw.get("sheet", 0)
            _pt = pt
            _kw = kw
            import copy as _copy
            _kw_snap = _copy.deepcopy(kw)
            self.after(0, lambda: self._embed_plot(
                fig, groups, kw=_kw_snap, tab_id=tab_id, job_id=job_id))
            # Re-enabled: populate results panel after plot embeds
            self.after(80, lambda: self._populate_results(_ep, _sh, _pt, _kw_snap))
            if hasattr(self, "_bus"):
                self._bus.emit("plot.completed", kw=_kw_snap)
        except Exception:
            err = traceback.format_exc()
            short = err.strip().splitlines()[-1]
            if hasattr(self, "_bus"):
                self._bus.emit("plot.failed", error=short)
            self.after(0, lambda: self._set_status(f"Error: {short}", err=True))
            self.after(0, lambda: messagebox.showerror("Runtime error", err))
            self.after(0, self._reset_btn)

    _WEBVIEW_CHART_TYPES = {"bar", "grouped_bar", "line", "scatter"}

    def _try_webview_embed(self, plot_frame: "tk.Frame", chart_type: str, kw: dict) -> bool:
        """Try to embed a Plotly chart via pywebview. Returns True on success."""
        if not getattr(self, "_web_server_running", False):
            return False
        if not getattr(self, "_use_webview", None) or not self._use_webview.get():
            return False
        if chart_type not in self._WEBVIEW_CHART_TYPES:
            return False
        try:
            from plotter_webview import PlotterWebView
            from plotter_server import get_port
            pv = PlotterWebView(plot_frame, port=get_port())
            if not pv.show():
                return False
            pv.render(chart_type, kw)
            self._plotly_views[id(plot_frame)] = pv
            return True
        except Exception:
            _log.debug("App._try_webview_embed: webview embed failed for %r", chart_type, exc_info=True)
            return False

    def _embed_plot(self, fig, groups=None, kw=None, tab_id=None, job_id=None):
        """
        Display *fig* in the right pane, routed to the correct tab.

        Parameters
        ----------
        fig    : matplotlib Figure
        groups : list of group names (for control-group dropdown)
        kw     : full plot kwargs dict (used for pick-event recoloring)
        tab_id : tab this render belongs to (thread-safety guard)
        job_id : render job id (superseded-job guard)
        """
        try:
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

            # ── Validate: discard stale or orphaned renders ───────────────────
            if self._tab_manager is not None and tab_id is not None:
                tab = self._tab_manager.get_tab(tab_id)
                if tab is None or tab.render_job_id != job_id:
                    import matplotlib.pyplot as _plt
                    _plt.close(fig)
                    return
                target_frame = tab.plot_frame
            else:
                # Fallback if tabs not yet initialised (should not happen in practice)
                target_frame = self._plot_frame
                tab = None

            # Always keep the mpl figure for export
            self._fig = fig
            if tab is not None:
                tab.fig = fig

            # Phase 3: try Plotly webview for priority chart types
            _pt_web = kw.get("plot_type", "") if kw else ""
            if not _pt_web and hasattr(self, "_plot_type"):
                _pt_web = self._plot_type.get()
            if hasattr(_pt_web, "get"):
                _pt_web = _pt_web.get()
            if kw and self._try_webview_embed(target_frame, str(_pt_web), kw):
                # Successfully embedded via pywebview — still fall through
                # to keep mpl figure for export, but skip Agg embedding
                self._plot_frame = target_frame
                if tab is not None:
                    tab.canvas_widget = None
                return

            # Hide the empty state overlay
            try:
                self._empty_state_frame.place_forget()
                self._plot_canvas.itemconfigure(self._empty_state_id, state="hidden")
            except Exception:
                pass

            # ── Render via FigureCanvasTkAgg ──────────────────────────────────
            for child in target_frame.winfo_children():
                child.destroy()
            self._canvas_widget = None
            self._zoom_level    = 1.0

            canvas = FigureCanvasTkAgg(fig, master=target_frame)
            canvas.draw()
            # Re-run tight_layout now that the figure is attached to a real
            # renderer so tick labels, titles, and axis labels are never clipped.
            try:
                fig.tight_layout(pad=1.2)
                canvas.draw()
            except Exception:
                _log.debug("App._embed_plot: tight_layout failed", exc_info=True)
            canvas.get_tk_widget().pack(padx=8, pady=(2, 8))
            self._canvas_widget = canvas
            self._plot_frame    = target_frame   # keep self._plot_frame pointing at active tab
            if tab is not None:
                tab.canvas_widget = canvas

            target_frame.update_idletasks()
            self._plot_canvas.configure(
                scrollregion=self._plot_canvas.bbox("all"))

            # ── mpl_connect pick-event recoloring (bar / grouped_bar) ────────
            _pt = self._plot_type.get()
            if kw is not None and _pt in ("bar", "grouped_bar"):
                ax = fig.axes[0] if fig.axes else None
                if ax is not None:
                    for patch in ax.patches:
                        patch.set_picker(True)

                    def _on_pick(event, _canvas=canvas):
                        patch = event.artist
                        from tkinter import colorchooser
                        import matplotlib.colors as _mc
                        try:
                            hex_color = _mc.to_hex(patch.get_facecolor())
                        except Exception:
                            hex_color = "#888888"
                        result = colorchooser.askcolor(
                            color=hex_color,
                            title="Choose bar color",
                            parent=self,
                        )
                        if result and result[1]:
                            patch.set_facecolor(result[1])
                            _canvas.draw_idle()

                    canvas.mpl_connect("pick_event", _on_pick)

            if groups:
                opts = list(groups)
                for cb in self._control_cbs:
                    cb.config(values=opts, state="readonly")
                for lbl in self._control_hint_lbls:
                    lbl.config(text=f"{len(groups)} group(s) found")
                if self._vars["control"].get() not in opts:
                    self._vars["control"].set(opts[0])
                self._tog_control()

            self._set_status("Done  ·  click a bar to recolor" if _pt in ("bar", "grouped_bar") else "Done")
            self._export_btn.config(state="normal")
            self._copy_btn.config(state="normal")
            self._copy_transparent_btn.config(state="normal")
            self._copy_svg_btn.config(state="normal")
            self._zoom_reset_btn.config(state="normal")
            self._zoom_in_btn.config(state="normal")
            self._zoom_out_btn.config(state="normal")

            # Save to recent files
            path = self._vars["excel_path"].get().strip()
            if path and os.path.exists(path):
                _add_recent(path)

        except Exception:
            err = traceback.format_exc()
            self._set_status(f"Embed error: {err.strip().splitlines()[-1]}", err=True)
        finally:
            self._reset_btn()

    def _zoom_plot(self, raw):
        """Zoom the rendered plot in/out using Ctrl+scroll."""
        self._zoom_plot_factor(1.1 if raw > 0 else 0.9)

    def _zoom_plot_factor(self, factor):
        """Apply a zoom multiplier directly (e.g. 1.1 = 10% zoom in)."""
        if self._fig is None or self._canvas_widget is None:
            return
        self._zoom_level = getattr(self, "_zoom_level", 1.0) * factor
        self._zoom_level = max(0.25, min(4.0, self._zoom_level))
        self._apply_zoom()
        self._set_status(f"Zoom: {self._zoom_level:.0%}  (Ctrl+scroll to zoom)")

    def _apply_zoom(self):
        """Resize the FigureCanvasTkAgg widget to match current zoom level.

        Only used in Agg-mode.  Canvas-mode zoom is handled in
        _zoom_plot_factor() via CanvasRenderer.rescale().
        """
        if self._canvas_widget is None or self._fig is None:
            return
        w, h   = self._fig.get_size_inches()
        dpi    = self._fig.get_dpi()
        new_w  = int(w * dpi * self._zoom_level)
        new_h  = int(h * dpi * self._zoom_level)
        widget = self._canvas_widget.get_tk_widget()
        widget.config(width=new_w, height=new_h)
        # Force matplotlib to redraw at the new pixel size
        self._canvas_widget.get_tk_widget().update_idletasks()
        self._canvas_widget.resize(new_w, new_h)
        self._canvas_widget.draw_idle()
        self._plot_frame.update_idletasks()
        self._plot_canvas.configure(scrollregion=self._plot_canvas.bbox("all"))

    def _reset_zoom(self):
        """Reset plot to 100% zoom and scroll back to top-left."""
        if self._fig is None:
            return
        self._zoom_level = 1.0
        if self._canvas_widget is not None:
            self._apply_zoom()
        self._plot_canvas.xview_moveto(0)
        self._plot_canvas.yview_moveto(0)
        self._set_status("Zoom reset to 100%")

    # ── Live preview ──────────────────────────────────────────────────────────

    # Variables that, when changed, should trigger a re-render without requiring
    # the user to click "Generate Plot".  Data-loading variables (excel_path,
    # sheet) are excluded — those go through the validate pipeline instead.
    _LIVE_PREVIEW_VARS = (
        "color", "axis_style", "tick_dir", "fig_bg", "minor_ticks",
        "show_points", "error", "yscale", "gridlines", "grid_style",
        "show_value_labels", "show_n_labels", "spine_width",
        "ref_line_enabled", "legend_pos", "bar_alpha",
        "title", "xlabel", "ytitle",
    )

    def _setup_live_preview(self):
        """Wire trace callbacks on display variables so any change triggers a
        debounced re-render (400 ms delay prevents re-running on every keystroke)."""
        for key in self._LIVE_PREVIEW_VARS:
            v = self._vars.get(key)
            if v is not None:
                v.trace_add("write", lambda *_, _k=key: self._schedule_preview())

        # Keep the active tab's label in sync with the chart title field.
        # Guarded by _switching_tabs so tab-switch var restores don't corrupt labels.
        title_v = self._vars.get("title")
        if title_v is not None:
            title_v.trace_add("write", lambda *_: self._sync_active_tab_label())

    def _schedule_preview(self):
        """Cancel any pending preview and schedule a new one in 400 ms."""
        if not self._live_preview_enabled:
            return
        if self._preview_after_id is not None:
            try:
                self.after_cancel(self._preview_after_id)
            except Exception:
                _log.debug("App._schedule_preview: after_cancel failed", exc_info=True)
        self._preview_after_id = self.after(400, self._live_preview_run)

    def _live_preview_run(self):
        """Fire a silent re-run if a plot is already showing and we are not busy."""
        self._preview_after_id = None
        if (self._fig is not None
                and self._validated
                and not self._running
                and self._pf is not None):
            excel = self._vars.get("excel_path", None)
            if excel and excel.get().strip():
                self._run()

    # ── (canvas-mode renderer removed in Phase 1; replaced by mpl_connect) ───



    def _copy_to_clipboard(self):
        """Copy the current plot to the macOS clipboard as a PNG image."""
        if self._fig is None:
            return
        try:
            import io, subprocess, tempfile
            from PIL import Image

            buf = io.BytesIO()
            self._fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
            buf.seek(0)
            png_bytes = buf.getvalue()

            img = Image.open(io.BytesIO(png_bytes))
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                tmp_path = tmp.name
                img.save(tmp_path, "PNG")
            subprocess.run(
                ["osascript", "-e",
                 f'set the clipboard to (read (POSIX file "{tmp_path}") as «class PNGf»)'],
                check=True, capture_output=True)
            os.unlink(tmp_path)
            self._set_status("Copied PNG to clipboard ✓")
        except Exception as ex:
            self._set_status(f"Copy failed: {ex}", err=True)

    def _copy_transparent(self):
        """Copy the current plot to the macOS clipboard as a transparent-background PNG."""
        if self._fig is None:
            return
        try:
            import io
            from PIL import Image
            buf = io.BytesIO()
            self._fig.savefig(buf, format="png", dpi=150, bbox_inches="tight",
                              transparent=True)
            buf.seek(0)
            img = Image.open(buf)
            import subprocess, tempfile
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                tmp_path = tmp.name
                img.save(tmp_path, "PNG")
            subprocess.run(
                ["osascript", "-e",
                 f'set the clipboard to (read (POSIX file "{tmp_path}") as «class PNGf»)'],
                check=True, capture_output=True)
            os.unlink(tmp_path)
            self._set_status("Copied transparent PNG to clipboard ✓")
        except Exception as ex:
            self._set_status(f"Copy failed: {ex}", err=True)

    def _copy_as_svg(self):
        """Copy the current plot to clipboard as SVG text."""
        if self._fig is None:
            return
        try:
            import io
            buf = io.StringIO()
            self._fig.savefig(buf, format="svg", bbox_inches="tight")
            svg_str = buf.getvalue()
            self.clipboard_clear()
            self.clipboard_append(svg_str)
            self._set_status("Copied SVG to clipboard ✓")
        except Exception as ex:
            self._set_status(f"SVG copy failed: {ex}", err=True)

    def _show_recent_files(self):
        """Show a popup menu of recently used files, anchored below the Recent button."""
        prefs  = _load_prefs()
        recent = [f for f in prefs.get("recent_files", []) if os.path.exists(f)]
        if not recent:
            self._set_status("No recent files")
            return
        menu = tk.Menu(self, tearoff=0)
        for path in recent[:10]:
            name = os.path.basename(path)
            menu.add_command(label=name,
                             command=lambda p=path: (
                                 self._vars["excel_path"].set(p),
                                 self._load_sheets(p)
                             ))
        try:
            btn = getattr(self, "_recent_btn", None)
            if btn and btn.winfo_exists():
                x = btn.winfo_rootx()
                y = btn.winfo_rooty() + btn.winfo_height()
            else:
                # Fallback: position near the cursor
                x = self.winfo_pointerx()
                y = self.winfo_pointery()
            menu.tk_popup(x, y)
        finally:
            menu.grab_release()

    def _load_last_file(self):
        """Restore the last used file path from preferences, silently if gone."""
        prefs = _load_prefs()
        last  = prefs.get("last_file", "")
        if last and os.path.exists(last):
            # Set the path var first so _validate_spreadsheet reads it correctly
            if "excel_path" in self._vars:
                self._vars["excel_path"].set(last)
            self.after(300, lambda: self._load_sheets(last))



    # ── Results Panel ─────────────────────────────────────────────────────────

    def _toggle_results_panel(self):
        """Slide the results panel open or closed."""
        if self._results_visible:
            self._results_strip.config(height=28)
            self._results_toggle_arrow.config(text="▲ Results")
            self._results_visible = False
        else:
            self._results_strip.config(height=220)
            self._results_toggle_arrow.config(text="▼ Results")
            self._results_visible = True

    def _populate_results(self, excel_path, sheet, plot_type, kw_snapshot):
        """Delegate to the standalone populate_results() in prism_results.py."""
        if _RESULTS_AVAILABLE:
            populate_results(self, excel_path, sheet, plot_type, kw_snapshot)

    def _export_results_csv(self):
        """Delegate to the standalone export_results_csv() in prism_results.py."""
        if _RESULTS_AVAILABLE:
            export_results_csv(self)
        else:
            self._set_status("Results module not available.", err=True)

    def _copy_results_tsv(self):
        """Delegate to the standalone copy_results_tsv() in prism_results.py."""
        if _RESULTS_AVAILABLE:
            copy_results_tsv(self)

    def _download_png(self):
        if self._fig is None: return
        from datetime import datetime
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        # Pre-fill with plot title or source filename
        title_str = self._vars.get("title", tk.StringVar()).get().strip()
        src_path  = self._vars.get("excel_path", tk.StringVar()).get().strip()
        if title_str:
            safe = "".join(c if c.isalnum() or c in " _-" else "_" for c in title_str).strip()
            default_name = safe[:60] or "claude_plotter"
        elif src_path:
            default_name = os.path.splitext(os.path.basename(src_path))[0][:60]
        else:
            default_name = f"claude_plotter_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        path = filedialog.asksaveasfilename(
            initialdir=desktop, initialfile=f"{default_name}.png",
            defaultextension=".png",
            filetypes=[("PNG image", "*.png"),
                       ("TIFF image", "*.tiff"),
                       ("EPS vector", "*.eps"),
                       ("SVG vector", "*.svg"),
                       ("PDF", "*.pdf"),
                       ("All files", "*.*")],
            title="Save plot")
        if path:
            ext = os.path.splitext(path)[1].lower()
            if ext == ".png":
                dpi = 600
            elif ext == ".tiff":
                dpi = 300
            else:
                dpi = 150
            self._fig.savefig(path, dpi=dpi, bbox_inches="tight")
            self._set_status(f"Saved  -  {os.path.basename(path)}")


    def _start_spinner(self):
        """Animate the Generate Plot button while running."""
        frames = ["Generating", "Generating.", "Generating..", "Generating..."]
        self._spinner_idx = 0
        def _tick():
            if not self._running: return
            self._run_btn.config(text=frames[self._spinner_idx % len(frames)])
            self._spinner_idx += 1
            self.after(400, _tick)
        _tick()

    def _reset_btn(self):
        self._running = False
        self._run_btn.config(state="normal", text="Generate Plot")

    def _set_status(self, msg, err=False):
        self._status_var.set(msg)
        if err:
            fg_color = "#cc0000"
        elif any(w in msg.lower() for w in ("done", "saved", "copied", "ready")):
            fg_color = "#2a8a2a"
        elif any(w in msg.lower() for w in ("warning", "validat")):
            fg_color = "#cc6600"
        else:
            fg_color = "#555555"
        self._status_lbl.config(foreground=fg_color)
        # Auto-fade success messages after 3 s
        if fg_color == "#2a8a2a":
            if hasattr(self, "_status_fade_id") and self._status_fade_id:
                try: self.after_cancel(self._status_fade_id)
                except Exception: pass
            self._status_fade_id = self.after(3000, lambda: (
                self._status_var.set(""),
                self._status_lbl.config(foreground="#555555")
            ))


    def _show_about(self):
        """About dialog."""
        from tkinter import messagebox
        messagebox.showinfo(
            "About Claude Plotter",
            "Claude Plotter\n\n"
            "A GraphPad Prism-style data visualization application.\n\n"
            "Designed and implemented by Claude (Anthropic).\n"
            "Commissioned by Ashwin Pasupathy.\n\n"
            "MIT License\n\n"
            "22 chart types · Statistical tests · Publication-ready export"
        )

    def _export_all_pdf(self):
        """Export all 22 chart types as a multi-page PDF showcase (P20)."""
        if self._pf is None:
            from tkinter import messagebox
            messagebox.showerror("Not ready", "Functions not loaded yet.")
            return
        excel = self._vars.get("excel_path", tk.StringVar()).get().strip()
        if not excel:
            from tkinter import messagebox
            messagebox.showwarning("No file", "Open an Excel file first so charts have data to render.")
            return
        from tkinter import filedialog
        out_path = filedialog.asksaveasfilename(
            title="Save Chart Showcase PDF",
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            initialfile="claude_plotter_showcase.pdf"
        )
        if not out_path:
            return
        self._set_status("Exporting PDF showcase…")
        import threading
        def _do_export():
            try:
                sr = self._vars["sheet"].get().strip()
                sheet = int(sr) if sr.lstrip("-").isdigit() else (sr if sr else 0)
                self._pf.export_all_charts_pdf(out_path, excel, sheet=sheet)
                self.after(0, lambda: self._set_status(f"PDF saved: {out_path}"))
            except Exception as exc:
                import traceback
                err = traceback.format_exc()
                self.after(0, lambda: self._set_status(f"PDF export error: {exc}", err=True))
        threading.Thread(target=_do_export, daemon=True).start()

    # ── Project file support (.cplot) ─────────────────────────────────────────

    def _save_project(self):
        """Save current app state as a .cplot project file."""
        from tkinter import filedialog, messagebox
        try:
            from plotter_project import save_project
        except ImportError:
            messagebox.showerror("Error", "plotter_project module not available.")
            return

        path = filedialog.asksaveasfilename(
            title="Save Project",
            defaultextension=".cplot",
            filetypes=[("Claude Plotter Project", "*.cplot"), ("All files", "*.*")],
        )
        if not path:
            return

        try:
            plot_type = self._plot_type.get() if hasattr(self, "_plot_type") else "bar"
            excel_path = self._vars.get("excel_path", tk.StringVar()).get().strip()
            sheet_var = self._vars.get("sheet")
            sheet = sheet_var.get() if sheet_var else 0

            # Capture thumbnail bytes if we have a figure
            thumbnail_bytes = None
            try:
                if hasattr(self, "_fig") and self._fig is not None:
                    import io
                    buf = io.BytesIO()
                    self._fig.savefig(buf, format="png", dpi=72, bbox_inches="tight")
                    thumbnail_bytes = buf.getvalue()
            except Exception:
                _log.debug("App._save_project: thumbnail capture failed", exc_info=True)

            save_project(
                path=path,
                app_vars=self._vars,
                plot_type=plot_type,
                excel_path=excel_path,
                sheet=sheet,
                thumbnail_bytes=thumbnail_bytes,
            )
            self._set_status(f"Project saved: {path}")
        except Exception as exc:
            messagebox.showerror("Save failed", str(exc))

    def _open_cplot(self):
        """Open a .cplot project file and restore app state."""
        from tkinter import filedialog, messagebox
        try:
            from plotter_project import load_project, extract_to_temp_excel
        except ImportError:
            messagebox.showerror("Error", "plotter_project module not available.")
            return

        path = filedialog.askopenfilename(
            title="Open Project",
            filetypes=[("Claude Plotter Project", "*.cplot"), ("All files", "*.*")],
        )
        if not path:
            return

        try:
            data = load_project(path)

            # Restore plot type
            pt = data.get("plot_type", "bar")
            if hasattr(self, "_plot_type"):
                self._plot_type.set(pt)
                self._on_chart_type_change(pt)

            # Restore UI state
            state = data.get("state", {})
            for key, value in state.items():
                var = self._vars.get(key)
                if var is not None:
                    try:
                        var.set(value)
                    except Exception:
                        _log.debug("App._open_cplot: could not restore var %r to %r", key, value, exc_info=True)

            # Extract embedded Excel data and load it
            try:
                temp_path = extract_to_temp_excel(path)
                if temp_path and hasattr(self, "_vars"):
                    ep_var = self._vars.get("excel_path")
                    if ep_var:
                        ep_var.set(temp_path)
                    self._load_sheets(temp_path)
            except Exception:
                _log.debug("App._open_cplot: extract_to_temp_excel failed for %r", path, exc_info=True)

            self._set_status(f"Project loaded: {path}")
        except Exception as exc:
            messagebox.showerror("Open failed", str(exc))

    def _open_pzfx(self):
        """Import a GraphPad Prism .pzfx file."""
        from tkinter import filedialog, messagebox
        try:
            from plotter_import_pzfx import import_pzfx
        except ImportError:
            messagebox.showerror("Error", "plotter_import_pzfx module not available.")
            return

        path = filedialog.askopenfilename(
            title="Import .pzfx File",
            filetypes=[("GraphPad Prism", "*.pzfx"), ("All files", "*.*")],
        )
        if not path:
            return

        try:
            result = import_pzfx(path)
            if result.errors:
                messagebox.showerror("Import Error",
                                     "\n".join(result.errors))
                return
            if result.warnings:
                messagebox.showwarning("Import Warnings",
                                       "\n".join(result.warnings))
            if result.temp_path:
                ep_var = self._vars.get("excel_path")
                if ep_var:
                    ep_var.set(result.temp_path)
                self._load_sheets(result.temp_path)
                # Set chart type if detected
                if result.chart_type and hasattr(self, "_plot_type"):
                    self._plot_type.set(result.chart_type)
                    self._on_chart_type_change(result.chart_type)
                self._set_status(f"Imported: {path}")
        except Exception as exc:
            messagebox.showerror("Import failed", str(exc))


if __name__ == "__main__":
    app = App()
    app.mainloop()
