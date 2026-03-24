#!/usr/bin/env python3
"""
plotter_barplot_app.py
======================
Refraction -- main application window (macOS, tabbed ttk layout).

Module structure
----------------
plotter_barplot_app.py  -- this file; App class + PLOT_REGISTRY + icon helpers
plotter_widgets.py      -- design-system tokens, PButton/PEntry/PCheckbox etc.
plotter_validators.py   -- standalone spreadsheet validation functions
plotter_results.py      -- results-panel population, export, and copy helpers
plotter_functions.py    -- matplotlib plot functions (29 chart types)
plotter_tabs.py         -- TabState, TabManager, TabBar (plot tab system)

The App class imports from all four companion modules so each can be
developed, tested, and documented independently.
"""
import collections, json, logging, math, os, threading, traceback, uuid

_log = logging.getLogger(__name__)

_pd_module = None
def _pd():
    """Return pandas, importing and caching it on first call."""
    global _pd_module
    if _pd_module is None:
        import pandas as _p
        _pd_module = _p
    return _pd_module

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# -- Companion-module imports --------------------------------------------------
import sys as _sys
_HERE = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(os.path.dirname(_HERE))
if _PROJECT_ROOT not in _sys.path:
    _sys.path.insert(0, _PROJECT_ROOT)

try:
    from refraction.app.widgets import (
        _DS, PButton, PCheckbox, PRadioGroup, PEntry, PCombobox,
        section_sep, _create_tooltip, add_placeholder, _bind_scroll_recursive,
        LABELS, HINTS, label, hint, tip,
        _is_num, _non_numeric_values, _scipy_summary, _sys_bg,
        PAD,
    )
except ImportError as _e:
    print(f"[refraction] warning: widgets not found ({_e})")

try:
    from refraction.core.validators import (
        validate_flat_header, validate_bar, validate_line,
        validate_grouped_bar, validate_kaplan_meier, validate_heatmap,
        validate_two_way_anova, validate_contingency, validate_chi_square_gof,
        validate_bland_altman, validate_forest_plot, validate_pyramid,
    )
    _VALIDATORS_AVAILABLE = True
except ImportError as _e:
    print(f"[refraction] warning: validators not found ({_e})")
    _VALIDATORS_AVAILABLE = False

try:
    from refraction.app.results import populate_results, export_results_csv, copy_results_tsv
    _RESULTS_AVAILABLE = True
except ImportError as _e:
    print(f"[refraction] warning: results not found ({_e})")
    _RESULTS_AVAILABLE = False

try:
    from refraction.app.icons import ICON_FN
    _ICON_FN = ICON_FN  # backward-compat alias; local icon functions removed
    _icon_bar = ICON_FN.get("bar", lambda *a: None)  # fallback for sidebar
except ImportError as _e:
    print(f"[refraction] warning: icons not found ({_e})")
    _ICON_FN = {}
    _icon_bar = lambda *a: None

# Heavy scientific imports are deferred to first use so the window
# appears immediately. They are loaded on a background thread via
# _import_functions() and also imported lazily inside each method
# that needs them (_do_run, _load_sheets, etc.).
# matplotlib and pandas are NOT imported here at module level.

# Try to use TkinterDnD for drag-and-drop support
try:
    from tkinterdnd2 import TkinterDnD, DND_FILES
    _DND_AVAILABLE = True
except ImportError:
    _DND_AVAILABLE = False
    DND_FILES      = None

SCRIPT_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
ICON_PNG   = os.path.join(SCRIPT_DIR, "assets", "AppIcon.png")
ICON_ICNS  = os.path.join(SCRIPT_DIR, "assets", "AppIcon.icns")

# Add the script directory to sys.path so plotter_functions is importable
import sys
if SCRIPT_DIR not in sys.path:
    sys.path.insert(0, SCRIPT_DIR)

def _load_icon_nsimage():
    """Return an NSImage loaded from bytes. The white border is baked into
    AppIcon.icns and AppIcon.png so no runtime compositing is needed.
    Tries .icns first (multi-resolution), then .png.
    """
    try:
        from AppKit import NSImage
        import Foundation
        for icon_path in (ICON_ICNS, ICON_PNG):
            if not os.path.exists(icon_path):
                continue
            with open(icon_path, "rb") as fh:
                raw = fh.read()
            data = Foundation.NSData.dataWithBytes_length_(raw, len(raw))
            img  = NSImage.alloc().initWithData_(data)
            if img and img.isValid():
                return img
    except Exception:
        _log.debug("_load_icon_nsimage: AppKit icon load failed", exc_info=True)
    return None


def set_dock_icon():
    """Set the macOS dock icon from bytes  -  immune to Cocoa path-resolution
    resets that happen when matplotlib/seaborn import NSApplication internally.

    Three-pronged approach:
      1. Load icon from raw bytes (not file path)  -  works inside bundles.
      2. Explicitly set NSApplicationActivationPolicyRegular so macOS treats
         the process as a proper GUI app.
      3. Called at startup, every 250 ms during heavy imports, and once more
         after all imports finish  -  see _watch_dock_icon.
    """
    try:
        from AppKit import NSApplication, NSApplicationActivationPolicyRegular
        app = NSApplication.sharedApplication()
        # Ensure the app is registered as a regular GUI application.
        # This prevents macOS from downgrading the dock entry after import.
        try:
            app.setActivationPolicy_(NSApplicationActivationPolicyRegular)
        except Exception:
            _log.debug("set_dock_icon: setActivationPolicy_ failed", exc_info=True)
        img = _load_icon_nsimage()
        if img:
            app.setApplicationIconImage_(img)
    except Exception:
        _log.debug("set_dock_icon: AppKit dock icon setup failed", exc_info=True)


def load_tk_icon():
    """Load the app icon as a Tkinter PhotoImage for wm_iconphoto().
    The white border is baked into AppIcon.png so no runtime compositing needed.
    """
    try:
        from PIL import Image, ImageTk
        img = Image.open(ICON_PNG).convert("RGBA")
        if max(img.size) > 256:
            img = img.resize((256, 256), Image.LANCZOS)
        return ImageTk.PhotoImage(img)
    except Exception:
        _log.debug("load_tk_icon: could not load Tk icon from %r", ICON_PNG, exc_info=True)
        return None


def _lbl(parent, row, key, font_size=13):
    """Create a bold section label with tooltip, grid it, return (label, next_row)."""
    l = ttk.Label(parent, text=label(key), font=("Helvetica Neue", font_size, "bold"))
    l.grid(row=row, column=0, sticky="w", padx=PAD, pady=(4, 2))
    tip(l, key)
    return l, row + 1


# ---------------------------------------------------------------------------
# Plot type registry  -  edit plotter_registry.py to add new chart types
# ---------------------------------------------------------------------------

from refraction.core.registry import (
    PlotTypeConfig, _REGISTRY_SPECS,
    ERROR_TYPE_MAP, STATS_TEST_MAP, MARKER_STYLE_MAP, PAD,
)
from refraction.core.tabs import TabState, TabManager, TabBar

PLOT_REGISTRY: list = []  # populated after App class definition so fn_name can be resolved


# ---------------------------------------------------------------------------
# Factory defaults for every UI variable.
# _reset_vars_to_defaults() reads this to restore state on chart-type switch.
# Sheet changes and in-pane edits must NOT call that method.
# ---------------------------------------------------------------------------
_VAR_DEFAULTS: dict = {
    # file / sheet
    "excel_path": "", "sheet": "",
    # data tab
    "error": "SEM (Standard Error)", "show_points": True,
    "show_n_labels": False, "show_value_labels": False, "color": "Default",
    "title": "", "xlabel": "", "ytitle": "",
    "jitter_amount": "0", "error_below_bar": False,
    # axes tab
    "yscale": "Linear", "ylim_lo": "", "ylim_hi": "",
    "figw": "", "figh": "", "font_size": "12",
    "ref_line_y": "0", "ref_line_enabled": False,
    "ylim_data_min": False, "ylim_none": True,
    "ylim_mode": 0,   # 0=Auto, 1=Data min, 2=Manual
    "xlim_mode": 0,   # 0=Auto, 1=Manual
    "xlim_lo": "", "xlim_hi": "",
    "gridlines": False, "open_points": False, "grid_style": "None",
    "horizontal": False, "show_median": False,
    "bar_alpha": "0.85", "xscale": "Linear", "ref_line_label": "",
    "bar_width": "0.6", "line_width": "1.5",
    "marker_style": "Different Markers", "marker_size": "7",
    "notch_box": False,
    # stats tab (shared)
    "show_stats": False, "show_ns": False, "show_p_values": False,
    "show_effect_size": False, "show_test_name": False,
    "show_normality_warning": True,
    "p_sig_threshold": "0.05",
    "bracket_style": "Lines",
    "stacked_horizontal": False,
    "xtick_labels_str": "",
    "twin_y_series_str": "",
    "ref_vline_enabled": False,
    "ref_vline_x": "0",
    "ref_vline_label": "",
    "stats_test": "Parametric", "n_permutations": "",
    "mc_correction": "Holm-Bonferroni", "posthoc": "Tukey HSD",
    "control": "",
    # scatter-specific
    "show_regression": False, "show_ci_band": False,
    "show_prediction_band": False, "show_correlation": False,
    "correlation_type": "Pearson",
    # kaplan-meier
    "show_ci": True, "show_censors": True, "show_at_risk": False,
    # heatmap
    "annotate": False, "cluster_rows": False, "cluster_cols": False,
    "robust": False, "heatmap_fmt": "",
    "heatmap_vmin": "", "heatmap_vmax": "", "heatmap_center": "",
    # two-way anova
    "show_posthoc": False,
    # grouped bar
    "show_anova_per_group": False,
    # histogram
    "hist_bins": "0", "hist_density": False, "hist_overlay_normal": False,
    # curve fit
    "curve_model": "4PL Sigmoidal (EC50/IC50)",
    "cf_show_ci": True, "cf_show_residuals": False,
    "cf_show_equation": True, "cf_show_r2": True, "cf_show_params": True,
    # column stats
    "cs_show_normality": True, "cs_show_ci": True, "cs_show_cv": True,
    # contingency
    "ct_show_pct": True, "ct_show_expected": False,
    # repeated measures
    "rm_show_lines": True, "rm_test_type": "Parametric",
    # ── Priority-1 styling (axes tab) ────────────────────────────────────────
    "axis_style":  "Open (default)",
    "tick_dir":    "Outward (default)",
    "minor_ticks": False,
    "point_size":  "6",
    "point_alpha": "0.80",
    "cap_size":    "4",
    "legend_pos":  "Upper right",
    "spine_width": "0.8",
    # ── Priority-2 stats ──────────────────────────────────────────────────────
    "show_regression_table": False,
    "one_sample_mu0": "0",
    # chi-square GoF
    "gof_expected_equal": True,
    # stacked bar
    "stacked_mode": "absolute", "stacked_value_labels": False,
    # bubble chart
    "bubble_scale": "1.0", "bubble_show_labels": False,
    # dot plot
    "dp_show_mean": True, "dp_show_median": False,
    # bland-altman
    "ba_show_ci": True,
    # forest plot
    "fp_ref_value": "0", "fp_show_weights": True, "fp_show_summary": True,
    # tick intervals
    "ytick_interval": "", "xtick_interval": "",
    # figure background
    "fig_bg": "White",
}


PREFS_PATH = os.path.expanduser("~/Library/Preferences/refraction.json")

def _load_prefs():
    try:
        with open(PREFS_PATH) as f:
            return json.load(f)
    except Exception:
        _log.debug("_load_prefs: could not load %r", PREFS_PATH, exc_info=True)
        return {}

def _save_prefs(data):
    try:
        os.makedirs(os.path.dirname(PREFS_PATH), exist_ok=True)
        with open(PREFS_PATH, "w") as f:
            json.dump(data, f, indent=2)
    except Exception:
        _log.debug("_save_prefs: could not write %r", PREFS_PATH, exc_info=True)

def _add_recent(path):
    prefs = _load_prefs()
    recent = prefs.get("recent_files", [])
    if path in recent:
        recent.remove(path)
    recent.insert(0, path)
    prefs["recent_files"] = recent[:10]
    prefs["last_file"] = path
    _save_prefs(prefs)


# Lightweight sheet/dataframe cache  -  avoids re-reading the same file twice
_SHEET_CACHE = {}   # path -> (mtime, sheetnames)
_DF_CACHE    = {}   # (path, sheet) -> (mtime, DataFrame)

def _cached_sheets(path):
    """Return sheet names, using cache if file unchanged."""
    try:
        mtime = os.path.getmtime(path)
        if path in _SHEET_CACHE and _SHEET_CACHE[path][0] == mtime:
            return _SHEET_CACHE[path][1]
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = wb.sheetnames; wb.close()
        _SHEET_CACHE[path] = (mtime, sheets)
        return sheets
    except Exception:
        _log.debug("_cached_sheets: could not read sheets from %r", path, exc_info=True)
        return []

def _cached_df(path, sheet):
    """Return DataFrame, using cache if file unchanged."""
    pd = _pd()
    try:
        mtime = os.path.getmtime(path)
        key   = (path, str(sheet))
        if key in _DF_CACHE and _DF_CACHE[key][0] == mtime:
            return _DF_CACHE[key][1]
        df = pd.read_excel(path, sheet_name=sheet, header=None)
        _DF_CACHE[key] = (mtime, df)
        return df
    except Exception as e:
        raise



_SB_ITEM_H    = 56   # height of each sidebar item in pixels
_SB_ICON_SIZE = 28   # icon canvas size
_SB_WIDTH     = 160  # total sidebar width


def _scipy_summary(fn, max_chars: int = 280) -> str:
    """Return the opening summary of a scipy function's docstring."""
    import inspect
    doc   = inspect.getdoc(fn) or ""
    paras = [p.strip() for p in doc.split("\n\n")]
    result = []
    for p in paras:
        if any(p.startswith(h) for h in
               ("Parameters", "Returns", "See Also", "Notes",
                "Warns", "References", "Examples", "Attributes")):
            break
        clean = " ".join(p.split())
        if clean:
            result.append(clean)
        if sum(len(r) for r in result) > max_chars:
            break
    return " ".join(result)


# ── Mixin classes (extracted from this file for readability) ─────────────────
from refraction.app.stats_tabs import StatsTabMixin
from refraction.app.validators import ValidationMixin
from refraction.app.collect import CollectMixin
from refraction.app.file_io import FileIOMixin
from refraction.app.execution import ExecutionMixin

_TkBase = TkinterDnD.Tk if _DND_AVAILABLE else tk.Tk


class App(StatsTabMixin, ValidationMixin, CollectMixin, FileIOMixin,
          ExecutionMixin, _TkBase):
    def __init__(self):
        super().__init__()
        # Hide window during build to avoid visible resize animation
        self.withdraw()
        self.title("Refraction")
        self.resizable(True, True)
        self._pf              = None
        self._pf_ready        = False
        self._running         = False
        # Excel parse cache — keyed by (path, sheet); avoids re-reading the
        # same file on consecutive renders (e.g. changing only style options).
        self._parse_cache_key = None
        self._parse_cache_df  = None
        self._vars            = {}
        self._file_selected   = False
        self._validated       = False
        self._validation_lock = False  # True while form is locked due to validation failure
        self._status_fade_id  = None   # handle for auto-clear timer
        self._kw_history      = collections.deque(maxlen=10)  # undo stack
        self._results_tsv_data = ""
        self._popup_count = 0   # number of open popup dialogs
        self._tab_manager:         TabManager | None = None   # set in _build_right_pane
        self._switching_tabs       = False   # True while TabManager.switch_to restores vars
        self._preview_after_id    = None   # pending after() id for live preview debounce
        self._live_preview_enabled = True  # can be toggled by user preference
        # --- Wave 2 infrastructure ---
        from refraction.core.events import EventBus
        self._bus = EventBus()
        from refraction.core.undo import UndoStack
        self._undo_stack = UndoStack(max_depth=50)
        from refraction.core.errors import reporter
        reporter.set_root(self)
        # --- end Wave 2 infrastructure ---
        self._use_webview = tk.BooleanVar(value=True)
        ttk.Style().theme_use("aqua")
        set_dock_icon()
        self._set_tk_icon()
        self._build()
        self.after(200, self._build_lockable_cache)  # build after window renders
        self._setup_dnd()  # register drag-and-drop targets
        # Start locked; unlock happens in _load_sheets after first file pick
        self._run_btn.config(state="disabled", text="Generate Plot")
        self._lock_form()
        self._import_functions()
        # Keep dock icon alive during background import (Cocoa resets it)
        self.after(250, self._watch_dock_icon)
        # Wire live-preview traces after the UI is fully built
        self.after(300, self._setup_live_preview)
        # Session persistence: auto-save every 30 s, offer restore on launch
        try:
            from refraction.core.session import Session
            self._session = Session()
            self.after(500, self._session_restore_prompt)
        except Exception:
            _log.debug("App.__init__: plotter_session import or init failed", exc_info=True)
            self._session = None
        self.protocol("WM_DELETE_WINDOW", self._on_quit)
        self.after(30000, self._auto_save)

        # Phase 3: start FastAPI server for Plotly rendering
        try:
            from refraction.server.api import start_server
            start_server(app_instance=self)
            self._web_server_running = True
        except Exception:
            _log.debug("App.__init__: plotter_server start failed", exc_info=True)
            self._web_server_running = False

        # Phase 3: web view instances per tab
        self._plotly_views: dict = {}

    def _auto_save(self):
        """Save session state to disk every 30 seconds."""
        try:
            if self._session is not None:
                pt = self._plot_type.get() if hasattr(self, "_plot_type") else "bar"
                state = self._session.capture(self._vars, pt, self.geometry())
                self._session.save_to_disk(state)
        except Exception:
            _log.debug("App._auto_save: session auto-save failed", exc_info=True)
        self.after(30000, self._auto_save)

    def _session_restore_prompt(self):
        """Offer to restore the previous session on startup."""
        try:
            if self._session is None:
                return
            saved = self._session.load_from_disk()
            if not saved or not saved.get("vars"):
                return
            from tkinter import messagebox
            if messagebox.askyesno("Restore Session",
                                   "Restore your previous session?"):
                self._session.restore(saved, self._vars)
        except Exception:
            _log.debug("App._session_restore_prompt: session restore failed", exc_info=True)

    def _on_quit(self):
        """Save session then destroy the window."""
        try:
            if self._session is not None:
                pt = self._plot_type.get() if hasattr(self, "_plot_type") else "bar"
                state = self._session.capture(self._vars, pt, self.geometry())
                self._session.save_to_disk(state)
        except Exception:
            _log.debug("App._on_quit: session save on quit failed", exc_info=True)
        self.destroy()

    def _set_tk_icon(self):
        photo = load_tk_icon()
        if photo:
            self._icon_photo = photo
            self.iconphoto(True, photo)

    def _watch_dock_icon(self):
        """Re-apply the dock icon on a timer while heavy imports run.

        matplotlib/seaborn/openpyxl each touch NSApplication on first import,
        resetting the dock icon.  The watcher fires every 200 ms until the
        module import background thread signals it is done (_pf_ready=True),
        then does one final restore after a short settling delay.

        Note: since matplotlib is now deferred to first export, the icon is
        only at risk from openpyxl during startup — the watcher still handles
        that correctly.
        """
        set_dock_icon()
        if not self._pf_ready:
            self.after(200, self._watch_dock_icon)
        else:
            # Imports just finished  -  do one more restore after a short delay
            # to catch any residual Cocoa activity from seaborn/openpyxl.
            self.after(800, set_dock_icon)

    def _setup_dnd(self):
        """Register file-path entry fields as drag-and-drop targets."""
        if not _DND_AVAILABLE:
            return
        def _on_drop(event):
            # macOS wraps paths in braces if they contain spaces; strip them
            raw = event.data.strip()
            # Handle multiple files  -  take only the first
            if raw.startswith("{"):
                raw = raw[1:raw.index("}")] if "}" in raw else raw[1:]
            path = raw.strip()
            if path.lower().endswith((".xlsx", ".xls")):
                self._vars["excel_path"].set(path)
                self._load_sheets(path)
            else:
                self._set_status("Drop an .xlsx or .xls file", err=True)

        # Register every file-path entry field
        for widget in getattr(self, "_drop_targets", []):
            try:
                widget.drop_target_register(DND_FILES)
                widget.dnd_bind("<<Drop>>", _on_drop)
            except Exception:
                _log.debug("App._setup_dnd: DND binding failed for widget %r", widget, exc_info=True)

    def _import_functions(self):
        threading.Thread(target=self._do_import, daemon=True).start()

    def _do_import(self):
        try:
            # Import plotter_functions module only — matplotlib/seaborn/scipy
            # are NOT loaded here.  They load lazily on first export call via
            # pf._ensure_imports().  This keeps startup fast and keeps matplotlib
            # completely out of memory during normal (Plotly-rendered) sessions.
            from refraction.core import chart_helpers as pf
            self._pf = pf
            # pandas is lightweight and needed immediately for file validation.
            import pandas as _pandas_mod
            self._pd = _pandas_mod
            # _plt stays None until first matplotlib export — callers must guard:
            #   if self._plt is not None: self._plt.close(...)
            self._plt = None
            self._pf_ready = True
            self.after(0, set_dock_icon)
            def _check_ready():
                if self._validated:
                    self._run_btn.config(state="normal", text="Generate Plot")
                else:
                    self._run_btn.config(state="disabled", text="Generate Plot")
                self._vars["color"].set(self._vars["color"].get())
                self._sync_analyze_btn()
            self.after(0, _check_ready)
        except Exception as e:
            m = str(e)
            self.after(0, lambda: self._set_status(f"Load error: {m}", err=True))

    # ── Form locking ──────────────────────────────────────────────────────────

    def _reset_vars_to_defaults(self):
        """Reset all UI variables to factory defaults.
        Called ONLY on chart-type switch  -  never on sheet change or in-pane edits."""
        for key, default in _VAR_DEFAULTS.items():
            if key in self._vars:
                try:
                    self._vars[key].set(default)
                except Exception:
                    pass
        # Clear the control group dropdown values so stale group names from a
        # previous chart type can't crash the next plot run.
        for cb in getattr(self, "_control_cbs", []):
            try:
                cb.config(values=[], state="disabled")
            except Exception:
                pass
        if "control" in self._vars:
            try: self._vars["control"].set("")
            except Exception: pass
        if "comparison_mode" in self._vars:
            try: self._vars["comparison_mode"].set(0)
            except Exception: pass
        for lbl in getattr(self, "_control_hint_lbls", []):
            try:
                lbl.config(text="Load a file to populate group names",
                           foreground="gray")
            except Exception:
                pass
        # Re-fire toggles that depend on var state
        if hasattr(self, "_yl_lo"):
            self._tog_ylim()
        if hasattr(self, "_xl_lo"):
            self._tog_xlim()

    def _reset_chart_type_state(self):
        """Called after a chart-type change (or new tab).
        Resets all UI variables to factory defaults, clears the cached spreadsheet,
        and locks the form until the user selects a new file for this chart type."""
        # Reset every UI variable to its factory default first
        self._reset_vars_to_defaults()
        # Clear file selection and validation state
        self._vars["excel_path"].set("")
        self._vars["sheet"].set("")
        for cb in getattr(self, "_sheet_cbs", []):
            cb.config(values=[], state="disabled")
        for lbl in getattr(self, "_sheet_hints", []):
            lbl.config(text="Browse for an Excel file to populate sheets")
        self._file_selected = False
        self._validated = False
        self._run_btn.config(state="disabled", text="Generate Plot")
        self._set_validate_text("", "gray")
        # Lock everything except the file browser so user must upload first
        if hasattr(self, "_file_row_frames"):
            self._lock_form()
        # Disable plot-output buttons
        for attr in ("_copy_btn", "_zoom_reset_btn", "_zoom_in_btn",
                     "_zoom_out_btn", "_export_btn", "_copy_svg_btn"):
            if hasattr(self, attr):
                getattr(self, attr).config(state="disabled")

    def _sb_select_silent(self, idx: int) -> None:
        """Update sidebar highlight and form pane without triggering a chart-type reset.

        Used by TabManager.switch_to so that switching tabs does not wipe the
        restored form state.
        """
        if hasattr(self, "_sb_select"):
            self._sb_select(idx)
        if hasattr(self, "_sb_show_pane"):
            self._sb_show_pane(idx)
        # Explicitly does NOT call _on_chart_type_change / _reset_chart_type_state

    def _sync_active_tab_label(self) -> None:
        """Keep the active tab's label in sync with the chart title field."""
        if getattr(self, "_switching_tabs", False):
            return
        if self._tab_manager is None:
            return
        tab = self._tab_manager.active
        if tab is None:
            return
        raw   = self._vars["title"].get().strip()
        label = raw if raw else "Untitled"
        tab.label = label
        self._tab_manager.update_label(tab.tab_id, label)

    def _lock_form(self):
        self._set_form_state("disabled")

    def _unlock_form(self):
        self._set_form_state("!disabled")
        self._validation_lock = False
        # _set_form_state re-enables every lockable widget unconditionally  - 
        # re-apply limit toggles to restore the correct enabled/disabled state.
        if hasattr(self, "_yl_lo"):
            self._tog_ylim()
        if hasattr(self, "_xl_lo"):
            self._tog_xlim()
        self._sync_analyze_btn()

    def _sync_analyze_btn(self):
        """Enable/disable the Help Analyze button based on validation state."""
        try:
            state = "normal" if (self._validated and self._pf_ready) else "disabled"
            self._help_analyze_btn.config(state=state)
        except Exception:
            pass

    def _build_lockable_cache(self):
        """Build a flat list of widgets that should be locked/unlocked.
        Called once after _build() and again after each lazy tab build."""
        skip_cls  = {"TLabel", "TSeparator", "TFrame", "Frame",
                     "Canvas", "Scrollbar", "TScrollbar",
                     "Listbox"}   # sidebar listbox must never be locked

        # Everything in the file row frames stays on (Browse button + path entry)
        always_on = set(self._file_row_frames)
        # All children of file row frames stay on too
        for frf in self._file_row_frames:
            try:
                for child in frf.winfo_children():
                    always_on.add(child)
            except Exception:
                pass
        # Sheet comboboxes must always stay enabled  -  user must be able to
        # switch sheets even after a validation failure on the current sheet
        for cb in getattr(self, "_sheet_cbs", []):
            always_on.add(cb)
        # The sidebar items and their parent are always on
        if hasattr(self, "_sb_sidebar_outer"):
            always_on.add(self._sb_sidebar_outer)
            try:
                def _walk_always(w):
                    always_on.add(w)
                    for ch in w.winfo_children():
                        _walk_always(ch)
                _walk_always(self._sb_sidebar_outer)
            except Exception:
                pass

        lockable = []

        def _walk(widget):
            cls = widget.winfo_class()
            if widget in always_on: return
            if cls == "Listbox": return   # never recurse into Listbox

            # Design-system custom widgets  -  lock via their .config(state=...)
            if getattr(widget, '_is_pwidget', False):
                if not getattr(widget, '_lock_exempt', False):
                    lockable.append(widget)
                return   # don't recurse into internals

            if cls in skip_cls:
                # recurse into container frames but don't add them
                for child in widget.winfo_children():
                    _walk(child)
                return

            if cls == "TRadiobutton": return
            if cls == "TButton":
                try:
                    txt = str(widget.cget("text"))
                except Exception:
                    txt = ""
                if "Browse" in txt or "Template" in txt:
                    return
            if cls not in skip_cls:
                lockable.append(widget)
            for child in widget.winfo_children():
                _walk(child)

        children = self.winfo_children()
        for child in children:
            if isinstance(child, ttk.Frame) and child is children[-1]:
                continue
            _walk(child)

        self._lockable_widgets = lockable

    def _set_form_state(self, state):
        if not hasattr(self, "_lockable_widgets"):
            self._build_lockable_cache()
        normal = (state == "!disabled")
        for widget in self._lockable_widgets:
            try:
                if getattr(widget, '_is_pwidget', False):
                    widget.config(state="normal" if normal else "disabled")
                else:
                    widget.state([state])
            except Exception:
                try:
                    widget.config(state="normal" if normal else "disabled")
                except Exception:
                    pass

    # ── Build ─────────────────────────────────────────────────────────────────

    def _build(self):
        # Build the native menu bar first (File / Edit / View / Window)
        self._build_menubar()

        # ── PanedWindow: sash disabled so user cannot accidentally drag it ────
        # We set sashwidth=0 and bind <Button-1> on the sash to absorb clicks.
        # The plot is always sized to fill the right pane via _embed_plot.
        paned = tk.PanedWindow(self, orient="horizontal",
                               sashwidth=0, sashrelief="flat", bg="#d1d1d6")
        self._paned = paned
        self._build_toolbar()
        paned.pack(fill="both", expand=True)
        self._left_pane = left = ttk.Frame(paned)
        paned.add(left, minsize=560)
        self._plot_type = tk.StringVar(value="bar")
        self._build_sidebar(left)
        self._build_right_pane(paned)

        # ── Suppress the spurious resize event at startup ─────────────────────
        # Tkinter fires a <Configure> event the moment the window is shown.
        # We gate _invalidate_pane_cache with a flag so the first event is
        # a no-op, preventing a flash of mismatched geometry.
        self._startup_done = False

        self.update_idletasks()
        self.geometry("1280x820")
        try:
            self.attributes("-zoomed", True)   # macOS / Linux
        except Exception:
            try:
                self.state("zoomed")           # Windows
            except Exception:
                pass
        self.deiconify()
        self._startup_done = True

        self.bind("<Command-r>", lambda e: self._run())
        self.bind("<Command-z>", lambda e: self._undo())
        self.bind("<Command-Z>", lambda e: self._do_redo())
        self.bind("<Command-s>", lambda e: self._save_project())
        self.bind("<Command-e>", lambda e: self._download_png())
        self.bind("<Command-o>", lambda e: self._browse_file())
        self.bind("<Command-c>", lambda e: self._copy_to_clipboard())
        self.bind("<Command-equal>", lambda e: self._zoom_plot_factor(1.1))
        self.bind("<Command-plus>",  lambda e: self._zoom_plot_factor(1.1))
        self.bind("<Command-minus>", lambda e: self._zoom_plot_factor(0.9))
        self.bind("<Command-0>",     lambda e: self._reset_zoom())

        # Set sash to 50/50 after the window settles  -  not during startup
        def _set_sash():
            w = self.winfo_width()
            try:
                paned.sash_place(0, w // 2, 0)
            except Exception:
                pass
        self.after(200, _set_sash)

    def _build_menubar(self):
        """Build the native macOS / cross-platform menu bar.

        Implements File, Edit, View, Window menus with reasonable actions.
        Falls back gracefully on platforms where some items are not applicable.
        """
        menubar = tk.Menu(self)

        # ── File menu ────────────────────────────────────────────────────────
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="Open Excel File…",
                              accelerator="⌘O",
                              command=self._browse_file)
        file_menu.add_command(label="Open .cplot Project…",
                              command=self._open_cplot)
        file_menu.add_command(label="Import .pzfx File…",
                              command=self._open_pzfx)
        file_menu.add_separator()
        file_menu.add_command(label="Save Project…",
                              accelerator="⌘S",
                              command=self._save_project)
        file_menu.add_separator()
        file_menu.add_command(label="Generate Plot",
                              accelerator="⌘R",
                              command=self._run)
        file_menu.add_command(label="Undo Last Plot",
                              accelerator="⌘Z",
                              command=self._undo)
        file_menu.add_separator()
        file_menu.add_command(label="Export Plot…",
                              accelerator="⌘E",
                              command=self._download_png)
        file_menu.add_command(label="Export All Charts (PDF Report)…",
                              command=self._export_all_pdf)
        file_menu.add_separator()
        file_menu.add_command(label="Copy PNG",
                              accelerator="⌘C",
                              command=self._copy_to_clipboard)
        file_menu.add_command(label="Copy Transparent PNG",
                              command=self._copy_transparent)
        file_menu.add_command(label="Copy SVG",
                              command=self._copy_as_svg)
        menubar.add_cascade(label="File", menu=file_menu)

        # ── Edit menu ─────────────────────────────────────────────────────────
        edit_menu = tk.Menu(menubar, tearoff=0)
        edit_menu.add_command(label="Select All",
                              accelerator="⌘A",
                              command=lambda: self.focus_get() and
                              self.focus_get().event_generate("<<SelectAll>>"))
        edit_menu.add_separator()
        edit_menu.add_command(label="Reset to Defaults",
                              command=self._reset_vars_to_defaults)
        menubar.add_cascade(label="Edit", menu=edit_menu)

        # ── View menu ─────────────────────────────────────────────────────────
        view_menu = tk.Menu(menubar, tearoff=0)
        view_menu.add_command(label="Zoom In",
                              accelerator="⌘+",
                              command=lambda: self._zoom_plot_factor(1.1))
        view_menu.add_command(label="Zoom Out",
                              accelerator="⌘−",
                              command=lambda: self._zoom_plot_factor(0.9))
        view_menu.add_command(label="Reset Zoom",
                              accelerator="⌘0",
                              command=self._reset_zoom)
        view_menu.add_separator()
        view_menu.add_checkbutton(label="Use Web Renderer (Plotly)",
                                  variable=self._use_webview)
        view_menu.add_separator()
        view_menu.add_command(label="Reference Wiki",
                              command=self._open_wiki_popup)
        menubar.add_cascade(label="View", menu=view_menu)

        # ── Window menu ───────────────────────────────────────────────────────
        window_menu = tk.Menu(menubar, tearoff=0)
        window_menu.add_command(label="Minimize",
                                accelerator="⌘M",
                                command=self.iconify)
        window_menu.add_command(label="Bring All to Front",
                                command=self.lift)
        menubar.add_cascade(label="Window", menu=window_menu)

        # ── Help menu ────────────────────────────────────────────────────────
        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="Help Analyze…",
                              command=self._help_analyze)
        help_menu.add_separator()
        help_menu.add_command(label="About Refraction",
                              command=self._show_about)
        menubar.add_cascade(label="Help", menu=help_menu)

        self.config(menu=menubar)

        # ── Chart-type keyboard shortcuts: ⌘1–⌘9 ─────────────────────────────
        try:
            from refraction.core.registry import KEYBOARD_SHORTCUTS, _REGISTRY_SPECS
            _key_to_idx = {spec.key: i for i, spec in enumerate(_REGISTRY_SPECS)}
            for num, chart_key in KEYBOARD_SHORTCUTS.items():
                idx = _key_to_idx.get(chart_key)
                if idx is not None:
                    self.bind_all(f"<Command-Key-{num}>",
                                  lambda e, i=idx: self._jump_to_chart(i))
                    self.bind_all(f"<Control-Key-{num}>",
                                  lambda e, i=idx: self._jump_to_chart(i))
        except Exception:
            _log.debug("App._build_menubar: keyboard shortcut binding failed", exc_info=True)

    def _jump_to_chart(self, idx: int):
        """Switch to chart type by registry index (used by ⌘1–⌘9)."""
        try:
            if hasattr(self, "_sb_select"):
                self._sb_select(idx)
            if hasattr(self, "_sb_show_pane"):
                self._sb_show_pane(idx)
        except Exception:
            _log.debug("App._jump_to_chart: failed to jump to chart index %d", idx, exc_info=True)

    def _build_toolbar(self):
        """Build the bottom status/button bar.

        Changes from v6:
          • Copy PNG and Copy (α) consolidated into one "Copy" button with a
            smaller "Copy α" ghost button for the transparent variant.
          • Keyboard hint strip shortened and less cluttered.
          • MIT license credit added bottom-left in gray.
        """
        ttk.Separator(self).pack(side="bottom", fill="x")
        bot = ttk.Frame(self)
        bot.pack(side="bottom", fill="x", padx=10, pady=5)

        # ── Left side: status + wiki + MIT note ───────────────────────────────
        left_grp = ttk.Frame(bot)
        left_grp.pack(side="left", fill="y")

        self._status_var = tk.StringVar(value="")
        self._status_lbl = ttk.Label(left_grp, textvariable=self._status_var,
                                     font=("Helvetica Neue", 11))
        self._status_lbl.pack(side="top", anchor="w")

        self._wiki_btn = PButton(left_grp, text="📖 Reference Wiki", style="ghost",
                                 command=self._open_wiki_popup)
        self._wiki_btn._lock_exempt = True
        self._wiki_btn.pack(side="top", anchor="w", pady=(1, 0))

        # MIT license note
        ttk.Label(left_grp,
                  text="MIT License · Designed & implemented by Claude (Anthropic) · Ashwin Pasupathy",
                  foreground="#cccccc", font=("Helvetica Neue", 9)
                  ).pack(side="top", anchor="w")

        # Compact keyboard shortcut strip
        ttk.Label(bot,
                  text="⌘R Generate  ·  ⌘Z Undo  ·  ⌘C Copy  ·  ⌘E Export  ·  ⌘O Open  ·  ⌘+/− Zoom",
                  foreground="#cccccc", font=("Helvetica Neue", 9)
                  ).pack(side="left", padx=(12, 0))

        # ── Right side: action buttons ────────────────────────────────────────
        self._export_btn = PButton(bot, text="Export…", style="secondary",
                                   command=self._download_png, state="disabled")
        self._export_btn.pack(side="right", padx=(4, 0))
        _create_tooltip(self._export_btn, "Export PNG / TIFF / EPS / SVG / PDF  (⌘E)")

        # Single consolidated "Copy" button
        self._copy_btn = PButton(bot, text="Copy", style="secondary",
                                 command=self._copy_to_clipboard, state="disabled")
        self._copy_btn.pack(side="right", padx=(4, 0))
        _create_tooltip(self._copy_btn, "Copy PNG to clipboard  (⌘C)")

        self._copy_transparent_btn = PButton(bot, text="Copy α", style="ghost",
                                             command=self._copy_transparent, state="disabled")
        self._copy_transparent_btn.pack(side="right", padx=(1, 0))
        _create_tooltip(self._copy_transparent_btn, "Copy with transparent background")

        self._copy_svg_btn = PButton(bot, text="SVG", style="ghost",
                                     command=self._copy_as_svg, state="disabled")
        self._copy_svg_btn.pack(side="right", padx=(1, 0))
        _create_tooltip(self._copy_svg_btn, "Copy as SVG")

        # Zoom controls
        self._zoom_reset_btn = PButton(bot, text="1:1", style="ghost",
                                       command=self._reset_zoom, state="disabled", width=3)
        self._zoom_reset_btn.pack(side="right", padx=(2, 0))
        self._zoom_out_btn = PButton(bot, text="−", style="ghost",
                                     command=lambda: self._zoom_plot_factor(0.9),
                                     state="disabled", width=2)
        self._zoom_out_btn.pack(side="right", padx=(1, 0))
        self._zoom_in_btn = PButton(bot, text="+", style="ghost",
                                    command=lambda: self._zoom_plot_factor(1.1),
                                    state="disabled", width=2)
        self._zoom_in_btn.pack(side="right", padx=(4, 0))

        self._run_btn = PButton(bot, text="Generate Plot", style="primary", command=self._run)
        self._run_btn.pack(side="right", padx=(6, 0))
        _create_tooltip(self._run_btn, "Generate Plot  (⌘R)")

        self._undo_btn = PButton(bot, text="Undo", style="ghost",
                                 command=self._undo, state="disabled")
        self._undo_btn.pack(side="right", padx=(2, 0))
        _create_tooltip(self._undo_btn, "Undo last plot  (⌘Z)")

        self._help_analyze_btn = PButton(bot, text="✦ Help Analyze",
                                         style="secondary", command=self._help_analyze,
                                         state="disabled")
        self._help_analyze_btn._lock_exempt = True   # managed separately via _sync_analyze_btn
        self._help_analyze_btn.pack(side="right", padx=(4, 8))

    def _build_sidebar(self, left):
        """Build the chart-type icon sidebar and the scrollable content panes."""
        SB_BG       = "#f0f0f0"
        SB_SEL_BG   = "#2274A5"
        SB_SEL_FG   = "white"
        SB_NORM_FG  = "#333333"
        SB_HOVER_BG = "#dde8f5"

        sidebar_outer = tk.Frame(left, width=_SB_WIDTH, bg=SB_BG)
        sidebar_outer.pack(side="left", fill="y")
        sidebar_outer.pack_propagate(False)
        self._sb_sidebar_outer = sidebar_outer

        hdr_lbl = tk.Label(sidebar_outer, text="Chart Type",
                           font=("Helvetica Neue", 11, "bold"),
                           fg="#555555", bg=SB_BG, anchor="w")
        hdr_lbl.pack(fill="x", padx=10, pady=(10, 4))
        tk.Frame(sidebar_outer, bg="#cccccc", height=1).pack(fill="x", padx=6)

        _sb_vsb    = ttk.Scrollbar(sidebar_outer, orient="vertical")
        _sb_vsb.pack(side="right", fill="y")
        _sb_canvas = tk.Canvas(sidebar_outer, bg=SB_BG, highlightthickness=0,
                               yscrollcommand=_sb_vsb.set, width=_SB_WIDTH)
        _sb_canvas.pack(side="left", fill="both", expand=True)
        _sb_vsb.config(command=_sb_canvas.yview)
        _sb_inner = tk.Frame(_sb_canvas, bg=SB_BG)
        _sb_win   = _sb_canvas.create_window((0, 0), window=_sb_inner, anchor="nw")
        _sb_inner.bind("<Configure>",
            lambda e: _sb_canvas.configure(scrollregion=_sb_canvas.bbox("all")))
        _sb_canvas.bind("<Configure>",
            lambda e: _sb_canvas.itemconfig(_sb_win, width=e.width))

        _sb_items    = []
        _sb_selected = [0]

        def _sb_paint(idx, selected):
            if idx >= len(_sb_items): return
            row, ic, lbl = _sb_items[idx]
            key = _SB_KEYS[idx]
            bg  = SB_SEL_BG if selected else SB_BG
            fg  = SB_SEL_FG if selected else SB_NORM_FG
            row.config(bg=bg); ic.config(bg=bg); lbl.config(bg=bg, fg=fg)
            ic.delete("all")
            _ICON_FN.get(key, _icon_bar)(ic, SB_SEL_FG if selected else "#2274A5", bg)

        def _sb_select(idx):
            old = _sb_selected[0]
            _sb_selected[0] = idx
            _sb_paint(old, False)
            _sb_paint(idx, True)

        def _sb_wheel(e):
            raw = e.delta if e.delta else 0
            if not raw: return
            bbox = _sb_canvas.bbox("all")
            if not bbox: return
            content_h = bbox[3] - bbox[1]
            view_h    = _sb_canvas.winfo_height()
            if content_h <= view_h: return
            import sys as _sys_sb
            _scale_sb = 8.0 if _sys_sb.platform == "darwin" and abs(raw) < 40 else 4.0
            frac = -raw * _scale_sb / content_h
            cur  = _sb_canvas.yview()[0]
            _sb_canvas.yview_moveto(max(0.0, min(1.0 - view_h/content_h, cur + frac)))
        _sb_canvas.bind("<MouseWheel>", _sb_wheel)
        _sb_inner.bind("<MouseWheel>",  _sb_wheel)

        _SB_KEYS = []

        def _sb_add_item(key, label_text, idx):
            _SB_KEYS.append(key)
            row = tk.Frame(_sb_inner, bg=SB_BG, cursor="hand2", height=_SB_ITEM_H)
            row.pack(fill="x", pady=0)
            row.pack_propagate(False)
            ic = tk.Canvas(row, width=_SB_ICON_SIZE, height=_SB_ICON_SIZE,
                           bg=SB_BG, highlightthickness=0)
            ic.place(x=8, y=(_SB_ITEM_H - _SB_ICON_SIZE) // 2)
            _ICON_FN.get(key, _icon_bar)(ic, "#2274A5", SB_BG)
            lbl = tk.Label(row, text=label_text, font=("Helvetica Neue", 11),
                           fg=SB_NORM_FG, bg=SB_BG, anchor="w",
                           justify="left", wraplength=_SB_WIDTH - _SB_ICON_SIZE - 22)
            lbl.place(x=_SB_ICON_SIZE + 14, y=0,
                      width=_SB_WIDTH - _SB_ICON_SIZE - 20, height=_SB_ITEM_H)

            def _enter(e, i=idx):
                if _sb_selected[0] != i:
                    for w in (_sb_items[i][0], _sb_items[i][1], _sb_items[i][2]):
                        w.config(bg=SB_HOVER_BG)
            def _leave(e, i=idx):
                if _sb_selected[0] != i:
                    for w in (_sb_items[i][0], _sb_items[i][1], _sb_items[i][2]):
                        w.config(bg=SB_BG)
            def _click(e, i=idx):
                _sb_select(i)
                _show_pane(i)
                _on_tab_change()

            for w in (row, ic, lbl):
                w.bind("<Enter>",    _enter)
                w.bind("<Leave>",    _leave)
                w.bind("<Button-1>", _click)
                w.bind("<MouseWheel>", _sb_wheel)

            tk.Frame(_sb_inner, bg="#e0e0e0", height=1).pack(fill="x")
            _sb_items.append((row, ic, lbl))

        class _SbShim:
            def curselection(self):     return (_sb_selected[0],)
            def selection_set(self, i): _sb_select(i)
            def activate(self, i):      pass
            def insert(self, pos, txt): pass
            @property
            def master(self):           return sidebar_outer

        self._chart_lb = _SbShim()

        ttk.Separator(left, orient="vertical").pack(side="left", fill="y")

        # Outer wrapper holds the navigator strip + scrollable content
        _content_outer = ttk.Frame(left)
        _content_outer.pack(side="left", fill="both", expand=True)

        # ── Section jump navigator ──────────────────────────────────────────
        _nav_bar = tk.Frame(_content_outer, bg="#f0f4fa", height=30)
        _nav_bar.pack(side="top", fill="x")
        _nav_bar.pack_propagate(False)
        tk.Label(_nav_bar, text="Jump to:", bg="#f0f4fa", fg="#666666",
                 font=("Helvetica Neue", 10)).pack(side="left", padx=(10, 6))
        self._nav_btns = {}
        for _sec_name, _sec_icon in (("Data", "📊"), ("Axes", "📐"), ("Statistics", "📈")):
            _b = tk.Label(_nav_bar, text=f"{_sec_icon} {_sec_name}",
                          bg="#f0f4fa", fg="#2274A5",
                          font=("Helvetica Neue", 10, "bold"),
                          cursor="hand2", padx=8, pady=4)
            _b.pack(side="left", padx=2)
            _b.bind("<Enter>", lambda e, b=_b: b.config(bg="#dde8f5"))
            _b.bind("<Leave>", lambda e, b=_b: b.config(bg="#f0f4fa"))
            self._nav_btns[_sec_name] = _b

        self._section_anchors = {}  # key -> {section_name: y_fraction}

        def _jump_to_section(section_name):
            idx = _current_idx[0]
            chart_key = _tab_map.get(idx, "")
            anchors = self._section_anchors.get(chart_key, {})
            canvas = self._tab_canvases.get((idx, 0))
            if canvas and section_name in anchors:
                canvas.yview_moveto(anchors[section_name])

        for _sec_name in ("Data", "Axes", "Statistics"):
            self._nav_btns[_sec_name].bind(
                "<Button-1>",
                lambda e, s=_sec_name: _jump_to_section(s)
            )

        content_frame = ttk.Frame(_content_outer)
        content_frame.pack(side="top", fill="both", expand=True)

        _panes       = {}
        _current_idx = [0]

        def _show_pane(idx):
            for p in _panes.values(): p.place_forget()
            if idx in _panes:
                _panes[idx].place(relx=0, rely=0, relwidth=1, relheight=1)
            _current_idx[0] = idx

        class _FakePlotNB:
            def select(self, idx=None):
                if idx is not None: _show_pane(idx)
            def index(self, what=None): return _current_idx[0]
            def bind(self, seq, fn): pass

        self._plot_nb = _FakePlotNB()

        def _make_scrollable(parent):
            vsb    = ttk.Scrollbar(parent, orient="vertical")
            vsb.pack(side="right", fill="y")
            canvas = tk.Canvas(parent, highlightthickness=0, yscrollcommand=vsb.set)
            canvas.pack(side="left", fill="both", expand=True)
            vsb.config(command=canvas.yview)
            inner = ttk.Frame(canvas)
            win   = canvas.create_window((0, 0), window=inner, anchor="nw")
            inner.bind("<Configure>",
                lambda e, c=canvas: c.configure(scrollregion=c.bbox("all")))
            canvas.bind("<Configure>",
                lambda e, c=canvas, w=win: c.itemconfig(w, width=e.width))
            return inner, canvas

        self._tab_canvases = {}

        def _add_plot_type(nb_attr, label_text, key, mode, main_idx):
            pane = ttk.Frame(content_frame)
            _panes[main_idx] = pane
            inner, canvas = _make_scrollable(pane)
            for sub in range(3):
                self._tab_canvases[(main_idx, sub)] = canvas
            class _StubNB:
                def select(self, i=None): pass
                def index(self, w=None):  return 0
            setattr(self, nb_attr, _StubNB())
            tabs = [(inner, canvas)] * 3
            _sb_add_item(key, label_text, main_idx)
            return tabs

        _all_tabs  = []
        _tab_inner = {}
        for spec in _REGISTRY_SPECS:
            tabs = _add_plot_type(f"_{spec.key.replace('-','_')}_nb",
                                  spec.label, spec.key, spec.tab_mode, len(_all_tabs))
            _all_tabs.append(tabs)
            _tab_inner[spec.key] = tabs

        _tab_map  = {i: spec.key for i, spec in enumerate(_REGISTRY_SPECS)}

        # Wiki is no longer a sidebar tab  -  it lives in a bottom-bar popup button.
        # Keep wiki_pane as a stub so old code paths don't crash.
        wiki_pane = ttk.Frame(content_frame)

        _show_pane(0)
        _sb_select(0)

        # Expose closures so _sb_select_silent can call them directly
        self._sb_select    = _sb_select
        self._sb_show_pane = _show_pane

        def _on_chart_type_change(e=None, from_tab_switch=False):
            idx = _current_idx[0]
            key = _tab_map.get(idx, "bar")
            self._plot_type.set(key)
            if key in self._all_tabs_map:
                self._build_tab_content(key)
            if not from_tab_switch:
                self.after(0, self._reset_chart_type_state)

        # Keep backward-compat alias used in a few other places
        _on_tab_change = _on_chart_type_change

        _pane_cache = {}
        def _invalidate_pane_cache(e=None):
            # Skip the spurious resize that fires during window creation
            if not getattr(self, "_startup_done", False):
                return
            _pane_cache.clear()
        self.bind("<Configure>", _invalidate_pane_cache)

        def _smart_wheel(e):
            """Route scroll events strictly to the pane the cursor is over.

            Guard: if any popup dialog is open, return immediately so the
            background pane does not scroll while a popup has focus.

            Fix: exact hit-test with fresh coords each time to prevent cross-pane
            scroll bleed that reappeared when the pane cache was stale.
            """
            if getattr(self, "_popup_count", 0) > 0:
                return
            wx, wy = e.x_root, e.y_root

            def _fresh_over(w):
                """Always query live geometry - never use stale cache for hit test."""
                try:
                    x1 = w.winfo_rootx()
                    y1 = w.winfo_rooty()
                    x2 = x1 + w.winfo_width()
                    y2 = y1 + w.winfo_height()
                    return x1 <= wx <= x2 and y1 <= wy <= y2
                except Exception:
                    return False

            def _over(w):
                """Cache geometry  -  only used for non-critical hit tests."""
                wid = id(w)
                if wid not in _pane_cache:
                    try:
                        _pane_cache[wid] = (w.winfo_rootx(), w.winfo_rooty(),
                                            w.winfo_rootx() + w.winfo_width(),
                                            w.winfo_rooty() + w.winfo_height())
                    except Exception:
                        return False
                x1, y1, x2, y2 = _pane_cache[wid]
                return x1 <= wx <= x2 and y1 <= wy <= y2

            def _scroll_canvas(canvas, raw):
                import sys
                bbox = canvas.bbox("all")
                if not bbox: return
                content_h = bbox[3] - bbox[1]
                view_h    = canvas.winfo_height()
                if content_h <= view_h: return
                # macOS trackpad sends tiny fractional deltas; scale up
                # macOS trackpad: small deltas need bigger scale for responsive feel
                scale = 8.0 if sys.platform == "darwin" and abs(raw) < 40 else 4.0
                frac = -raw * scale / content_h
                cur  = canvas.yview()[0]
                canvas.yview_moveto(max(0.0, min(1.0 - view_h/content_h, cur + frac)))

            raw = e.delta if e.delta else 0
            if not raw: return

            # Use _fresh_over for the two main panes to get exact real-time geometry
            in_right = _fresh_over(self._right_pane)
            in_left  = _fresh_over(self._left_pane)

            if in_right:
                if e.state & 0x8 or e.state & 0x4:
                    self._zoom_plot(raw); return
                if e.state & 0x1:
                    import sys as _sys
                    bbox = self._plot_canvas.bbox("all")
                    if bbox:
                        content_w = bbox[2] - bbox[0]
                        view_w    = self._plot_canvas.winfo_width()
                        if content_w > view_w:
                            _hscale = 8.0 if _sys.platform == "darwin" and abs(raw) < 40 else 4.0
                            cur = self._plot_canvas.xview()[0]
                            self._plot_canvas.xview_moveto(
                                max(0.0, min(1.0, cur + (-raw * _hscale / content_w))))
                else:
                    bbox = self._plot_canvas.bbox("all")
                    if bbox:
                        content_h = bbox[3] - bbox[1]
                        view_h    = self._plot_canvas.winfo_height()
                        if content_h > view_h:
                            cur = self._plot_canvas.yview()[0]
                            self._plot_canvas.yview_moveto(
                                max(0.0, min(1.0, cur + (-raw * 4 / content_h))))
                return

            if in_left and not _fresh_over(sidebar_outer):
                main_idx = _current_idx[0]
                canvas   = self._tab_canvases.get((main_idx, 0))
                if canvas:
                    _scroll_canvas(canvas, raw)
            # If cursor is over sidebar, do nothing (sidebar has fixed content)

        self.bind_all("<MouseWheel>", _smart_wheel)

        try:
            def _on_magnify(e):
                wx, wy = e.x_root, e.y_root
                try:
                    x1 = self._right_pane.winfo_rootx()
                    y1 = self._right_pane.winfo_rooty()
                    x2 = x1 + self._right_pane.winfo_width()
                    y2 = y1 + self._right_pane.winfo_height()
                except Exception:
                    return
                if x1 <= wx <= x2 and y1 <= wy <= y2:
                    self._zoom_plot_factor(1.0 + e.delta)
            self.bind_all("<Magnify>", _on_magnify)
        except Exception:
            _log.debug("App._build: <Magnify> event binding failed (non-macOS?)", exc_info=True)

        self._tabs_built   = set()
        self._all_tabs_map = {spec.key: (spec, tabs)
                              for spec, tabs in zip(_REGISTRY_SPECS, _all_tabs)}

        def _build_tab_content(key):
            if key in self._tabs_built: return
            self._tabs_built.add(key)
            spec, tabs = self._all_tabs_map[key]
            outer = tabs[0][0]

            def _section(title):
                hdr = tk.Frame(outer, bg="#dde7f3")
                hdr.pack(fill="x", padx=0, pady=(18, 2))
                tk.Frame(hdr, bg="#2274A5", width=4).pack(side="left", fill="y", padx=(0, 10))
                tk.Label(hdr, text=title.upper(), bg="#dde7f3", fg="#1a5a8a",
                         font=("Helvetica Neue", 10, "bold"), anchor="w"
                         ).pack(side="left", pady=6)
                sub = ttk.Frame(outer)
                sub.pack(fill="x", expand=False)
                sub.columnconfigure(0, weight=1)
                return sub

            data_frame = ttk.Frame(outer)
            data_frame.pack(fill="x", expand=False)
            data_frame.columnconfigure(0, weight=1)
            self._tab_data(data_frame, mode=spec.tab_mode)

            # Record section anchor for the navigator
            if key not in self._section_anchors:
                self._section_anchors[key] = {}
            self._section_anchors[key]["Data"] = 0.0  # always at top

            axes_frame  = _section("Axes")
            self._tab_axes(axes_frame, mode=spec.tab_mode)

            stats_frame = _section("Statistics")
            _stats_dispatch = {
                "grouped_bar":      self._tab_stats_grouped_bar,
                "scatter":          self._tab_stats_scatter,
                "kaplan_meier":     self._tab_stats_kaplan_meier,
                "heatmap":          self._tab_stats_heatmap,
                "two_way_anova":    self._tab_stats_two_way_anova,
                "before_after":     self._tab_stats_before_after,
                "histogram":        self._tab_stats_histogram,
                "curve_fit":        self._tab_stats_curve_fit,
                "column_stats":     self._tab_stats_column_stats,
                "contingency":      self._tab_stats_contingency,
                "repeated_measures":self._tab_stats_repeated_measures,
                "chi_square_gof":   self._tab_stats_chi_square_gof,
                "stacked_bar":      self._tab_stats_stacked_bar,
                "bubble":           self._tab_stats_bubble,
                "dot_plot":         self._tab_stats_dot_plot,
                "bland_altman":     self._tab_stats_bland_altman,
                "forest_plot":      self._tab_stats_forest_plot,
            }
            _stats_dispatch.get(spec.stats_tab, self._tab_stats)(stats_frame)

            ttk.Frame(outer, height=20).pack()
            self._build_lockable_cache()

            # Measure Axes and Statistics section y-positions after layout settles
            def _measure_anchors(key=key, axes_frame=axes_frame, stats_frame=stats_frame, outer=outer):
                canvas = self._tab_canvases.get((_tab_map.get(list(_tab_map.keys())[
                    list(_tab_map.values()).index(key)], -1) 
                    if key in _tab_map.values() else -1, 0))
                # Find pane index for this key
                for idx, k in _tab_map.items():
                    if k == key:
                        canvas = self._tab_canvases.get((idx, 0))
                        break
                if canvas is None:
                    return
                try:
                    canvas.update_idletasks()
                    total_h = canvas.bbox("all")
                    if not total_h: return
                    content_h = total_h[3] - total_h[1]
                    if content_h <= 0: return
                    # axes_frame y relative to the canvas window
                    axes_y = axes_frame.winfo_y()
                    stats_y = stats_frame.winfo_y()
                    self._section_anchors[key]["Axes"] = max(0.0, (axes_y - 10) / content_h)
                    self._section_anchors[key]["Statistics"] = max(0.0, (stats_y - 10) / content_h)
                except Exception:
                    _log.debug("App._build: _measure_anchors failed for key %r", key, exc_info=True)
            self.after(300, _measure_anchors)

        self._build_tab_content = _build_tab_content
        _build_tab_content("bar")

    def _build_right_pane(self, paned):
        """Build the tab bar and scrollable plot canvas on the right side of the split."""
        self._right_pane = right_outer = ttk.Frame(paned)
        paned.add(right_outer, minsize=300)

        # ── Tab bar ───────────────────────────────────────────────────────────
        self._tab_bar_widget = TabBar(
            right_outer,
            on_select=lambda tab_id: self._tab_manager.switch_to(tab_id),
            on_close=lambda tab_id: self._tab_manager.close_tab(tab_id),
            on_new=lambda: self._tab_manager.new_tab("bar"),
            on_reorder=lambda f, t: self._tab_manager.reorder(f, t),
        )
        self._tab_bar_widget.pack(fill="x", side="top")

        # ── Results panel  -  pinned at the bottom, hidden until first run ──────
        self._results_visible = False
        self._results_strip   = tk.Frame(right_outer, bg="#f4f7fb", height=0)
        self._results_strip.pack(side="bottom", fill="x")
        self._results_strip.pack_propagate(False)

        # Toggle header bar
        _rh = tk.Frame(self._results_strip, bg="#dde8f5", height=28, cursor="hand2")
        _rh.pack(fill="x")
        _rh.pack_propagate(False)
        self._results_toggle_arrow = tk.Label(_rh, text="▲ Results", bg="#dde8f5",
                                              fg="#2274A5", font=("Helvetica Neue", 11, "bold"),
                                              cursor="hand2")
        self._results_toggle_arrow.pack(side="left", padx=10)
        self._results_copy_btn = tk.Label(_rh, text="Export CSV", bg="#dde8f5",
                                          fg="#2274A5", font=("Helvetica Neue", 10),
                                          cursor="hand2")
        self._results_copy_btn.pack(side="right", padx=10)
        self._results_copy_btn.bind("<Button-1>", lambda e: self._export_results_csv())

        for w in (_rh, self._results_toggle_arrow):
            w.bind("<Button-1>", lambda e: self._toggle_results_panel())

        # Scrollable body
        _rb = tk.Frame(self._results_strip, bg="#f4f7fb")
        _rb.pack(fill="both", expand=True)
        _vsb = ttk.Scrollbar(_rb, orient="vertical")
        _vsb.pack(side="right", fill="y")
        _hsb = ttk.Scrollbar(_rb, orient="horizontal")
        _hsb.pack(side="bottom", fill="x")
        self._results_canvas = tk.Canvas(_rb, bg="#f4f7fb", highlightthickness=0,
                                         yscrollcommand=_vsb.set, xscrollcommand=_hsb.set)
        self._results_canvas.pack(side="left", fill="both", expand=True)
        _vsb.config(command=self._results_canvas.yview)
        _hsb.config(command=self._results_canvas.xview)
        self._results_inner = tk.Frame(self._results_canvas, bg="#f4f7fb")
        self._results_inner_id = self._results_canvas.create_window(
            (0, 0), window=self._results_inner, anchor="nw")
        self._results_inner.bind("<Configure>", lambda e: self._results_canvas.configure(
            scrollregion=self._results_canvas.bbox("all")))
        self._results_canvas.bind("<Configure>", lambda e: self._results_canvas.itemconfig(
            self._results_inner_id, width=e.width))
        # mousewheel on results
        def _rscroll(e):
            self._results_canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        self._results_canvas.bind("<MouseWheel>", _rscroll)
        self._results_inner.bind("<MouseWheel>", _rscroll)

        # ── Main plot scrollable canvas ───────────────────────────────────────
        right_vsb = ttk.Scrollbar(right_outer, orient="vertical")
        right_vsb.pack(side="right", fill="y")
        right_hsb = ttk.Scrollbar(right_outer, orient="horizontal")
        right_hsb.pack(side="bottom", fill="x")
        self._plot_canvas = tk.Canvas(right_outer, highlightthickness=0,
                                      yscrollcommand=right_vsb.set,
                                      xscrollcommand=right_hsb.set)
        self._plot_canvas.pack(side="left", fill="both", expand=True)
        right_vsb.config(command=self._plot_canvas.yview)
        right_hsb.config(command=self._plot_canvas.xview)

        self._plot_frame    = None   # points to active tab's plot_frame; set by TabManager
        self._canvas_widget = None   # FigureCanvasTkAgg; updated on tab switch
        self._fig           = None   # matplotlib Figure; updated on tab switch
        self._last_kw       = None   # kw dict from last successful render (for export)
        self._last_chart_type = None # chart type from last successful render
        self._zoom_level    = 1.0

        self._empty_state_frame = ttk.Frame(self._plot_canvas)
        self._empty_state_id    = self._plot_canvas.create_window(
            0, 0, window=self._empty_state_frame, anchor="center")

        def _centre_empty(e=None):
            self._plot_canvas.coords(
                self._empty_state_id,
                self._plot_canvas.winfo_width() // 2,
                self._plot_canvas.winfo_height() // 2)
        self._plot_canvas.bind("<Configure>", _centre_empty)

        ttk.Label(self._empty_state_frame, text="📊",
                  font=("Helvetica Neue", 48)).pack(pady=(0, 8))
        ttk.Label(self._empty_state_frame, text="No plot yet",
                  font=("Helvetica Neue", 18, "bold"),
                  foreground="#aaaaaa").pack()
        ttk.Label(self._empty_state_frame,
                  text="Browse for an Excel file on the left,\nthen click Generate Plot",
                  font=("Helvetica Neue", 13), foreground="#bbbbbb",
                  justify="center").pack(pady=(6, 0))

        # ── Tab manager ───────────────────────────────────────────────────────
        # Created after _plot_canvas exists; opens with one bar-chart tab.
        self._tab_manager = TabManager(self, self._tab_bar_widget, self._plot_canvas)
        self._tab_manager.new_tab("bar")

    # ── Tab builders ──────────────────────────────────────────────────────────

    # ── Wiki ──────────────────────────────────────────────────────────────────

    def _track_popup(self, dlg):
        """Increment popup counter when dlg opens, decrement when it closes.
        While _popup_count > 0 the main-window scroll handler is suspended,
        preventing the background pane from scrolling while a popup is active.
        """
        self._popup_count = getattr(self, "_popup_count", 0) + 1
        def _on_close():
            self._popup_count = max(0, getattr(self, "_popup_count", 1) - 1)
        dlg.protocol("WM_DELETE_WINDOW", lambda: (_on_close(), dlg.destroy()))
        dlg.bind("<Destroy>", lambda e: _on_close() if e.widget is dlg else None)

    def _open_wiki_popup(self):
        """Open the Reference Wiki in a floating popup window."""
        try:
            from refraction.app.wiki import open_wiki_popup
            open_wiki_popup(self,
                            track_popup_fn=self._track_popup,
                            bind_scroll_fn=None)
        except Exception as e:
            from tkinter import messagebox
            messagebox.showinfo("Wiki", f"Wiki unavailable: {e}")

    def _open_wiki_popup_inline(self):
        """Inline fallback wiki — kept for reference, not called directly."""
        dlg = tk.Toplevel(self)
        dlg.title("Reference Wiki")
        dlg.resizable(True, True)
        dlg.geometry("680x700")
        dlg.configure(bg="#ffffff")
        # non-modal so the user can keep the app visible alongside
        dlg.transient(self)
        self._track_popup(dlg)  # suspend background scroll while open

        # Header
        hdr = tk.Frame(dlg, bg="#2274A5", height=52)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="📖  Reference Wiki",
                 bg="#2274A5", fg="white",
                 font=("Helvetica Neue", 15, "bold")).pack(side="left", padx=18, pady=12)
        tk.Label(hdr, text="scipy · Statistics Reference",
                 bg="#2274A5", fg="#a8cce0",
                 font=("Helvetica Neue", 11)).pack(side="left", padx=(0, 18), pady=12)

        # Two-pane layout: sidebar TOC (left) + scrollable content (right)
        body_outer = tk.Frame(dlg, bg="#ffffff")
        body_outer.pack(fill="both", expand=True)

        # TOC sidebar
        toc_frame = tk.Frame(body_outer, bg="#f0f4fa", width=180)
        toc_frame.pack(side="left", fill="y")
        toc_frame.pack_propagate(False)
        tk.Label(toc_frame, text="CONTENTS", bg="#f0f4fa", fg="#2274A5",
                 font=("Helvetica Neue", 9, "bold"), anchor="w"
                 ).pack(fill="x", padx=10, pady=(12, 4))
        tk.Frame(toc_frame, bg="#c8d8ea", height=1).pack(fill="x", padx=6)
        toc_listbox = tk.Listbox(toc_frame, bg="#f0f4fa", fg="#333333",
                                 font=("Helvetica Neue", 11), selectmode="single",
                                 selectbackground="#2274A5", selectforeground="white",
                                 activestyle="none", relief="flat",
                                 highlightthickness=0, borderwidth=0, cursor="hand2")
        toc_listbox.pack(fill="both", expand=True, padx=4, pady=4)
        # Disable scroll on TOC listbox from bleeding to content
        toc_listbox.bind("<MouseWheel>", lambda e: "break")
        tk.Frame(body_outer, bg="#dde4ee", width=1).pack(side="left", fill="y")

        # Content area
        content_area = tk.Frame(body_outer, bg="#ffffff")
        content_area.pack(side="left", fill="both", expand=True)
        vsb    = ttk.Scrollbar(content_area, orient="vertical")
        vsb.pack(side="right", fill="y")
        canvas = tk.Canvas(content_area, highlightthickness=0,
                           yscrollcommand=vsb.set, background="white")
        canvas.pack(side="left", fill="both", expand=True)
        vsb.config(command=canvas.yview)
        frame = ttk.Frame(canvas)
        win   = canvas.create_window((0, 0), window=frame, anchor="nw")
        frame.bind("<Configure>", lambda e: canvas.configure(
            scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(win, width=e.width))

        # TOC anchor registry  -  {section_title: y_pixel_offset}
        _toc_anchors = {}
        _toc_entries = []  # ordered list of section titles

        def _scroll(e):
            import sys
            raw = getattr(e, "delta", 0) or (120 if getattr(e, "num", 0) == 4 else -120)
            if not raw: return
            bbox = canvas.bbox("all")
            if not bbox: return
            content_h = bbox[3] - bbox[1]
            view_h    = canvas.winfo_height()
            if content_h <= view_h: return
            # macOS trackpad sends small deltas (±3–10); mouse sends ±120
            scale = 8.0 if sys.platform == "darwin" and abs(raw) < 40 else 4.0
            frac = -raw * scale / content_h
            cur  = canvas.yview()[0]
            canvas.yview_moveto(max(0.0, min(1.0 - view_h/content_h, cur + frac)))

        def _scroll_linux(e):
            raw = 120 if e.num == 4 else -120
            _scroll(type("E", (), {"delta": raw, "num": e.num})())

        # Bind only to this popup's widget tree (not bind_all which leaks to parent)
        dlg.after(50, lambda: _bind_scroll_recursive(dlg, _scroll, _scroll_linux, _scroll_linux))

        def _toc_select(e):
            sel = toc_listbox.curselection()
            if not sel: return
            title = _toc_entries[sel[0]]
            lbl_widget = _toc_anchors.get(title)
            if lbl_widget is None: return
            try:
                # Get the label's y position relative to the scrolled canvas content
                lbl_widget.update_idletasks()
                lbl_y = lbl_widget.winfo_y()
                bbox = canvas.bbox("all")
                if bbox:
                    content_h = bbox[3] - bbox[1]
                    if content_h > 0:
                        canvas.yview_moveto(max(0.0, (lbl_y - 10) / content_h))
            except Exception:
                pass
        toc_listbox.bind("<<ListboxSelect>>", _toc_select)

        # ── Content helpers ───────────────────────────────────────────────────
        def h1(text):
            lbl = ttk.Label(frame, text=text,
                      font=("Helvetica Neue", 16, "bold"),
                      foreground="#2274A5", background="white")
            lbl.pack(anchor="w", padx=20, pady=(20, 4))
            ttk.Separator(frame).pack(fill="x", padx=20, pady=(0, 8))
            # Register TOC entry
            _toc_entries.append(text)
            toc_listbox.insert("end", f"  {text}")
            # Store a reference to measure y-position after layout
            _toc_anchors[text] = lbl

        def h2(text):
            ttk.Label(frame, text=text,
                      font=("Helvetica Neue", 13, "bold"),
                      foreground="#444444", background="white"
                      ).pack(anchor="w", padx=20, pady=(12, 2))

        def body(text):
            import re
            text = re.sub(r'\*\*(.+?)\*\*', r'\1', text)
            text = re.sub(r'\*(.+?)\*',     r'\1', text)
            text = re.sub(r'``(.+?)``',     r'\1', text)
            text = re.sub(r':math:`(.+?)`', r'\1', text)
            text = re.sub(r'.. versionadded.*', '', text)
            text = re.sub(r'.. versionchanged.*', '', text)
            text = text.strip()
            if not text: return
            ttk.Label(frame, text=text,
                      font=("Helvetica Neue", 12),
                      foreground="#333333", background="white",
                      wraplength=580, justify="left"
                      ).pack(anchor="w", padx=28, pady=(0, 6))

        def tag(text, color="#2274A5"):
            lbl = tk.Label(frame, text=f" {text} ",
                           font=("Helvetica Neue", 10, "bold"),
                           fg="white", bg=color, padx=4, pady=1)
            lbl.pack(anchor="w", padx=28, pady=(0, 4))

        def plotter_link(page_slug, link_text):
            """Reference label (non-clickable) for statistics documentation."""
            lbl = tk.Label(frame, text=f"📎 {link_text}",
                           font=("Helvetica Neue", 11),
                           fg="#6B7280", bg="white")
            lbl.pack(anchor="w", padx=28, pady=(0, 6))

        def scipy_entry(fn, label_text, use_sections=("Notes",)):
            import inspect, re
            doc = inspect.getdoc(fn) or ""
            paras = doc.split("\n\n")
            if paras:
                body(paras[0].replace("\n", " "))
            current_section = None
            section_buf     = []
            for para in paras[1:]:
                lines = para.strip().splitlines()
                if (len(lines) >= 2 and
                        set(lines[1].strip()) <= set("-=~^") and
                        len(lines[1].strip()) >= 3):
                    if current_section in use_sections and section_buf:
                        sentences = re.split(r'(?<=[.!?])\s+', " ".join(section_buf).strip())
                        body(" ".join(sentences[:6]))
                    current_section = lines[0].strip()
                    section_buf     = []
                elif current_section in use_sections:
                    clean = " ".join(para.replace("\n", " ").split())
                    if clean and not clean.startswith(">>>"):
                        section_buf.append(clean)
            if current_section in use_sections and section_buf:
                sentences = re.split(r'(?<=[.!?])\s+', " ".join(section_buf).strip())
                body(" ".join(sentences[:6]))

        from scipy import stats as _st

        # ── Statistical Tests ─────────────────────────────────────────────────
        h1("Statistical Tests")

        h2("Independent Samples t-test  (Parametric, 2 groups)")
        tag("scipy.stats.ttest_ind", "#2274A5")
        scipy_entry(_st.ttest_ind, "ttest_ind")
        body("Welch's t-test (unequal variances) does not assume equal standard "
             "deviations and is preferred over Student's t-test in most practical "
             "situations.")
        plotter_link("stat_qa_choosing_a_test_to_compare_.htm",
                   "Prism 11: Choosing a test to compare two groups")

        h2("Paired t-test  (Parametric, 2 matched groups)")
        tag("scipy.stats.ttest_rel", "#2274A5")
        scipy_entry(_st.ttest_rel, "ttest_rel")
        plotter_link("stat_checklist_pairedt.htm",
                   "Prism 11: Checklist for paired t-test")

        h2("One-Way ANOVA  (Parametric, 3+ groups)")
        tag("scipy.stats.f_oneway", "#2274A5")
        scipy_entry(_st.f_oneway, "f_oneway")
        body("Prism 8+ also offers Welch's ANOVA and Brown-Forsythe ANOVA when equal "
             "variances cannot be assumed. These adjust the F-ratio and degrees of freedom "
             "for heteroscedasticity  -  recommended when Levene's or Bartlett's test is significant.")
        plotter_link("stat_checklist_1wayanova.htm",
                   "Prism 11: Checklist for one-way ANOVA")
        plotter_link("stat_how_to_multiple_comparisons_af.htm",
                   "Prism 11: Multiple comparisons after one-way ANOVA")

        h2("Repeated-Measures One-Way ANOVA  (Parametric, matched)")
        tag("pingouin.rm_anova", "#2274A5")
        body("Tests differences across 3+ conditions when the same subjects are measured "
             "in each condition. More powerful than one-way ANOVA when matching is effective. "
             "Prism 11 applies the Geisser-Greenhouse correction by default when sphericity "
             "cannot be assumed.")
        plotter_link("stat_checklist_1wayanova_rm.htm",
                   "Prism 11: Repeated-measures one-way ANOVA")

        h2("Two-Way ANOVA  (Parametric, 2 factors)")
        tag("pingouin.anova (Type II SS)", "#2274A5")
        body("Tests main effects of Factor A and Factor B, plus their interaction  -  "
             "whether the effect of one factor depends on the level of the other. "
             "Type II Sum of Squares tests each main effect after controlling for the other. "
             "Post-hoc pairwise comparisons are corrected with Holm-Bonferroni.")
        plotter_link("stat_howto_twowayanova.htm",
                   "Prism 11: Two-way ANOVA")

        h2("Mann-Whitney U  (Non-parametric, 2 groups)")
        tag("scipy.stats.mannwhitneyu", "#32936F")
        scipy_entry(_st.mannwhitneyu, "mannwhitneyu")
        plotter_link("stat_checklist_mwt.htm",
                   "Prism 11: Checklist for Mann-Whitney test")

        h2("Wilcoxon Signed-Rank  (Non-parametric, paired)")
        tag("scipy.stats.wilcoxon", "#32936F")
        scipy_entry(_st.wilcoxon, "wilcoxon")
        plotter_link("stat_checklist_wilcoxon_signed_rank.htm",
                   "Prism 11: Wilcoxon signed-rank test")

        h2("Kruskal-Wallis H  (Non-parametric, 3+ groups)")
        tag("scipy.stats.kruskal", "#32936F")
        scipy_entry(_st.kruskal, "kruskal")
        plotter_link("stat_checklist_kw.htm",
                   "Prism 11: Checklist for Kruskal-Wallis test")

        h2("Friedman Test  (Non-parametric, repeated measures)")
        tag("scipy.stats.friedmanchisquare", "#32936F")
        body("Non-parametric analogue of repeated-measures one-way ANOVA. Ranks values "
             "within each subject (row) and tests whether ranks differ across conditions. "
             "Follow with Dunn's post-hoc + Holm correction for pairwise comparisons.")
        plotter_link("stat_checklist_friedman.htm",
                   "Prism 11: Friedman test")

        h2("Tukey HSD  (Post-hoc, all-pairwise, parametric)")
        tag("scipy.stats.tukey_hsd", "#F18F01")
        try:
            scipy_entry(_st.tukey_hsd, "tukey_hsd")
        except Exception:
            body("Tukey's Honestly Significant Difference test compares all pairs of means "
                 "while maintaining the family-wise error rate. The Prism 11 default for "
                 "all-pairwise comparisons after one-way ANOVA when equal SDs are assumed.")
        body("Holm-Sidak is more powerful than Tukey HSD (can find significant differences "
             "where Tukey cannot) but does not produce confidence intervals. Prism 11 "
             "recommends Tukey for its combination of power and interpretability.")
        plotter_link("stat_options_tab_1wayanova.htm",
                   "Prism 11: Post-hoc test options for one-way ANOVA")

        h2("Dunn's Test  (Post-hoc, non-parametric)")
        tag("scikit_posthocs.posthoc_dunn", "#F18F01")
        body("Performs pairwise comparisons on ranked data after a significant "
             "Kruskal-Wallis or Friedman test. Holm-Bonferroni correction controls the "
             "family-wise error rate across all comparisons. Prism 11 default for "
             "non-parametric pairwise post-hoc analysis.")
        plotter_link("stat_checklist_kw.htm",
                   "Prism 11: Multiple comparisons after Kruskal-Wallis")

        # ── Normality & Variance Tests ────────────────────────────────────────
        h1("Normality & Variance Tests")

        h2("Shapiro-Wilk Normality Test")
        tag("scipy.stats.shapiro", "#A846A0")
        scipy_entry(_st.shapiro, "shapiro")
        body("Prism 11 runs Shapiro-Wilk automatically. However, Prism explicitly warns "
             "that 'it is not a good idea to base your decision solely on the normality test'  -  "
             "small samples have low power to detect non-normality, and large samples "
             "detect trivially small deviations. Use it as one input among several.")
        plotter_link("stat_checklist_1wayanova.htm",
                   "Prism 11: Normality assumption in ANOVA")

        h2("Levene's Test for Equal Variances")
        tag("scipy.stats.levene", "#A846A0")
        scipy_entry(_st.levene, "levene")
        body("Prism 11 uses Brown-Forsythe and Bartlett's tests for homoscedasticity, "
             "not Levene's. Brown-Forsythe (median-centred) is more robust to non-normality. "
             "Refraction uses Levene's as an equivalent diagnostic.")
        plotter_link("stat_checklist_1wayanova.htm",
                   "Prism 11: Equal variance assumption (Brown-Forsythe / Bartlett)")

        # ── Effect Sizes ─────────────────────────────────────────────────────
        h1("Effect Sizes")

        h2("Cohen's d  (2-group standardised difference)")
        tag("d = (μ₁ − μ₂) / s_pooled", "#6B4226")
        body("Standardised mean difference. d = 0.2: small, 0.5: medium, 0.8: large "
             "(Cohen 1988). Prism 11 reports Cohen's d for t-tests and partial η² for ANOVA.")
        plotter_link("stat_options_tab_one-way_anova.htm",
                   "Prism 11: Effect size for ANOVA")

        h2("Eta Squared / Partial Eta Squared  (ANOVA)")
        tag("η² = SS_effect / SS_total", "#6B4226")
        body("Proportion of total variance explained by a factor. η² = 0.01: small, "
             "0.06: medium, 0.14: large (Cohen). For one-way ANOVA η² = partial η². "
             "For two-way ANOVA Refraction reports partial η² (effect SS / (effect SS + error SS)).")
        plotter_link("stat_options_tab_one-way_anova.htm",
                   "Prism 11: Eta squared and omega squared")

        # ── Choosing the Right Test ───────────────────────────────────────────
        h1("Choosing the Right Test")

        body("Use parametric tests (t-test, ANOVA) when data are approximately normally "
             "distributed. Shapiro-Wilk p > 0.05 is one indicator, but also inspect "
             "histograms and Q-Q plots for small samples.")
        body("Use non-parametric tests (Mann-Whitney, Kruskal-Wallis) when normality "
             "fails, samples are small (n < 5), or data are ordinal. They compare "
             "ranks rather than means, losing less power than commonly assumed.")
        body("Use paired tests (paired t-test, Wilcoxon) when observations are matched  -  "
             "before/after on the same subject, or left-vs-right measurements. Pairing "
             "removes between-subject variability and greatly increases power.")
        body("After a significant ANOVA, always apply a post-hoc test with "
             "multiple-comparison correction. Prism 11 recommends Tukey HSD for "
             "all-pairwise comparisons and Dunnett for comparing groups to a control. "
             "Fisher's LSD provides no correction and inflates Type I error  -  "
             "only use it with a very specific justification.")
        plotter_link("stat_how_to_multiple_comparisons_af.htm",
                   "Prism 11: Guide to multiple comparisons")
        plotter_link("stat_---_principles_of_statistics_-.htm",
                   "Prism 11: Principles of Statistics (overview)")

        # ── Chart Types ───────────────────────────────────────────────────────
        h1("Chart Types")
        chart_info = [
            ("Bar Chart", "#2274A5",
             "Displays group means as vertical bars with error bars (SEM, SD, or 95% CI). "
             "Significance brackets show pairwise statistical comparisons. "
             "Best for comparing means across independent groups.",
             "stat_choosing_test_one-way_anova.htm",
             "Prism 11: One-way ANOVA (the analysis behind bar charts)"),
            ("Grouped Bar", "#F18F01",
             "Bar chart with two grouping factors  -  bars grouped by Category, colored "
             "by Subgroup. Best for 2×2 or 2×3 factorial designs. Stats compare subgroups "
             "within each category cluster.",
             "stat_howto_twowayanova.htm",
             "Prism 11: Two-way ANOVA"),
            ("Scatter Plot", "#048A81",
             "Plots individual (X, Y) data points. Optional linear regression with 95% CI "
             "band and Pearson or Spearman correlation annotation.",
             "stat_checklist_linearregression.htm",
             "Prism 11: Checklist for linear regression"),
            ("Survival Curve", "#D4AC0D",
             "Kaplan-Meier survival curves. Log-rank (Mantel-Cox) test compares groups. "
             "Tick marks show censored observations.",
             "stat_howto_km.htm",
             "Prism 11: Kaplan-Meier survival analysis"),
            ("Before / After", "#32936F",
             "Paired dot plot. Each subject appears in every condition connected by a line. "
             "Paired t-test or Wilcoxon signed-rank when statistics are enabled.",
             "stat_checklist_pairedt.htm",
             "Prism 11: Paired t-test"),
            ("Repeated Measures", "#D4AC0D",
             "Within-subjects design. Parametric: RM one-way ANOVA (Geisser-Greenhouse "
             "correction available). Non-parametric: Friedman + Dunn's post-hoc.",
             "stat_checklist_1wayanova_rm.htm",
             "Prism 11: Repeated-measures one-way ANOVA"),
            ("Bland-Altman", "#6B4226",
             "Assesses agreement between two measurement methods. Mean difference = bias; "
             "±1.96 SD lines = limits of agreement. Not a hypothesis test  -  "
             "clinical acceptability is the criterion.",
             "stat_howto_bland_altman.htm",
             "Prism 11: Bland-Altman plot"),
        ]
        for chart_name, color, desc, slug, link_text in chart_info:
            h2(chart_name)
            tag(chart_name, color)
            body(desc)
            plotter_link(slug, link_text)

        ttk.Frame(frame, height=40).pack()

        # Footer with single Close button
        ttk.Separator(dlg).pack(fill="x")
        foot = tk.Frame(dlg, bg="#ffffff")
        foot.pack(fill="x", padx=20, pady=10)
        PButton(foot, text="Close", style="primary",
                command=dlg.destroy).pack(side="right")

    # Keep _build_wiki as a no-op stub so nothing crashes if called
    def _build_wiki(self, parent):
        pass

    def _tab_data(self, f, mode="bar"):
        for k, v in [("excel_path", ""), ("sheet", ""),
                     ("error", "SEM (Standard Error)"), ("show_points", True),
                     ("show_n_labels", False), ("show_value_labels", False),
                     ("stacked_value_labels", False), ("error_below_bar", False),
                     ("jitter_amount", "0"),
                     ("color", "Default"), ("title", ""),
                     ("xlabel", ""), ("ytitle", "")]:
            if k not in self._vars:
                self._vars[k] = (tk.BooleanVar(value=v)
                                 if isinstance(v, bool) else tk.StringVar(value=v))

        # Initialise per-tab widget lists on first call
        if not hasattr(self, "_drop_targets"): self._drop_targets = []
        for attr in ("_file_row_frames", "_validate_lbls", "_sheet_hints", "_sheet_cbs"):
            if not hasattr(self, attr):
                setattr(self, attr, [])

        g = f; r = 0

        _l = ttk.Label(g, text=label("excel_path"), font=("Helvetica Neue", 13, "bold"))
        _l.grid(row=r, column=0, sticky="w", padx=PAD, pady=(12, 2)); tip(_l, "excel_path"); r += 1

        # Single-file row
        file_row_frame = ttk.Frame(g)
        file_row_frame.grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=2)
        _e_path = PEntry(file_row_frame, textvariable=self._vars["excel_path"],
                  font=("Menlo", 12)); _e_path.pack(side="left", fill="x", expand=True)
        add_placeholder(_e_path, self._vars["excel_path"], "Browse or drop an Excel file…")
        PButton(file_row_frame, text="Browse", style="ghost", lock_exempt=True,
                   command=self._browse_file).pack(side="left", padx=(8, 0))
        self._recent_btn = PButton(file_row_frame, text="Recent", style="ghost", lock_exempt=True,
                   command=self._show_recent_files)
        self._recent_btn.pack(side="left", padx=(4, 0))
        self._file_row_frames.append(file_row_frame)
        self._file_row_frame = file_row_frame
        self._drop_targets.append(_e_path)
        r += 1

        # Format hint + template button
        _FORMAT_HINTS = {
            "bar":         "Row 1: group names  |  Rows 2+: numeric values",
            "line":        "Row 1: col 1 = X label, cols 2+ = series names  |  Rows 2+: col 1 = X value, cols 2+ = Y values",
            "grouped_bar": "Row 1: category names  |  Row 2: subgroup names  |  Rows 3+: numeric values",
            "box":         "Row 1: group names  |  Rows 2+: numeric values",
            "scatter":     "Row 1: col 1 = X label, cols 2+ = series names  |  Rows 2+: col 1 = X value, cols 2+ = Y values",
            "violin":      "Row 1: group names  |  Rows 2+: numeric values",
            "kaplan_meier":"Row 1: group names (each spans 2 cols)  |  Row 2: 'Time' / 'Event'  |  Rows 3+: time value, event (1=occurred, 0=censored)",
            "heatmap":     "Row 1: blank/label in A1, then column labels  |  Rows 2+: row label in col A, then numeric values",
            "two_way_anova": "Row 1: column headers (e.g. Factor_A, Factor_B, Value)  |  Rows 2+: one observation per row (long format)",
            "before_after":  "Row 1: condition names (e.g. Before, After)  |  Rows 2+: one subject per row (matched by row order)",
            "histogram":     "Row 1: series names  |  Rows 2+: raw numeric values (columns may differ in length)",
            "subcolumn_scatter": "Row 1: group names  |  Rows 2+: numeric values",
            "curve_fit":   "Row 1: col 1 = X label, cols 2+ = series names  |  Rows 2+: X value, Y replicates",
            "column_stats":"Row 1: group names  |  Rows 2+: numeric values",
            "contingency": "Row 1: col A blank, cols B+ = outcome labels  |  Rows 2+: col A = group name, then counts",
            "repeated_measures": "Row 1: condition/timepoint names  |  Rows 2+: one row per subject",
        }
        ttk.Label(g, text=_FORMAT_HINTS.get(mode, ""),
                  foreground="#888", font=("Helvetica Neue", 10),
                  wraplength=400, justify="left"
                  ).grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=(2, 2)); r += 1
        PButton(g, text="Template", style="ghost", lock_exempt=True,
                   command=lambda m=mode: self._download_template(m)
                   ).grid(row=r, column=0, sticky="w", padx=PAD, pady=(0, 4)); r += 1

        vr = ttk.Frame(g); vr.grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=(4, 2)); r += 1
        validate_lbl = ttk.Label(vr, text="", foreground="#aaaaaa",
                                 font=("Helvetica Neue", 11))
        validate_lbl.pack(side="left")
        self._validate_lbls.append(validate_lbl)
        self._validate_lbl = validate_lbl

        section_sep(g, r, "Sheet"); r += 1

        _l, r = _lbl(g, r, "sheet")
        sheet_hint = ttk.Label(g, text="Browse for an Excel file to populate sheets",
                               foreground="gray")
        sheet_hint.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD); r += 1
        self._sheet_hints.append(sheet_hint)
        self._sheet_hint = sheet_hint

        sheet_cb = PCombobox(g, textvariable=self._vars["sheet"],
                                state="disabled", width=36, font=("Menlo", 12))
        sheet_cb.grid(row=r, column=0, columnspan=2, sticky="w", padx=PAD, pady=4); r += 1
        self._sheet_cbs.append(sheet_cb)
        self._sheet_cb = sheet_cb
        # When user picks a different sheet, revalidate without resetting sheet selection
        sheet_cb.bind("<<ComboboxSelected>>", lambda e: self._load_sheets(
            self._vars["excel_path"].get().strip(), reset_sheet=False)
            if self._vars["excel_path"].get().strip() else None)

        section_sep(g, r, "Error Bars"); r += 1

        _l, r = _lbl(g, r, "error")
        PCombobox(g, textvariable=self._vars["error"],
                     values=["SEM (Standard Error)", "SD (Standard Deviation)", "95% CI"],
                     state="readonly", width=26, font=("Helvetica Neue", 12)
                     ).grid(row=r, column=0, sticky="w", padx=PAD, pady=4); r += 1

        fr = ttk.Frame(g); fr.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=(6, 2)); r += 1
        PCheckbox(fr, variable=self._vars["show_points"],
                        text=f" {label('show_points')}").pack(side="left")

        r = self._checkbox(g, r, "show_n_labels", " Show n= sample size on x-axis")
        if mode in ("bar", "grouped_bar"):
            r = self._checkbox(g, r, "show_value_labels", " Show value label on each bar")
        if mode == "stacked_bar":
            r = self._checkbox(g, r, "stacked_value_labels", " Show value label in each segment")

        r = self._checkbox(g, r, "error_below_bar", " Extend error bar below bar (to zero)")
        section_sep(g, r, "Data Points"); r += 1

        _l, r = _lbl(g, r, "jitter_amount")
        jrow = ttk.Frame(g); jrow.grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=4); r += 1
        jlbl = ttk.Label(jrow, text="0.00", width=5, anchor="e"); jlbl.pack(side="right")
        def _ju(val): v=round(float(val),2); self._vars["jitter_amount"].set(str(v)); jlbl.config(text=f"{v:.2f}")
        ttk.Scale(jrow, from_=0.0, to=0.5, orient="horizontal",
                  variable=self._vars["jitter_amount"], command=_ju
                  ).pack(side="left", fill="x", expand=True, padx=(0, 8))
        section_sep(g, r, "Color"); r += 1

        _l, r = _lbl(g, r, "color")
        _color_cb = PCombobox(g, textvariable=self._vars["color"],
                     values=[
                         # ── Prism default ──────────────────────────────────
                         "Default",
                         # ── Curated presets ────────────────────────────────
                         "Pastel",          # soft pastels, great for posters
                         "Vivid",           # saturated, high-contrast
                         "CB-safe",         # colorblind-safe (Wong 2011)
                         "Grayscale",       # publication B&W
                         "Blues",           # sequential blue ramp
                         "Reds",            # sequential red ramp
                         "Greens",          # sequential green ramp
                         "RdBu",            # diverging red-blue
                         "mako",            # perceptually uniform blue-green (restored)
                         "rocket",          # dark-to-light red sequential
                         "flare",           # warm sequential
                         "crest",           # cool sequential
                         # ── Seaborn qualitative ────────────────────────────
                         "deep", "muted", "pastel", "bright", "dark",
                         "colorblind", "tab10",
                         # ── Matplotlib named sets ──────────────────────────
                         "Set1", "Set2", "Set3", "husl", "hls",
                         # ── Single colors ─────────────────────────────────
                         "steelblue", "tomato", "mediumseagreen", "mediumpurple",
                         "coral", "goldenrod", "slategray", "darkorange",
                     ],
                     state="normal", width=26, font=("Menlo", 12))
        _color_cb.grid(row=r, column=0, columnspan=2, sticky="w", padx=PAD, pady=4)
        add_placeholder(_color_cb, self._vars["color"], "Default")

        # Color swatch  -  5 squares that update live as palette name changes
        _swatch_frame = ttk.Frame(g)
        _swatch_frame.grid(row=r, column=2, sticky="w", padx=(4, PAD), pady=4)
        _swatch_canvases = []
        for _si in range(5):
            _sc = tk.Canvas(_swatch_frame, width=16, height=16,
                            highlightthickness=1, highlightbackground="#cccccc")
            _sc.pack(side="left", padx=2)
            _swatch_canvases.append(_sc)

        def _update_swatches(*_):
            import sys
            pal_name = self._vars["color"].get().strip()
            # Curated presets defined inline (mirrors _assign_colors in plotter_functions)
            _PRESET_SWATCHES = {
                "Pastel":    ["#AEC6CF", "#FFD1DC", "#B5EAD7", "#FFDAC1", "#C9C9FF"],
                "Vivid":     ["#E8453C", "#2274A5", "#32936F", "#F18F01", "#A846A0"],
                "CB-safe":   ["#E69F00", "#56B4E9", "#009E73", "#F0E442", "#0072B2"],
                "Grayscale": ["#222222", "#555555", "#888888", "#aaaaaa", "#cccccc"],
                "Blues":     ["#084594", "#2171b5", "#4292c6", "#6baed6", "#9ecae1"],
                "Reds":      ["#99000d", "#cb181d", "#ef3b2c", "#fb6a4a", "#fc9272"],
                "Greens":    ["#005a32", "#238b45", "#41ae76", "#74c476", "#a1d99b"],
                "RdBu":      ["#d73027", "#f46d43", "#e0f3f8", "#74add1", "#4575b4"],
            }
            if not pal_name or pal_name.lower() in ("none", "default (prism)"):
                pf_mod = sys.modules.get("plotter_functions")
                hex_cols = (pf_mod.PRISM_PALETTE[:5] if pf_mod
                            else ["#E8453C", "#2274A5", "#32936F", "#F18F01", "#A846A0"])
            elif pal_name in _PRESET_SWATCHES:
                hex_cols = _PRESET_SWATCHES[pal_name]
            else:
                hex_cols = ["#cccccc"] * 5
            for _sci, _sc in enumerate(_swatch_canvases):
                _sc.delete("all")
                _sc.create_rectangle(1, 1, 15, 15,
                                     fill=hex_cols[_sci % len(hex_cols)],
                                     outline="")

        self._vars["color"].trace_add("write", _update_swatches)
        _color_cb.bind("<<ComboboxSelected>>", _update_swatches)
        self.after(100, _update_swatches)
        r += 1

        # ── Style Presets ──────────────────────────────────────────────────────
        section_sep(g, r, "Style Presets"); r += 1
        try:
            from refraction.core.presets import BUILT_IN_PRESETS
            _preset_names = ["— none —"] + list(BUILT_IN_PRESETS.keys())
        except Exception:
            _log.debug("App._tab_data: plotter_presets import failed", exc_info=True)
            _preset_names = ["— none —"]
            BUILT_IN_PRESETS = {}
        if "preset" not in self._vars:
            self._vars["preset"] = tk.StringVar(value="— none —")
        _preset_row = ttk.Frame(g)
        _preset_row.grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=4)
        _preset_cb = PCombobox(_preset_row, textvariable=self._vars["preset"],
                               values=_preset_names, state="readonly", width=26,
                               font=("Menlo", 12))
        _preset_cb.pack(side="left")

        def _apply_preset(*_):
            name = self._vars["preset"].get()
            preset = BUILT_IN_PRESETS.get(name, {})
            for k, v in preset.items():
                if k in self._vars:
                    try:
                        self._vars[k].set(v)
                    except Exception:
                        pass

        PButton(_preset_row, text="Apply", style="ghost",
                command=_apply_preset).pack(side="left", padx=(6, 0))
        _preset_cb.bind("<<ComboboxSelected>>", _apply_preset)
        r += 1

        section_sep(g, r, "Labels"); r += 1

        # Label fields  -  deliberately NO placeholder text.
        # Empty = no label on the chart, which is the correct default behavior.
        for key in ("title", "xlabel", "ytitle"):
            _l = ttk.Label(g, text=label(key), font=("Helvetica Neue", 13, "bold"))
            _l.grid(row=r, column=0, sticky="w", padx=PAD, pady=(4, 2)); tip(_l, key); r += 1
            # Row frame holds entry + Clear × button side-by-side
            _row_f = ttk.Frame(g)
            _row_f.grid(row=r, column=0, columnspan=2, sticky="ew", padx=PAD, pady=4)
            _row_f.columnconfigure(0, weight=1)
            _e = PEntry(_row_f, textvariable=self._vars[key], width=34, font=("Menlo", 12))
            _e.grid(row=0, column=0, sticky="ew")
            # Clear button — clears the StringVar with one click
            _var_ref = self._vars[key]
            def _make_clear(v):
                return lambda e=None: v.set("")
            _clr = tk.Label(_row_f, text="×", bg=_DS.BG, fg="#aaaaaa",
                             font=("Helvetica Neue", 14), cursor="hand2",
                             padx=4, pady=0)
            _clr.grid(row=0, column=1, sticky="e", padx=(4, 0))
            _clr.bind("<Button-1>", _make_clear(_var_ref))
            _clr.bind("<Enter>",  lambda e, w=_clr: w.config(fg="#cc0000"))
            _clr.bind("<Leave>",  lambda e, w=_clr: w.config(fg="#aaaaaa"))
            r += 1

        g.columnconfigure(0, weight=1)
        ttk.Frame(g).grid(row=r, pady=8)

    def _tab_axes(self, f, mode="bar"):
        self._init_vars({
            "yscale": "Linear", "ylim_lo": "", "ylim_hi": "",
            "figw": "", "figh": "", "font_size": "12", "ref_line_y": "0",
            "ref_line_enabled": False, "ylim_data_min": False, "ylim_none": True,
            "ylim_mode": 0, "xlim_mode": 0, "xlim_lo": "", "xlim_hi": "",
            "gridlines": False, "open_points": False,
            "horizontal": False, "show_median": False,
            "bar_alpha": "0.85", "xscale": "Linear", "ref_line_label": "",
        })
        # Resolve capability flags from the registry (single source of truth).
        # This replaces the hardcoded _POINTS_MODES / _CAP_MODES / etc. sets.
        _spec = next((s for s in _REGISTRY_SPECS if s.tab_mode == mode), None)
        _has_points    = _spec.has_points     if _spec else (mode not in {"heatmap", "kaplan_meier", "histogram", "column_stats", "contingency", "two_way_anova", "chi_square_gof"})
        _has_error_bars= _spec.has_error_bars if _spec else (mode in {"bar", "line", "grouped_bar", "box", "before_after", "subcolumn_scatter", "repeated_measures"})
        _has_legend    = _spec.has_legend     if _spec else (mode in {"line", "grouped_bar", "scatter", "kaplan_meier", "histogram"})
        _x_continuous  = _spec.x_continuous   if _spec else (mode in {"line", "scatter", "kaplan_meier", "histogram"})
        if mode in ("bar", "grouped_bar", "box"):
            self._init_vars({"bar_width": "0.6"})
        if mode in ("line", "scatter"):
            self._init_vars({"line_width": "1.5", "marker_style": "Different Markers",
                             "marker_size": "7"})
        if mode == "box":
            self._init_vars({"notch_box": False})

        g = f; r = 0

        _l = ttk.Label(g, text=label("yscale"), font=("Helvetica Neue", 13, "bold"))
        _l.grid(row=r, column=0, sticky="w", padx=PAD, pady=(12, 2)); tip(_l, "yscale"); r += 1
        PCombobox(g, textvariable=self._vars["yscale"], values=["Linear", "Log"],
                     state="readonly", width=14, font=("Helvetica Neue", 12)
                     ).grid(row=r, column=0, sticky="w", padx=PAD, pady=4); r += 1
        ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=8); r += 1
        section_sep(g, r, "Y Limits"); r += 1

        _l, r = _lbl(g, r, "ylim")
        # ── Y limit radio group ───────────────────────────────────────────────
        _ylim_rg = PRadioGroup(g, variable=self._vars["ylim_mode"],
                               options=["Auto", "Start at data min", "Manual range"],
                               command=self._tog_ylim)
        _ylim_rg.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=4); r += 1

        fr2 = ttk.Frame(g); fr2.grid(row=r, column=0, columnspan=3, sticky="w",
                                     padx=(PAD + 4, PAD), pady=(0, 6)); r += 1
        ttk.Label(fr2, text="from").pack(side="left")
        self._yl_lo = PEntry(fr2, textvariable=self._vars["ylim_lo"],
                             width=10, font=("Menlo", 12))
        self._yl_lo.pack(side="left", padx=6)
        add_placeholder(self._yl_lo, self._vars["ylim_lo"], "e.g. 0")
        ttk.Label(fr2, text="to").pack(side="left")
        self._yl_hi = PEntry(fr2, textvariable=self._vars["ylim_hi"],
                             width=10, font=("Menlo", 12))
        self._yl_hi.pack(side="left", padx=6)
        add_placeholder(self._yl_hi, self._vars["ylim_hi"], "e.g. 100")
        self._tog_ylim()
        ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=8); r += 1

        # ── X Limits (only meaningful for continuous-x charts) ────────────────
        section_sep(g, r, "X Limits"); r += 1
        if _x_continuous:
            _xlim_rg = PRadioGroup(g, variable=self._vars["xlim_mode"],
                                   options=["Auto", "Manual range"],
                                   command=self._tog_xlim)
            _xlim_rg.grid(row=r, column=0, columnspan=3, sticky="w",
                          padx=PAD, pady=4); r += 1
            frx = ttk.Frame(g)
            frx.grid(row=r, column=0, columnspan=3, sticky="w",
                     padx=(PAD + 4, PAD), pady=(0, 6)); r += 1
            ttk.Label(frx, text="from").pack(side="left")
            self._xl_lo = PEntry(frx, textvariable=self._vars["xlim_lo"],
                                 width=10, font=("Menlo", 12))
            self._xl_lo.pack(side="left", padx=6)
            add_placeholder(self._xl_lo, self._vars["xlim_lo"], "e.g. 0")
            ttk.Label(frx, text="to").pack(side="left")
            self._xl_hi = PEntry(frx, textvariable=self._vars["xlim_hi"],
                                 width=10, font=("Menlo", 12))
            self._xl_hi.pack(side="left", padx=6)
            add_placeholder(self._xl_hi, self._vars["xlim_hi"], "e.g. 100")
            self._tog_xlim()
        else:
            ttk.Label(g, text="X axis is categorical  -  limits not applicable",
                      foreground="#aaaaaa", font=("Helvetica Neue", 11)
                      ).grid(row=r, column=0, columnspan=3, sticky="w",
                             padx=PAD, pady=4); r += 1
        ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew",
                              padx=PAD, pady=8); r += 1
        section_sep(g, r, "Figure Size"); r += 1

        _l, r = _lbl(g, r, "figsize")
        fr3 = ttk.Frame(g); fr3.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=4); r += 1
        _e_figw = PEntry(fr3, textvariable=self._vars["figw"], width=6, font=("Menlo", 12))
        _e_figw.pack(side="left")
        ttk.Label(fr3, text="  x  ").pack(side="left")
        _e_figh = PEntry(fr3, textvariable=self._vars["figh"], width=6, font=("Menlo", 12))
        _e_figh.pack(side="left")
        ttk.Label(fr3, text="  inches").pack(side="left")
        ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=8); r += 1

        if mode in ("bar", "grouped_bar", "box"):
            r = self._slider(g, r, "bar_width",  label("bar_width"),  0.1, 1.0, ".2f", 0.6)
            ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=8); r += 1

        if mode == "line":
            r = self._slider(g, r, "line_width", label("line_width"), 0.5, 5.0, ".1f", 1.5)
            ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=8); r += 1

        if mode in ("line", "scatter"):
            _l = ttk.Label(g, text=label("marker_style"), font=("Helvetica Neue", 13, "bold"))
            _l.grid(row=r, column=0, sticky="w", padx=PAD, pady=(4, 2)); tip(_l, "marker_style"); r += 1
            PCombobox(g, textvariable=self._vars["marker_style"],
                         values=["Different Markers", "Circle (o)", "Square (s)", "Triangle (^)",
                                 "Diamond (D)", "Down-Triangle (v)", "Star (*)", "Plus (P)",
                                 "Cross (X)", "Hexagon (h)"],
                         state="readonly", width=22, font=("Menlo", 12)
                         ).grid(row=r, column=0, sticky="w", padx=PAD, pady=4); r += 1
            ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=8); r += 1
            r = self._slider(g, r, "marker_size", label("marker_size"), 2, 20, "d", 7)
            ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=8); r += 1

        section_sep(g, r, "Font"); r += 1
        r = self._slider(g, r, "font_size", label("font_size"), 6, 24, "d", 12)
        ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=8); r += 1
        section_sep(g, r, "Reference Line"); r += 1

        ref_row = ttk.Frame(g)
        ref_row.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=(4, 2)); r += 1
        PCheckbox(ref_row, variable=self._vars["ref_line_enabled"],
                        text=" Show at Y =").pack(side="left")
        _e_ref = PEntry(ref_row, textvariable=self._vars["ref_line_y"],
                           width=8, font=("Menlo", 12))
        _e_ref.pack(side="left", padx=(6, 0))
        add_placeholder(_e_ref, self._vars["ref_line_y"], "e.g. 0")
        ref_lbl_row = ttk.Frame(g)
        ref_lbl_row.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=(0, 4)); r += 1
        ttk.Label(ref_lbl_row, text="Label:", font=("Helvetica Neue", 11)).pack(side="left")
        _e_ref_lbl = PEntry(ref_lbl_row, textvariable=self._vars["ref_line_label"],
                               width=20, font=("Menlo", 12))
        _e_ref_lbl.pack(side="left", padx=(6, 0))
        add_placeholder(_e_ref_lbl, self._vars["ref_line_label"], "blank = show y=…")

        # ── Vertical reference line (where meaningful: line, scatter, curve_fit) ──
        if mode in ("line", "scatter", "curve_fit"):
            self._init_vars({"ref_vline_enabled": False, "ref_vline_x": "0",
                             "ref_vline_label": ""})
            section_sep(g, r, "Vertical Reference Line"); r += 1
            ref_v_row = ttk.Frame(g)
            ref_v_row.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=(4, 2)); r += 1
            PCheckbox(ref_v_row, variable=self._vars["ref_vline_enabled"],
                      text=" Show at X =").pack(side="left")
            _e_vref = PEntry(ref_v_row, textvariable=self._vars["ref_vline_x"],
                             width=8, font=("Menlo", 12))
            _e_vref.pack(side="left", padx=(6, 0))
            add_placeholder(_e_vref, self._vars["ref_vline_x"], "e.g. 50")
            ref_vl_row = ttk.Frame(g)
            ref_vl_row.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=(0, 4)); r += 1
            ttk.Label(ref_vl_row, text="Label:", font=("Helvetica Neue", 11)).pack(side="left")
            _e_vref_lbl = PEntry(ref_vl_row, textvariable=self._vars["ref_vline_label"],
                                 width=20, font=("Menlo", 12))
            _e_vref_lbl.pack(side="left", padx=(6, 0))
            add_placeholder(_e_vref_lbl, self._vars["ref_vline_label"], "blank = show x=…")

        # ── P19: Twin Y-axis (line only) ──────────────────────────────────────
        if mode == "line":
            self._init_vars({"twin_y_series_str": ""})
            section_sep(g, r, "Secondary Y-Axis (P19)"); r += 1
            ttk.Label(g, text="Series on right Y-axis (comma-separated series names):",
                      font=("Helvetica Neue", 12)).grid(row=r, column=0, columnspan=3,
                      sticky="w", padx=PAD); r += 1
            _tw_e = PEntry(g, textvariable=self._vars["twin_y_series_str"], width=38)
            _tw_e.grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=2); r += 1
            add_placeholder(_tw_e, self._vars["twin_y_series_str"], "e.g. Temperature, Pressure")
            r += 1  # spacing

        # ── P18: Custom x-tick labels (categorical charts) ────────────────────
        if mode in ("bar", "box", "violin", "dot_plot", "subcolumn_scatter",
                    "before_after", "repeated_measures"):
            self._init_vars({"xtick_labels_str": ""})
            section_sep(g, r, "Custom Category Labels (P18)"); r += 1
            ttk.Label(g, text="Override labels (comma-separated):",
                      font=("Helvetica Neue", 12)).grid(row=r, column=0, columnspan=3,
                      sticky="w", padx=PAD); r += 1
            _xt_e = PEntry(g, textvariable=self._vars["xtick_labels_str"], width=38)
            _xt_e.grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=2); r += 1
            add_placeholder(_xt_e, self._vars["xtick_labels_str"],
                            "e.g. Control, Drug A, Drug B")
            r += 1

        display_opts = [
            ("open_points", " Open/hollow data points"),
        ]
        if mode == "bar":
            display_opts += [("horizontal", " Horizontal bars"), ("show_median", " Show median line on bar")]
        if mode == "grouped_bar":
            display_opts += [("horizontal", " Horizontal bars")]
        if mode == "box":
            display_opts.append(("notch_box", " Notched box (95% CI on median)"))

        for key, text in display_opts:
            r = self._checkbox(g, r, key, text)

        # Grid style  -  radio group replaces old gridlines checkbox
        self._init_vars({"grid_style": "None"})
        ttk.Label(g, text="Grid Lines", font=("Helvetica Neue", 13, "bold")
                  ).grid(row=r, column=0, sticky="w", padx=PAD, pady=(8, 2)); r += 1
        PRadioGroup(g, variable=self._vars["grid_style"],
                    options=["None", "Horizontal", "Full (H + V)"],
                    ).grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=4); r += 1

        if mode == "bar":
            ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=6); r += 1
            r = self._slider(g, r, "bar_alpha", "Bar Transparency", 0.1, 1.0, ".2f", 0.85)
            ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=6); r += 1
            ttk.Label(g, text="X Scale", font=("Helvetica Neue", 13, "bold")
                      ).grid(row=r, column=0, sticky="w", padx=PAD, pady=(4, 2)); r += 1
            _xscale_row = ttk.Frame(g)
            _xscale_row.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=4); r += 1
            PCombobox(_xscale_row, textvariable=self._vars["xscale"], values=["Linear", "Log"],
                         state="disabled", width=14, font=("Helvetica Neue", 12)
                         ).pack(side="left")
            ttk.Label(_xscale_row, text="  (categorical axis  -  not applicable)",
                      foreground="#aaaaaa", font=("Helvetica Neue", 11)
                      ).pack(side="left")

        # ── Priority-1: Axis Frame / Tick / Point / Cap styling ──────────────
        ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=8); r += 1
        r = self._section_label(g, r, "Axis & Tick Style")

        # Axis Frame
        self._init_vars({"axis_style": "Open (default)"})
        ttk.Label(g, text="Axis Frame", font=("Helvetica Neue", 13, "bold")
                  ).grid(row=r, column=0, sticky="w", padx=PAD, pady=(4, 2)); r += 1
        PCombobox(g, textvariable=self._vars["axis_style"],
                     values=["Open (default)", "Closed box", "Floating", "None"],
                     state="readonly", width=22, font=("Helvetica Neue", 12)
                     ).grid(row=r, column=0, sticky="w", padx=PAD, pady=4); r += 1

        # Background color
        self._init_vars({"fig_bg": "White"})
        ttk.Label(g, text="Background", font=("Helvetica Neue", 13, "bold")
                  ).grid(row=r, column=0, sticky="w", padx=PAD, pady=(8, 2)); r += 1
        PCombobox(g, textvariable=self._vars["fig_bg"],
                     values=["White", "Light gray", "Transparent", "Black"],
                     state="readonly", width=22, font=("Helvetica Neue", 12)
                     ).grid(row=r, column=0, sticky="w", padx=PAD, pady=4); r += 1

        # Tick Direction (radio group)
        self._init_vars({"tick_dir": "Outward (default)"})
        ttk.Label(g, text="Tick Direction", font=("Helvetica Neue", 13, "bold")
                  ).grid(row=r, column=0, sticky="w", padx=PAD, pady=(8, 2)); r += 1
        PRadioGroup(g, variable=self._vars["tick_dir"],
                    options=["Outward (default)", "Inward", "Both", "None"],
                    ).grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=4); r += 1

        # Minor ticks
        self._init_vars({"minor_ticks": False})
        r = self._checkbox(g, r, "minor_ticks", " Show minor tick marks")

        # Spine / axis line width
        self._init_vars({"spine_width": "0.8"})
        r = self._slider(g, r, "spine_width", "Axis Line Width", 0.3, 3.0, ".1f", 0.8)

        ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=8); r += 1
        r = self._section_label(g, r, "Data Points & Error Bars")

        # Point size slider (for chart types that show points)
        if _has_points:
            self._init_vars({"point_size": "6", "point_alpha": "0.80"})
            r = self._slider(g, r, "point_size",  "Point Size (pts)",        2.0, 20.0, ".0f", 6.0)
            ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=4); r += 1
            r = self._slider(g, r, "point_alpha", "Point Transparency",      0.1,  1.0, ".2f", 0.80)
            ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=4); r += 1

        # Cap width slider (for chart types with error bars)
        if _has_error_bars:
            self._init_vars({"cap_size": "4"})
            r = self._slider(g, r, "cap_size", "Error Bar Cap Width", 0.0, 12.0, ".1f", 4.0)
            ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=4); r += 1

        # Legend position (for chart types that have a legend)
        if _has_legend:
            self._init_vars({"legend_pos": "Upper right"})
            ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=8); r += 1
            r = self._section_label(g, r, "Legend")
            ttk.Label(g, text="Legend Position", font=("Helvetica Neue", 13, "bold")
                      ).grid(row=r, column=0, sticky="w", padx=PAD, pady=(4, 2)); r += 1
            PCombobox(g, textvariable=self._vars["legend_pos"],
                         values=["Auto (best fit)", "Upper right", "Upper left",
                                 "Lower right", "Lower left", "Outside right", "None (hidden)"],
                         state="readonly", width=22, font=("Helvetica Neue", 12)
                         ).grid(row=r, column=0, sticky="w", padx=PAD, pady=4); r += 1

        # ── Tick Interval controls ────────────────────────────────────────────
        ttk.Separator(g).grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=8); r += 1
        r = self._section_label(g, r, "Tick Intervals")
        self._init_vars({"ytick_interval": "", "xtick_interval": ""})

        _tick_hint = ttk.Label(g, text="Leave blank for automatic ticks.",
                               foreground="#888888", font=("Helvetica Neue", 11))
        _tick_hint.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=(0, 6)); r += 1

        fr_ytick = ttk.Frame(g)
        fr_ytick.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=(0, 4)); r += 1
        ttk.Label(fr_ytick, text="Y-axis interval:", font=("Helvetica Neue", 12)
                  ).pack(side="left")
        _e_ytick = PEntry(fr_ytick, textvariable=self._vars["ytick_interval"],
                          width=10, font=("Menlo", 12))
        _e_ytick.pack(side="left", padx=(8, 0))
        add_placeholder(_e_ytick, self._vars["ytick_interval"], "e.g. 10")

        if _x_continuous:
            fr_xtick = ttk.Frame(g)
            fr_xtick.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=(0, 4)); r += 1
            ttk.Label(fr_xtick, text="X-axis interval:", font=("Helvetica Neue", 12)
                      ).pack(side="left")
            _e_xtick = PEntry(fr_xtick, textvariable=self._vars["xtick_interval"],
                              width=10, font=("Menlo", 12))
            _e_xtick.pack(side="left", padx=(8, 0))
            add_placeholder(_e_xtick, self._vars["xtick_interval"], "e.g. 5")

        g.columnconfigure(0, weight=1)
        ttk.Frame(g).grid(row=r + 1, pady=8)

    def _slider(self, g, r: int, var_key: str, label_text: str,
                lo: float, hi: float, fmt: str, default) -> int:
        """Grid a label + Scale + live value label row; return next row index.

        fmt   -  Python format spec for the value label, e.g. '.2f', '.1f', 'd'
        """
        ttk.Label(g, text=label_text, font=("Helvetica Neue", 13, "bold")
                  ).grid(row=r, column=0, sticky="w", padx=PAD, pady=(4, 2))
        r += 1
        row_fr = ttk.Frame(g)
        row_fr.grid(row=r, column=0, columnspan=3, sticky="ew", padx=PAD, pady=4)
        r += 1
        val_lbl = ttk.Label(row_fr, text=format(default, fmt), width=5, anchor="e")
        val_lbl.pack(side="right")

        def _on_change(raw, _vk=var_key, _fmt=fmt, _lbl=val_lbl):
            v = round(float(raw), 6)
            self._vars[_vk].set(str(v))
            _lbl.config(text=format(v, _fmt))

        ttk.Scale(row_fr, from_=lo, to=hi, orient="horizontal",
                  variable=self._vars[var_key], command=_on_change
                  ).pack(side="left", fill="x", expand=True, padx=(0, 8))
        return r

    def _init_vars(self, defaults: dict):
        """Ensure each key exists in self._vars; create it from its default if not.
        Booleans  ->  BooleanVar, ints  ->  IntVar, everything else  ->  StringVar."""
        for k, v in defaults.items():
            if k not in self._vars:
                if isinstance(v, bool):
                    self._vars[k] = tk.BooleanVar(value=v)
                elif isinstance(v, int):
                    self._vars[k] = tk.IntVar(value=v)
                else:
                    self._vars[k] = tk.StringVar(value=str(v))

    def _sep(self, g, r: int) -> int:
        """Grid a horizontal separator and return the next row index."""
        ttk.Separator(g).grid(row=r, column=0, columnspan=3,
                              sticky="ew", padx=PAD, pady=8)
        return r + 1

    def _section_label(self, g, r: int, text: str) -> int:
        """Grid a styled section-band header and return the next row index."""
        outer = tk.Frame(g, bg="#edf2f8")
        outer.grid(row=r, column=0, columnspan=3, sticky="ew", padx=0, pady=(14, 2))
        tk.Frame(outer, bg="#2274A5", width=3).pack(side="left", fill="y", padx=(0, 9))
        tk.Label(outer, text=text.upper(), bg="#edf2f8", fg="#2274A5",
                 font=("Helvetica Neue", 9, "bold"), anchor="w").pack(side="left", pady=5)
        return r + 1

    def _checkbox(self, g, r: int, key: str, text: str) -> int:
        """Grid a single PCheckbox row and return the next row index.
        Auto-creates a BooleanVar(False) if the key is not yet in self._vars,
        so tab-build order cannot cause a KeyError.
        """
        if key not in self._vars:
            self._vars[key] = tk.BooleanVar(value=False)
        cb = PCheckbox(g, variable=self._vars[key], text=text)
        cb.grid(row=r, column=0, columnspan=3, sticky="w", padx=PAD, pady=(0, 4))
        return r + 1

    def _hint(self, g, r: int, text: str) -> int:
        """Grid a small gray explanatory label and return the next row index."""
        ttk.Label(g, text=text,
                  foreground="gray", wraplength=380, justify="left",
                  font=("Helvetica Neue", 11)
                  ).grid(row=r, column=0, columnspan=3, sticky="w",
                         padx=PAD, pady=(0, 8))
        return r + 1

    def _tab_footer(self, g, r: int):
        """Finish a tab frame: configure column weight and add bottom spacer."""
        g.columnconfigure(0, weight=1)
        ttk.Frame(g).grid(row=r + 1, pady=10)

    # ── Toggle helpers ─────────────────────────────────────────────────────────

    def _tog_ylim(self):
        """Enable the manual from/to fields only when mode==2 (Manual range)."""
        manual = (self._get_var("ylim_mode", 0) == 2)
        s = ["!disabled"] if manual else ["disabled"]
        if hasattr(self, "_yl_lo"):
            self._yl_lo.state(s); self._yl_hi.state(s)

    def _tog_xlim(self):
        """Enable the x manual from/to fields only when xlim_mode==1."""
        manual = (self._get_var("xlim_mode", 0) == 1)
        s = ["!disabled"] if manual else ["disabled"]
        if hasattr(self, "_xl_lo"):
            self._xl_lo.state(s); self._xl_hi.state(s)

    def _tog_posthoc(self):
        """Show post-hoc selector for Parametric only.
        Non-parametric always uses Dunn's test  -  no choice needed, just show a hint.
        Paired and Permutation don't use post-hoc selectors."""
        test = self._vars["stats_test"].get()
        show_parametric = (test == "Parametric")
        show_nonparam   = (test == "Non-parametric")
        r = self._posthoc_r
        if show_parametric and not self._posthoc_visible:
            self._posthoc_label.grid(row=r,   column=0, sticky="w",  padx=PAD, pady=(4, 2))
            self._posthoc_cb   .grid(row=r+1, column=0, sticky="w",  padx=PAD, pady=4)
            self._posthoc_sep  .grid(row=r+2, column=0, columnspan=3, sticky="ew", padx=PAD, pady=6)
            if hasattr(self, "_posthoc_dunn_hint"):
                self._posthoc_dunn_hint.grid_remove()
            self._posthoc_visible = True
        elif not show_parametric and self._posthoc_visible:
            self._posthoc_label.grid_remove()
            self._posthoc_cb   .grid_remove()
            self._posthoc_sep  .grid_remove()
            self._posthoc_visible = False

        # Non-parametric: always Dunn's test  -  show informational hint only
        if show_nonparam:
            if hasattr(self, "_posthoc_dunn_hint"):
                self._posthoc_dunn_hint.grid(row=r, column=0, columnspan=3,
                                             sticky="w", padx=PAD, pady=(4, 6))
                self._posthoc_sep.grid(row=r+2, column=0, columnspan=3,
                                       sticky="ew", padx=PAD, pady=6)
        else:
            if hasattr(self, "_posthoc_dunn_hint"):
                self._posthoc_dunn_hint.grid_remove()
            # Only hide sep when not parametric either
            if not show_parametric:
                self._posthoc_sep.grid_remove()

    def _tog_control(self):
        """Synchronize the Comparison Mode radio buttons and Control Group dropdown
        so every visible state is consistent and logically valid.

        Consistency rules:

        HIDDEN entirely (comparison_mode radio + control group section invisible):
          - One-sample test  ->  each group vs μ₀, no group-to-group comparison
          - Paired test       ->  direct pair comparison, no control concept
          - Only 1 or 2 groups loaded  ->  no choice needed (single pair or nothing)

        FORCED "vs control" (radio locked, required):
          - Dunnett (vs control) posthoc selected  ->  inherently a control test

        FORCED "all pairwise" (radio + dropdown both disabled):
          - Non-parametric with k>2  ->  Dunn's test is always all-pairwise
          - Permutation test          ->  all-pairwise only implementation

        NORMAL (radio choosable, dropdown enabled when mode=vs-control):
          - Parametric, k>=3, non-Dunnett posthoc

        DROPDOWN CONTENT:
          - Mode=all-pairwise  ->  dropdown disabled, value cleared to ""
          - Mode=vs-control, groups loaded  ->  shows ONLY real group names
            (no "(none)" sentinel — that would be a contradictory state)
          - No file loaded  ->  disabled with "Load a file…" hint
        """
        if not hasattr(self, "_control_cbs"):
            return

        # ── Gather current state ──────────────────────────────────────────────
        test       = self._get_var("stats_test", "Parametric")
        posthoc    = self._get_var("posthoc", "Tukey HSD")
        is_dunnett = (posthoc == "Dunnett (vs control)")
        cmode      = self._get_var("comparison_mode", 0)  # 0=all-pairs, 1=vs-ctrl

        # Extract actual group names from the first control combobox
        # The list was populated by _populate_control_dropdown during validation.
        raw_values = []
        try:
            raw_values = list(self._control_cbs[0].cget("values"))
        except Exception:
            pass
        # Filter out any sentinel "(none…)" entries to get real group names
        group_names = [v for v in raw_values if v and not v.startswith("(none")]
        n_groups    = len(group_names)
        has_groups  = n_groups > 0

        # ── Decide visibility / enabled state ─────────────────────────────────
        # Tests where comparison mode + control make no sense at all
        hide_section = test in ("One-sample", "Paired")

        # 2-group case: nothing to choose — always a single pair
        if has_groups and n_groups <= 2:
            hide_section = True

        # Tests that are all-pairwise-only (no control comparison implemented)
        force_pairwise = (test in ("Permutation",) or
                          (test == "Non-parametric" and n_groups > 2))

        # Dunnett forces vs-control mode
        if is_dunnett and not hide_section:
            force_pairwise = False
            # Silently set comparison_mode to 1 (vs control) if it isn't already
            if self._get_var("comparison_mode", 0) != 1:
                try:
                    self._vars["comparison_mode"].set(1)
                    cmode = 1
                except Exception:
                    pass

        # ── Apply visibility to the comparison mode radio group ───────────────
        if hasattr(self, "_cmode_rg_widget"):
            try:
                if hide_section or force_pairwise:
                    self._cmode_rg_widget.config(state="disabled")
                else:
                    self._cmode_rg_widget.config(state="normal")
            except Exception:
                pass

        # ── Apply state to control dropdown ───────────────────────────────────
        # Decide whether the dropdown should be enabled
        want_control_enabled = (
            not hide_section
            and not force_pairwise
            and has_groups
            and cmode == 1       # "vs control" mode selected
        )

        for cb in self._control_cbs:
            try:
                if want_control_enabled:
                    # Populate with ONLY real group names (no sentinel)
                    cb.config(values=group_names, state="readonly")
                    # If current value is a sentinel or empty, auto-select first group
                    cur = self._vars["control"].get()
                    if not cur or cur.startswith("(none") or cur not in group_names:
                        self._vars["control"].set(group_names[0])
                else:
                    # Clear to empty and disable
                    cb.config(values=group_names if group_names else [""], state="disabled")
                    if not want_control_enabled:
                        self._vars["control"].set("")
            except Exception:
                pass

        # ── Update hint label ─────────────────────────────────────────────────
        for lbl in getattr(self, "_control_hint_lbls", []):
            try:
                if hide_section:
                    lbl.config(
                        text="Not applicable for this test type.",
                        foreground="#aaaaaa")
                elif force_pairwise:
                    lbl.config(
                        text="This test always compares all groups pairwise.",
                        foreground="#aaaaaa")
                elif not has_groups:
                    lbl.config(text="Load a file to populate group names.",
                               foreground="#888888")
                elif cmode == 0:  # all pairwise
                    lbl.config(
                        text=f"{n_groups} groups — all pairwise brackets.",
                        foreground="#666666")
                elif want_control_enabled:
                    cur = self._vars["control"].get()
                    if is_dunnett:
                        lbl.config(
                            text=f"Dunnett test: comparing all groups vs. '{cur}'  "
                                 f"(control required).",
                            foreground="#2274A5")
                    else:
                        lbl.config(
                            text=f"Showing brackets vs. '{cur}' only.  "
                                 f"Other pairs hidden.",
                            foreground="#2274A5")
            except Exception:
                pass

        # Re-fire when control selection changes so hint stays live
        for cb in self._control_cbs:
            try:
                cb.bind("<<ComboboxSelected>>",
                        lambda e: self._tog_control(), add="+")
            except Exception:
                pass

    def _tog_perm(self):
        is_perm = self._vars["stats_test"].get() == "Permutation"
        r = self._perm_r_start
        if is_perm and not self._perm_visible:
            self._perm_label.grid(row=r,   column=0, sticky="w",  padx=PAD, pady=(4, 2))
            self._perm_hint .grid(row=r+1, column=0, columnspan=3, sticky="w", padx=PAD)
            self._perm_entry.grid(row=r+2, column=0, sticky="w",  padx=PAD, pady=4)
            self._perm_sep  .grid(row=r+3, column=0, columnspan=3, sticky="ew", padx=PAD, pady=6)
            self._perm_visible = True
        elif not is_perm and self._perm_visible:
            for w in (self._perm_label, self._perm_hint, self._perm_entry, self._perm_sep):
                w.grid_remove()
            self._perm_visible = False

    # ── File loading ──────────────────────────────────────────────────────────

    def _show_about(self):
        """About dialog."""
        from tkinter import messagebox
        messagebox.showinfo(
            "About Refraction",
            "Refraction\n\n"
            "A publication-quality scientific plotting application.\n\n"
            "Designed and implemented by Claude (Anthropic).\n"
            "Commissioned by Ashwin Pasupathy.\n\n"
            "MIT License\n\n"
            "29 chart types · Statistical tests · Publication-ready export"
        )

