"""File I/O, export, and clipboard mixin."""

import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox


class FileIOMixin:
    """Extracted from plotter_barplot_app.py."""

    def _browse_file(self):
        path = filedialog.askopenfilename(
            title="Select Excel file",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")])
        if path:
            self._vars["excel_path"].set(path)
            self._load_sheets(path)

    def _download_template(self, mode):
        """Write an example Excel template for the given plot type to the Desktop."""
        import os
        from datetime import datetime
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            messagebox.showerror("Missing package", "openpyxl is required to create templates.")
            return

        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
        fname   = f"refraction_template_{mode}_{ts}.xlsx"
        path    = os.path.join(desktop, fname)

        wb = openpyxl.Workbook()
        ws = wb.active

        # Styles
        hdr_font    = Font(bold=True, color="FFFFFF")
        hdr_fill    = PatternFill("solid", fgColor="4472C4")
        sub_fill    = PatternFill("solid", fgColor="8EAADB")
        hint_font   = Font(italic=True, color="888888", size=9)
        center      = Alignment(horizontal="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"))

        def hdr(ws, row, col, val):
            c = ws.cell(row=row, column=col, value=val)
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = center; c.border = thin_border

        def sub(ws, row, col, val):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(bold=True); c.fill = sub_fill
            c.alignment = center; c.border = thin_border

        def dat(ws, row, col, val):
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = center; c.border = thin_border

        def hint(ws, row, col, val):
            c = ws.cell(row=row, column=col, value=val)
            c.font = hint_font

        if mode in ("bar", "box"):
            ws.title = "Bar_Box Template"
            hint(ws, 1, 1, "Row 1: group names | Rows 2+: numeric replicates")
            for ci, grp in enumerate(["Control", "Treatment A", "Treatment B"], 1):
                hdr(ws, 2, ci, grp)
            data = [[2.1, 3.4, 4.2], [1.9, 3.8, 4.6], [2.3, 3.1, 3.9],
                    [2.0, 3.5, 4.0], [2.2, 3.6, 4.4]]
            for ri, row_vals in enumerate(data, 3):
                for ci, v in enumerate(row_vals, 1):
                    dat(ws, ri, ci, v)
            for ci in range(1, 4): ws.column_dimensions[chr(64+ci)].width = 14

        elif mode == "grouped_bar":
            ws.title = "Grouped Bar Template"
            hint(ws, 1, 1, "Row 1: category names | Row 2: subgroup names | Rows 3+: replicates")
            cats = ["Condition A", "Condition A", "Condition B", "Condition B"]
            subs = ["Male", "Female", "Male", "Female"]
            for ci, (cat, s) in enumerate(zip(cats, subs), 1):
                hdr(ws, 2, ci, cat); sub(ws, 3, ci, s)
            data = [[2.1, 2.8, 3.4, 4.1], [1.9, 2.6, 3.6, 3.9],
                    [2.3, 2.9, 3.1, 4.3], [2.0, 2.7, 3.5, 4.0]]
            for ri, row_vals in enumerate(data, 4):
                for ci, v in enumerate(row_vals, 1):
                    dat(ws, ri, ci, v)
            for ci in range(1, 5): ws.column_dimensions[chr(64+ci)].width = 14

        elif mode in ("line", "scatter"):
            ws.title = "Line_Scatter Template"
            hint(ws, 1, 1, "Row 1: col 1=X label, cols 2+ = series names (repeat per replicate) | Rows 2+: X value then Y replicates")
            headers = ["Time (h)", "Control", "Control", "Control", "Drug", "Drug", "Drug"]
            for ci, h in enumerate(headers, 1):
                hdr(ws, 2, ci, h)
            data = [
                [0,  1.0, 1.1, 0.9,  1.0, 1.2, 0.8],
                [6,  1.5, 1.4, 1.6,  2.1, 2.3, 1.9],
                [12, 2.0, 2.2, 1.9,  3.4, 3.6, 3.2],
                [24, 2.8, 2.7, 3.0,  5.1, 4.9, 5.3],
                [48, 3.1, 3.3, 2.9,  6.2, 6.0, 6.4],
            ]
            for ri, row_vals in enumerate(data, 3):
                for ci, v in enumerate(row_vals, 1):
                    dat(ws, ri, ci, v)
            ws.column_dimensions["A"].width = 12
            for ci in range(2, 8): ws.column_dimensions[chr(64+ci)].width = 10

        elif mode == "before_after":
            ws.title = "Before_After Template"
            hint(ws, 1, 1, "Row 1: condition names | Rows 2+: one subject per row (matched by row order)")
            for ci, grp in enumerate(["Before", "After"], 1):
                hdr(ws, 2, ci, grp)
            data = [[2.1, 3.4], [1.9, 3.8], [2.3, 2.9], [2.0, 3.5], [2.2, 4.1],
                    [1.8, 3.2], [2.5, 4.4], [2.1, 3.7]]
            for ri, row_vals in enumerate(data, 3):
                for ci, v in enumerate(row_vals, 1):
                    dat(ws, ri, ci, v)
            for ci in range(1, 3): ws.column_dimensions[chr(64+ci)].width = 12

        elif mode == "histogram":
            ws.title = "Histogram Template"
            hint(ws, 1, 1, "Row 1: series names | Rows 2+: raw numeric values (columns may have different lengths)")
            for ci, grp in enumerate(["Group A", "Group B"], 1):
                hdr(ws, 2, ci, grp)
            import random; random.seed(42)
            a_vals = [round(random.gauss(5, 1), 2) for _ in range(30)]
            b_vals = [round(random.gauss(6.5, 1.2), 2) for _ in range(30)]
            for ri, (a, b) in enumerate(zip(a_vals, b_vals), 3):
                dat(ws, ri, 1, a); dat(ws, ri, 2, b)
            for ci in range(1, 3): ws.column_dimensions[chr(64+ci)].width = 12

        elif mode in ("subcolumn_scatter", "violin"):
            ws.title = "Subcolumn_Violin Template"
            hint(ws, 1, 1, "Row 1: group names | Rows 2+: numeric replicates")
            for ci, grp in enumerate(["Control", "Low Dose", "High Dose"], 1):
                hdr(ws, 2, ci, grp)
            data = [[2.1, 3.4, 4.8], [1.9, 3.1, 5.2], [2.4, 3.7, 4.5],
                    [2.0, 2.9, 5.0], [2.3, 3.5, 4.7], [1.8, 3.2, 5.3]]
            for ri, row_vals in enumerate(data, 3):
                for ci, v in enumerate(row_vals, 1):
                    dat(ws, ri, ci, v)
            for ci in range(1, 4): ws.column_dimensions[chr(64+ci)].width = 14

        elif mode == "kaplan_meier":
            ws.title = "Survival Template"
            hint(ws, 1, 1, "Row 1: group names (each spans 2 cols) | Row 2: 'Time' / 'Event' | Rows 3+: time, event (1=occurred, 0=censored)")
            groups_km = ["Control", "Treatment"]
            for gi, grp in enumerate(groups_km):
                c1 = gi * 2 + 1
                hdr(ws, 2, c1, grp); hdr(ws, 2, c1 + 1, grp)
                sub(ws, 3, c1, "Time"); sub(ws, 3, c1 + 1, "Event")
            km_data = [
                [5,  1,  3,  1],
                [8,  0,  6,  1],
                [10, 1,  8,  0],
                [12, 1, 10,  1],
                [15, 0, 12,  1],
                [18, 1, 14,  0],
                [20, 1, 16,  1],
                [22, 0, 18,  1],
            ]
            for ri, row_vals in enumerate(km_data, 4):
                for ci, v in enumerate(row_vals, 1):
                    dat(ws, ri, ci, v)
            for ci in range(1, 5): ws.column_dimensions[chr(64+ci)].width = 12

        elif mode == "heatmap":
            ws.title = "Heatmap Template"
            hint(ws, 1, 1, "Row 1: blank in A1, then column labels | Rows 2+: row label in col A, then numeric values")
            col_labels = ["Sample 1", "Sample 2", "Sample 3", "Sample 4"]
            row_labels = ["Gene A", "Gene B", "Gene C", "Gene D", "Gene E"]
            for ci, lbl in enumerate(col_labels, 2):
                hdr(ws, 2, ci, lbl)
            import random; random.seed(7)
            for ri, rlbl in enumerate(row_labels, 3):
                sub(ws, ri, 1, rlbl)
                for ci in range(2, 6):
                    dat(ws, ri, ci, round(random.gauss(0, 2), 3))
            ws.column_dimensions["A"].width = 12
            for ci in range(2, 6): ws.column_dimensions[chr(64+ci)].width = 11

        elif mode == "two_way_anova":
            ws.title = "Two-Way ANOVA Template"
            hint(ws, 1, 1, "Row 1: column headers | Rows 2+: one observation per row (long format)")
            for ci, h in enumerate(["Factor_A", "Factor_B", "Value"], 1):
                hdr(ws, 2, ci, h)
            twa_data = [
                ["Drug",    "Male",   3.2], ["Drug",    "Male",   2.9], ["Drug",    "Male",   3.5],
                ["Drug",    "Female", 4.1], ["Drug",    "Female", 3.8], ["Drug",    "Female", 4.4],
                ["Control", "Male",   1.8], ["Control", "Male",   2.1], ["Control", "Male",   1.6],
                ["Control", "Female", 2.3], ["Control", "Female", 2.0], ["Control", "Female", 2.5],
            ]
            for ri, row_vals in enumerate(twa_data, 3):
                for ci, v in enumerate(row_vals, 1):
                    dat(ws, ri, ci, v)
            for ci in range(1, 4): ws.column_dimensions[chr(64+ci)].width = 14

        elif mode == "curve_fit":
            ws.title = "Curve Fit Template"
            hint(ws, 1, 1, "Row 1: col 1=X label, cols 2+ = series names (repeat per replicate) | Rows 2+: X value then Y replicates")
            headers = ["Concentration", "Control", "Control", "Control", "Drug", "Drug", "Drug"]
            for ci, h in enumerate(headers, 1):
                hdr(ws, 2, ci, h)
            cf_data = [
                [0.001, 0.02, 0.03, 0.01,  0.03, 0.02, 0.04],
                [0.01,  0.08, 0.07, 0.09,  0.12, 0.14, 0.11],
                [0.1,   0.31, 0.29, 0.33,  0.55, 0.52, 0.58],
                [1,     0.62, 0.60, 0.65,  0.88, 0.85, 0.91],
                [10,    0.81, 0.79, 0.83,  0.97, 0.95, 0.98],
                [100,   0.91, 0.89, 0.93,  0.99, 0.98, 1.00],
                [1000,  0.95, 0.94, 0.96,  1.00, 0.99, 1.00],
            ]
            for ri, row_vals in enumerate(cf_data, 3):
                for ci, v in enumerate(row_vals, 1):
                    dat(ws, ri, ci, v)
            ws.column_dimensions["A"].width = 15
            for ci in range(2, 8): ws.column_dimensions[chr(64+ci)].width = 10

        elif mode == "column_stats":
            ws.title = "Column Stats Template"
            hint(ws, 1, 1, "Row 1: group names | Rows 2+: numeric replicates")
            for ci, grp in enumerate(["Group A", "Group B", "Group C"], 1):
                hdr(ws, 2, ci, grp)
            import random; random.seed(99)
            cs_data = [[round(random.gauss(5+ci*1.5, 1), 2) for ci in range(3)]
                       for _ in range(12)]
            for ri, row_vals in enumerate(cs_data, 3):
                for ci, v in enumerate(row_vals, 1):
                    dat(ws, ri, ci, v)
            for ci in range(1, 4): ws.column_dimensions[chr(64+ci)].width = 12

        elif mode == "contingency":
            ws.title = "Contingency Template"
            hint(ws, 1, 1, "Row 1: col A blank, cols B+ = outcome labels | Rows 2+: col A = group name, then counts")
            for ci, lbl in enumerate(["Survived", "Died"], 2):
                hdr(ws, 2, ci, lbl)
            ct_data = [["Drug", 45, 5], ["Control", 20, 30]]
            for ri, row_vals in enumerate(ct_data, 3):
                sub(ws, ri, 1, row_vals[0])
                for ci, v in enumerate(row_vals[1:], 2):
                    dat(ws, ri, ci, v)
            ws.column_dimensions["A"].width = 12
            for ci in range(2, 4): ws.column_dimensions[chr(64+ci)].width = 12

        elif mode == "repeated_measures":
            ws.title = "Repeated Measures Template"
            hint(ws, 1, 1, "Row 1: condition/timepoint names | Rows 2+: one row per subject")
            for ci, cond in enumerate(["Baseline", "Week 2", "Week 4", "Week 8"], 1):
                hdr(ws, 2, ci, cond)
            rm_data = [
                [2.1, 2.8, 3.5, 4.2],
                [1.9, 2.5, 3.1, 3.8],
                [2.4, 3.0, 3.7, 4.5],
                [1.8, 2.4, 2.9, 3.6],
                [2.2, 2.9, 3.4, 4.0],
                [2.0, 2.6, 3.2, 3.9],
            ]
            for ri, row_vals in enumerate(rm_data, 3):
                for ci, v in enumerate(row_vals, 1):
                    dat(ws, ri, ci, v)
            for ci in range(1, 5): ws.column_dimensions[chr(64+ci)].width = 12

        wb.save(path)
        self._set_status(f"Template saved to Desktop: {fname}")
        messagebox.showinfo("Template saved",
                            f"Template saved to your Desktop:\n{fname}\n\n"
                            "Open it in Excel, fill in your data, then browse for it in the app.")

    def _load_sheets(self, path, reset_sheet=True):
        """Load sheet names for the given file.
        reset_sheet=True   ->  set sheet var to first sheet (new file load)
        reset_sheet=False  ->  keep current sheet selection (sheet change trigger)
        """
        # Invalidate the Excel parse cache whenever the file or sheet changes
        # so _do_run always reads fresh data for the new file.
        self._parse_cache_key = None
        self._parse_cache_df  = None
        # openpyxl first import can trigger a Cocoa tick that resets the dock icon.
        self.after(100, set_dock_icon)
        try:
            import openpyxl
            sheets = _cached_sheets(path)
            if not sheets: raise ValueError("No sheets found")
            for cb in self._sheet_cbs:
                cb.config(values=sheets, state="readonly")
            if reset_sheet:
                self._vars["sheet"].set(sheets[0])
            elif self._vars["sheet"].get() not in sheets:
                # Current selection no longer valid  -  fall back to first
                self._vars["sheet"].set(sheets[0])
            for lbl in self._sheet_hints:
                lbl.config(text=f"{len(sheets)} sheet(s) found")
        except Exception as e:
            for cb in self._sheet_cbs:
                cb.config(values=[], state="readonly")
            for lbl in self._sheet_hints:
                lbl.config(text=f"Could not read sheets: {e}")

        if not self._file_selected:
            self._file_selected = True
            self._unlock_form()

        # Always keep sheet comboboxes enabled regardless of validation outcome
        for cb in self._sheet_cbs:
            try: cb.config(state="readonly")
            except Exception: pass

        self._validated = False
        self._run_btn.config(state="disabled", text="Generate Plot")
        self._set_validate_text("Validating...", "gray")
        # Run validation automatically after a short delay so the UI updates first
        self.after(50, self._validate_spreadsheet)

    def _get_spec_json(self):
        """Build a Plotly spec JSON string from the last render state."""
        if self._last_kw is None or self._last_chart_type is None:
            return None
        try:
            import json
            from refraction.server.api import _build_spec
            spec_json = _build_spec(self._last_chart_type, self._last_kw)
            obj = json.loads(spec_json)
            if "error" in obj:
                return None
            return spec_json
        except Exception:
            return None

    def _copy_to_clipboard(self):
        """Copy the current plot to the macOS clipboard as a PNG image."""
        spec_json = self._get_spec_json()
        if spec_json is None:
            self._set_status("No chart to copy — generate a plot first")
            return
        try:
            import subprocess, tempfile
            from refraction.io import export as _pe
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                tmp_path = tmp.name
            _pe.export_plotly(spec_json, tmp_path, dpi=150)
            subprocess.run(
                ["osascript", "-e",
                 f'set the clipboard to (read (POSIX file "{tmp_path}") as «class PNGf»)'],
                check=True, capture_output=True)
            os.unlink(tmp_path)
            self._set_status("Copied PNG to clipboard")
        except Exception as ex:
            self._set_status(f"Copy failed: {ex}", err=True)

    def _copy_transparent(self):
        """Copy the current plot to the macOS clipboard as a transparent-background PNG."""
        spec_json = self._get_spec_json()
        if spec_json is None:
            self._set_status("No chart to copy — generate a plot first")
            return
        try:
            import subprocess, tempfile, json
            import plotly.io as pio
            fig = pio.from_json(spec_json)
            fig.update_layout(paper_bgcolor="rgba(0,0,0,0)",
                              plot_bgcolor="rgba(0,0,0,0)")
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                tmp_path = tmp.name
            pio.write_image(fig, tmp_path, scale=150/72)
            subprocess.run(
                ["osascript", "-e",
                 f'set the clipboard to (read (POSIX file "{tmp_path}") as «class PNGf»)'],
                check=True, capture_output=True)
            os.unlink(tmp_path)
            self._set_status("Copied transparent PNG to clipboard")
        except Exception as ex:
            self._set_status(f"Copy failed: {ex}", err=True)

    def _copy_as_svg(self):
        """Copy the current plot to clipboard as SVG text."""
        spec_json = self._get_spec_json()
        if spec_json is None:
            self._set_status("No chart to copy — generate a plot first")
            return
        try:
            import plotly.io as pio
            fig = pio.from_json(spec_json)
            svg_str = pio.to_image(fig, format="svg").decode("utf-8")
            self.clipboard_clear()
            self.clipboard_append(svg_str)
            self._set_status("Copied SVG to clipboard")
        except Exception as ex:
            self._set_status(f"SVG copy failed: {ex}", err=True)

    def _show_recent_files(self):
        """Show a popup menu of recently used files, anchored below the Recent button."""
        prefs  = _load_prefs()
        recent = [f for f in prefs.get("recent_files", []) if os.path.exists(f)]
        if not recent:
            self._set_status("No recent files")
            return
        menu = tk.Menu(self, tearoff=0)
        for path in recent[:10]:
            name = os.path.basename(path)
            menu.add_command(label=name,
                             command=lambda p=path: (
                                 self._vars["excel_path"].set(p),
                                 self._load_sheets(p)
                             ))
        try:
            btn = getattr(self, "_recent_btn", None)
            if btn and btn.winfo_exists():
                x = btn.winfo_rootx()
                y = btn.winfo_rooty() + btn.winfo_height()
            else:
                # Fallback: position near the cursor
                x = self.winfo_pointerx()
                y = self.winfo_pointery()
            menu.tk_popup(x, y)
        finally:
            menu.grab_release()

    def _load_last_file(self):
        """Restore the last used file path from preferences, silently if gone."""
        prefs = _load_prefs()
        last  = prefs.get("last_file", "")
        if last and os.path.exists(last):
            # Set the path var first so _validate_spreadsheet reads it correctly
            if "excel_path" in self._vars:
                self._vars["excel_path"].set(last)
            self.after(300, lambda: self._load_sheets(last))



    # ── Results Panel ─────────────────────────────────────────────────────────

    def _toggle_results_panel(self):
        """Slide the results panel open or closed."""
        if self._results_visible:
            self._results_strip.config(height=28)
            self._results_toggle_arrow.config(text="▲ Results")
            self._results_visible = False
        else:
            self._results_strip.config(height=220)
            self._results_toggle_arrow.config(text="▼ Results")
            self._results_visible = True

    def _populate_results(self, excel_path, sheet, plot_type, kw_snapshot):
        """Delegate to the standalone populate_results() in plotter_results.py."""
        if _RESULTS_AVAILABLE:
            populate_results(self, excel_path, sheet, plot_type, kw_snapshot)

    def _export_results_csv(self):
        """Delegate to the standalone export_results_csv() in plotter_results.py."""
        if _RESULTS_AVAILABLE:
            export_results_csv(self)
        else:
            self._set_status("Results module not available.", err=True)

    def _copy_results_tsv(self):
        """Delegate to the standalone copy_results_tsv() in plotter_results.py."""
        if _RESULTS_AVAILABLE:
            copy_results_tsv(self)

    def _download_png(self):
        """Open the journal export dialog, then save the figure."""
        if self._last_kw is None:
            from tkinter import messagebox
            messagebox.showwarning("Nothing to export", "Generate a plot first.")
            return
        self._show_export_dialog()

    def _show_export_dialog(self):
        """Show a journal-preset export dialog and save the figure."""
        import importlib
        from tkinter import messagebox, filedialog

        try:
            from refraction.io import export as _pe
        except ImportError:
            messagebox.showerror("Export error", "plotter_export module not found.")
            return

        # ── Build dialog ───────────────────────────────────────────────────────
        dlg = tk.Toplevel(self)
        dlg.title("Export Figure")
        dlg.resizable(False, False)
        dlg.grab_set()

        pad = dict(padx=10, pady=5)

        # Journal row
        tk.Label(dlg, text="Journal preset:", anchor="w").grid(
            row=0, column=0, sticky="w", **pad)
        journal_var = tk.StringVar(value="Custom")
        journal_cb  = ttk.Combobox(
            dlg, textvariable=journal_var, state="readonly", width=22,
            values=["Custom"] + list(_pe.JOURNAL_PRESETS.keys()))
        journal_cb.grid(row=0, column=1, sticky="w", **pad)

        # Column width row (only active when journal ≠ Custom)
        tk.Label(dlg, text="Column width:", anchor="w").grid(
            row=1, column=0, sticky="w", **pad)
        col_var = tk.StringVar()
        col_cb  = ttk.Combobox(dlg, textvariable=col_var, state="disabled", width=22)
        col_cb.grid(row=1, column=1, sticky="w", **pad)

        # Info label
        info_var = tk.StringVar(value="")
        tk.Label(dlg, textvariable=info_var, foreground="#666", font=("Arial", 9),
                 anchor="w", wraplength=300).grid(
            row=2, column=0, columnspan=2, sticky="w", padx=10, pady=2)

        def _update_cols(*_):
            j = journal_var.get()
            if j == "Custom":
                col_cb.config(state="disabled", values=[])
                col_var.set("")
                info_var.set("")
                dpi_cb.config(state="readonly")
            else:
                opts = list(_pe.JOURNAL_PRESETS[j]["columns"].keys())
                col_cb.config(state="readonly", values=opts)
                if not col_var.get() or col_var.get() not in opts:
                    col_var.set(opts[0])
                p = _pe.JOURNAL_PRESETS[j]
                dpi_var.set(str(p["dpi"]))
                dpi_cb.config(state="disabled")
                info_var.set(
                    f"{j}: {p['font']} ≥{p['min_font']}pt · "
                    f"{p['dpi']} DPI · max height {p['max_h_mm']} mm")

        journal_var.trace_add("write", _update_cols)

        # Format row
        tk.Label(dlg, text="Format:", anchor="w").grid(
            row=3, column=0, sticky="w", **pad)
        fmt_var = tk.StringVar(value="PNG (high-res)")
        ttk.Combobox(
            dlg, textvariable=fmt_var, state="readonly", width=22,
            values=["PNG (high-res)", "SVG (vector)", "PDF", "HTML (interactive)"]
        ).grid(row=3, column=1, sticky="w", **pad)

        # DPI row (disabled when a journal preset is active — preset DPI takes precedence)
        tk.Label(dlg, text="DPI (PNG only):", anchor="w").grid(
            row=4, column=0, sticky="w", **pad)
        dpi_var = tk.StringVar(value="300")
        dpi_cb = ttk.Combobox(
            dlg, textvariable=dpi_var, state="readonly", width=22,
            values=["150", "300", "600"])
        dpi_cb.grid(row=4, column=1, sticky="w", **pad)

        # ── Export action ──────────────────────────────────────────────────────
        def _do_export():
            journal = journal_var.get()
            col     = col_var.get() if journal != "Custom" else None
            fmt     = fmt_var.get()
            try:
                dpi = int(dpi_var.get())
            except ValueError:
                dpi = 300

            _fmt_map = {
                "PNG (high-res)":      (".png",  "PNG image",  "*.png"),
                "SVG (vector)":        (".svg",  "SVG vector", "*.svg"),
                "PDF":                 (".pdf",  "PDF",        "*.pdf"),
                "HTML (interactive)":  (".html", "HTML",       "*.html"),
            }
            ext, desc, glob = _fmt_map.get(fmt, (".png", "PNG image", "*.png"))

            from datetime import datetime
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            title_str = self._vars.get("title", tk.StringVar()).get().strip()
            src_path  = self._vars.get("excel_path", tk.StringVar()).get().strip()
            if title_str:
                safe = "".join(c if c.isalnum() or c in " _-" else "_"
                               for c in title_str).strip()
                default_name = safe[:60] or "refraction"
            elif src_path:
                default_name = os.path.splitext(os.path.basename(src_path))[0][:60]
            else:
                default_name = f"refraction_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

            path = filedialog.asksaveasfilename(
                parent=dlg,
                initialdir=desktop,
                initialfile=f"{default_name}{ext}",
                defaultextension=ext,
                filetypes=[(desc, glob), ("All files", "*.*")],
                title="Save figure")
            if not path:
                return

            dlg.destroy()

            # Export via Plotly + kaleido
            spec_json = self._get_spec_json()
            if spec_json is None:
                messagebox.showerror("Export failed",
                    "No rendered chart available. Generate a plot first.")
                return
            try:
                _pe.export_plotly(
                    spec_json, path,
                    journal=journal if journal != "Custom" else None,
                    col_label=col,
                    dpi=dpi)
                self._set_status(f"Saved  ·  {os.path.basename(path)}")
            except Exception as exc:
                messagebox.showerror("Export failed", str(exc))

        # ── Buttons ────────────────────────────────────────────────────────────
        btn_frame = tk.Frame(dlg)
        btn_frame.grid(row=5, column=0, columnspan=2, pady=(10, 8))
        ttk.Button(btn_frame, text="Export", command=_do_export).pack(
            side="left", padx=6)
        ttk.Button(btn_frame, text="Cancel", command=dlg.destroy).pack(
            side="left", padx=6)

        # Centre dialog over main window
        dlg.update_idletasks()
        x = self.winfo_x() + (self.winfo_width()  - dlg.winfo_width())  // 2
        y = self.winfo_y() + (self.winfo_height() - dlg.winfo_height()) // 2
        dlg.geometry(f"+{x}+{y}")


    def _export_all_pdf(self):
        """Export the current chart as a PDF via Plotly + kaleido."""
        spec_json = self._get_spec_json()
        if spec_json is None:
            messagebox.showwarning("No chart", "Generate a plot first.")
            return
        from tkinter import filedialog
        out_path = filedialog.asksaveasfilename(
            title="Save Chart as PDF",
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            initialfile="chart.pdf"
        )
        if not out_path:
            return
        try:
            from refraction.io import export as _pe
            _pe.export_plotly(spec_json, out_path, dpi=300)
            self._set_status(f"PDF saved: {out_path}")
        except Exception as exc:
            messagebox.showerror("Export failed", str(exc))

    # ── Project file support (.cplot) ─────────────────────────────────────────

    def _save_project(self):
        """Save current app state as a .cplot project file."""
        from tkinter import filedialog, messagebox
        try:
            from refraction.io.project import save_project
        except ImportError:
            messagebox.showerror("Error", "plotter_project module not available.")
            return

        path = filedialog.asksaveasfilename(
            title="Save Project",
            defaultextension=".cplot",
            filetypes=[("Refraction Project", "*.cplot"), ("All files", "*.*")],
        )
        if not path:
            return

        try:
            plot_type = self._plot_type.get() if hasattr(self, "_plot_type") else "bar"
            excel_path = self._vars.get("excel_path", tk.StringVar()).get().strip()
            sheet_var = self._vars.get("sheet")
            sheet = sheet_var.get() if sheet_var else 0

            # Capture thumbnail bytes via Plotly + kaleido
            thumbnail_bytes = None
            try:
                spec_json = self._get_spec_json()
                if spec_json is not None:
                    import tempfile, io
                    from refraction.io import export as _pe
                    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                        tmp_path = tmp.name
                    _pe.export_plotly(spec_json, tmp_path, dpi=72)
                    with open(tmp_path, "rb") as f:
                        thumbnail_bytes = f.read()
                    os.unlink(tmp_path)
            except Exception:
                _log.debug("App._save_project: thumbnail capture failed", exc_info=True)

            save_project(
                path=path,
                app_vars=self._vars,
                plot_type=plot_type,
                excel_path=excel_path,
                sheet=sheet,
                thumbnail_bytes=thumbnail_bytes,
            )
            self._set_status(f"Project saved: {path}")
        except Exception as exc:
            messagebox.showerror("Save failed", str(exc))

    def _open_cplot(self):
        """Open a .cplot project file and restore app state."""
        from tkinter import filedialog, messagebox
        try:
            from refraction.io.project import load_project, extract_to_temp_excel
        except ImportError:
            messagebox.showerror("Error", "plotter_project module not available.")
            return

        path = filedialog.askopenfilename(
            title="Open Project",
            filetypes=[("Refraction Project", "*.cplot"), ("All files", "*.*")],
        )
        if not path:
            return

        try:
            data = load_project(path)

            # Restore plot type
            pt = data.get("plot_type", "bar")
            if hasattr(self, "_plot_type"):
                self._plot_type.set(pt)
                self._on_chart_type_change(pt)

            # Restore UI state
            state = data.get("state", {})
            for key, value in state.items():
                var = self._vars.get(key)
                if var is not None:
                    try:
                        var.set(value)
                    except Exception:
                        _log.debug("App._open_cplot: could not restore var %r to %r", key, value, exc_info=True)

            # Extract embedded Excel data and load it
            try:
                temp_path = extract_to_temp_excel(path)
                if temp_path and hasattr(self, "_vars"):
                    ep_var = self._vars.get("excel_path")
                    if ep_var:
                        ep_var.set(temp_path)
                    self._load_sheets(temp_path)
            except Exception:
                _log.debug("App._open_cplot: extract_to_temp_excel failed for %r", path, exc_info=True)

            self._set_status(f"Project loaded: {path}")
        except Exception as exc:
            messagebox.showerror("Open failed", str(exc))

    def _open_pzfx(self):
        """Import a .pzfx file."""
        from tkinter import filedialog, messagebox
        try:
            from refraction.io.import_pzfx import import_pzfx
        except ImportError:
            messagebox.showerror("Error", "plotter_import_pzfx module not available.")
            return

        path = filedialog.askopenfilename(
            title="Import .pzfx File",
            filetypes=[("Prism (.pzfx)", "*.pzfx"), ("All files", "*.*")],
        )
        if not path:
            return

        try:
            result = import_pzfx(path)
            if result.errors:
                messagebox.showerror("Import Error",
                                     "\n".join(result.errors))
                return
            if result.warnings:
                messagebox.showwarning("Import Warnings",
                                       "\n".join(result.warnings))
            if result.temp_path:
                ep_var = self._vars.get("excel_path")
                if ep_var:
                    ep_var.set(result.temp_path)
                self._load_sheets(result.temp_path)
                # Set chart type if detected
                if result.chart_type and hasattr(self, "_plot_type"):
                    self._plot_type.set(result.chart_type)
                    self._on_chart_type_change(result.chart_type)
                self._set_status(f"Imported: {path}")
        except Exception as exc:
            messagebox.showerror("Import failed", str(exc))


if __name__ == "__main__":
    app = App()
    app.mainloop()
