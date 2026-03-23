"""
plotter_project.py
==================
Save/load .cplot project files for Refraction.
A .cplot file is a ZIP archive containing:
  manifest.json    - version, timestamp, app version
  state.json       - all UI variable values
  plot_type.json   - active chart type
  comparisons.json - custom comparison selections (optional)
  data/            - embedded spreadsheet data as CSV
    _sheets.json   - list of sheet names
    Sheet1.csv     - data for each sheet
  thumbnail.png    - last rendered plot (optional)

No external dependencies beyond stdlib + openpyxl (already installed).
"""
import zipfile
import json
import csv
import io
import logging
import os
import time
import tempfile

_log = logging.getLogger(__name__)


EXTENSION = ".cplot"
VERSION = 1


def save_project(path, app_vars, plot_type, excel_path,
                 sheet=None, comparisons=None,
                 thumbnail_bytes=None):
    """Save complete app state + embedded data to a .cplot file.

    Parameters:
        path: str - output file path (should end with .cplot)
        app_vars: dict - {key: StringVar/BooleanVar/IntVar}
        plot_type: str - current chart type key
        excel_path: str - path to the source Excel file
        sheet: str or int - active sheet
        comparisons: dict or None - serialized ComparisonSet
        thumbnail_bytes: bytes or None - PNG image data
    """
    with zipfile.ZipFile(path, 'w', zipfile.ZIP_DEFLATED) as zf:
        # Manifest
        manifest = {
            "version": VERSION,
            "created": time.time(),
            "app_version": "2.0.0",
        }
        zf.writestr("manifest.json", json.dumps(manifest, indent=2))

        # UI state - read .get() from each var
        state = {}
        for key, var in app_vars.items():
            try:
                state[key] = var.get()
            except Exception:
                _log.debug("save_project: could not read var %r", key, exc_info=True)
        zf.writestr("state.json", json.dumps(state, indent=2))

        # Plot type
        zf.writestr("plot_type.json", json.dumps({
            "plot_type": plot_type,
            "sheet": sheet,
        }, indent=2))

        # Comparisons
        if comparisons:
            zf.writestr("comparisons.json",
                        json.dumps(comparisons, indent=2))

        # Embedded data
        if excel_path and os.path.exists(excel_path):
            try:
                import pandas as pd
                xls = pd.ExcelFile(excel_path)
                sheet_names = xls.sheet_names
                zf.writestr("data/_sheets.json",
                           json.dumps(sheet_names))
                for sname in sheet_names:
                    df = pd.read_excel(xls, sheet_name=sname,
                                       header=None)
                    buf = io.StringIO()
                    df.to_csv(buf, index=False, header=False)
                    zf.writestr(f"data/{sname}.csv",
                               buf.getvalue())
            except Exception as e:
                zf.writestr("data/_error.txt", str(e))

        # Thumbnail
        if thumbnail_bytes:
            zf.writestr("thumbnail.png", thumbnail_bytes)


def load_project(path):
    """Load a .cplot file.

    Returns dict with keys:
        manifest: dict
        state: dict of {var_key: value}
        plot_type: str
        sheet: str or int or None
        comparisons: dict or None
        sheet_names: list of str
        has_thumbnail: bool
    """
    result = {
        "manifest": {},
        "state": {},
        "plot_type": "bar",
        "sheet": None,
        "comparisons": None,
        "sheet_names": [],
        "has_thumbnail": False,
    }

    with zipfile.ZipFile(path, 'r') as zf:
        names = zf.namelist()

        if "manifest.json" in names:
            result["manifest"] = json.loads(
                zf.read("manifest.json").decode())

        if "state.json" in names:
            result["state"] = json.loads(
                zf.read("state.json").decode())

        if "plot_type.json" in names:
            pt = json.loads(zf.read("plot_type.json").decode())
            result["plot_type"] = pt.get("plot_type", "bar")
            result["sheet"] = pt.get("sheet")

        if "comparisons.json" in names:
            result["comparisons"] = json.loads(
                zf.read("comparisons.json").decode())

        if "data/_sheets.json" in names:
            result["sheet_names"] = json.loads(
                zf.read("data/_sheets.json").decode())

        result["has_thumbnail"] = "thumbnail.png" in names

    return result


def extract_to_temp_excel(cplot_path):
    """Extract embedded CSV data to a temporary .xlsx file.

    Returns path to temp .xlsx file that can be passed to plot functions.
    The caller is responsible for cleanup (or let OS handle via tempdir).
    """
    import openpyxl

    with zipfile.ZipFile(cplot_path, 'r') as zf:
        names = zf.namelist()

        # Read sheet names
        sheet_names = ["Sheet1"]
        if "data/_sheets.json" in names:
            sheet_names = json.loads(
                zf.read("data/_sheets.json").decode())

        wb = openpyxl.Workbook()
        # Remove default sheet
        default_ws = wb.active

        for i, sname in enumerate(sheet_names):
            csv_name = f"data/{sname}.csv"
            if csv_name not in names:
                continue

            if i == 0:
                ws = default_ws
                ws.title = sname
            else:
                ws = wb.create_sheet(title=sname)

            csv_data = zf.read(csv_name).decode()
            reader = csv.reader(io.StringIO(csv_data))
            for row_idx, row in enumerate(reader, 1):
                for col_idx, val in enumerate(row, 1):
                    # Try to write as number
                    try:
                        ws.cell(row=row_idx, column=col_idx,
                                value=float(val))
                    except (ValueError, TypeError):
                        ws.cell(row=row_idx, column=col_idx,
                                value=val if val else "")

        temp_dir = tempfile.mkdtemp(prefix="refraction_")
        temp_path = os.path.join(temp_dir, "project_data.xlsx")
        wb.save(temp_path)
        return temp_path


def get_thumbnail(cplot_path):
    """Extract thumbnail PNG bytes from a .cplot file.
    Returns bytes or None.
    """
    try:
        with zipfile.ZipFile(cplot_path, 'r') as zf:
            if "thumbnail.png" in zf.namelist():
                return zf.read("thumbnail.png")
    except Exception:
        _log.debug("get_thumbnail: could not read thumbnail from %r", cplot_path, exc_info=True)
    return None
