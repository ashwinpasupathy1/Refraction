"""
plotter_import_pzfx.py
======================
Import .pzfx (XML) files into Spectra.

Extracts data tables, group names, titles, and maps TableType
to Spectra chart types. Writes data to a temp .xlsx file.

Does NOT import (with user warning):
- Color/styling (Prism uses internal palette indices)
- Statistical analysis configurations
- Curve fit model parameters

Uses only xml.etree.ElementTree (stdlib) + openpyxl (already installed).
"""
import xml.etree.ElementTree as ET
import os
import tempfile
from dataclasses import dataclass, field


@dataclass
class PzfxTable:
    """One data table extracted from a .pzfx file."""
    title: str = ""
    table_type: str = ""
    x_format: str = "none"
    y_format: str = "replicates"
    replicates: int = 1
    row_titles: list = field(default_factory=list)
    columns: list = field(default_factory=list)
    # Each column: {"title": str, "type": "x"|"y",
    #               "subcolumns": [[val, val, ...], ...]}


@dataclass
class PzfxImportResult:
    """Result of importing a .pzfx file."""
    success: bool = False
    temp_excel_path: str = None
    chart_type: str = "bar"
    title: str = ""
    group_count: int = 0
    data_points: int = 0
    warnings: list = field(default_factory=list)
    errors: list = field(default_factory=list)


# Prism TableType to Spectra chart type mapping
TABLE_TYPE_MAP = {
    "OneWay": "bar",
    "TwoWay": "grouped_bar",
    "XY": "scatter",
    "Survival": "kaplan_meier",
    "Contingency": "contingency",
    "NestedOneWay": "subcolumn_scatter",
    "Parts of 100": "stacked_bar",
}


def parse_pzfx(file_path):
    """Parse a .pzfx XML file into PzfxTable objects.

    Returns (tables, metadata) where:
        tables: list of PzfxTable
        metadata: dict with title, notes, version
    Raises ValueError if not valid Prism XML.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    tree = ET.parse(file_path)
    root = tree.getroot()

    if root.tag != "GraphPadPrismFile":
        raise ValueError(
            "Not a valid GraphPad Prism file "
            "(expected <GraphPadPrismFile> root element)")

    metadata = {"version": root.get("PrismXMLVersion", "unknown")}

    # Extract info
    for info in root.iter("Info"):
        title_el = info.find("Title")
        if title_el is not None and title_el.text:
            metadata["title"] = title_el.text.strip()
        notes_el = info.find("Notes")
        if notes_el is not None and notes_el.text:
            metadata["notes"] = notes_el.text.strip()

    # Extract tables
    tables = []
    for table_el in root.iter("Table"):
        t = PzfxTable()
        title_el = table_el.find("Title")
        if title_el is not None and title_el.text:
            t.title = title_el.text.strip()

        t.table_type = table_el.get("TableType", "OneWay")
        t.x_format = table_el.get("XFormat", "none")
        t.y_format = table_el.get("YFormat", "replicates")
        try:
            t.replicates = int(table_el.get("Replicates", "1"))
        except ValueError:
            t.replicates = 1

        # Row titles
        row_titles_col = table_el.find("RowTitlesColumn")
        if row_titles_col is not None:
            for subcol in row_titles_col.findall("Subcolumn"):
                for d in subcol.findall("d"):
                    if d.text and d.text.strip():
                        t.row_titles.append(d.text.strip())

        # X columns
        for xcol in table_el.findall("XColumn"):
            col = {"title": "", "type": "x", "subcolumns": []}
            xt = xcol.find("Title")
            if xt is not None and xt.text:
                col["title"] = xt.text.strip()
            for subcol in xcol.findall("Subcolumn"):
                values = []
                for d in subcol.findall("d"):
                    values.append(d.text.strip() if d.text else "")
                col["subcolumns"].append(values)
            t.columns.append(col)

        # Y columns
        for ycol in table_el.findall("YColumn"):
            col = {"title": "", "type": "y", "subcolumns": []}
            yt = ycol.find("Title")
            if yt is not None and yt.text:
                col["title"] = yt.text.strip()
            for subcol in ycol.findall("Subcolumn"):
                values = []
                for d in subcol.findall("d"):
                    values.append(d.text.strip() if d.text else "")
                col["subcolumns"].append(values)
            t.columns.append(col)

        tables.append(t)

    return tables, metadata


def _table_to_rows(table):
    """Convert a PzfxTable to row-major data for Excel.

    Returns (rows, chart_type) where rows[0] is the header.
    Layout matches Spectra's expected Excel format.
    """
    chart_type = TABLE_TYPE_MAP.get(table.table_type, "bar")

    if table.table_type in ("OneWay", "NestedOneWay"):
        # Flat: row 1 = group names, rows 2+ = data
        headers = list(table.row_titles) if table.row_titles else []
        y_cols = [c for c in table.columns if c["type"] == "y"]

        if not headers and y_cols:
            for yc in y_cols:
                if yc["title"]:
                    headers.append(yc["title"])

        group_data = {}
        for yc in y_cols:
            for subcol in yc["subcolumns"]:
                for i, val in enumerate(subcol):
                    if i < len(headers):
                        group_data.setdefault(headers[i], []).append(val)

        if not group_data:
            return [headers], chart_type

        max_rows = max(len(v) for v in group_data.values())
        rows = [headers]
        for ri in range(max_rows):
            row = []
            for h in headers:
                vals = group_data.get(h, [])
                row.append(vals[ri] if ri < len(vals) else "")
            rows.append(row)
        return rows, chart_type

    elif table.table_type == "XY":
        x_cols = [c for c in table.columns if c["type"] == "x"]
        y_cols = [c for c in table.columns if c["type"] == "y"]

        x_title = x_cols[0]["title"] if x_cols else "X"
        x_vals = (x_cols[0]["subcolumns"][0]
                  if x_cols and x_cols[0]["subcolumns"] else [])

        headers = [x_title]
        data_cols = []
        for yc in y_cols:
            name = yc["title"] or "Y"
            for subcol in yc["subcolumns"]:
                headers.append(name)
                data_cols.append(subcol)

        max_rows = max(len(x_vals),
                       max((len(dc) for dc in data_cols), default=0))
        rows = [headers]
        for ri in range(max_rows):
            row = [x_vals[ri] if ri < len(x_vals) else ""]
            for dc in data_cols:
                row.append(dc[ri] if ri < len(dc) else "")
            rows.append(row)
        return rows, chart_type

    elif table.table_type == "Survival":
        y_cols = [c for c in table.columns if c["type"] == "y"]
        headers = []
        subcol_pairs = []
        group_names = table.row_titles or []

        gi = 0
        for yc in y_cols:
            name = (group_names[gi] if gi < len(group_names)
                    else yc["title"] or f"Group {gi + 1}")
            if len(yc["subcolumns"]) >= 2:
                headers.extend([name, name])
                subcol_pairs.append(
                    (yc["subcolumns"][0], yc["subcolumns"][1]))
            gi += 1

        row2 = []
        for _ in subcol_pairs:
            row2.extend(["Time", "Event"])

        max_rows = max(
            (max(len(t), len(e)) for t, e in subcol_pairs),
            default=0)
        rows = [headers, row2]
        for ri in range(max_rows):
            row = []
            for time_col, event_col in subcol_pairs:
                row.append(
                    time_col[ri] if ri < len(time_col) else "")
                row.append(
                    event_col[ri] if ri < len(event_col) else "")
            rows.append(row)
        return rows, "kaplan_meier"

    else:
        # Fallback: flat dump
        headers = table.row_titles or []
        return [headers], chart_type


def import_pzfx(file_path):
    """Import a .pzfx file and write data to a temporary Excel file.

    Returns PzfxImportResult with temp .xlsx path and chart type.
    """
    result = PzfxImportResult()

    try:
        tables, metadata = parse_pzfx(file_path)
    except Exception as e:
        result.errors.append(f"Failed to parse: {e}")
        return result

    if not tables:
        result.errors.append("No data tables found.")
        return result

    table = tables[0]
    if len(tables) > 1:
        result.warnings.append(
            f"File has {len(tables)} tables. "
            f"Only '{table.title}' was imported.")

    try:
        rows, chart_type = _table_to_rows(table)
    except Exception as e:
        result.errors.append(f"Failed to convert: {e}")
        return result

    result.chart_type = chart_type
    result.title = metadata.get("title", table.title)

    if rows:
        result.group_count = len(rows[0])
        result.data_points = sum(
            1 for row in rows[1:]
            for cell in row
            if cell and str(cell).strip())

    # Write to temp Excel
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = table.title or "Imported"
        for ri, row in enumerate(rows, 1):
            for ci, val in enumerate(row, 1):
                try:
                    ws.cell(row=ri, column=ci, value=float(val))
                except (ValueError, TypeError):
                    ws.cell(row=ri, column=ci,
                            value=str(val) if val else "")

        temp_dir = tempfile.mkdtemp(prefix="spectra_import_")
        temp_path = os.path.join(temp_dir, "imported_data.xlsx")
        wb.save(temp_path)
        result.temp_excel_path = temp_path
        result.success = True
    except Exception as e:
        result.errors.append(f"Failed to write Excel: {e}")

    result.warnings.append(
        "Styling was not imported. Configure in the Axes tab.")
    result.warnings.append(
        "Statistical tests were not imported. Configure in Stats tab.")

    return result
